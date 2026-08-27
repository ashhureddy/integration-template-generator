import streamlit as st
import sys
import pandas as pd
import re
import io
import time
import zipfile
from datetime import date
from pathlib import Path
# Note: reportlab is imported lazily inside build_parameter_verification_pdf() below, not here —
# a new feature's dependency must never be able to crash the whole app on startup if it isn't
# installed yet. Every other page (including Generate Report) must keep working regardless.

# ============================================================
# CONFIG
# ============================================================
st.set_page_config(page_title="Integration Template Generator", page_icon="📡", layout="wide")

TDIR = Path(__file__).parent / "templates" / "MCA"

def resolve_template(exact_name, keyword):
    """Prefer the exact expected filename; if it's missing (e.g. uploaded with a slightly
    different name), fall back to any file in templates/MCA containing `keyword`."""
    exact_path = TDIR / exact_name
    if exact_path.exists():
        return exact_path
    if TDIR.exists():
        candidates = [p for p in TDIR.glob("*.txt") if keyword.lower() in p.name.lower()]
        if candidates:
            return candidates[0]
    return exact_path  # falls through to a clear FileNotFoundError naming what was expected

TPL_MMBB = resolve_template("LTE+5G_MMBB_Integration_Pre-existing_Procedure_with_LTE_or_5G_Node_as_Primary_CMCLI_Updated_V11.txt", "MMBB_Integration")
TPL_TMBB = resolve_template("TRIMODE_Integration_Pre-existing_Procedure_with_LTE_or_5G_Node_as_Primary_CMCLI_Updated_V10.txt", "TRIMODE_Integration")
TPL_CENM = resolve_template("cENM_TRIMODE_Integration_Pre-existing_Procedure_with_LTE_or_5G_Node_as_Primary_CMCLI_Updated_V4.txt", "cENM_TRIMODE")
TPL_CENM_MMBB = resolve_template("cENM_MMBB_Integration_Pre-existing_Procedure_with_LTE_or_5G_Node_as_Primary_CMCLI_Updated_V4.txt", "cENM_MMBB")
# SMBB (LTE-only, LTE primary) — same source file shared by MCA and CENM per the blueprint.
# NSB has its own file (no Pre-checks section, different EDP-field legend).
TPL_SMBB_LTE = resolve_template("LTE_Integration_Pre-existing_Procedure_with_LTE_as_Primary_CMCLI_Updated_1.txt", "LTE_as_Primary_CMCLI")
# Deleted-node / board-swap node install+delete commands — universal across MCA/CENM/CRAN.
TPL_NODE_DELETION = resolve_template("Site_Install_Generation_and_Node_Deletion_commands.txt", "Node_Deletion_commands")
TPL_6610 = resolve_template("6610 Controller Integration Procedure_25Q3_Updated_V12.txt", "6610")
TPL_PORT_CONVERSION = resolve_template("Template_Port_Conversion_1G_to_10G_BBU_V1_1.txt", "Port_Conversion")
TPL_CRAN_TRIP1 = resolve_template("CRAN_TO_CRAN_Rehome_Pre-integration_Trip-1_Procedure_for_SA_Sites_V2.txt", "Trip-1")
TPL_CRAN_TRIP2 = resolve_template("CRAN_TO_CRAN_Rehome_and_6673_Sidehaul_Change_With_MPST_Trip-2_Procedure_for_SA_Sites_V1.txt", "Trip-2")
TPL_CRAN_NSA = resolve_template("CRAN_TO_CRAN_Rehome_Integration_and_Cutover_Procedure_for_NSA_Sites_V2.txt", "NSA_Sites")

def resolve_dss_template(exact_stem):
    """stand/standard were uploaded with no .txt extension — try both forms.
    Exact stem match only (no fuzzy 'contains' search) since 'stand' is a substring of
    'standard' and a fuzzy match could silently load the wrong DSS template."""
    for candidate in (TDIR / f"{exact_stem}.txt", TDIR / exact_stem):
        if candidate.exists():
            return candidate
    return TDIR / f"{exact_stem}.txt"

TPL_DSS_4SECTOR = resolve_dss_template("standard")
TPL_DSS_3SECTOR = resolve_dss_template("stand")

TDIR_N2E = Path(__file__).parent / "templates" / "N2E"
TPL_N2E_LTE = TDIR_N2E / "N2E_LTE_Integration_Procedure_with_LTE_Node_as_Primary_V4.txt"
TPL_N2E_5G = TDIR_N2E / "N2E_5G_Integration_Procedure_with_5G_Node_as_Primary_V4.txt"
TPL_N2E_MMBB = TDIR_N2E / "MMBB_N2E_Integration_Procedure_with_LTE_or_5G_Node_as_Primary_CMCLI_Updated_V6.txt"
TPL_N2E_TRIMODE = TDIR_N2E / "N2E_TRIMODE_Integration_Procedure_with_LTE_or_5G_Node_as_Primary_CMCLI_Updated_V6.txt"

TDIR_NSB = Path(__file__).parent / "templates" / "NSB"
TDIR_STATIC = Path(__file__).parent / "templates" / "Static"
TDIR_MCA_IDL_CRAN = Path(__file__).parent / "templates" / "MCA" / "IDL_CRAN"
TPL_NSB_MMBB = TDIR_NSB / "LTE+5G_MMBB_Integration_NSB_Procedure_with_LTE_or_5G_Node_as_Primary_CMCLI_Updated_V13.txt"
TPL_NSB_TRIMODE = TDIR_NSB / "TRIMODE_Integration_NSB_Procedure_with_LTE_or_5G_Node_as_Primary_CMCLI_Updated_V6.txt"
TPL_NSB_SMBB_LTE = TDIR_NSB / "LTE_Integration_NSB_Procedure_with_LTE_as_Primary_CMCLI_Updated_1.txt"

# ============================================================
# SHARED HELPERS
# ============================================================

def load_workbook_any(file_bytes, filename):
    """openpyxl can only read real .xlsx (zip-based OOXML) — legacy .xls (OLE2/CFB binary) needs xlrd.
    Some files have a mismatched extension (e.g. an old .xls saved/renamed with a .xlsx name), so this
    doesn't trust the filename alone: it tries the format the extension suggests first, then the other
    one on failure. It also repairs a common 'could not read stylesheet' crash — some non-Microsoft
    export tools produce a malformed xl/styles.xml even though the actual cell data is fine — by
    swapping in a minimal valid stylesheet and retrying (verified against a deliberately-corrupted
    styles.xml: openpyxl's read_only mode does NOT sidestep this, but replacing the styles part does)."""
    import openpyxl, zipfile

    def via_xlrd():
        import pandas as pd
        all_sheets = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, engine="xlrd", header=None)
        out_wb = openpyxl.Workbook()
        out_wb.remove(out_wb.active)
        for sheet_name, df in all_sheets.items():
            ws = out_wb.create_sheet(title=str(sheet_name)[:31])  # Excel sheet name limit
            for row in df.itertuples(index=False, name=None):
                ws.append(list(row))
        return out_wb

    def via_openpyxl():
        return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    def via_repaired_styles():
        MINIMAL_STYLES = b'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>
<fills count="1"><fill><patternFill patternType="none"/></fill></fills>
<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>
<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>
</styleSheet>'''
        zin = zipfile.ZipFile(io.BytesIO(file_bytes), "r")
        repaired_buf = io.BytesIO()
        zout = zipfile.ZipFile(repaired_buf, "w")
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/styles.xml":
                data = MINIMAL_STYLES
            zout.writestr(item, data)
        zout.close()
        return openpyxl.load_workbook(io.BytesIO(repaired_buf.getvalue()), data_only=True)

    looks_like_xls = filename.lower().endswith(".xls") and not filename.lower().endswith(".xlsx")
    attempts = [via_xlrd, via_openpyxl, via_repaired_styles] if looks_like_xls else [via_openpyxl, via_repaired_styles, via_xlrd]

    first_error = None
    for attempt in attempts:
        try:
            return attempt()
        except Exception as e:
            if first_error is None:
                first_error = e
    raise first_error  # surface the first (most likely relevant) error

def sheet_objs(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    objs = []
    for r in rows[1:]:
        if not any(str(c).strip() for c in r if c is not None):
            continue
        objs.append({headers[i]: (r[i].strip() if isinstance(r[i], str) else r[i]) if i < len(r) else "" for i in range(len(headers))})
    return objs


def is_populated(v):
    if v is None:
        return False
    s = str(v).strip().upper()
    return s not in ("", "N/A")


def locate_edp_header_row(rows):
    for i, row in enumerate(rows):
        if any(str(c).strip().upper() == "EDP_SITE_ID" for c in row if c is not None):
            return i
    return -1


def build_edp_index(edp_wb):
    for sn in edp_wb.sheetnames:
        ws = edp_wb[sn]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        hidx = locate_edp_header_row(rows)
        if hidx >= 0:
            headers = rows[hidx]
            header_map = {str(h).strip().upper(): i for i, h in enumerate(headers) if h is not None}
            data_rows = [r for r in rows[hidx + 1:] if any(str(c).strip() for c in r if c is not None)]
            return {"header_map": header_map, "data_rows": data_rows}
    return None


def edp_row_for(edp_index, site_name):
    if not edp_index or not site_name:
        return None
    idx = edp_index["header_map"].get("SITE_NAME")
    if idx is None:
        return None
    for r in edp_index["data_rows"]:
        val = r[idx] if idx < len(r) else None
        if str(val or "").strip().upper() == str(site_name).strip().upper():
            return r
    return None


def edp_get(edp_index, row, header_name):
    if not edp_index or row is None:
        return None
    idx = edp_index["header_map"].get(header_name.upper())
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    if v is None or (isinstance(v, float) and v != v):  # v != v is the standard NaN check
        return None
    s = str(v).strip()
    return None if s == "" or s.lower() == "nan" else s


def find_row_by_name(ciq_wb, sheet_name, name_header, name_value):
    if sheet_name not in ciq_wb.sheetnames or not name_value:
        return None
    for r in sheet_objs(ciq_wb[sheet_name]):
        if str(r.get(name_header, "")).strip().upper() == str(name_value).strip().upper():
            return r
    return None


def hw_string(row):
    if not row:
        return None
    du = row.get("DU type") or row.get("1st DU type")
    if not du:
        return None
    xmu_count = sum(1 for k in ("1st XMU", "2nd XMU") if str(row.get(k, "")).strip().upper() == "YES")
    suffix = "" if xmu_count == 0 else " + XMU" if xmu_count == 1 else f" + {xmu_count} XMU"
    return f"{du}{suffix}"


def extract_pre_hw(text, node_name):
    if not text or not node_name:
        return None
    esc = re.escape(node_name)
    # some Pre-checks PDF versions insert an extra ISO-timestamp token between ENABLED and the
    # actual product name (seen on newer AAS/5216-style hardware rows) — (?:\S+\s+)? skips it if present
    # some Pre-checks PDF versions insert an extra isSharedWithExternalMe column (true/false)
    # between faultIndicator and operationalIndicator — (?:(?:true|false)\s+)? skips it if present
    m = re.search(esc + r"\s+1\s+UNLOCKED\s+OFF\s+(?:(?:true|false)\s+)?STEADY_ON\s+ENABLED\s+(?:\S+\s+)?([A-Za-z0-9 ]+?)\s+\d{6,8}", text, re.I)
    if not m:
        return None
    # take the LAST token as the model number — handles "Baseband 6630", "RAN Processor 6651",
    # "Baseband R503", or any future hardware family name, matching the CIQ side's bare model number
    tokens = m.group(1).strip().split()
    return tokens[-1] if tokens else None


def extract_pre_xmu_count(text, node_name):
    if not text or not node_name:
        return 0
    esc = re.escape(node_name)
    return len(re.findall(esc + r"\s+XMU\S*\s+UNLOCKED\s+OFF\s+(?:(?:true|false)\s+)?STEADY_ON\s+ENABLED", text, re.I))


def pre_hw_string(text, node_name):
    base = extract_pre_hw(text, node_name)
    if not base:
        return None
    xmu_count = extract_pre_xmu_count(text, node_name)
    suffix = "" if xmu_count == 0 else " + XMU" if xmu_count == 1 else f" + {xmu_count} XMU"
    return f"{base}{suffix}"


def extract_pdf_text(pdf_bytes):
    """Some 'PDF' uploads are actually zip archives (confirmed this session, e.g. real
    ECL02586 Pre/Post-checks files) — a zip containing per-page N.jpeg + N.txt + a
    manifest.json, not a real PDF at all. pdfplumber correctly rejects these
    ("No /Root object! - Is this really a PDF?"). Try the zip-bundle path first (cheap,
    exact check via the zip magic number), fall back to pdfplumber for genuine PDFs —
    same two-format resilience pattern already used for load_workbook_any's xlsx/xls
    fallback."""
    import zipfile
    import json as _json

    if pdf_bytes[:2] == b"PK":  # zip magic number
        try:
            with zipfile.ZipFile(io.BytesIO(pdf_bytes)) as zf:
                manifest = _json.loads(zf.read("manifest.json"))
                pages = sorted(manifest["pages"], key=lambda p: p["page_number"])
                text = ""
                for p in pages:
                    text += zf.read(p["text"]["path"]).decode("utf-8", errors="replace") + "\n\n"
                return text
        except Exception:
            pass  # fall through to pdfplumber, which will raise its own clear error

    import pdfplumber
    text = ""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text += (page.extract_text() or "") + "\n\n"
    return text


def push_siad_row(rows, edp_index, node_built_as):
    row = edp_row_for(edp_index, node_built_as)
    rows.append({
        "Node": node_built_as,
        "SIAD CLLI": edp_get(edp_index, row, "SIAD_CLLI") or "NOT FOUND",
        "Port Size": edp_get(edp_index, row, "SIAD_PORT_SIZE_BBU") or "NOT FOUND",
        "Port Facing BBU": edp_get(edp_index, row, "SIAD_PORT_FACING_BBU") or "NOT FOUND",
    })


def push_controller_siad_row(rows, edp_index, controller_id):
    """6610 controller rows in EDP use a different column set (ANCEQ_*) than regular BBU nodes —
    same SITE_NAME match, but the port lives in ANCEQ_SIAD_PORT, not SIAD_PORT_FACING_BBU.
    Returns True if the controller was actually found published in EDP, False otherwise."""
    row = edp_row_for(edp_index, controller_id)
    anceq_type = edp_get(edp_index, row, "ANCEQ_TYPE")
    found = row is not None and anceq_type and "6610" in str(anceq_type)
    rows.append({
        "Node": controller_id,
        "SIAD CLLI": (edp_get(edp_index, row, "SIAD_CLLI") or "NOT FOUND") if found else "NOT FOUND",
        "Port Size": "1G" if found else "NOT FOUND",
        "Port Facing BBU": (edp_get(edp_index, row, "ANCEQ_SIAD_PORT") or "NOT FOUND") if found else "EDP not published for controller",
    })
    return found


def highlight_unresolved(text):
    cands = re.findall(r"xx[A-Za-z0-9_]+xx|(?<!#)##[A-Za-z0-9_]+##(?!#)", text)
    return sorted(set(c for c in cands if not re.fullmatch(r"x+", c, re.I)))


def has_6610(controller_objs):
    return any(str(r.get("Controller", "")).strip() == "6610" for r in controller_objs)


# ============================================================
# BAND / SECTOR LABEL SYSTEM (Scope of Work display)
# Confirmed against real sites — NOT the same convention as DSS's Greek naming.
# ============================================================

SECTOR_NAME = {'A': 'Alpha', 'B': 'Beta', 'C': 'Gamma', 'D': 'Delta', 'E': 'Epsilon', 'F': 'Foxtrot'}
SECTOR_ORDER = ['Alpha', 'Beta', 'Gamma', 'Delta', 'Epsilon', 'Foxtrot']

def lte_band_label(cell_name):
    """e.g. ECL00043_2A_1 -> ('AWS_1', 'Alpha') ; DXL04049_7A_2_F -> ('FNET', 'Alpha')"""
    if not cell_name:
        return None, None
    m = re.search(r'_(\d)([A-F])_(\d+)(_[EF])?$', str(cell_name))
    if not m:
        return None, None
    digit, letter, carrier, suffix = m.group(1), m.group(2), m.group(3), m.group(4)
    sector = SECTOR_NAME.get(letter, letter)
    if digit == '9':
        return f"PCS_{carrier}", sector
    if digit == '2':
        return f"AWS_{carrier}", sector
    if digit == '8':
        return f"850_{carrier}", sector
    if digit == '3':
        return "WCS", sector
    if digit == '7':
        if suffix == '_F':
            return "FNET", sector
        if suffix == '_E':
            return "LTE_700_E", sector
        return "LTE_700", sector
    return f"BAND{digit}_{carrier}", sector

def nr_band_label(cell_name):
    """e.g. NCRN002376_N066A_1 -> ('5G_AWS_1', 'Alpha') ; ..._N077A_2 -> ('DOD', 'Alpha')"""
    if not cell_name:
        return None, None
    m = re.search(r'_N(\d{3})([A-F])_(\d+)$', str(cell_name))
    if not m:
        return None, None
    band, letter, carrier = m.group(1), m.group(2), m.group(3)
    sector = SECTOR_NAME.get(letter, letter)
    if band == '005':
        return "5G_850", sector
    if band == '002':
        return f"5G_PCS_{carrier}", sector
    if band == '066':
        return f"5G_AWS_{carrier}", sector
    if band == '077':
        return {'1': 'CBAND', '2': 'DOD', '3': 'DOD_BWE'}.get(carrier, f"N077_{carrier}"), sector
    return f"N{band}_{carrier}", sector

def band_label(cell_name):
    """Dispatch to LTE or 5G labeler based on whether the cell name contains an 'N0xx' 5G marker."""
    if re.search(r'_N\d{3}[A-F]_\d+$', str(cell_name or '')):
        return nr_band_label(cell_name)
    return lte_band_label(cell_name)

def is_5g_cell(cell_name):
    return bool(re.search(r'_N\d{3}[A-F]_\d+$', str(cell_name or '')))

def is_wll_node_name(name):
    """Confirmed rule: any node name ending in 'L' (Pre-checks, Post-checks, or CIQ alike)
    is a WLL node — a co-located logical entity, not a real radio node. Reused everywhere
    a node needs to be recognized as WLL: excluded from Pre/Post Configuration's node
    lists, folded into "WLL node :" in the report header, and treated as deleted from ENM
    rather than a genuine site node."""
    return bool(name) and str(name).strip().upper().endswith("L")


def dedupe_labels(cell_names, lte_first=True):
    """Classify a list of cell names into unique band labels, LTE group first then 5G group,
    preserving first-seen order within each group."""
    lte_labels, fiveg_labels = [], []
    for c in cell_names:
        label, _ = band_label(c)
        if not label:
            continue
        target = fiveg_labels if is_5g_cell(c) else lte_labels
        if label not in target:
            target.append(label)
    return (lte_labels + fiveg_labels) if lte_first else (fiveg_labels + lte_labels)


def extract_precheck_sectors(text):
    """Parse the Pre-checks PDF's 'Summary Status' table: Node | Technology | Cell | ...
    Returns (set of (node, cell) tuples, set of node names)."""
    if not text:
        return set(), set()
    pairs = set()
    nodes = set()
    for m in re.finditer(r'(\S+)\s+(LTE|5G)\s+(\S+)\s+(UNLOCKED|LOCKED)', text):
        node, tech, cell = m.group(1), m.group(2), m.group(3)
        pairs.add((node, cell))
        nodes.add(node)
    return pairs, nodes


# ============================================================
# CARRIER ADD / DELETE / MOVE / RETUNE CLASSIFICATION
# ============================================================

def build_node_alias_map(mm_objs):
    """A node's secondary identity (eNodeB or gNodeB name) can appear in Sector Del_Movement's
    Source/Target columns instead of its Primary ID — happens specifically when the moving cell's
    own technology matches the secondary identity (e.g. a 5G cell moving into a dual-identity node
    records the target using that node's gNodeB name, not its Primary 'Node to be built as')."""
    alias = {}
    for row in mm_objs:
        primary = row.get("Node to be built as")
        if not primary:
            continue
        for secondary in (row.get("eNodeB Name"), row.get("gNodeB Name")):
            if secondary and str(secondary).strip() and str(secondary).strip() != str(primary).strip():
                alias[str(secondary).strip()] = str(primary).strip()
    return alias


def classify_carriers(ciq_wb, mm_objs, precheck_text):
    """Returns a dict: added (per node), moved, deleted_sectors, deleted_nodes, retuned."""
    result = {"added": {}, "moved": [], "deleted_sectors": {}, "deleted_nodes": [], "retuned": [], "node_band_sectors": {}}
    alias_map = build_node_alias_map(mm_objs)

    def normalize(name):
        return alias_map.get(str(name).strip(), name) if name else name

    pre_pairs, pre_nodes = extract_precheck_sectors(precheck_text)
    pre_cells = {cell for (_, cell) in pre_pairs}

    # per (node, band label) sector inventory — used to tell "whole band moved" from "partial move"
    node_band_sectors = {}
    for (node, cell) in pre_pairs:
        label, sector = band_label(cell)
        if label and sector:
            node_band_sectors.setdefault((node, label), set()).add(sector)

    ciq_nodes = {str(r.get("Node to be built as", "")).strip() for r in mm_objs if r.get("Node to be built as")}
    if pre_nodes:
        result["deleted_nodes"] = sorted(pre_nodes - ciq_nodes)

    # Confirmed WLL rule: any node ending in "L" is a WLL node, not a real radio node — it
    # must NEVER appear in "Deleted Node from ENM" (only in the "WLL node :" field,
    # elsewhere). Without this, a WLL node found only in Pre-checks would naturally fall
    # into the pre_nodes-ciq_nodes diff above and get flagged as deleted, so it's actively
    # filtered back out here rather than just not being added.
    result["deleted_nodes"] = [n for n in result["deleted_nodes"] if not is_wll_node_name(n)]

    delmove_objs = sheet_objs(ciq_wb["Sector Del_Movement"]) if "Sector Del_Movement" in ciq_wb.sheetnames else []
    handled_cells = set()

    for r in delmove_objs:
        src_node, src_sector = normalize(r.get("Source Node name")), r.get("Source Sector")
        tgt_node_raw, tgt_sector = r.get("Target Node name"), r.get("Target Sector")
        tgt_node = tgt_node_raw if str(tgt_node_raw).strip().upper() == "DELETE" else normalize(tgt_node_raw)
        handled_cells.add(src_sector)
        if str(tgt_node).strip().upper() == "DELETE":
            result["deleted_sectors"].setdefault(src_node, []).append(src_sector)
            continue
        handled_cells.add(tgt_sector)
        src_dl, tgt_dl = str(r.get("Source channelNumberDL", "")).strip(), str(r.get("Target channelNumberDL", "")).strip()
        src_bw, tgt_bw = str(r.get("Source Bandwidth", "")).strip(), str(r.get("Target Bandwidth", "")).strip()
        retuned = (src_dl != tgt_dl) or (src_bw != tgt_bw)
        if str(src_node).strip().upper() == str(tgt_node).strip().upper():
            if retuned:
                label, _ = lte_band_label(src_sector)
                if not label:
                    label, _ = nr_band_label(src_sector)
                result["retuned"].append({"label": label, "from": f"{src_dl}/{src_bw}", "to": f"{tgt_dl}/{tgt_bw}"})
        else:
            result["moved"].append({"cell": src_sector, "from_node": src_node, "to_node": tgt_node})
            if retuned:
                label, _ = lte_band_label(tgt_sector)
                if not label:
                    label, _ = nr_band_label(tgt_sector)
                result["retuned"].append({"label": label, "from": f"{src_dl}/{src_bw}", "to": f"{tgt_dl}/{tgt_bw}"})

    # ADD: any CIQ cell (LTE or 5G) not present in Pre-checks and not already accounted for as moved/deleted
    eutran_objs = sheet_objs(ciq_wb["eUtran Parameters"]) if "eUtran Parameters" in ciq_wb.sheetnames else []
    fiveg_objs = sheet_objs(ciq_wb["5G Info"]) if "5G Info" in ciq_wb.sheetnames else []
    # Confirmed real rule: Integration should reflect actual sectors, not just the band —
    # if a band's full CIQ target sector set (Alpha/Beta/Gamma/... whichever this node
    # actually has for that band) is ALL newly added, show just the band label; if only
    # some of that band's target sectors are new (the rest already pre-existing/untouched),
    # show the band label plus the specific newly-added sector name(s). target_band_sectors
    # is the FULL target inventory per (node, band label) — every sector this node has for
    # that band in the CIQ, added or not — the same "whole vs partial" pattern
    # node_band_sectors already uses for Moved Sectors, just sourced from the CIQ target
    # instead of Pre-checks.
    target_band_sectors = {}
    for r in mm_objs:
        node = r.get("Node to be built as")
        if is_wll_node_name(node):
            continue  # WLL node — not a real radio node, never gets Integration/added-cell entries
        e_name, g_name = r.get("eNodeB Name"), r.get("gNodeB Name")
        added_here = []
        for row in eutran_objs:
            cell = row.get("EutranCellFDDId")
            if not cell:
                continue
            if e_name and str(cell).startswith(str(e_name)):
                label, sector = band_label(cell)
                if label and sector:
                    target_band_sectors.setdefault((node, label), set()).add(sector)
            if cell in handled_cells or cell in pre_cells:
                continue
            if e_name and str(cell).startswith(str(e_name)):
                added_here.append(cell)
        for row in fiveg_objs:
            cell = row.get("NRCellDU")
            if not cell:
                continue
            if g_name and str(cell).startswith(str(g_name)):
                label, sector = band_label(cell)
                if label and sector:
                    target_band_sectors.setdefault((node, label), set()).add(sector)
            if cell in handled_cells or cell in pre_cells:
                continue
            if g_name and str(cell).startswith(str(g_name)):
                added_here.append(cell)
        if added_here:
            result["added"][node] = added_here

    result["node_band_sectors"] = node_band_sectors
    result["target_band_sectors"] = target_band_sectors
    return result


def format_scope_of_work(classification, controller_objs, dss_outputs_meta=None, controller_edp_found=None, radio_swaps=None):
    """Turn the classification dict into the confirmed display lines.
    controller_edp_found: dict of {controller_id: bool} — False means the 6610 shows in the CIQ
    but isn't published in EDP yet."""
    lines = []
    target_band_sectors = classification.get("target_band_sectors", {})
    for node, cells in classification.get("added", {}).items():
        labels = dedupe_labels(cells)
        added_by_label = {}
        for c in cells:
            label, sector = band_label(c)
            if label and sector:
                added_by_label.setdefault(label, set()).add(sector)
        parts = []
        for label in labels:
            added_sectors = added_by_label.get(label, set())
            target_sectors = target_band_sectors.get((node, label))
            if not target_sectors or added_sectors >= target_sectors:
                parts.append(label)  # whole band added — target unknown or fully covered
            else:
                sector_names = sorted(added_sectors, key=lambda s: SECTOR_ORDER.index(s) if s in SECTOR_ORDER else 99)
                parts.append(f"{label} {', '.join(sector_names)}")
        lines.append(f"Integration:\t{'/'.join(parts)}\t{node}")

    ctrl_rows = [r for r in controller_objs if str(r.get("Controller", "")).strip() == "6610"]
    for r in ctrl_rows:
        ctrl_id = r.get('Controller ID')
        # Confirmed real bug: this loop iterates EVERY row matching Controller=="6610" — if
        # controller_objs has more than one such row (e.g. a duplicate/blank leftover row
        # alongside the real one), a "6610 Controller Integration:" line with no ID at all
        # got generated for each blank row too, producing "6610 Controller Integration: ."
        # in the report. _get_controller_id() (used for the report header) sidesteps this
        # by only ever reading the first matching row — skip blank rows here too instead.
        if not ctrl_id or not str(ctrl_id).strip():
            continue
        if controller_edp_found is not None and controller_edp_found.get(ctrl_id) is False:
            lines.append(f"EDP is not published for the controller — {ctrl_id}")
        else:
            lines.append(f"6610 Controller Integration:\t{ctrl_id}")

    moved_by_pair = {}
    for m in classification.get("moved", []):
        key = (m["from_node"], m["to_node"])
        moved_by_pair.setdefault(key, []).append(m["cell"])
    WHOLE_BAND_SET = {"Alpha", "Beta", "Gamma"}
    for (from_node, to_node), cells in moved_by_pair.items():
        if not is_populated(to_node) or not any(is_populated(c) for c in cells):
            # Malformed Sector Del_Movement row (missing Target Node name / Source Sector) —
            # flag it plainly instead of emitting a garbled "Moved Sectors: [] ... To: None" line.
            lines.append(f"Moved Sectors:\tCHECK CIQ — incomplete Sector Del_Movement row\tFrom:\t{from_node or 'NOT FOUND'}\tTo:\t{to_node or 'NOT FOUND'}")
            continue
        labels = dedupe_labels(cells)
        label_str = labels[0] if len(labels) == 1 else f"[{'/'.join(labels)}]"
        per_label_moved = {}
        for c in cells:
            label, sector = band_label(c)
            if label and sector:
                per_label_moved.setdefault(label, set()).add(sector)
        # "whole" = every band in this move brought all of Alpha+Beta+Gamma together — not
        # about what the source node happened to have historically, just this move itself
        is_whole = bool(per_label_moved) and all(WHOLE_BAND_SET <= sset for sset in per_label_moved.values())
        sector_names = sorted({s for sset in per_label_moved.values() for s in sset}, key=lambda s: SECTOR_ORDER.index(s) if s in SECTOR_ORDER else 99)
        sectors_str = "" if is_whole else (f" {', '.join(sector_names)}" if sector_names else "")
        lines.append(f"Moved Sectors:\t{label_str}{sectors_str}\tFrom:\t{from_node}\tTo:\t{to_node}")

    deleted_nodes = classification.get("deleted_nodes", [])
    if deleted_nodes:
        lines.append(f"Deleted Node from ENM:\t{'|'.join(deleted_nodes)}")

    for node, cells in classification.get("deleted_sectors", {}).items():
        labels = dedupe_labels(cells)
        lines.append(f"Deleted Sector:\t{'/'.join(labels)}\t{node}")

    retune_seen = set()
    for r in classification.get("retuned", []):
        sig = (r["label"], r["from"], r["to"])
        if sig in retune_seen:
            continue
        retune_seen.add(sig)
        lines.append(f"Retune on:\t{r['label']}\tFrom:\t{r['from']}\tTo:\t{r['to']}")

    # Confirmed real bug found this session: "Radio Swap on:" generation used to live here
    # too — a flat, undifferentiated list with no Completed/Pending distinction and no
    # stakeholder concept at all. It's fully superseded by classify_radio_swap_placement()
    # + format_radio_swaps() in mca_completed_logic.py (real Post-checks-driven placement,
    # correct stakeholder tagging on Pending only) — mca_report_ui.py already stripped this
    # old version's lines back out of scope_lines before ever reaching the checklist, but
    # that stripping happened too late to stop it from leaking into any earlier-stage
    # output built directly from this function's own return value. Removed at the source
    # instead of relying on a later filter to catch it.

    if dss_outputs_meta:
        lines.append(f"DSS Activation:\t{' & '.join(dss_outputs_meta)}")

    return lines


def scope_lines_to_table(scope_lines):
    """Parse the tab-separated Scope of Work lines into a clean table: Category | Details | From | To.
    Plain monospace text can't align cleanly since 'Integration:' and '6610 Controller Integration:'
    are very different lengths — a real table sidesteps that entirely."""
    rows = []
    for line in scope_lines:
        parts = line.split("\t")
        category = parts[0].rstrip(":")
        if "From:" in parts:
            fi = parts.index("From:")
            details = " ".join(p for p in parts[1:fi] if p)
            from_val = parts[fi + 1] if fi + 1 < len(parts) else ""
            ti = parts.index("To:") if "To:" in parts else None
            to_val = parts[ti + 1] if ti is not None and ti + 1 < len(parts) else ""
            rows.append({"Category": category, "Details": details, "From": from_val, "To": to_val})
        else:
            details = " — ".join(p for p in parts[1:] if p)
            rows.append({"Category": category, "Details": details, "From": "", "To": ""})
    return rows


def scope_lines_to_readable_text(scope_lines):
    """Same parsed fields as scope_lines_to_table, but rendered as compact readable sentences —
    single spaces, no raw tab characters (which jump to wide tab-stops in monospace display)."""
    out = []
    for row in scope_lines_to_table(scope_lines):
        if row["From"] or row["To"]:
            out.append(f"{row['Category']}: {row['Details']}  From: {row['From']}  To: {row['To']}")
        else:
            out.append(f"{row['Category']}: {row['Details']}" if row["Details"] else row["Category"])
    return out


# ============================================================
# GENERATOR: shared node-template fill (used by MMBB / TMBB / CENM alike —
# they all share the identical placeholder set, confirmed against the source templates)
# ============================================================

def fill_node_template(base_tpl, row, edp_index, user_id, date_str, summary_rows, log):
    site_id = row.get("Node to be built as")
    e_name, g_name, g_id = row.get("eNodeB Name"), row.get("gNodeB Name"), row.get("gNBId")
    is_primary_lte = str(site_id).strip().upper() == str(e_name or "").strip().upper()
    lte_row = edp_row_for(edp_index, e_name)
    fiveg_row = edp_row_for(edp_index, g_name)
    primary_row = lte_row if is_primary_lte else fiveg_row
    secondary_row = fiveg_row if is_primary_lte else lte_row

    lte_bearer = edp_get(edp_index, lte_row, "IPV6_ENODEB_BEARER_IP")
    fiveg_bearer = edp_get(edp_index, fiveg_row, "IPV6_ENODEB_BEARER_IP")
    ptp_server = edp_get(edp_index, primary_row, "BBU_PTP_SERVER_IP") or edp_get(edp_index, secondary_row, "BBU_PTP_SERVER_IP")
    ptp_siad = edp_get(edp_index, primary_row, "BBU_PTP_SIAD_IP") or edp_get(edp_index, secondary_row, "BBU_PTP_SIAD_IP")

    tpl = base_tpl
    fills = [
        ("xxxSiteIdxxx", site_id, "CIQ · Mixed Mode Info · Node to be built as (triple-x variant)"),
        ("xxSiteIDxx", site_id, "CIQ · Mixed Mode Info · Node to be built as"),
        ("xxSiteIdxx", site_id, "CIQ · Mixed Mode Info · Node to be built as"),
        ("xxDatexx", date_str, "manual input"),
        ("xxUserIDxx", user_id, "manual input"),
        ("x5G_gNBIdx", g_id, "CIQ · Mixed Mode Info · gNBId"),
        ("xgNB_Namex", g_name, "CIQ · Mixed Mode Info · gNodeB Name"),
        ("xxBBU_PTP_SERVER_IPxx", ptp_server, "EDP · BBU_PTP_SERVER_IP (primary→secondary fallback)"),
        ("xxBBU_PTP_SIAD_IPxx", ptp_siad, "EDP · BBU_PTP_SIAD_IP (primary→secondary fallback)"),
        ("xLTE_IPV6_ENODEB_BEARER_IPx", lte_bearer, "EDP · IPV6_ENODEB_BEARER_IP (row matched by eNodeB Name)"),
        ("x5G_IPV6_ENODEB_BEARER_IPx", fiveg_bearer, "EDP · IPV6_ENODEB_BEARER_IP (row matched by gNodeB Name)"),
    ]
    for token, val, src in fills:
        if val:
            tpl = tpl.replace(token, str(val))
            summary_rows.append({"Item": f"{site_id} · {token}", "Source": src, "Value": val, "Note": ""})
        else:
            summary_rows.append({"Item": f"{site_id} · {token}", "Source": src, "Value": "NOT FOUND", "Note": "left as placeholder"})
        log(f"{'✓' if val else '✗'} {site_id} · {token} -> {val or 'NOT FOUND'}")
    return tpl


def is_smbb_lte_primary(row):
    """SMBB (single-identity) row whose only/primary identity is LTE — the only SMBB variant
    a template exists for so far (SMBB-5G-primary is not yet supported, stays skipped)."""
    bbu_mode = str(row.get("BBU Mode", "")).strip().upper()
    has_lte = is_populated(row.get("eNBId")) or is_populated(row.get("eNodeB Name"))
    has_5g = is_populated(row.get("gNBId")) or is_populated(row.get("gNodeB Name"))
    return bbu_mode == "SMBB" and has_lte and not has_5g


def fill_node_template_smbb_lte(base_tpl, row, mm_objs, edp_index, user_id, date_str, summary_rows, log):
    """SMBB/LTE-primary placeholder fill — shared by MCA, CENM, and NSB (same CIQ/EDP source
    data; only the template *file* differs per scope). Confirmed real mapping:
    xLTE_SiteID2x/xLTE_SiteID3x = the other Mixed Mode Info rows in this same CIQ, in row order
    (not a distinct 'co-located' concept — just 'this row's siblings').
    xgNBIdx/xgNB_Namex/x5G_IPV6_ENODEB_BEARER_IPx and the xxSharing_*/xxFDD_NAMExx tokens (plus
    the literal 'xx' in RiLink=xx) are confirmed manual fill-in — deliberately NOT substituted
    here, left as-is in the output for the engineer."""
    site_id = row.get("Node to be built as")
    e_name = row.get("eNodeB Name")
    lte_row = edp_row_for(edp_index, e_name)
    lte_bearer = edp_get(edp_index, lte_row, "IPV6_ENODEB_BEARER_IP")
    ptp_server = edp_get(edp_index, lte_row, "BBU_PTP_SERVER_IP")
    ptp_siad = edp_get(edp_index, lte_row, "BBU_PTP_SIAD_IP")
    siad_bearer_router = edp_get(edp_index, lte_row, "ipv6_siad_bearer_ip_def_router")
    vlan_id = edp_get(edp_index, lte_row, "bearer_enodeb_sb_vlan_id")

    siblings = [r.get("Node to be built as") for r in mm_objs
                if r.get("Node to be built as") and r.get("Node to be built as") != site_id]
    site_id2 = siblings[0] if len(siblings) > 0 else None
    site_id3 = siblings[1] if len(siblings) > 1 else None

    tpl = base_tpl
    fills = [
        ("xxSiteIdxx", site_id, "CIQ · Mixed Mode Info · Node to be built as"),
        ("xLTE_SiteID1x", site_id, "CIQ · Mixed Mode Info · Node to be built as (same node, MCA amos-login token)"),
        ("xxUserIDxx", user_id, "manual input"),
        ("xxDatexx", date_str, "manual input"),
        ("xDatex", date_str, "manual input"),
        ("xLTE_IPV6_ENODEB_BEARER_IPx", lte_bearer, "EDP · IPV6_ENODEB_BEARER_IP (row matched by eNodeB Name)"),
        ("xxBBU_PTP_SERVER_IPxx", ptp_server, "EDP · BBU_PTP_SERVER_IP"),
        ("xxBBU_PTP_SIAD_IPxx", ptp_siad, "EDP · BBU_PTP_SIAD_IP"),
        ("xsecondary_IPV6_ENODEB_BEARER_IPx", lte_bearer, "EDP · IPV6_ENODEB_BEARER_IP (row matched by eNodeB Name)"),
        ("xLTE_IPV6_SIAD_BEARER_IPx", siad_bearer_router, "EDP · ipv6_siad_bearer_ip_def_router"),
        ("xLTE_Vlan_IDx", vlan_id, "EDP · bearer_enodeb_sb_vlan_id"),
        ("xLTE_SiteID2x", site_id2, "CIQ · Mixed Mode Info · other row (2nd)"),
        ("xLTE_SiteID3x", site_id3, "CIQ · Mixed Mode Info · other row (3rd)"),
    ]
    for token, val, src in fills:
        if val:
            tpl = tpl.replace(token, str(val))
            summary_rows.append({"Item": f"{site_id} · {token}", "Source": src, "Value": val, "Note": ""})
        else:
            summary_rows.append({"Item": f"{site_id} · {token}", "Source": src, "Value": "NOT FOUND", "Note": "left as placeholder"})
        log(f"{'✓' if val else '✗'} {site_id} · {token} -> {val or 'NOT FOUND'}")
    return tpl


def generate_6610(controller_objs, user_id, date_str, log, edp_found=None):
    """Universal add-on: generate the 6610 controller template if Controller Info shows 6610.
    Applies to ALL scopes (MCA, CENM, CRAN) per the blueprint's 'For ALL SCOPES' rule.
    edp_found: {controller_id: bool} — a controller not confirmed published in EDP gets no
    IX template at all (nothing reliable to fill it with), just a summary note explaining why."""
    outputs, summary_rows = [], []
    ctrl_rows = [r for r in controller_objs if str(r.get("Controller", "")).strip() == "6610"]
    if not ctrl_rows:
        return outputs, summary_rows
    base_tpl = TPL_6610.read_text(encoding="utf-8")
    for r in ctrl_rows:
        ctrl_id = r.get("Controller ID")
        if edp_found is not None and edp_found.get(ctrl_id) is False:
            summary_rows.append({"Item": "6610 Controller ID", "Source": "CIQ · Controller Info", "Value": ctrl_id, "Note": "EDP not published — 6610 IX template skipped"})
            log(f"✗ 6610 present but EDP not published for Controller ID {ctrl_id} — IX template skipped")
            continue
        tpl = base_tpl.replace("##Controller_id##", str(ctrl_id))
        tpl = tpl.replace("xSite_IDx", str(ctrl_id))
        tpl = tpl.replace("xxUserIDxx", user_id)
        tpl = tpl.replace("xDatex", date_str)
        outputs.append((f"{ctrl_id}_6610_Controller_Integration_Filled.txt", tpl))
        summary_rows.append({"Item": "6610 Controller ID", "Source": "CIQ · Controller Info", "Value": ctrl_id, "Note": "6610 IX template generated (applies across all scopes)"})
        summary_rows.append({"Item": "xSite_IDx", "Source": "same as Controller ID — no other node in scope for this template", "Value": ctrl_id, "Note": "VERIFY if ever wrong"})
        log(f"✓ 6610 present -> generated for Controller ID {ctrl_id}")
    return outputs, summary_rows


# ============================================================
# GENERATOR: DSS checks (ported from ashhureddy/TRYDSS — DSS Extractor Tool)
# Universal add-on across ALL scopes (MCA, CENM, CRAN), same as 6610.
# ============================================================

DSS_SECTOR_MAP = {'A': 'alpha', 'B': 'beta', 'C': 'gamma', 'D': 'delta', 'E': 'epsilon', 'F': 'zeta'}
DSS_KEEP_PARAMS = ['gNBId', 'gNB Name', 'SectorEquipmentFunction', 'cellLocalId', 'Carrier', 'ssbFrequency']
DSS_ESS_SC_LOOKUP = {
    "N066A_1": {"essScPairId": 2222, "essScLocalId": 20}, "N066B_1": {"essScPairId": 2223, "essScLocalId": 21},
    "N066C_1": {"essScPairId": 2224, "essScLocalId": 22}, "N066D_1": {"essScPairId": 2225, "essScLocalId": 23},
    "N066A_2": {"essScPairId": 2226, "essScLocalId": 24}, "N066B_2": {"essScPairId": 2227, "essScLocalId": 25},
    "N066C_2": {"essScPairId": 2228, "essScLocalId": 26}, "N066D_2": {"essScPairId": 2229, "essScLocalId": 27},
    "N002A_1": {"essScPairId": 3322, "essScLocalId": 30}, "N002B_1": {"essScPairId": 3323, "essScLocalId": 31},
    "N002C_1": {"essScPairId": 3324, "essScLocalId": 32}, "N002D_1": {"essScPairId": 3325, "essScLocalId": 33},
    "N002A_2": {"essScPairId": 3326, "essScLocalId": 34}, "N002B_2": {"essScPairId": 3327, "essScLocalId": 35},
    "N002C_2": {"essScPairId": 3328, "essScLocalId": 36}, "N002D_2": {"essScPairId": 3329, "essScLocalId": 37},
    "N005A_1": {"essScPairId": 1122, "essScLocalId": 10}, "N005B_1": {"essScPairId": 1123, "essScLocalId": 11},
    "N005C_1": {"essScPairId": 1124, "essScLocalId": 12}, "N005D_1": {"essScPairId": 1125, "essScLocalId": 13},
    "N005A_2": {"essScPairId": 1126, "essScLocalId": 14}, "N005B_2": {"essScPairId": 1127, "essScLocalId": 15},
    "N005C_2": {"essScPairId": 1128, "essScLocalId": 16}, "N005D_2": {"essScPairId": 1129, "essScLocalId": 17},
}
DSS_PLACEHOLDERS = {
    "primary_node": "xxMMBB_Primary_Node_Namexx", "lte_site_id": "xxLTE_Site_IDxx",
    "nr_node_name": "xx5G_NR_Node_Namexx", "lte_enbid": "xxLTE_eNBIDxx", "nr_gnbid": "xx5G_NR_gNBIDxx",
    "lte_cellid_a": "LTE_cellidA", "lte_cellid_b": "LTE_cellidB", "lte_cellid_c": "LTE_cellidC", "lte_cellid_d": "LTE_cellidD",
    "nr_celllocalid_a": "xx5G_celllocalidAxx", "nr_celllocalid_b": "xx5G_celllocalidBxx",
    "nr_celllocalid_c": "xx5G_celllocalidCxx", "nr_celllocalid_d": "xx5G_celllocalidDxx",
    "nr_ssbfrequency_a": "xx5G_ssbfrequencyAxx",
    "nr_sector_carrier_alpha": "xx5G_NRSectorCarrier_Alphaxx", "nr_sector_carrier_beta": "xx5G_NRSectorCarrier_Betaxx",
    "nr_sector_carrier_gamma": "xx5G_NRSectorCarrier_Gammaxx", "nr_sector_carrier_delta": "xx5G_NRSectorCarrier_Deltaxx",
    "lte_sector_carrier_alpha": "xxLTE_SectorCarrier_No_Alphaxx", "lte_sector_carrier_beta": "xxLTE_SectorCarrier_No_Betaxx",
    "lte_sector_carrier_gamma": "xxLTE_SectorCarrier_No_Gammaxx", "lte_sector_carrier_delta": "xxLTE_SectorCarrier_No_Deltaxx",
    "lte_site_xa_1": "xxLTE_Site_IDxx_XA_1", "lte_site_xb_1": "xxLTE_Site_IDxx_XB_1",
    "lte_site_xc_1": "xxLTE_Site_IDxx_XC_1", "lte_site_xd_1": "xxLTE_Site_IDxx_XD_1",
    "nr_node_n00xa_1": "xx5G_NR_Node_Namexx_N00XA_1", "nr_node_n00xb_1": "xx5G_NR_Node_Namexx_N00XB_1",
    "nr_node_n00xc_1": "xx5G_NR_Node_Namexx_N00XC_1", "nr_node_n00xd_1": "xx5G_NR_Node_Namexx_N00XD_1",
    "n00xa": "N00XA", "n00xb": "N00XB", "n00xc": "N00XC", "n00xd": "N00XD", "n00x": "N00X",
    "ess_sc_pair_id_a": "essScPairId_A", "ess_sc_pair_id_b": "essScPairId_B",
    "ess_sc_pair_id_c": "essScPairId_C", "ess_sc_pair_id_d": "essScPairId_D",
    "ess_sc_local_id_a": "essScLocalId_A", "ess_sc_local_id_b": "essScLocalId_B",
    "ess_sc_local_id_c": "essScLocalId_C", "ess_sc_local_id_d": "essScLocalId_D",
    "nr_node_n00x": "xx5G_NR_Node_Namexx_N00X",
}

def dss_extract_band_carrier_pattern(nrcelldu):
    if not nrcelldu:
        return "UNKNOWN"
    parts = str(nrcelldu).strip().split('_')
    if len(parts) < 3:
        return "UNKNOWN"
    middle_part, carrier_num = parts[1], parts[2]
    m = re.match(r'^([A-Z]\d+)[A-Z]?$', middle_part)
    return f"{m.group(1)}_{carrier_num}" if m else "UNKNOWN"

def dss_extract_sector(value):
    if not value or str(value) == 'nan':
        return None
    s = str(value)
    m = re.search(r'_([A-Z]\d+)([A-Z])_', s)
    if m:
        return m.group(2)
    m = re.search(r'_(\d+)([A-Z])_', s)
    return m.group(2) if m else None

def dss_get_greek_name(sector, counts):
    if sector not in DSS_SECTOR_MAP:
        return f"sector_{sector.lower()}"
    greek = DSS_SECTOR_MAP[sector]
    if sector in counts:
        counts[sector] += 1
        return f"{greek}{counts[sector]}"
    counts[sector] = 0
    return greek

def dss_filter_row(row):
    row_map = {str(k).strip().upper(): k for k in row.keys()}
    out = {}
    for p in DSS_KEEP_PARAMS:
        ku = p.strip().upper()
        if ku in row_map:
            out[p] = row[row_map[ku]]
    return out

def dss_get_primary_node_info(mm_objs, gnb_name, gnb_id):
    # Confirmed bug: Mixed Mode Info's gNBId is often text while 5G Info's gNBId loads as a
    # number (different cell formatting between sheets) — a strict == comparison silently
    # never matches, leaving MMBB_Primary_Node_Name/LTE_Site_ID/LTE_eNBID unresolved even
    # when the row is right there. Compare both sides as strings.
    for r in mm_objs:
        if str(r.get("gNodeB Name", "")).strip() == str(gnb_name).strip() and str(r.get("gNBId", "")).strip() == str(gnb_id).strip():
            out = {}
            if r.get("Node to be built as"):
                out["primary_node"] = r.get("Node to be built as")
            if is_populated(r.get("eNBId")):
                out["eNBId"] = r.get("eNBId")
            if r.get("eNodeB Name"):
                out["lte_siteID"] = r.get("eNodeB Name")
            return out
    return {}

def dss_get_sector_cell_ids(eutran_objs, dss_value, greek_name):
    search_val = str(dss_value).strip()
    for row in eutran_objs:
        if str(row.get("EutranCellFDDId", row.get("EUtranCellFDDId", ""))).strip() == search_val:
            out = {}
            if is_populated(row.get("sectorId")) or is_populated(row.get("SectorId")):
                out[f"{greek_name}_sectorId"] = row.get("sectorId", row.get("SectorId"))
            if is_populated(row.get("cellId")) or is_populated(row.get("CellId")):
                out[f"{greek_name}_cellId"] = row.get("cellId", row.get("CellId"))
            return out
    return {}

def dss_extract_pattern_for_ess(nr_value):
    if not nr_value:
        return None
    m = re.search(r'_(N\d{3}[A-D]_\d)$', str(nr_value))
    return m.group(1) if m else None

def dss_extract_n00x_from_node(nr_node_value):
    if not nr_node_value:
        return None
    m = re.match(r'^(.+_N\d{3})[A-D]_\d$', str(nr_node_value))
    return m.group(1) if m else None


def generate_dss(ciq_wb, mm_objs, user_id, date_str, log):
    """Ported from ashhureddy/TRYDSS DSS Extractor Tool — 6-step pipeline:
    extract -> group -> clean -> populate -> map -> generate.
    Also returns dss_activation_labels: ["5G_PCS_1|PCS_1", ...] for the Scope of Work summary."""
    outputs, summary_rows, dss_activation_labels = [], [], []

    if "5G Info" not in ciq_wb.sheetnames:
        return outputs, summary_rows, dss_activation_labels
    fiveg_objs = sheet_objs(ciq_wb["5G Info"])
    dss_col = next((k for k in (fiveg_objs[0].keys() if fiveg_objs else []) if k.strip().upper() == "DSS"), None)
    if not dss_col:
        return outputs, summary_rows, dss_activation_labels

    # Step 1: extract rows where DSS != "NO"
    dss_rows = [r for r in fiveg_objs if r.get(dss_col) is not None and str(r.get(dss_col)).strip().upper() != "NO"]
    if not dss_rows:
        summary_rows.append({"Item": "DSS checks", "Source": "CIQ · 5G Info · DSS column", "Value": "no DSS-active cells found", "Note": ""})
        return outputs, summary_rows, dss_activation_labels
    log(f"✓ DSS: found {len(dss_rows)} DSS-active cell(s) in 5G Info")

    # Step 2: group by band+carrier pattern
    groups = {}
    for row in dss_rows:
        pattern = dss_extract_band_carrier_pattern(row.get("NRCellDU"))
        groups.setdefault(pattern, []).append(row)

    eutran_objs = sheet_objs(ciq_wb["eUtran Parameters"]) if "eUtran Parameters" in ciq_wb.sheetnames else []

    outputs_count = {"4_sector": 0, "3_sector": 0}
    for i, (pattern, rows) in enumerate(sorted(groups.items()), start=1):
        var_name = pattern if pattern != "UNKNOWN" else f"DSS{i}"

        # Step 3: clean — Greek sector names + filtered rows
        cleaned = {}
        dss_counts, nr_counts = {}, {}
        for row in rows:
            sector = dss_extract_sector(row.get(dss_col))
            if sector:
                greek = dss_get_greek_name(sector, dss_counts)
                cleaned[f"DSS_{greek}"] = row.get(dss_col)
        for row in rows:
            sector = dss_extract_sector(row.get("NRCellDU"))
            if sector:
                greek = dss_get_greek_name(sector, nr_counts)
                cleaned[f"NR_{greek}"] = row.get("NRCellDU")
        cleaned["rows"] = [dss_filter_row(r) for r in rows]

        # Step 4: populate — Mixed Mode Info + eUtran Parameters
        if cleaned["rows"]:
            gnb_name = cleaned["rows"][0].get("gNB Name")
            gnb_id = cleaned["rows"][0].get("gNBId")
            if gnb_name and gnb_id:
                cleaned.update(dss_get_primary_node_info(mm_objs, gnb_name, gnb_id))
        for dss_key in sorted(k for k in cleaned if k.startswith("DSS_")):
            greek = dss_key.replace("DSS_", "")
            cleaned.update(dss_get_sector_cell_ids(eutran_objs, cleaned[dss_key], greek))
        for idx, row in enumerate(rows, start=1):
            sec_eq = row.get("SectorEquipmentFunction")
            if sec_eq:
                parts = str(sec_eq).split("_")
                if len(parts) >= 2:
                    cleaned[f"row{idx}"] = parts[-1]

        # Step 5: map to placeholders
        r = cleaned.get("rows", [])
        mapped = {}
        mapped[DSS_PLACEHOLDERS["primary_node"]] = cleaned.get("primary_node")
        mapped[DSS_PLACEHOLDERS["lte_site_id"]] = cleaned.get("lte_siteID")
        mapped[DSS_PLACEHOLDERS["nr_node_name"]] = r[0].get("gNB Name") if r else None
        mapped[DSS_PLACEHOLDERS["lte_enbid"]] = cleaned.get("eNBId")
        mapped[DSS_PLACEHOLDERS["nr_gnbid"]] = r[0].get("gNBId") if r else None
        for i2, letter in enumerate(["a", "b", "c", "d"]):
            mapped[DSS_PLACEHOLDERS[f"nr_celllocalid_{letter}"]] = r[i2].get("cellLocalId") if len(r) > i2 else None
        mapped[DSS_PLACEHOLDERS["nr_ssbfrequency_a"]] = r[0].get("ssbFrequency") if r else None
        nr_vals = {}
        for letter, key in [("a", "nr_alpha"), ("b", "nr_beta"), ("c", "nr_gamma"), ("d", "nr_delta")]:
            greek = DSS_SECTOR_MAP[letter.upper()]
            nr_vals[key] = cleaned.get(f"NR_{greek}")
        mapped[DSS_PLACEHOLDERS["nr_sector_carrier_alpha"]] = nr_vals["nr_alpha"]
        mapped[DSS_PLACEHOLDERS["nr_sector_carrier_beta"]] = nr_vals["nr_beta"]
        mapped[DSS_PLACEHOLDERS["nr_sector_carrier_gamma"]] = nr_vals["nr_gamma"]
        mapped[DSS_PLACEHOLDERS["nr_sector_carrier_delta"]] = nr_vals["nr_delta"]
        for letter, greek in [("a", "alpha"), ("b", "beta"), ("c", "gamma"), ("d", "delta")]:
            mapped[DSS_PLACEHOLDERS[f"lte_sector_carrier_{greek}"]] = cleaned.get(f"{greek}_sectorId")
            mapped[DSS_PLACEHOLDERS[f"lte_cellid_{letter}"]] = cleaned.get(f"{greek}_cellId")
            mapped[DSS_PLACEHOLDERS[f"lte_site_x{letter}_1"]] = cleaned.get(f"DSS_{greek}")
            mapped[DSS_PLACEHOLDERS[f"nr_node_n00x{letter}_1"]] = cleaned.get(f"NR_{greek}")
        mapped[DSS_PLACEHOLDERS["n00xa"]] = cleaned.get("row1")
        mapped[DSS_PLACEHOLDERS["n00xb"]] = cleaned.get("row2")
        mapped[DSS_PLACEHOLDERS["n00xc"]] = cleaned.get("row3")
        mapped[DSS_PLACEHOLDERS["n00xd"]] = cleaned.get("row4")
        for letter, greek in [("a", "alpha"), ("b", "beta"), ("c", "gamma"), ("d", "delta")]:
            ess = DSS_ESS_SC_LOOKUP.get(dss_extract_pattern_for_ess(nr_vals.get(f"nr_{greek}")) or "", {})
            mapped[DSS_PLACEHOLDERS[f"ess_sc_pair_id_{letter}"]] = ess.get("essScPairId")
            mapped[DSS_PLACEHOLDERS[f"ess_sc_local_id_{letter}"]] = ess.get("essScLocalId")
        nr_node_ref = nr_vals.get("nr_gamma") or nr_vals.get("nr_delta")
        mapped[DSS_PLACEHOLDERS["nr_node_n00x"]] = dss_extract_n00x_from_node(nr_node_ref)
        mapped[DSS_PLACEHOLDERS["n00x"]] = pattern.split('_')[0] if pattern != "UNKNOWN" else None
        mapped["xxDatexx"] = date_str

        # Step 6: pick template (4-sector if Delta present, else 3-sector) and generate
        has_delta = any(mapped.get(DSS_PLACEHOLDERS[k]) is not None for k in
                         ["lte_cellid_d", "nr_celllocalid_d", "nr_sector_carrier_delta", "ess_sc_pair_id_d"])
        tpl_path = TPL_DSS_4SECTOR if has_delta else TPL_DSS_3SECTOR
        tpl_key = "4_sector" if has_delta else "3_sector"
        outputs_count[tpl_key] += 1
        tpl_text = tpl_path.read_text(encoding="utf-8")

        for placeholder, val in sorted(mapped.items(), key=lambda x: len(x[0]), reverse=True):
            if val is not None:
                tpl_text = tpl_text.replace(placeholder, str(val))
                summary_rows.append({"Item": f"{var_name} · {placeholder}", "Source": "DSS pipeline", "Value": val, "Note": ""})
            else:
                summary_rows.append({"Item": f"{var_name} · {placeholder}", "Source": "DSS pipeline", "Value": "NOT FOUND", "Note": "left as placeholder"})

        summary_rows.append({"Item": f"{var_name} · noOfRxAntennas/noOfTxAntennas", "Source": "not computed by source tool", "Value": "n/a", "Note": "gap in the original DSS Extractor — never mapped there either"})
        outputs.append((f"{var_name}_DSS_output.txt", tpl_text))
        log(f"✓ DSS group {var_name} -> {tpl_key} template, {sum(1 for v in mapped.values() if v is not None)} placeholders resolved")

        nr_cell_for_label = nr_vals.get("nr_alpha")
        lte_cell_for_label = cleaned.get("DSS_alpha")
        nr_label, _ = nr_band_label(nr_cell_for_label) if nr_cell_for_label else (None, None)
        lte_label, _ = lte_band_label(lte_cell_for_label) if lte_cell_for_label else (None, None)
        if nr_label and lte_label:
            dss_activation_labels.append(f"{nr_label}|{lte_label}")

    return outputs, summary_rows, dss_activation_labels


# ============================================================
# FINAL CONNECTIONS (universal — same across MCA/CENM/CRAN/N2E/NSB)
# ============================================================

def sheet_objs_dedup_first(ws):
    """Like sheet_objs but keeps the FIRST occurrence of a duplicate header name (eUtran Parameters
    has DUS/XMU columns repeated near the end — the two sets hold identical values, first wins)."""
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    seen, header_idx = {}, []
    for i, h in enumerate(headers):
        hs = str(h).strip() if h is not None else ''
        if hs and hs not in seen:
            seen[hs] = i
            header_idx.append((hs, i))
    objs = []
    for r in rows[1:]:
        if not any(str(c).strip() for c in r if c is not None):
            continue
        objs.append({h: ((r[i].strip() if isinstance(r[i], str) else r[i]) if i < len(r) else None) for h, i in header_idx})
    return objs


def generate_final_connections(ciq_wb, mm_objs):
    """One Excel file per CIQ (not per node) — Mixed Mode Info + conditional per-node XMU rows,
    then all 5G Info rows, then all eUtran Parameters rows. Styled: yellow/red title, blue/white headers, borders."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side

    fiveg_objs = sheet_objs(ciq_wb["5G Info"]) if "5G Info" in ciq_wb.sheetnames else []
    eutran_objs = sheet_objs_dedup_first(ciq_wb["eUtran Parameters"]) if "eUtran Parameters" in ciq_wb.sheetnames else []
    enb_objs = sheet_objs(ciq_wb["eNB Info"]) if "eNB Info" in ciq_wb.sheetnames else []
    gnb_objs = sheet_objs(ciq_wb["gNB Info"]) if "gNB Info" in ciq_wb.sheetnames else []

    MM_COLS = ['Node to be built as', 'eNBId', 'eNodeB Name', 'gNBId', 'gNodeB Name', 'IDLA', 'Connected To Node', 'Connected From Port']
    XMU_ENB_COLS = ['eNBId', 'eNodeB Name', '1st DU type', '1st XMU', '1st XMU Port 1', '1st XMU Port 2', '1st XMU Port 3', '2nd DU type', '2nd XMU', '2nd XMU Port 1', '2nd XMU Port 2', '2nd XMU Port 3']
    XMU_GNB_COLS = ['gNBId', 'gNodeB Name', 'DU type', '1st XMU', '1st XMU Port 1', '1st XMU Port 2', '1st XMU Port 3', '2nd XMU', '2nd XMU Port 1', '2nd XMU Port 2', '2nd XMU Port 3']
    FIVEG_COLS = ['gNBId', 'gNB Name', 'NRCellDU', 'Operating Band', 'RRU Type', 'BB/XMU', 'Port 1', 'Port 2', 'Port 3', 'Port 4', 'Radio Port', 'Cascaded From Radio', 'BBU/XMU End SFP', 'Radio End SFP']
    LTE_COLS = ['EutranCellFDDId', 'eUTRA operating band', 'RRU type', 'DUS / XMU', 'DUS / XMU Port', 'DUS / XMU Port Expansion', 'Cascaded From Radio', 'Radio Port', 'BBU/XMU End SFP', 'Radio End SFP']
    NCOLS = 14

    TITLE_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    TITLE_FONT = Font(bold=True, color="FF0000", size=14)
    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="000000")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_header_row(ws, r, cols, ncols):
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = HEADER_FILL
            cell.border = BORDER
            cell.value = cols[c - 1]
            cell.font = HEADER_FONT

    def write_data_row(ws, r, row, cols, ncols):
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.value = row.get(cols[c - 1])

    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = "Sheet1"
    r = 1
    title_cell = ws.cell(row=r, column=2, value="Final connections")
    title_cell.font = TITLE_FONT
    for c in range(2, 9):
        ws.cell(row=r, column=c).fill = TITLE_FILL
    r += 2

    write_header_row(ws, r, MM_COLS, len(MM_COLS)); r += 1
    xmu_rows_to_add = []
    for row in mm_objs:
        clean_row = dict(row)
        for k in ("eNBId", "eNodeB Name", "gNBId", "gNodeB Name"):
            if not is_populated(clean_row.get(k)):
                clean_row[k] = None
        write_data_row(ws, r, clean_row, MM_COLS, len(MM_COLS)); r += 1
        primary = row.get("Node to be built as")
        e_name, g_name = row.get("eNodeB Name"), row.get("gNodeB Name")
        is_lte_primary = str(primary).strip().upper() == str(e_name or "").strip().upper()
        if is_lte_primary:
            match = next((x for x in enb_objs if str(x.get("eNodeB Name")) == str(e_name)), None)
            if match and (str(match.get("1st XMU")).strip().upper() == "YES" or str(match.get("2nd XMU")).strip().upper() == "YES"):
                xmu_rows_to_add.append((XMU_ENB_COLS, match))
        else:
            match = next((x for x in gnb_objs if str(x.get("gNodeB Name")) == str(g_name)), None)
            if match and (str(match.get("1st XMU")).strip().upper() == "YES" or str(match.get("2nd XMU")).strip().upper() == "YES"):
                xmu_rows_to_add.append((XMU_GNB_COLS, match))
    for cols, match in xmu_rows_to_add:
        r += 1
        write_header_row(ws, r, cols, len(cols)); r += 1
        write_data_row(ws, r, match, cols, len(cols)); r += 1

    r += 2
    write_header_row(ws, r, FIVEG_COLS, NCOLS); r += 1
    for row in fiveg_objs:
        write_data_row(ws, r, row, FIVEG_COLS, NCOLS); r += 1

    r += 2
    write_header_row(ws, r, LTE_COLS, NCOLS); r += 1
    for row in eutran_objs:
        write_data_row(ws, r, row, LTE_COLS, NCOLS); r += 1

    widths = [14, 10, 12, 10, 14, 10, 10, 10, 10, 10, 10, 16, 16, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    out_wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============================================================
# IDL CONNECTIONS (shared by MCA / CENM / NSB — CIQ-only, no Pre-checks PDF)
#
# Confirmed logic:
#  - Trigger: site has 2+ BBU rows in Mixed Mode Info.
#  - Per node: type (MMBB/TMBB vs LTE-standalone vs 5G-standalone) is read off eNBId/gNBId
#    presence in Mixed Mode Info; board generation comes from the "DU type" column in
#    eNB Info (LTE-standalone), gNB Info (5G-standalone), or either (MMBB/TMBB) —
#    "1st DU type"/"2nd DU type" are explicitly ignored.
#  - DU type -> generation: 6630/5216 -> G2, 6648/6651 -> G3, 6672 -> G4.
#  - The site's generation combination (order-independent) is matched against the confirmed
#    15-template registry. Combinations with both a Preferred and Alternate variant
#    (G2+G3, G3+G3, G3+G3+G3) generate BOTH files. The 4 known-unsupported 3-BBU
#    combinations (G2+G2+G4, G2+G3+G4, G3+G3+G4, G3+G4+G4) return "IDL template not found".
#  - Node ordering for same-generation nodes follows CIQ row order, top = 1st.
#  - Placeholder filling is generic rather than hardcoded per template: for each node we build
#    a set of candidate slot-prefixes (global row-position ordinal, per-generation-group
#    ordinal, and the plain generation label when that generation is a singleton at the site)
#    and only replace whichever placeholder actually appears in that specific template file —
#    per your instruction to "just fill whatever placeholders the template has", since the 15
#    files don't all share one exact naming convention (e.g. G2+G4+G4 uses Node_ID/ENBID
#    instead of NODE_ID/Node_eNBId).
# ============================================================

TDIR_IDL = Path(__file__).parent / "templates" / "IDL"
TDIR_N2E_IDL = Path(__file__).parent / "templates" / "N2E" / "IDL"

DU_TYPE_TO_GEN = {"6630": "G2", "5216": "G2", "6648": "G3", "6651": "G3", "6672": "G4"}

# Confirmed real reference table (uploaded IDL_Connections.xlsx, "IDL Connections_MCA" tab) —
# Build Type letter -> combination pattern + IDLe/IDLy cable part number(s) with their own
# combination pattern. Patterns use bare "G2"/"G3"/"G4" for a single node of that generation,
# or "G2(1)"/"G2(2)" etc. when more than one node shares the same generation (e.g. Build Type D
# has three separate G2 nodes) — each substituted with that specific node's own real label.
IDL_CABLE_REFERENCE = {
    "A": {"idle": [("RPM 777 417", "G2(1)+G2(2)")], "idly": []},
    "B": {"idle": [], "idly": [("RPM 777 098", "G2+G3")]},
    "BB": {"idle": [], "idly": [("RPM 777 544", "G2+G3")]},
    "C": {"idle": [("RPM 777 052", "G3(1)+G3(2)")], "idly": []},
    "CC": {"idle": [("RPM 777 053", "G3(1)+G3(2)")], "idly": []},
    "R": {"idle": [("RPM 777 052", "G4(1)+G4(2)")], "idly": []},
    "S": {"idle": [("RPM 777 543", "G4+G2")], "idly": []},
    "T": {"idle": [("RPM 777 052", "G4+G3")], "idly": []},
    "D": {"idle": [("RPM 777 417", "G2(1)+G2(2)"), ("RPM 777 417", "G2(2)+G2(3)"), ("RPM 777 417", "G2(3)+G2(1)")], "idly": []},
    "E": {"idle": [("RPM 777 417", "G2(1)+G2(2)")], "idly": [("RPM 777 098", "G2(1)+G2(2)+G3")]},
    "EE": {"idle": [("RPM 777 417", "G2(1)+G2(2)")], "idly": [("RPM 777 098", "G2(2)+G3")]},
    "F": {"idle": [("RPM 777 053", "G3(1)+G3(2)")], "idly": [("RPM 777 544", "G2+G3(1)"), ("RPM 777 544", "G2+G3(2)")]},
    "FF": {"idle": [("RPM 777 053", "G3(1)+G3(2)")], "idly": [("RPM 777 544", "G2+G3(1)")]},
    "G": {"idle": [("RPM 777 053", "G3(1)+G3(2)")], "idly": [("RPM 777 054", "G3(1)+G3(2)+G3(3)")]},
    "GG": {"idle": [("RPM 777 052", "G3(1)+G3(2)"), ("RPM 777 052", "G3(2)+G3(3)"), ("RPM 777 052", "G3(3)+G3(1)")], "idly": []},
    "RR": {"idle": [("RPM 777 052", "G4(1)+G4(2)"), ("RPM 777 052", "G4(2)+G4(3)"), ("RPM 777 052", "G4(3)+G4(1)")], "idly": []},
    "TT": {"idle": [("RPM 777 052", "G3+G4(1)"), ("RPM 777 052", "G4(1)+G4(2)"), ("RPM 777 052", "G3+G4(2)")], "idly": []},
    "U": {"idle": [("RPM 777 053", "G4(1)+G4(2)"), ("RPM 777 543", "G2+G4(1)"), ("RPM 777 543", "G2+G4(2)")], "idly": []},
    "UU": {"idle": [("RPM 777 052", "G3+G4"), ("RPM 777 543", "G2+G3"), ("RPM 777 543", "G2+G4")], "idly": []},
}


def _idl_cable_node_label(ciq_wb, row):
    """Confirmed format for IDL cable substitution — same (P)/(S) dual-identity tagging and
    hardware as Post Configuration, but WITHOUT the BBU mode tag (TMBB/SMBB/etc.): just
    "{primary}(P)/{secondary}(S)({hw})" or "{primary}({hw})" for a single-identity node."""
    primary = row.get("Node to be built as")
    e_name, g_name = row.get("eNodeB Name"), row.get("gNodeB Name")
    is_lte_primary = str(primary).strip().upper() == str(e_name or "").strip().upper()
    r = find_row_by_name(ciq_wb, "eNB Info", "eNodeB Name", e_name) if is_lte_primary else \
        find_row_by_name(ciq_wb, "gNB Info", "gNodeB Name", g_name)
    if not r:
        r = find_row_by_name(ciq_wb, "eNB Info", "eNodeB Name", e_name) or \
            find_row_by_name(ciq_wb, "gNB Info", "gNodeB Name", g_name)
    hw = hw_string(r) or "NOT FOUND"
    # Confirmed: XMU presence isn't relevant for IDL cable reporting — hw_string() appends
    # " + XMU"/" + N XMU" when XMU boards are present, strip that here specifically.
    hw = re.sub(r"\s*\+\s*\d*\s*XMU\s*$", "", hw).strip()
    if is_populated(e_name) and is_populated(g_name):
        secondary = g_name if is_lte_primary else e_name
        return f"{primary}(P)/{secondary}(S)({hw})"
    return f"{primary}({hw})"


def _idl_cable_columns(ciq_wb, nodes_by_gen, entries, third_col=False):
    """Confirmed real .xlsm gap: the template has FIXED columns per row (C=Cable P/N,
    D=1st Node ID, E=2nd Node ID, [F=3rd Node ID for IDLy only]) -- not a place to dump the
    whole combined text line, and only one row exists (no overflow rows like
    Integration/Transport SFP have). Multiple cable connections for the same build type get
    pipe-joined per column (C: part1|part2, D: node1a|node2a, E: node1b|node2b); within a
    single connection with more nodes than there are columns, the excess pipe-joins into the
    LAST available column. third_col=True (IDLy only -- confirmed real template difference:
    row 21 has a genuine 4th value column F for a 3rd node, row 20/IDLe does not) uses D/E/F
    directly for a single 3-node connection (e.g. Build Type G's "G3(1)+G3(2)+G3(3)") instead
    of pipe-joining nodes 2+3 into E. Returns (C, D, E) normally, or (C, D, E, F) when
    third_col=True."""
    def substitute_nodes(pattern):
        nodes, used = [], {}
        for gen, idx in re.findall(r"(G\d)(?:\((\d+)\))?", pattern):
            candidates = nodes_by_gen.get(gen, [])
            pos = (int(idx) - 1) if idx else used.get(gen, 0)
            used[gen] = used.get(gen, 0) + 1
            nodes.append(_idl_cable_node_label(ciq_wb, candidates[pos]) if 0 <= pos < len(candidates) else f"{gen} (node not found)")
        return nodes

    if third_col and len(entries) == 1:
        part, combo = entries[0]
        nodes = substitute_nodes(combo)
        if len(nodes) == 3:
            return part, nodes[0], nodes[1], nodes[2]
        if len(nodes) > 3:
            return part, nodes[0], nodes[1], "|".join(nodes[2:])

    parts, firsts, rests = [], [], []
    for part, combo in entries:
        nodes = substitute_nodes(combo)
        parts.append(part)
        firsts.append(nodes[0] if nodes else "")
        if len(nodes) > 1:
            rests.append("|".join(nodes[1:]))
    c, d, e = "|".join(parts), "|".join(firsts), "|".join(rests)
    return (c, d, e, "") if third_col else (c, d, e)


def build_type_idl_slots(build_type_letter):
    """Confirmed real UI gap: a build type can structurally have NO IDLe entry at all (e.g.
    Type BB only ever uses IDLy) — that's different from "auto-derivation found nothing but
    this build type could plausibly need one," and showing a manual IDLe fallback field in
    the former case is just wrong, not merely unhelpful. Returns (has_idle, has_idly) —
    whether IDL_CABLE_REFERENCE has ANY entries for that slot, regardless of whether
    substitution later succeeds."""
    entry = IDL_CABLE_REFERENCE.get(str(build_type_letter or "").strip().upper())
    if not entry:
        return True, True  # unknown build type — can't rule either slot out, so allow both
    return bool(entry.get("idle")), bool(entry.get("idly"))


def idl_cable_columns_for_build_type(ciq_wb, mm_objs, build_type_letter):
    """(C, D, E) column version of idl_cable_lines_for_build_type(), for the .xlsm write —
    see _idl_cable_columns() for the combination rule. Returns (idle_cols, idly_cols),
    each an (C, D, E) tuple."""
    entry = IDL_CABLE_REFERENCE.get(str(build_type_letter or "").strip().upper())
    if not entry:
        return ("", "", ""), ("", "", "")
    nodes_by_gen = {}
    for row in mm_objs:
        gen = get_node_generation(ciq_wb, row)
        if gen:
            nodes_by_gen.setdefault(gen, []).append(row)
    return _idl_cable_columns(ciq_wb, nodes_by_gen, entry["idle"]), _idl_cable_columns(ciq_wb, nodes_by_gen, entry["idly"], third_col=True)


def idl_cable_lines_for_build_type(ciq_wb, mm_objs, build_type_letter):
    """Confirmed rule: match nodes to each combination pattern's generation slots using
    get_node_generation() — same mapping already used for Build Type detection — then
    substitute each slot with that specific real node's own label (_idl_cable_node_label).
    Bare "G2" means the single node of that generation; "G2(1)"/"G2(2)" means the 1st/2nd
    node of that generation, in CIQ row order. Returns (idle_lines, idly_lines), each a list
    of "{part number} : {substituted combination}" strings."""
    entry = IDL_CABLE_REFERENCE.get(str(build_type_letter or "").strip().upper())
    if not entry:
        return [], []

    nodes_by_gen = {}
    for row in mm_objs:
        gen = get_node_generation(ciq_wb, row)
        if gen:
            nodes_by_gen.setdefault(gen, []).append(row)

    def substitute(pattern):
        parts = []
        used_counts = {}
        for gen, idx in re.findall(r"(G\d)(?:\((\d+)\))?", pattern):
            candidates = nodes_by_gen.get(gen, [])
            pos = (int(idx) - 1) if idx else used_counts.get(gen, 0)
            used_counts[gen] = used_counts.get(gen, 0) + 1
            if 0 <= pos < len(candidates):
                parts.append(_idl_cable_node_label(ciq_wb, candidates[pos]))
            else:
                parts.append(f"{gen} (node not found)")
        return " + ".join(parts)

    idle_lines = [f"IDLe : {part} : {substitute(combo)}" for part, combo in entry["idle"]]
    idly_lines = [f"IDLy : {part} : {substitute(combo)}" for part, combo in entry["idly"]]
    return idle_lines, idly_lines


# combo (sorted tuple of generations) -> list of (filename, variant label)
IDL_TEMPLATE_REGISTRY = {
    ("G2", "G2"): [("G2+G2_Buildtype_A.txt", "")],
    ("G2", "G3"): [("G2+ G3_Buildtype_BB.txt", "Preferred"), ("G2+ G3_Buildtype_B.txt", "Alternate")],
    ("G2", "G4"): [("G4+G2_Buildtype_S.txt", "IDLe")],
    ("G3", "G3"): [("G3+G3_Buildtype_C.txt", "Preferred"), ("G3+G3_Buildtype_CC.txt", "Alternate")],
    ("G3", "G4"): [("G4+G3_Buildtype_T.txt", "IDLe")],
    ("G4", "G4"): [("G4+G4_Buildtype_R.txt", "Preferred")],
    ("G2", "G2", "G2"): [("G2+G2+G2_Buildtype_D.txt", "")],
    ("G2", "G2", "G3"): [("G2+ G2+G3_Buildtype_E.txt", "")],
    ("G2", "G3", "G3"): [("G2+G3+G3_Buildtype_F.txt", "")],
    ("G3", "G3", "G3"): [("G3+ G3+ G3_Buildtype_GG.txt", "Preferred"), ("G3+ G3+ G3_Buildtype_G.txt", "Alternate")],
    ("G2", "G4", "G4"): [("G2+G4+G4_Buildtype_U.txt", "")],
    ("G4", "G4", "G4"): [("G4+G4+G4_Buildtype_RR.txt", "")],
    ("G3", "G4", "G4"): [("G3 + G4 + G4_Buildtype_TT.txt", "")],
    # ("G2","G2","G4"), ("G2","G3","G4"), ("G3","G3","G4") -> no template exists;
    # falls through to the "IDL Template not found" branch below.
}

# N2E confirmed to support these combinations — reuses the same file content/naming as the
# shared set, just from its own templates/N2E/IDL/ folder. Originally just C/CC/R; T, G, GG,
# RR, TT added this session (files copied into templates/N2E/IDL/ to match).
N2E_IDL_TEMPLATE_REGISTRY = {
    ("G3", "G3"): [("G3+G3_Buildtype_C.txt", "Preferred"), ("G3+G3_Buildtype_CC.txt", "Alternate")],
    ("G4", "G4"): [("G4+G4_Buildtype_R.txt", "Preferred")],
    ("G3", "G4"): [("G4+G3_Buildtype_T.txt", "IDLe")],
    ("G3", "G3", "G3"): [("G3+ G3+ G3_Buildtype_GG.txt", "Preferred"), ("G3+ G3+ G3_Buildtype_G.txt", "Alternate")],
    ("G4", "G4", "G4"): [("G4+G4+G4_Buildtype_RR.txt", "")],
    ("G3", "G4", "G4"): [("G3 + G4 + G4_Buildtype_TT.txt", "")],
    # every other combination -> "IDL Template not found" for N2E specifically, even though
    # MCA/CENM/NSB support it via the full registry above.
}

# MCA sites with a node ending in "F" (CRAN-styled node present, but NOT going through an actual
# CRAN rehome) use this SEPARATE registry entirely, replacing the standard IDL_TEMPLATE_REGISTRY
# for that site — confirmed. Every combo here includes exactly one G3 node (the "F" node).
# Filenames confirmed against actual GitHub uploads.
# Confirmed real rule (uploaded IDL_Connections.xlsx, "CRAN TRACKING" tab) — CRAN Build
# Type is fully derivable from the CIQ, no ambiguity/dropdown needed. Each non-hub node
# contributes (generation, mode) where mode is "M" (MMBB/TMBB — dual LTE+5G identity) or
# "L" (SMBB — single identity), sourced from the CIQ's own "BBU Mode" column. The hub is
# the node ending in "F" (always G3-class hardware), excluded from the signature — its own
# connection type (RPM coax vs SM fiber jumper) is a separate, independent choice (the "-1"
# suffix), not derivable from the CIQ. Keyed by the SORTED tuple of (gen, mode) pairs, since
# the reference sheet's own column order isn't a reliable real-node ordering to match against.
CRAN_BUILD_TYPE_REGISTRY = {
    (("G2", "M"),): "L-1",
    (("G2", "M"), ("G2", "M")): "L-2",
    (("G2", "L"), ("G2", "M")): "L-2B",
    (("G2", "M"), ("G2", "M"), ("G2", "M")): "L-3B",
    (("G3", "M"),): "L-4",
    (("G2", "M"), ("G3", "M")): "L-5",
    (("G2", "L"), ("G3", "M")): "L-5B",
    (("G3", "M"), ("G3", "M")): "L-6",
    (("G2", "M"), ("G2", "M"), ("G2", "M"), ("G2", "M")): "L-8",
    (("G4", "M"),): "L-9",
    (("G2", "L"), ("G4", "M")): "L-10",
    (("G4", "M"), ("G4", "M")): "L-11",
    (("G2", "L"), ("G4", "M"), ("G4", "M")): "L-12",
}


def cran_build_type(ciq_wb, mm_objs):
    """Returns (base_build_type_letter_or_None, hub_row_or_None). The hub node is identified
    by its name ending in "F" (same convention already used elsewhere to detect a CRAN-styled
    node present in an MCA site) — confirmed to be excluded from the tech signature, since its
    own hub-cable type is chosen separately (the "-1" variant), not part of what determines
    which Build Type letter applies."""
    hub_row = None
    regular_pairs = []
    for row in mm_objs:
        name = row.get("Node to be built as")
        if str(name or "").strip().upper().endswith("F"):
            hub_row = row
            continue
        gen = get_node_generation(ciq_wb, row)
        if not gen:
            continue
        bbu_mode = str(row.get("BBU Mode") or "").strip().upper()
        mode = "L" if bbu_mode == "SMBB" else "M"
        regular_pairs.append((gen, mode))
    signature = tuple(sorted(regular_pairs))
    return CRAN_BUILD_TYPE_REGISTRY.get(signature), hub_row


# Confirmed real reference tables (uploaded IDL_Connections.xlsx, "CRAN" tab) — one entry per
# BASE build type (the "-1" suffix only ever changes the hub cable, handled separately below).
# CRAN_SLOT_CABLE_REFERENCE: (generation, position-index-among-same-generation) -> cable
# description for that node's DIRECT hub connection. A generation/index NOT present here means
# that node connects via IDL instead (see CRAN_IDL_CABLE_REFERENCE) — confirmed real pattern,
# e.g. L-10's SMBB G2 node has no direct hub slot (all "NA" in the sheet), reaching the hub only
# via the G2+G4 IDL connection.
CRAN_SLOT_CABLE_REFERENCE = {
    "L-1": {("G2", 1): "RPM 777 811 + 8300 DaFi"},
    "L-2": {("G2", 1): "RPM 777 811 + 8300 DaFi", ("G2", 2): "RPM 777 811 + 8300 DaFi"},
    "L-2B": {("G2", 1): "RPM 777 811 + 8300 DaFi"},
    "L-3B": {("G2", 1): "RPM 777 811 + 8300 DaFi", ("G2", 2): "RPM 777 811 + 8300 DaFi", ("G2", 3): "RPM 777 811 + 8300 DaFi"},
    "L-4": {("G3", 1): "Approved MM\nfiber jumper + RDH 102 65/1 + 8300 DaFi"},
    "L-5": {("G2", 1): "RPM 777 811 + 8300 DaFi", ("G3", 1): "Approved MM\nfiber jumper + RDH 102 65/1 + 8300 DaFi"},
    "L-5B": {("G3", 1): "Approved MM\nfiber jumper + RDH 102 65/1 + 8300 DaFi"},
    "L-6": {("G3", 1): "Approved MM\nfiber jumper + RDH 102 65/1 + 8300 DaFi", ("G3", 2): "Approved MM\nfiber jumper + RDH 102 65/1 + 8300 DaFi"},
    "L-8": {("G2", 1): "RPM 777 811 + 8300 DaFi", ("G2", 2): "RPM 777 811 + 8300 DaFi", ("G2", 3): "RPM 777 811 + 8300 DaFi", ("G2", 4): "RPM 777 811 + 8300 DaFi"},
    "L-9": {("G4", 1): "Approved MM\nfiber jumper + RDH 102 65/1 + 8300 DaFi"},
    "L-10": {("G4", 1): "Approved MM\nfiber jumper + RDH 102 65/1 + 8300 DaFi"},
    "L-11": {("G4", 1): "Approved MM\nfiber jumper + RDH 102 65/1 + 8300 DaFi", ("G4", 2): "Approved MM\nfiber jumper + RDH 102 65/1 + 8300 DaFi"},
    "L-12": {("G4", 1): "Approved MM\nfiber jumper + RDH 102 65/1 + 8300 DaFi", ("G4", 2): "Approved MM\nfiber jumper + RDH 102 65/1 + 8300 DaFi"},
}

# Hub's own cable, keyed by (base_build_type, is_dash1_variant) — the ONLY thing the "-1"
# suffix actually changes.
CRAN_HUB_CABLE_REFERENCE = {
    ("L-1", False): "RPM 777 052", ("L-1", True): "RDH 102 75/3 + Approved SM \nfiber jumper ",
    ("L-2", False): "RPM 777 052", ("L-2", True): "RDH 102 75/3 + Approved SM \nfiber jumper ",
    ("L-2B", False): "RPM 777 052", ("L-2B", True): "RDH 102 75/3 + Approved SM \nfiber jumper ",
    ("L-3B", False): "RPM 777 052", ("L-3B", True): "RDH 102 75/3 + Approved SM \nfiber jumper ",
    ("L-4", False): "RPM 777 052", ("L-4", True): "RDH 102 75/3 + Approved SM \nfiber jumper ",
    ("L-5", False): "RPM 777 052", ("L-5", True): "RDH 102 75/3 + Approved SM \nfiber jumper ",
    ("L-5B", False): "RPM 777 052", ("L-5B", True): "RDH 102 75/3 + Approved SM \nfiber jumper ",
    ("L-6", False): "RPM 777 052", ("L-6", True): "RDH 102 75/3 + Approved SM \nfiber jumper ",
    ("L-8", False): "RPM 777 052", ("L-8", True): "RDH 102 75/3 + Approved SM \nfiber jumper ",
    ("L-9", False): "RPM 777 052", ("L-9", True): "RDH 102 75/3 + Approved SM \nfiber jumper ",
    ("L-10", False): "RPM 777 052", ("L-10", True): "RDH 102 75/3 + Approved SM \nfiber jumper ",
    ("L-11", False): "RPM 777 052", ("L-11", True): "RDH 102 75/3 + Approved SM \nfiber jumper ",
    ("L-12", False): "RPM 777 052", ("L-12", True): "RDH 102 75/3 + Approved SM \nfiber jumper ",
}

# IDLe/IDLy cross-node connections, same shape as MCA's IDL_CABLE_REFERENCE — only build
# types that actually need one appear here (most CRAN combos don't).
CRAN_IDL_CABLE_REFERENCE = {
    "L-2": {"idle": [("RPM 777 417", "G2(1)+G2(2)")], "idly": []},
    "L-2B": {"idle": [("RPM 777 417", "G2(1)+G2(2)")], "idly": []},
    "L-5": {"idle": [], "idly": [("RPM 777 098", "G2+G3")]},
    "L-5B": {"idle": [], "idly": [("RPM 777 098", "G2+G3")]},
    "L-6": {"idle": [("RPM 777 053", "G3(1)+G3(2)")], "idly": []},
    "L-10": {"idle": [("RPM 777 543", "G2+G4")], "idly": []},
    "L-11": {"idle": [("RPM 777 052", "G4(1)+G4(1)")], "idly": []},
    "L-12": {"idle": [("RPM 777 543", "G2+G4(1)"), ("RPM 777 543", "G2+G4(2)")], "idly": []},
}


def cran_idl_cable_lines(ciq_wb, mm_objs, base_build_type):
    """Same substitution mechanism as idl_cable_lines_for_build_type() (MCA), reused here
    with CRAN's own reference table."""
    entry = CRAN_IDL_CABLE_REFERENCE.get(base_build_type)
    if not entry:
        return [], []
    nodes_by_gen = {}
    for row in mm_objs:
        name = row.get("Node to be built as")
        if str(name or "").strip().upper().endswith("F"):
            continue  # hub excluded — it's never part of an inter-node IDL connection
        gen = get_node_generation(ciq_wb, row)
        if gen:
            nodes_by_gen.setdefault(gen, []).append(row)

    def substitute(pattern):
        parts, used = [], {}
        for gen, idx in re.findall(r"(G\d)(?:\((\d+)\))?", pattern):
            cands = nodes_by_gen.get(gen, [])
            pos = (int(idx) - 1) if idx else used.get(gen, 0)
            used[gen] = used.get(gen, 0) + 1
            parts.append(_idl_cable_node_label(ciq_wb, cands[pos]) if 0 <= pos < len(cands) else f"{gen} (node not found)")
        return " + ".join(parts)

    idle_lines = [f"IDLe : {part} : {substitute(combo)}" for part, combo in entry["idle"]]
    idly_lines = [f"IDLy : {part} : {substitute(combo)}" for part, combo in entry["idly"]]
    return idle_lines, idly_lines


def cran_idl_cable_columns(ciq_wb, mm_objs, base_build_type):
    """(C, D, E) column version of cran_idl_cable_lines(), for the .xlsm write — see
    _idl_cable_columns() for the combination rule. Returns (idle_cols, idly_cols)."""
    entry = CRAN_IDL_CABLE_REFERENCE.get(base_build_type)
    if not entry:
        return ("", "", ""), ("", "", "")
    nodes_by_gen = {}
    for row in mm_objs:
        name = row.get("Node to be built as")
        if str(name or "").strip().upper().endswith("F"):
            continue
        gen = get_node_generation(ciq_wb, row)
        if gen:
            nodes_by_gen.setdefault(gen, []).append(row)
    return _idl_cable_columns(ciq_wb, nodes_by_gen, entry["idle"]), _idl_cable_columns(ciq_wb, nodes_by_gen, entry["idly"], third_col=True)


def _cran_slot_port_data(ciq_wb, mm_objs, base_build_type, is_dash1, sidehaul_rows):
    """Shared matching logic for cran_slot_port_lines() (display text) and
    cran_slot_port_rows() (structured, for .xlsm columns) — returns
    (switch_ids, [{"slot_port", "cable", "node_display"}, ...])."""
    sidehaul_by_node = {str(r["node_id"]).strip().upper(): r for r in sidehaul_rows if r.get("node_id")}

    hub_row = None
    nodes_by_gen = {}
    for row in mm_objs:
        name = row.get("Node to be built as")
        if str(name or "").strip().upper().endswith("F"):
            hub_row = row
            continue
        gen = get_node_generation(ciq_wb, row)
        if gen:
            nodes_by_gen.setdefault(gen, []).append(row)

    switch_ids = sorted({str(r["switch_id"]) for r in sidehaul_rows if r.get("switch_id")})

    def sidehaul_match(row):
        primary = str(row.get("Node to be built as") or "").strip().upper()
        e_name = str(row.get("eNodeB Name") or "").strip().upper()
        g_name = str(row.get("gNodeB Name") or "").strip().upper()
        for key in (primary, e_name, g_name):
            if key and key in sidehaul_by_node:
                return sidehaul_by_node[key]
        return None

    entries = []
    for gen, rows in nodes_by_gen.items():
        for idx, row in enumerate(rows, start=1):
            cable = CRAN_SLOT_CABLE_REFERENCE.get(base_build_type, {}).get((gen, idx))
            if not cable:
                continue  # reaches the hub via IDL instead, not a direct slot connection
            cable = re.sub(r"\s+", " ", cable).strip()
            sh = sidehaul_match(row)
            slot_port = sh["slot_port"] if sh else "Slot/Port not found in Sidehaul Info"
            node_display = _idl_cable_node_label(ciq_wb, row)
            # Strip the trailing "(hw)" tag — confirmed format for this line uses (P)/(S)
            # only, hardware isn't shown (unlike the IDLe/IDLy line, which always includes it).
            node_display = re.sub(r"\([^()]*\)$", "", node_display).rstrip()
            entries.append({"slot_port": slot_port, "cable": cable, "node_display": node_display})

    if hub_row:
        hub_cable = CRAN_HUB_CABLE_REFERENCE.get((base_build_type, bool(is_dash1)))
        hub_cable = re.sub(r"\s+", " ", hub_cable).strip() if hub_cable else hub_cable
        sh = sidehaul_match(hub_row)
        slot_port = sh["slot_port"] if sh else "Slot/Port not found in Sidehaul Info"
        entries.append({"slot_port": slot_port, "cable": hub_cable, "node_display": hub_row.get("Node to be built as")})

    return switch_ids, entries


def cran_slot_port_lines(ciq_wb, mm_objs, base_build_type, is_dash1, sidehaul_rows):
    """Builds the Switch + Slot/Port report DISPLAY lines for a CRAN site — for the report
    text / on-screen caption. See _cran_slot_port_data() for the matching rule. Returns
    (switch_lines, slot_port_lines)."""
    switch_ids, entries = _cran_slot_port_data(ciq_wb, mm_objs, base_build_type, is_dash1, sidehaul_rows)
    switch_lines = [f"Switch : {sid}" for sid in switch_ids]
    slot_port_lines = [f"{e['slot_port']} -> {e['cable']} -> {e['node_display']}" for e in entries]
    return switch_lines, slot_port_lines


def cran_slot_port_rows(ciq_wb, mm_objs, base_build_type, is_dash1, sidehaul_rows):
    """Structured version of cran_slot_port_lines(), for the .xlsm write — the template has
    real Slot/Port and Cable P/N columns (rows 23/24, then overflow rows 25-39), not a place
    to dump pre-formatted "->" display strings. Returns (switch_ids, [{"slot_port", "cable",
    "node_display"}, ...])."""
    return _cran_slot_port_data(ciq_wb, mm_objs, base_build_type, is_dash1, sidehaul_rows)


MCA_CRAN_IDL_REGISTRY = {
    ("G2", "G3"): [("G2 BBU+G3 BBU  Dafi 6673 Connections Build Type (L-1) and (L-1-1).txt", "")],
    ("G2", "G2", "G3"): [
        ("G2 BBU+G2 BBU+G3 BBU  Dafi 6673 Connections Build Type (L-2) and (L-2-1).txt", "L-2"),
        ("G2 BBU+G2 BBU+G3 BBU  Dafi 6673 Connections Build Type (L-2B) and (L-2B-1).txt", "L-2B"),
    ],
    ("G2", "G2", "G2", "G3"): [
        ("G2 BBU+G2 BBU+G2 BBU+G3 BBU  Dafi 6673 Connections Build Type (L-3) and (L-3-1).txt", "L-3"),
        ("G2 BBU+G2 BBU+G2 BBU+G3 BBU  Dafi 6673 Connections Build Type (L-3B) and (L-3B-1).txt", "L-3B"),
    ],
    ("G2", "G3", "G3"): [("G2 BBU + G3 BBU + G3 BBU  6673 Connections Via Dafi Build Type (L-5B) and (L-5B-1).txt", "")],
    ("G2", "G3", "G4"): [("G2 BBU+G4 BBU+ G3 BBU 6673 Connections Via Dafi Build Type (L-10) and (L-10-1)].txt", "")],
    ("G3", "G4", "G4"): [("G4 BBU+G4 BBU+G3 BBU 6673 Connections Via Dafi Build Type (L-11) and (L-11-1)].txt", "")],
    ("G2", "G3", "G4", "G4"): [("G2 BBU + G4 BBU + G4 BBU + G3 BBU 6673 Connections Via Dafi Build Type (L-12) and (L-12-1).txt", "")],
}

IDL_SUFFIX_CANDIDATES = {
    "NODE_ID": ["NODE_ID", "Node_ID", "BBU_Node_ID"],
    "5G_NODE_ID": ["5G_NODE_ID", "5G_NodeID"],
    "GNB_ID": ["NODE_GNB_ID", "GNBID"],
    # Confirmed casing gap: L-5B's own template uses "Node_eNBId" for one node but
    # "NODE_eNBId" for another (same conceptual field) — exact-case matching missed the
    # all-caps form entirely. Added as an extra candidate; purely additive.
    "eNBId": ["Node_eNBId", "NODE_eNBId", "ENBID", "BBU_ENBID"],
}


def _ordinal(n):
    return {1: "1st", 2: "2nd", 3: "3rd"}.get(n, f"{n}th")


def get_node_generation(ciq_wb, row):
    """Board generation (G2/G3/G4) for one Mixed Mode Info row, per confirmed rule."""
    has_enb = is_populated(row.get("eNBId"))
    has_gnb = is_populated(row.get("gNBId"))
    e_name, g_name = row.get("eNodeB Name"), row.get("gNodeB Name")

    du_type = None
    if has_enb and has_gnb:  # MMBB/TMBB — either tab carries it
        r = find_row_by_name(ciq_wb, "eNB Info", "eNodeB Name", e_name)
        du_type = r.get("DU type") if r else None
        if not is_populated(du_type):
            r = find_row_by_name(ciq_wb, "gNB Info", "gNodeB Name", g_name)
            du_type = r.get("DU type") if r else None
    elif has_enb:  # LTE standalone
        r = find_row_by_name(ciq_wb, "eNB Info", "eNodeB Name", e_name)
        du_type = r.get("DU type") if r else None
    elif has_gnb:  # 5G standalone
        r = find_row_by_name(ciq_wb, "gNB Info", "gNodeB Name", g_name)
        du_type = r.get("DU type") if r else None

    if not is_populated(du_type):
        return None
    return DU_TYPE_TO_GEN.get(str(du_type).strip())


def _idl_node_values(row):
    return {
        "NODE_ID": row.get("Node to be built as"),
        "5G_NODE_ID": row.get("gNodeB Name"),
        "GNB_ID": row.get("gNBId"),
        "eNBId": row.get("eNBId"),
    }


def fill_idl_template(template_text, node_slots, summary_rows, log, template_name):
    """node_slots: list of (candidate_prefixes, row). For each node/concept, tries every
    candidate-prefix x suffix-variant combination and fills whichever placeholder actually
    exists in this template — templates only get the placeholders they actually reference.

    Divider/hash handling (per confirmed instruction: placeholders are filled ALONG WITH their
    surrounding hashes, leaving just the bare value):
    - 4-hash forms like ####G2_NODE_ID#### are replaced entirely with the value.
    - Bare ##<prefix>_NODE## tokens are replaced entirely with the node's ID."""
    tpl = template_text
    for prefixes, row in node_slots:
        values = _idl_node_values(row)
        node_label = row.get("Node to be built as")
        for concept, value in values.items():
            # Confirmed real bug: without stopping at the first candidate that actually exists
            # in this template, a node with multiple candidate prefixes could keep matching and
            # overwriting placeholders meant for a DIFFERENT node of the same generation (e.g. one
            # G4 node's absolute-position candidate colliding with another G4 node's group-rank
            # candidate) — one node's data silently ended up in two slots, the other node's data
            # nowhere. Stop as soon as one real placeholder is found and handled for this node.
            matched = False
            for prefix in prefixes:
                if matched:
                    break
                for suffix in IDL_SUFFIX_CANDIDATES[concept]:
                    placeholder = f"##{prefix}_{suffix}##"
                    divider_form = f"####{prefix}_{suffix}####"
                    # Confirmed real bug: the 4-hash and 2-hash forms can both exist as
                    # genuinely separate occurrences for the SAME prefix+suffix (e.g. one
                    # template has a "####G2_NODE_ID####" section header AND many
                    # "##G2_NODE_ID##" references throughout its cmedit body) — breaking
                    # immediately after the first one matched left the other unfilled. Check
                    # and replace both before deciding whether to move on.
                    found_any = False
                    if divider_form in tpl and is_populated(value):
                        tpl = tpl.replace(divider_form, str(value))
                        found_any = True
                    if placeholder in tpl:
                        if is_populated(value):
                            tpl = tpl.replace(placeholder, str(value))
                            summary_rows.append({"Item": f"IDL · {node_label} · {placeholder}", "Source": template_name, "Value": value, "Note": ""})
                            log(f"✓ IDL {template_name}: {placeholder} -> {value}")
                        else:
                            summary_rows.append({"Item": f"IDL · {node_label} · {placeholder}", "Source": template_name, "Value": "NOT FOUND", "Note": ""})
                            log(f"✗ IDL {template_name}: {placeholder} -> NOT FOUND")
                        found_any = True
                    if found_any:
                        matched = True
                        break
        # bare "_NODE" tokens (e.g. ##1st_G3_NODE##) — filled entirely with the node's ID.
        # Same early-exit reasoning: stop at the first candidate prefix that actually exists.
        # Confirmed real bug: L-5B has a 4-hash section-header form (####2nd_G3_NODE####) for
        # this same bare-NODE case, which the 2-hash-only check never handled — and since
        # "##2nd_G3_NODE##" is literally a SUBSTRING of "####2nd_G3_NODE####", replacing just
        # the substring left the outer "##" pair behind (e.g. "####2nd_G3_NODE####" became
        # "##B##" instead of "B"). Check the 4-hash divider form first, exactly like the
        # concept-suffix loop above already does.
        if is_populated(node_label):
            for prefix in prefixes:
                # Both hash-forms can exist as genuinely SEPARATE occurrences for the same
                # prefix (e.g. L-5B has both a "####2nd_G3_NODE####" section header AND a
                # separate "##2nd_G3_NODE##" reference inside the cmedit body) — check and
                # replace both before deciding whether this prefix matched anything, so the
                # first one found doesn't short-circuit the second.
                node_divider = f"##{prefix}_NODE##"
                node_divider_4hash = f"####{prefix}_NODE####"
                found_any = False
                if node_divider_4hash in tpl:
                    tpl = tpl.replace(node_divider_4hash, str(node_label))
                    found_any = True
                if node_divider in tpl:
                    tpl = tpl.replace(node_divider, str(node_label))
                    found_any = True
                if found_any:
                    break
    return tpl


def generate_idl_connections(ciq_wb, mm_objs, user_id, date_str, log, template_dir=None, registry=None):
    """Returns (outputs, summary_rows, scope_lines) — same shape as the other generate_* add-ons.
    Shared by MCA / CENM / NSB (default template_dir/registry). N2E passes its own restricted
    registry (only G3+G3, G4+G4) and its own templates/N2E/IDL/ folder. No-ops for single-BBU sites."""
    template_dir = template_dir if template_dir is not None else TDIR_IDL
    registry = registry if registry is not None else IDL_TEMPLATE_REGISTRY
    outputs, summary_rows, scope_lines = [], [], []

    if len(mm_objs) < 2:
        return outputs, summary_rows, scope_lines

    nodes = [{"row": row, "gen": get_node_generation(ciq_wb, row)} for row in mm_objs]

    unresolved = [n["row"].get("Node to be built as") for n in nodes if not n["gen"]]
    if unresolved:
        note = f"Could not determine board generation (DU type) for: {', '.join(str(u) for u in unresolved)}"
        summary_rows.append({"Item": "IDL Connections", "Source": "DU type lookup", "Value": "NOT FOUND", "Note": note})
        log(f"✗ IDL Connections: {note}")
        scope_lines.append(f"IDL Connections:\tCould not determine board generation\t{', '.join(str(u) for u in unresolved)}")
        return outputs, summary_rows, scope_lines

    combo = tuple(sorted(n["gen"] for n in nodes))
    matches = registry.get(combo)

    if not matches:
        summary_rows.append({"Item": "IDL Connections", "Source": f"combination {'+'.join(combo)}", "Value": "IDL Template not found", "Note": ""})
        log(f"✗ IDL Connections: IDL Template not found for combination {'+'.join(combo)}")
        scope_lines.append(f"IDL Connections:\tIDL Template not found\t{'+'.join(combo)}")
        return outputs, summary_rows, scope_lines

    # Confirmed real bug (two parts) — every CRAN IDL template treats the "F"-ending hub node
    # as the LAST/highest-numbered node in its generation, and numbers all OTHER generations in
    # site order (e.g. G2 before G3 before G4) — never based on raw Mixed Mode Info row order.
    # A real CIQ can list a G3 node before its G2 node, or a hub before its peers, which shifted
    # absolute-position numbers and caused two different nodes to compute the SAME candidate
    # prefix (confirmed: e.g. an L-5B site with the hub or relay out of "expected" order, an
    # L-11 site with the hub listed first). Canonicalizing the order before computing any
    # ordinal — by generation, then hub-last within a generation — removes the ambiguity at
    # its source; a no-op for single-node generations and non-CRAN sites with no F-node.
    GEN_ORDER_RANK = {"G2": 0, "G3": 1, "G4": 2}

    def _idl_sort_key(n):
        is_f = str(n["row"].get("Node to be built as") or "").strip().upper().endswith("F")
        return (GEN_ORDER_RANK.get(n["gen"], 99), is_f)

    nodes = sorted(nodes, key=_idl_sort_key)

    gen_counts = {}
    for n in nodes:
        gen_counts[n["gen"]] = gen_counts.get(n["gen"], 0) + 1
    group_seen = {}
    for i, n in enumerate(nodes, start=1):
        g = n["gen"]
        group_seen[g] = group_seen.get(g, 0) + 1
        if gen_counts[g] > 1:
            # Confirmed real bug: different templates use different conventions for a
            # multi-node generation — some number by absolute site position (e.g. L-5B's two
            # G3 nodes: "2nd_G3"/"3rd_G3"), others by rank within the generation itself (e.g.
            # L-11's two G4 nodes: "1st_G4"/"2nd_G4", restarting regardless of the hub's
            # position). Trying absolute position first broke the latter case whenever another
            # generation's node shifted everyone's position. Trying the group-relative rank
            # first, falling back to absolute position, satisfies both: whichever a specific
            # template doesn't actually use simply doesn't match anything and is skipped, and
            # the early-exit in fill_idl_template stops each node from later reusing whatever
            # candidate the correct node was supposed to claim.
            candidates = [f"{_ordinal(group_seen[g])}_{g}", f"{_ordinal(i)}_{g}"]
        else:
            candidates = [f"{_ordinal(i)}_{g}", f"{_ordinal(group_seen[g])}_{g}"]
            candidates.append(g)
            # Confirmed gap: some templates label the sole node of a generation by "how many
            # other nodes are at the site" rather than its absolute position — e.g. the lone G3
            # in a 3xG2+1xG3 site is written as "3rd_G3" (3 = the other 3 nodes), not "4th_G3"
            # (its actual position). Purely additive: only ever adds a candidate, never removes
            # or changes one, so this can't affect any placeholder that already fills correctly.
            if len(nodes) > 1:
                candidates.append(f"{_ordinal(len(nodes) - 1)}_{g}")
        n["prefixes"] = list(dict.fromkeys(candidates))  # dedupe, preserve order

    site_id = mm_objs[0].get("Node to be built as", "site")
    node_slots = [(n["prefixes"], n["row"]) for n in nodes]

    for fname, variant in matches:
        tpl_path = template_dir / fname
        if not tpl_path.exists() and template_dir != TDIR_IDL:
            # Confirmed rule (N2E specifically): only C/CC/R have dedicated files in N2E's
            # own folder — everything else (T, G, GG, RR, TT, etc.) falls back to the same
            # shared templates/IDL/ folder MCA/CENM/NSB use, rather than needing its own copy.
            tpl_path = TDIR_IDL / fname
        if not tpl_path.exists():
            summary_rows.append({"Item": "IDL Connections", "Source": f"template {fname}", "Value": "NOT FOUND", "Note": f"expected file not in {template_dir}/ or {TDIR_IDL}/: {fname}"})
            log(f"✗ IDL Connections: template file not found: {fname}")
            scope_lines.append(f"IDL Connections:\ttemplate file missing from repo\t{fname}")
            continue
        tpl_text = tpl_path.read_text(encoding="utf-8")
        filled = fill_idl_template(tpl_text, node_slots, summary_rows, log, fname)
        label = "+".join(combo) + (f"_{variant}" if variant else "")
        outputs.append((f"{site_id}_IDL_Connections_{label}.txt", filled))
        scope_lines.append(f"IDL Connections:\t{'+'.join(combo)}" + (f" ({variant})" if variant else "") + f"\t{fname}")

    return outputs, summary_rows, scope_lines


# ============================================================
# NGS CHECKS (all scopes — CIQ-only, no template output)
#
# Confirmed logic:
#  - Trigger: site has 2+ BBUs (same as IDL Connections).
#  - Data source: eUtran Parameters tab, "Co-Located Technology Cell" column (comma-separated
#    list of cell names sharing the same physical radio as that row's cell).
#  - Detection: build a cellname -> owning-node map (from eUtran Parameters' EutranCellFDDId and
#    5G Info's cell id, both matched back to a node via eNBId/gNBId), then for every pair of
#    different nodes at the site, check whether a cell on Node A references a cell on Node B
#    AND a cell on Node B references a cell on Node A (bidirectional, per your original
#    description — "mapped for the BBU2 sectors... & vice versa"). If both directions are
#    confirmed, the two nodes are sharing a physical radio -> NGS applies.
#  - No template file is generated — this only ever contributes a line to the Scope of Work
#    (and, by extension, the Checks Performed panel, since that reads Scope of Work lines).
# ============================================================

def _ngs_build_cell_node_map(ciq_wb, mm_objs):
    """cell name (LTE or 5G) -> owning node's 'Node to be built as', via eNBId/gNBId match."""
    enbid_to_node = {str(r.get("eNBId")).strip(): r.get("Node to be built as") for r in mm_objs if is_populated(r.get("eNBId"))}
    gnbid_to_node = {str(r.get("gNBId")).strip(): r.get("Node to be built as") for r in mm_objs if is_populated(r.get("gNBId"))}

    cell_to_node = {}
    if "eUtran Parameters" in ciq_wb.sheetnames:
        for row in sheet_objs(ciq_wb["eUtran Parameters"]):
            cell_name = row.get("EutranCellFDDId")
            node = enbid_to_node.get(str(row.get("eNBId")).strip())
            if cell_name and node:
                cell_to_node[str(cell_name).strip()] = node
    if "5G Info" in ciq_wb.sheetnames:
        for row in sheet_objs(ciq_wb["5G Info"]):
            cell_name = row.get("NRCellDU") or row.get("gNodeB Name")
            node = gnbid_to_node.get(str(row.get("gNBId")).strip())
            if cell_name and node:
                cell_to_node[str(cell_name).strip()] = node
    return cell_to_node


def _ngs_cell_band(cell_name):
    """Band label for either an LTE or a 5G cell name, whichever pattern matches."""
    label, _sector = lte_band_label(cell_name)
    if label:
        return label
    label, _sector = nr_band_label(cell_name)
    return label


TDIR_NGS = Path(__file__).parent / "templates" / "NGS"
TPL_NGS_LTE_LTE = TDIR_NGS / "LTE-LTE_NGS_Template.txt"
TPL_NGS_LTE_5G = TDIR_NGS / "LTE-5G_NGS_Templates.txt"


def _carrier_sort_key(carrier_str):
    """'2C' -> 2, '5C BWE' -> 5, '3C' -> 3 — leading number only, ignoring BWE suffix."""
    m = re.match(r'(\d+)', str(carrier_str or '').strip())
    return int(m.group(1)) if m else 999


def _xmu_trunk_for_port(port_num):
    """Confirmed splitter groups: XMU trunk Port 1 powers 13-16, Port 2 powers 9-12, Port 3 powers 4-7."""
    if 13 <= port_num <= 16:
        return 1
    if 9 <= port_num <= 12:
        return 2
    if 4 <= port_num <= 7:
        return 3
    return None


def _lookup_xmu_bbu_port(ciq_wb, node, trunk_num, info_sheet_name):
    """eNB Info / gNB Info's '1st XMU Port N' column tells us where that XMU trunk port
    actually connects on the BBU — that's the real RiPort value."""
    if info_sheet_name not in ciq_wb.sheetnames:
        return None
    name_col = "eNodeB Name" if info_sheet_name == "eNB Info" else "gNodeB Name"
    for row in sheet_objs(ciq_wb[info_sheet_name]):
        if str(row.get(name_col, "")).strip().upper() == str(node).strip().upper():
            return row.get(f"1st XMU Port {trunk_num}")
    return None


def _resolve_riport(ciq_wb, node, raw_port_value, info_sheet_name):
    """Confirmed rule: a numeric DUS/XMU Port means the radio is on an XMU — resolve via the
    splitter group + eNB/gNB Info's XMU-trunk-to-BBU-port mapping. A non-numeric (letter) value
    is a direct baseband connection — use it as the RiPort value directly."""
    if not is_populated(raw_port_value):
        return None
    raw = str(raw_port_value).strip()
    if raw.isdigit():
        trunk = _xmu_trunk_for_port(int(raw))
        if trunk is None:
            return None
        return _lookup_xmu_bbu_port(ciq_wb, node, trunk, info_sheet_name)
    return raw


def _lte_cluster_rru_riport(ciq_wb, cell_to_node, node, seed_cell):
    """seed_cell: one cell already CONFIRMED to be co-located with the NGS partner (from the
    detection pass), belonging to `node`. Finds every other same-node cell sharing that exact
    physical radio (via seed_cell's own Co-Located Technology Cell list) and picks the lowest
    carrier among that cluster — not a blind node+sector scan, since a site can have multiple
    unrelated LTE carriers sharing the same sector letter that have nothing to do with this
    specific shared radio (confirmed via SCL04291's standalone 1C carrier)."""
    if "eUtran Parameters" not in ciq_wb.sheetnames:
        return None, None
    all_rows = {str(row.get("EutranCellFDDId")).strip(): row for row in sheet_objs(ciq_wb["eUtran Parameters"]) if is_populated(row.get("EutranCellFDDId"))}
    seed_row = all_rows.get(str(seed_cell).strip())
    if not seed_row:
        return None, None
    cluster = {str(seed_cell).strip()}
    colo_raw = seed_row.get("Co-Located Technology Cell")
    if is_populated(colo_raw) and str(colo_raw).strip().upper() not in ("NA", "N/A"):
        for c in str(colo_raw).split(","):
            c = c.strip()
            if c in all_rows and cell_to_node.get(c) == node:
                cluster.add(c)
    candidates = [all_rows[c] for c in cluster]
    candidates.sort(key=lambda r: _carrier_sort_key(r.get("Carrier")))
    winner = candidates[0]
    sector_id = str(winner.get("sectorId", "")).strip()
    rru = sector_id.split("_")[0] if sector_id else None
    riport = _resolve_riport(ciq_wb, node, winner.get("DUS / XMU Port"), "eNB Info")
    return (rru or None), riport


def _5g_cell_riport(ciq_wb, node, seed_cell):
    """seed_cell: the exact 5G cell name already confirmed via the co-location reference —
    no need to scan by sector at all, we know precisely which cell. Reads Port 1-4 (whichever
    is populated) and resolves the same XMU-or-direct way as the LTE side."""
    if "5G Info" not in ciq_wb.sheetnames:
        return None
    for row in sheet_objs(ciq_wb["5G Info"]):
        if str(row.get("NRCellDU", "")).strip() != str(seed_cell).strip():
            continue
        for col in ("Port 1", "Port 2", "Port 3", "Port 4"):
            val = row.get(col)
            if is_populated(val):
                return _resolve_riport(ciq_wb, node, val, "gNB Info")
    return None


def _node_priority(ciq_wb, cell_to_node, node, is_lte):
    """Confirmed: Radio Port = DATA1 -> syncNodePriority 1, DATA2 -> priority 2. Read from any
    of the node's own involved cells (LTE: eUtran Parameters; 5G: 5G Info)."""
    sheet_name = "eUtran Parameters" if is_lte else "5G Info"
    cell_col = "EutranCellFDDId" if is_lte else "NRCellDU"
    if sheet_name not in ciq_wb.sheetnames:
        return None
    for row in sheet_objs(ciq_wb[sheet_name]):
        cell = row.get(cell_col)
        if not is_populated(cell) or cell_to_node.get(str(cell).strip()) != node:
            continue
        radio_port = str(row.get("Radio Port", "")).strip().upper()
        if radio_port == "DATA1":
            return "1"
        if radio_port == "DATA2":
            return "2"
    return None


def _dedupe_preserve_order(values):
    """A,A,A -> [A]. A,B,B -> [A,B]. Confirmed: extras beyond the unique count stay unfilled."""
    seen, out = set(), []
    for v in values:
        if v and v not in seen:
            seen.add(v)
            out.append(v)
    return out


NGS_VARIANT_RE = re.compile(
    r'(?P<label>WHen Only Port ([12]) link is used:[ \t]*\r?\n)?'
    r'(?P<stmt>cmedit create SubNetwork=ONRM_ROOT_MO,MeContext=##Site_ID_[12]##[^\n]*?'
    r'NodeGroupSyncMember=1 NodeGroupSyncMemberid=1;[^\n]*?syncRiPortCandidate=\[.*?\])',
    re.DOTALL
)


def _strip_ngs_variant_blocks(tpl_text, site1_link_count, site2_link_count):
    """Each site section has 3 alternative cmedit-create blocks (default = 3 links, Port-1-only,
    Port-2-only). Keep only the one matching that site's actual unique RiPort count; remove the
    other two (including their 'WHen Only Port N link is used:' label, if any)."""
    matches = list(NGS_VARIANT_RE.finditer(tpl_text))
    to_remove = []
    for m in matches:
        label = m.group("label") or ""
        stmt = m.group("stmt")
        site_num = 1 if "site1" in stmt else (2 if "site2" in stmt else None)
        if site_num is None:
            continue
        target = site1_link_count if site_num == 1 else site2_link_count
        variant = 1 if "Port 1" in label else (2 if "Port 2" in label else 3)
        if variant != target:
            to_remove.append((m.start(), m.end()))
    for start, end in sorted(to_remove, reverse=True):
        tpl_text = tpl_text[:start] + tpl_text[end:]
    return tpl_text


def _fill_ngs_site_block(tpl_text, slot, site_id, rru_list, riport_list, priority):
    """slot: 1 or 2. Fills this site's Site_ID/RRU/Riport/Priority placeholders.
    RiPort slots beyond the actual unique count are blanked (safe — their cmedit block was
    already removed by variant-stripping, so they only remain in the header reference list).
    RRU is NOT blanked when missing — it also appears in the always-present
    isSharedWithExternalMe commands, so a genuinely missing sector should stay visibly
    unresolved as a correct signal for manual attention, not silently blanked."""
    tpl_text = tpl_text.replace(f"##Site_ID_{slot}##", str(site_id) if site_id else "")
    rru_placeholder_nums = (1, 2, 3) if slot == 1 else (4, 5, 6)
    for i, ph_num in enumerate(rru_placeholder_nums):
        val = rru_list[i] if i < len(rru_list) else None
        if val is not None:
            tpl_text = tpl_text.replace(f"##RRU_{ph_num}##", str(val))
    for i in range(1, 4):
        val = riport_list[i - 1] if i - 1 < len(riport_list) else None
        tpl_text = tpl_text.replace(f"##Riport{i}_site{slot}##", str(val) if val is not None else "")
    tpl_text = tpl_text.replace(f"##Priority_Site_{slot}##", str(priority) if priority is not None else "")
    return tpl_text


def _ngs_pair_is_pure_lte(a_to_b, b_to_a):
    """Ported from main branch (confirmed real bug fix, was missing from report-feature):
    own_cell is always an LTE cell by construction (only eUtran Parameters is scanned for
    Co-Located Technology Cell references) — so whether a pair is genuinely LTE-LTE (needing
    bidirectional confirmation) vs. mixed LTE-5G (one-directional suffices) depends on whether
    any ref_cell matches the 5G naming pattern — NOT on whether the target node separately has
    its own eNBId. A dual-tech TMBB node (its own LTE side unrelated to this specific shared
    radio) would otherwise be wrongly classified as a pure-LTE pair, requiring an impossible
    bidirectional confirmation and causing a false negative on a real shared-radio site."""
    for _own_cell, ref_cell in a_to_b + b_to_a:
        if nr_band_label(ref_cell)[0] is not None:
            return False
    return True


def generate_ngs_template_output(ciq_wb, mm_objs, user_id, date_str, log):
    """Returns (outputs, summary_rows). For every confirmed NGS pair, generates the fully filled
    activation template (LTE-LTE or LTE-5G, whichever applies), keeping only the correct
    3-link/Port-1-only/Port-2-only variant per site based on each site's actual unique RiPort
    count. Universal — same pattern as the display-only generate_ngs_checks, shared by all scopes."""
    outputs, summary_rows = [], []
    if len(mm_objs) < 2 or "eUtran Parameters" not in ciq_wb.sheetnames:
        return outputs, summary_rows

    cell_to_node = _ngs_build_cell_node_map(ciq_wb, mm_objs)
    node_names = [r.get("Node to be built as") for r in mm_objs if r.get("Node to be built as")]
    has_lte = {r.get("Node to be built as"): is_populated(r.get("eNBId")) for r in mm_objs}

    directional_refs = {}
    for row in sheet_objs(ciq_wb["eUtran Parameters"]):
        own_cell = row.get("EutranCellFDDId")
        raw = row.get("Co-Located Technology Cell")
        if not is_populated(own_cell) or not is_populated(raw) or str(raw).strip().upper() in ("NA", "N/A"):
            continue
        own_node = cell_to_node.get(str(own_cell).strip())
        if not own_node:
            continue
        for ref_cell in str(raw).split(","):
            ref_cell = ref_cell.strip()
            ref_node = cell_to_node.get(ref_cell)
            if ref_node and ref_node != own_node:
                directional_refs.setdefault((own_node, ref_node), []).append((own_cell, ref_cell))

    checked_pairs = set()
    for i, node_a in enumerate(node_names):
        for node_b in node_names[i + 1:]:
            pair_key = frozenset((node_a, node_b))
            if pair_key in checked_pairs:
                continue
            checked_pairs.add(pair_key)
            a_to_b = directional_refs.get((node_a, node_b), [])
            b_to_a = directional_refs.get((node_b, node_a), [])
            both_lte = _ngs_pair_is_pure_lte(a_to_b, b_to_a)
            confirmed = (a_to_b and b_to_a) if both_lte else (a_to_b or b_to_a)
            if not confirmed:
                continue

            # Collect each node's own confirmed cells (the "own_cell" from that node's perspective),
            # deduped by sector — one seed cell per sector present.
            own_a = [oc for oc, rc in a_to_b] + [rc for oc, rc in b_to_a]
            own_b = [oc for oc, rc in b_to_a] + [rc for oc, rc in a_to_b]

            def seed_by_sector(cells, is_lte_node):
                out = {}
                for c in cells:
                    label, sector = (lte_band_label(c) if is_lte_node else nr_band_label(c))
                    if sector and sector not in out:
                        out[sector] = c
                return out

            a_is_lte, b_is_lte = has_lte.get(node_a), has_lte.get(node_b)
            seeds_a = seed_by_sector(own_a, a_is_lte)
            seeds_b = seed_by_sector(own_b, b_is_lte)
            sectors_present = [s for s in SECTOR_ORDER if s in seeds_a or s in seeds_b]

            def derive(node, other_node, seeds, is_lte_node, other_seeds, other_is_lte):
                rru_list, riport_list = [], []
                for sector in sectors_present:
                    seed = seeds.get(sector)
                    if is_lte_node and seed:
                        rru, riport = _lte_cluster_rru_riport(ciq_wb, cell_to_node, node, seed)
                    elif is_lte_node:
                        rru, riport = None, None
                    else:
                        # 5G: RRU reused from the LTE partner's SAME sector; RiPort derived independently
                        other_seed = other_seeds.get(sector)
                        if other_is_lte and other_seed:
                            rru, _ = _lte_cluster_rru_riport(ciq_wb, cell_to_node, other_node, other_seed)
                        else:
                            rru = None
                        riport = _5g_cell_riport(ciq_wb, node, seed) if seed else None
                    rru_list.append(rru)
                    riport_list.append(riport)
                return rru_list, riport_list

            rru_a, riport_a = derive(node_a, node_b, seeds_a, a_is_lte, seeds_b, b_is_lte)
            rru_b, riport_b = derive(node_b, node_a, seeds_b, b_is_lte, seeds_a, a_is_lte)
            riport_a = _dedupe_preserve_order(riport_a)
            riport_b = _dedupe_preserve_order(riport_b)
            priority_a = _node_priority(ciq_wb, cell_to_node, node_a, a_is_lte)
            priority_b = _node_priority(ciq_wb, cell_to_node, node_b, b_is_lte)

            tpl_path = TPL_NGS_LTE_LTE if both_lte else TPL_NGS_LTE_5G
            if not tpl_path.exists():
                summary_rows.append({"Item": "NGS Template", "Source": f"template {tpl_path.name}", "Value": "NOT FOUND", "Note": f"expected file not in templates/NGS/: {tpl_path.name}"})
                log(f"\u2717 NGS Template: file not found for {node_a} <-> {node_b}")
                continue
            tpl_text = tpl_path.read_text(encoding="utf-8")

            link_count_a = len(riport_a) or 1
            link_count_b = len(riport_b) or 1
            tpl_text = _strip_ngs_variant_blocks(tpl_text, link_count_a, link_count_b)
            tpl_text = _fill_ngs_site_block(tpl_text, 1, node_a, rru_a, riport_a, priority_a)
            tpl_text = _fill_ngs_site_block(tpl_text, 2, node_b, rru_b, riport_b, priority_b)
            tpl_text = tpl_text.replace("xxDatexx", str(date_str))

            unresolved = highlight_unresolved(tpl_text)
            summary_rows.append({
                "Item": "NGS Template", "Source": f"{node_a} <-> {node_b}",
                "Value": "generated", "Note": f"unresolved: {len(unresolved)}" if unresolved else "fully resolved",
            })
            log(f"\u2713 NGS Template generated: {node_a} <-> {node_b} ({'LTE-LTE' if both_lte else 'LTE-5G'})" + (f" — {len(unresolved)} unresolved" if unresolved else ""))
            outputs.append((f"{node_a}_{node_b}_NGS_Activation.txt", tpl_text))

    return outputs, summary_rows


def generate_ngs_checks(ciq_wb, mm_objs, log):
    """Returns (summary_rows, scope_lines). No file outputs — pure detection."""
    summary_rows, scope_lines = [], []

    if len(mm_objs) < 2 or "eUtran Parameters" not in ciq_wb.sheetnames:
        return summary_rows, scope_lines

    cell_to_node = _ngs_build_cell_node_map(ciq_wb, mm_objs)
    node_names = [r.get("Node to be built as") for r in mm_objs if r.get("Node to be built as")]
    # Whether each node has an LTE side at all — only eUtran Parameters carries the
    # "Co-Located Technology Cell" column, so a 5G-only node can never declare a reference
    # back; requiring bidirectional confirmation for an LTE<->5G pair would be structurally
    # impossible to satisfy and produce a false negative on every genuinely-shared-radio site.
    has_lte = {r.get("Node to be built as"): is_populated(r.get("eNBId")) for r in mm_objs}

    # directional_refs[(from_node, to_node)] = list of every (from_cell, to_cell) pair seen
    directional_refs = {}
    for row in sheet_objs(ciq_wb["eUtran Parameters"]):
        own_cell = row.get("EutranCellFDDId")
        raw = row.get("Co-Located Technology Cell")
        if not is_populated(own_cell) or not is_populated(raw) or str(raw).strip().upper() in ("NA", "N/A"):
            continue
        own_node = cell_to_node.get(str(own_cell).strip())
        if not own_node:
            continue
        for ref_cell in str(raw).split(","):
            ref_cell = ref_cell.strip()
            ref_node = cell_to_node.get(ref_cell)
            if ref_node and ref_node != own_node:
                directional_refs.setdefault((own_node, ref_node), []).append((own_cell, ref_cell))

    checked_pairs = set()
    confirmed_nodes = set()
    for i, node_a in enumerate(node_names):
        for node_b in node_names[i + 1:]:
            pair_key = frozenset((node_a, node_b))
            if pair_key in checked_pairs:
                continue
            checked_pairs.add(pair_key)
            a_to_b = directional_refs.get((node_a, node_b), [])
            b_to_a = directional_refs.get((node_b, node_a), [])
            both_lte = _ngs_pair_is_pure_lte(a_to_b, b_to_a)
            confirmed = (a_to_b and b_to_a) if both_lte else (a_to_b or b_to_a)
            if confirmed:
                bands_a, bands_b = set(), set()
                for own_cell, ref_cell in a_to_b:
                    ba, bb = _ngs_cell_band(own_cell), _ngs_cell_band(ref_cell)
                    if ba: bands_a.add(ba)
                    if bb: bands_b.add(bb)
                for own_cell, ref_cell in b_to_a:
                    bb, ba = _ngs_cell_band(own_cell), _ngs_cell_band(ref_cell)
                    if bb: bands_b.add(bb)
                    if ba: bands_a.add(ba)
                a_str = ", ".join(sorted(bands_a)) if bands_a else "band not determined"
                b_str = ", ".join(sorted(bands_b)) if bands_b else "band not determined"
                band_list = f"{a_str} & {b_str}"
                summary_rows.append({
                    "Item": "NGS Checks", "Source": f"{node_a} <-> {node_b}",
                    "Value": "radio shared", "Note": f"bands: {band_list}",
                })
                log(f"\u2713 NGS Checks: {node_a} <-> {node_b} share a radio (bands: {band_list})")
                scope_lines.append(f"NGS Activation on :\t{band_list}\t{node_a} <-> {node_b}")
                confirmed_nodes.add(node_a)
                confirmed_nodes.add(node_b)

    # Safety-net (ported from main branch — confirmed missing from report-feature entirely
    # until now): some CIQs carry an explicit "NodeGroupSync" = "Y" column marking cells
    # that have NGS. This never drives detection on its own — it only flags a cell whose
    # node wasn't already covered by a confirmed co-location pair above, so a genuine miss
    # doesn't silently go unnoticed. Log/summary-row only, per explicit instruction — never
    # feeds the Warnings tab, since the primary detection logic is trusted as-is.
    for sheet_name, cell_col in (("eUtran Parameters", "EutranCellFDDId"), ("5G Info", "NRCellDU")):
        if sheet_name not in ciq_wb.sheetnames:
            continue
        for row in sheet_objs(ciq_wb[sheet_name]):
            if str(row.get("NodeGroupSync", "")).strip().upper() != "Y":
                continue
            cell = row.get(cell_col)
            if not is_populated(cell):
                continue
            node = cell_to_node.get(str(cell).strip())
            if node and node not in confirmed_nodes:
                log(f"\u26a0 NodeGroupSync=Y flagged for {cell} ({node}) but no confirmed NGS pair was detected \u2014 check this cell manually.")
                summary_rows.append({
                    "Item": "NGS Checks", "Source": cell,
                    "Value": "NodeGroupSync=Y, not confirmed", "Note": f"node {node} not part of any detected NGS pair \u2014 check manually",
                })

    return summary_rows, scope_lines


# ============================================================
# PORT CONVERSION (MCA / CENM / CRAN — CIQ + Pre-checks + EDP)
#
# Confirmed logic:
#  - Rule 1: no board swap. Pre-checks' Hardware Status Information reports the node's actual
#    baseband model (via extract_pre_hw, already used elsewhere for Pre/Post configuration) —
#    if its generation doesn't match the CIQ's DU-type-derived generation, a board swap is
#    already in progress and Port Conversion doesn't apply (the OpMode difference is explained
#    by the swap, not a pure port-speed change).
#  - Rule 2: Pre-checks' "Transport Fiber link Status" table shows the board's relevant port
#    (below) at OpMode = 1G_FULL, while the EDP's SIAD_PORT_SIZE_BBU already shows 10G for that
#    same node — i.e. the port hasn't been converted yet, but the EDP already calls for it.
#  - Rule 3: which port to check depends on board generation:
#      G2 -> TN_A or TN_B      G3 -> TN_IDL_B      G4 -> TN_IDL_C
#  - Output: "Port speed 1G to 10G conversion with MPST: <NodeID>." — plain sentence, not
#    tab-separated like the other Scope of Work lines (confirmed).
# ============================================================

PORT_BY_GEN = {"G2": ["TN_A", "TN_B"], "G3": ["TN_IDL_B"], "G4": ["TN_IDL_C"]}

TRANSPORT_FIBER_ROW_RE = re.compile(
    r'(\S+)\s+(\S+)\s+(\d+)\s+(TN_A|TN_B|TN_IDL_B|TN_IDL_C)\s+\S+\s+\d+\s+(\S+)\s+(?:true|false)'
)


def extract_transport_fiber_opmode(precheck_text, node, port_labels):
    """OpMode string for the first Transport Fiber link Status row matching this node and
    one of its generation's relevant port labels, or None if no such row exists."""
    if not precheck_text or not node:
        return None
    node_u = str(node).strip().upper()
    for m in TRANSPORT_FIBER_ROW_RE.finditer(precheck_text):
        row_node, board, lnh, port, opmode = m.groups()
        if row_node.strip().upper() == node_u and port in port_labels:
            return opmode
    return None


# ============================================================
# DATA2 FIBER'S TESTING (MCA / CENM / CRAN — Pre-checks + CIQ)
#
# Confirmed logic:
#  - "Pre-existing" = the cell is found in Pre-checks' Summary Status table at all.
#  - Its Pre-checks state comes from the DL/UL Loss table's RRU description — ending in
#    "Port D1" means it's currently on DATA1.
#  - If the CIQ's own Radio Port column shows DATA2 (the target state) for that same
#    pre-existing/DATA1 cell, DATA2 fiber testing is required.
#  - Output: grouped by band, no node ID — "DATA2 Fiber's testing on: <band>." if all of
#    Alpha/Beta/Gamma are flagged, otherwise "DATA2 Fiber's testing on: <band> <sectors>."
# ============================================================

def generate_data2_testing_checks(ciq_wb, mm_objs, precheck_text, log):
    scope_lines = []
    if not precheck_text:
        return scope_lines

    pre_pairs, _ = extract_precheck_sectors(precheck_text)
    pre_existing_cells = {cell for _node, cell in pre_pairs}

    cell_port_state = {}
    for r in extract_dl_ul_loss_rows(precheck_text):
        cell, desc = r["Cells"], r["DUS/XMU (S.No) - RRU"].strip()
        if cell in cell_port_state:
            continue
        if desc.endswith("Port D1"):
            cell_port_state[cell] = "D1"
        elif desc.endswith("Port D2"):
            cell_port_state[cell] = "D2"

    band_sectors = {}

    def check_cell(cell, radio_port, band_label_fn):
        if not is_populated(cell) or cell not in pre_existing_cells:
            return
        if cell_port_state.get(cell) != "D1":
            return
        if str(radio_port or "").strip().upper() != "DATA2":
            return
        label, sector = band_label_fn(cell)
        if label and sector:
            band_sectors.setdefault(label, set()).add(sector)

    if "eUtran Parameters" in ciq_wb.sheetnames:
        for row in sheet_objs(ciq_wb["eUtran Parameters"]):
            check_cell(row.get("EutranCellFDDId"), row.get("Radio Port"), lte_band_label)
    if "5G Info" in ciq_wb.sheetnames:
        for row in sheet_objs(ciq_wb["5G Info"]):
            check_cell(row.get("NRCellDU"), row.get("Radio Port"), nr_band_label)

    WHOLE_BAND_SET = {"Alpha", "Beta", "Gamma"}
    for label in sorted(band_sectors):
        sectors = band_sectors[label]
        if WHOLE_BAND_SET <= sectors:
            scope_lines.append(f"DATA2 Fiber's testing on: {label}.")
        else:
            names = sorted(sectors, key=lambda s: SECTOR_ORDER.index(s) if s in SECTOR_ORDER else 99)
            scope_lines.append(f"DATA2 Fiber's testing on: {label} {', '.join(names)}.")
        log(f"\u2713 DATA2 Fiber's testing flagged: {label} ({', '.join(sorted(sectors))})")

    return scope_lines


def generate_port_conversion_checks(ciq_wb, mm_objs, edp_index, precheck_text, log):
    """Returns (outputs, summary_rows, scope_lines). Shared by MCA / CENM / CRAN.
    Generation is read from Pre-checks' Hardware Status (the board that CURRENTLY exists),
    not the CIQ's target/post board — the template applies to whatever board is actually in
    Pre-checks right now, regardless of what it's being swapped to (confirmed: this template is
    for the board that's in Pre, not the board in the CIQ's post state). G4 is excluded outright
    since it's the newest board (can never be a Pre-checks-side board here) and the template has
    no G4/TN_IDL_C content."""
    outputs, summary_rows, scope_lines = [], [], []
    if not precheck_text:
        return outputs, summary_rows, scope_lines

    tpl_text = TPL_PORT_CONVERSION.read_text(encoding="utf-8") if TPL_PORT_CONVERSION.exists() else None

    for row in mm_objs:
        node = row.get("Node to be built as")
        if not node:
            continue

        pre_model = extract_pre_hw(precheck_text, node)
        pre_gen = DU_TYPE_TO_GEN.get(str(pre_model).strip()) if pre_model else None
        if pre_gen not in ("G2", "G3"):
            continue  # G4 (or undetectable) in Pre-checks — template doesn't apply here at all
        port_labels = PORT_BY_GEN[pre_gen]

        opmode = extract_transport_fiber_opmode(precheck_text, node, port_labels)
        if not opmode or "1G" not in opmode.upper():
            continue  # not currently 1G in Pre-checks — nothing pending

        edp_row = edp_row_for(edp_index, node)
        siad_port_size = edp_get(edp_index, edp_row, "SIAD_PORT_SIZE_BBU") if edp_row else None
        if not is_populated(siad_port_size) or "10G" not in str(siad_port_size).upper():
            continue  # EDP doesn't call for 10G — nothing pending

        # Confirmed mismatch on a G2/G3 Pre-checks board — always show display line and always
        # generate the template, regardless of what the CIQ's target board ends up being.
        summary_rows.append({
            "Item": "Port Conversion", "Source": node,
            "Value": "1G -> 10G pending", "Note": f"Pre-checks board: {pre_gen}, port: {'/'.join(port_labels)}, EDP SIAD_PORT_SIZE_BBU: {siad_port_size}",
        })
        log(f"\u2713 Port Conversion: {node} — 1G in Pre-checks ({pre_gen} board), EDP calls for 10G")
        scope_lines.append(f"Port speed 1G to 10G conversion with MPST: {node}.")

        if tpl_text is None:
            summary_rows.append({"Item": "Port Conversion", "Source": f"template {TPL_PORT_CONVERSION.name}", "Value": "NOT FOUND", "Note": f"expected file not in templates/MCA/: {TPL_PORT_CONVERSION.name}"})
            log(f"\u2717 Port Conversion: template file not found for {node}")
            continue

        filled = tpl_text.replace("xxSiteIdxx", str(node)).replace("xSiteIDx", str(node))
        outputs.append((f"{node}_Port_Conversion_1G_to_10G.txt", filled))

    return outputs, summary_rows, scope_lines


# ============================================================
# PRE FIBERS (universal — pulled from Pre-checks' DL/UL Loss table)
# ============================================================

DL_UL_LOSS_ROW_RE = re.compile(
    r'(\S+)\s+(?:Up|Down)\s+\d+\s+(?:(?!\S+\s+(?:Up|Down)\s+\d+).)*?'
    r'((?:Baseband|XMU|RAN Processor|:\s*RRU)\S*(?:(?!\S+\s+(?:Up|Down)\s+\d+).)*Port\s+D\d)', re.DOTALL
)

def extract_dl_ul_loss_rows(precheck_text):
    if not precheck_text:
        return []
    seen, rows = set(), []
    for m in DL_UL_LOSS_ROW_RE.finditer(precheck_text):
        cell, dus_xmu_rru = m.group(1), m.group(2).strip()
        key = (cell, dus_xmu_rru)  # dedupe by cell+description, not cell alone — a dual-band
        if key not in seen:        # radio cell can legitimately have two distinct entries (Port D1/D2)
            seen.add(key)
            rows.append({"Cells": cell, "DUS/XMU (S.No) - RRU": dus_xmu_rru})
    return rows


# ============================================================
# UNIVERSAL STATIC OUTPUTS (all scopes — no filling, pure passthrough)
# ============================================================

STATIC_OUTPUT_FILES = [
    "Integration_Checklist_v3.xlsx",
    "Global Local Script Execution Order.xlsx",
]


# ============================================================
# MCA INTEGRATION REPORT (Report_MCA-style output) — auto-fill what QUICKIX already knows,
# manual entry for everything else, matching the Legacy_MCA_Macro_Template checklist.
# ============================================================

def derive_idl_build_type_label(ciq_wb, mm_objs):
    """Re-derives which IDL template combo/filename would be used (same lookup + the same
    get_node_generation() the real IDL Connections generator uses), purely to extract the
    'Buildtype_X' suffix already embedded in the real template filenames (confirmed: e.g.
    'G3+G3_Buildtype_C.txt' -> 'Type C'). Returns None if no IDL combo applies (e.g. single-BBU
    site) or the combo isn't in the registry.
    Confirmed: when a combo has BOTH a Preferred and Alternate variant (e.g. G2+G3 -> Type BB
    and Type B), show BOTH joined by '/' rather than silently picking just the first — which
    physical ports are actually free on the board can't be known from the CIQ, so the engineer
    picks which applies by deleting whichever didn't."""
    if len(mm_objs) < 2:
        return None
    gens = [get_node_generation(ciq_wb, row) for row in mm_objs]
    if not all(gens):
        return None
    combo = tuple(sorted(gens))
    entries = IDL_TEMPLATE_REGISTRY.get(combo)
    if not entries:
        return None
    labels = []
    for fname, _variant in entries:
        m = re.search(r'Buildtype_([A-Z]+)', fname)
        if m:
            labels.append(f"Type {m.group(1)}")
    return "/".join(labels) if labels else None


def build_mca_integration_report(pre_line, post_line, controller_objs, mm_objs, manual):
    """manual: dict of engineer-provided values from the UI —
    {mic, market, site_name, sow, iwm_details, current_config, wll_node, software_version,
     gs_version, completed_extra, pending, pre_existing_issues, notes, idl_cable_details,
     switch_details, slot_port_details}.
    Status (ATP/STF) is derived: ATP only if 'pending' is empty, STF otherwise — confirmed rule.
    Returns the final Report_MCA-style plain-text block."""
    site_ids = "/".join(r.get("Node to be built as") for r in mm_objs if r.get("Node to be built as"))
    status = "ATP" if not (manual.get("pending") or "").strip() else "STF"
    mic = manual.get("mic") or "MIC"

    lines = []
    lines.append("Subject")
    lines.append(f"{mic} | {manual.get('market','')} | {status} | {manual.get('site_name','')} | "
                  f"{manual.get('fa_code','')} | {site_ids} | {manual.get('sow','')}")
    lines.append("")
    lines.append("IWM Details")
    lines.append(manual.get("iwm_details", ""))
    lines.append("")
    lines.append("Configuration")
    lines.append(f"Pre Configuration : {pre_line}")
    if (manual.get("current_config") or "").strip():
        lines.append(f"Current Configuration : {manual['current_config']}")
    lines.append(f"Post Configuration : {post_line}")
    if (manual.get("wll_node") or "").strip():
        lines.append(f"WLL  node : {manual['wll_node']}")
    controller_id = controller_objs[0].get("Controller") if controller_objs else ""
    lines.append(f"6610 Controller : {controller_id}")
    lines.append(f"Software version: {manual.get('software_version','')}")
    lines.append(f"GS Version: {manual.get('gs_version','')}")
    lines.append("")
    lines.append("IDL Connections")
    build_type = manual.get("idl_build_type")
    if build_type:
        lines.append(f"Build Type : {build_type}")
    if (manual.get("idl_cable_details") or "").strip():
        lines.append(manual["idl_cable_details"])
    if (manual.get("switch_details") or "").strip():
        lines.append("Switch")
        lines.append(manual["switch_details"])
    if (manual.get("slot_port_details") or "").strip():
        lines.append("Slot/Port")
        lines.append(manual["slot_port_details"])
    lines.append("")
    lines.append("Completed:")
    for auto_line in manual.get("completed_auto_lines", []):
        lines.append(auto_line)
    if (manual.get("completed_extra") or "").strip():
        lines.append(manual["completed_extra"])
    lines.append("")
    lines.append("Pending:")
    lines.append(manual.get("pending", ""))
    lines.append("")
    lines.append("Pre-Existing Issues:")
    lines.append(manual.get("pre_existing_issues", ""))
    lines.append("")
    lines.append("Notes:")
    lines.append(manual.get("notes", ""))
    return "\n".join(lines)


def _xml_escape(text):
    return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _resolve_cell_text(sheet_xml, shared_strings_list, cell_ref):
    """Returns (text, style_attr, kind) for a cell — text/kind are None if the cell holds a
    non-text value (boolean/numeric) or doesn't exist at all; style_attr is None only if the
    cell doesn't exist. Handles both shared-string (t="s") and inline-string (t="inlineStr")
    cells — every text cell in these static templates is one or the other."""
    m = re.search(r'<c r="' + re.escape(cell_ref) + r'"([^>]*?)(?:/>|>(.*?)</c>)', sheet_xml, re.S)
    if not m:
        return None, None, None
    attrs, inner = m.group(1), m.group(2) or ""
    style_m = re.search(r's="(\d+)"', attrs)
    style_attr = f' s="{style_m.group(1)}"' if style_m else ""
    if 't="s"' in attrs:
        idx = int(re.search(r'<v>(\d+)</v>', inner).group(1))
        text_m = re.search(r'<t[^>]*>(.*?)</t>', shared_strings_list[idx], re.S)
        return (text_m.group(1) if text_m else ""), style_attr, "s"
    elif 't="inlineStr"' in attrs:
        text_m = re.search(r'<t[^>]*>(.*?)</t>', inner, re.S)
        return (text_m.group(1) if text_m else ""), style_attr, "inlineStr"
    return None, style_attr, None  # boolean/numeric/other — a cell we'd never touch here


def _patch_text_cell(sheet_xml, cell_ref, style_attr, new_text):
    """Replaces an EXISTING cell's content with new_text, preserving its style ('s=')
    attribute exactly. Always re-emits as t="inlineStr" — simplest and safest, since it never
    touches sharedStrings.xml (which other untouched cells may still reference)."""
    m = re.search(r'<c r="' + re.escape(cell_ref) + r'"([^>]*?)(?:/>|>.*?</c>)', sheet_xml, re.S)
    new_cell = f'<c r="{cell_ref}"{style_attr} t="inlineStr"><is><t xml:space="preserve">{_xml_escape(new_text)}</t></is></c>'
    return sheet_xml[:m.start()] + new_cell + sheet_xml[m.end():]


def _col_letters(cell_ref):
    return re.match(r'([A-Z]+)(\d+)', cell_ref).group(1)


def _col_to_num(letters):
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - 64)
    return n


def _insert_cell_in_row(sheet_xml, row_num, cell_ref, new_text):
    """Inserts a brand-new cell that doesn't exist in the row XML at all (Excel omits truly
    empty cells from sheetData entirely) at the correct sorted column position among the
    row's existing cells — confirmed necessary: Global Local Script Order's B1 cells are
    empty in the pristine template and simply absent from the XML, not present-but-blank."""
    row_m = re.search(r'(<row r="' + str(row_num) + r'"[^>]*>)(.*?)(</row>)', sheet_xml, re.S)
    open_tag, body, close_tag = row_m.group(1), row_m.group(2), row_m.group(3)
    new_cell = f'<c r="{cell_ref}" t="inlineStr"><is><t xml:space="preserve">{_xml_escape(new_text)}</t></is></c>'
    target_col = _col_to_num(_col_letters(cell_ref))
    insert_pos = len(body)
    for m2 in re.finditer(r'<c r="([A-Z]+)\d+"[^/]*?(?:/>|>.*?</c>)', body):
        if _col_to_num(m2.group(1)) > target_col:
            insert_pos = m2.start()
            break
    new_body = body[:insert_pos] + new_cell + body[insert_pos:]
    return sheet_xml[:row_m.start()] + open_tag + new_body + close_tag + sheet_xml[row_m.end():]


def _patch_or_insert_cell(sheet_xml, cell_ref, new_text):
    _, style_attr, kind = _resolve_cell_text(sheet_xml, [], cell_ref)
    if style_attr is not None or kind is not None:
        if re.search(r'<c r="' + re.escape(cell_ref) + r'"[^/]*?(?:/>|>.*?</c>)', sheet_xml, re.S):
            return _patch_text_cell(sheet_xml, cell_ref, style_attr or "", new_text)
    row_num = int(re.match(r'[A-Z]+(\d+)', cell_ref).group(1))
    return _insert_cell_in_row(sheet_xml, row_num, cell_ref, new_text)


def get_fa_code(ciq_wb):
    """FA Code lives in the CIQ's 5G Info sheet only (same value repeated on every 5G cell row).
    LTE-only (SMBB) sites have no 5G Info rows at all — confirmed to just leave it blank there
    rather than guess a fallback column."""
    if "5G Info" not in ciq_wb.sheetnames:
        return None
    ws = ciq_wb["5G Info"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if "FA Code" not in header:
        return None
    idx = header.index("FA Code")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if idx < len(row) and is_populated(row[idx]):
            return row[idx]
    return None


def fill_integration_checklist(fpath, ciq_wb, mm_objs, log):
    """Integration_Checklist_v3.xlsx: the real template already has labeled, highlighted cells
    (A1="Site ID's :", A2='FA Code :', A3='Support Engineer Name : ') — value appended directly
    onto the label in the SAME cell (both labels already end in ':' with no trailing space, so
    plain concatenation reproduces e.g. "Site ID's :KYL06026/KYL07626R" exactly).

    CRITICAL — never use openpyxl.load_workbook()+save() on this file. Confirmed by direct
    testing: this template's column-B checklist cells are wired to Excel 365's native
    Checkbox feature (xl/featurePropertyBag/featurePropertyBag.xml declares a "Checkbox" bag;
    each boolean cell's style (s="15") has an <extLst><xfpb:xfComplement i="0"/></extLst>
    pointing at it) — a plain openpyxl round-trip silently drops BOTH the featurePropertyBag
    part AND the style's extLst (confirmed: neither survives), which is why the checkboxes
    were showing as literal "TRUE"/"FALSE" text. Same category of destructive-round-trip bug
    as the ActiveX macro checkboxes in mca_xlsm_surgical.py, different Excel feature. Fix here
    is the same philosophy: raw zip/XML surgery, only patch the two label cells' text, copy
    every other byte through untouched — this preserves the native checkbox wiring perfectly,
    confirmed working with a single click-tested real checkbox before scaling to all 102."""
    site_ids = "/".join(r.get("Node to be built as") for r in mm_objs if r.get("Node to be built as"))
    fa_code = get_fa_code(ciq_wb)

    with zipfile.ZipFile(fpath) as zin:
        sheet1 = zin.read("xl/worksheets/sheet1.xml").decode("utf-8")
        shared_strings = zin.read("xl/sharedStrings.xml").decode("utf-8")
        infos = {i.filename: i for i in zin.infolist()}
        other_files = {n: zin.read(n) for n in zin.namelist() if n != "xl/worksheets/sheet1.xml"}

    shared_strings_list = re.findall(r'<si>(.*?)</si>', shared_strings, re.S)
    site_label, style_a1, _ = _resolve_cell_text(sheet1, shared_strings_list, "A1")
    fa_label, style_a2, _ = _resolve_cell_text(sheet1, shared_strings_list, "A2")
    site_label = site_label or "Site ID's :"
    fa_label = fa_label or "FA Code :"

    sheet1 = _patch_text_cell(sheet1, "A1", style_a1, site_label + site_ids)
    sheet1 = _patch_text_cell(sheet1, "A2", style_a2, fa_label + (str(fa_code) if fa_code is not None else ""))

    log(f"{'✓' if site_ids else '✗'} Integration Checklist · Site ID's -> {site_ids or 'NOT FOUND'}")
    log(f"{'✓' if fa_code is not None else '✗'} Integration Checklist · FA Code -> {fa_code if fa_code is not None else 'NOT FOUND'}")

    out_buf = io.BytesIO()
    with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in other_files.items():
            zout.writestr(infos[name], data)
        zout.writestr(infos["xl/worksheets/sheet1.xml"], sheet1.encode("utf-8"))
    return out_buf.getvalue()


def fill_global_local_script_order(fpath, mm_objs, log):
    """Global Local Script Execution Order.xlsx: Node_1/Node_2/Node_3 sheets, one per Mixed Mode
    Info row in row order (same sibling-order convention as SMBB's xLTE_SiteID2x/3x). Each
    sheet's own 'Node ID :' label (B1) and every 'XXSITEIDXX' placeholder inside that sheet's
    script filenames get that sheet's node ID. A site with fewer than 3 nodes leaves the
    unused Node_2/Node_3 sheet(s) untouched (confirmed — not deleted).

    Same native-checkbox fragility as fill_integration_checklist (this file has the identical
    featurePropertyBag Checkbox wiring across all 3 sheets) — raw zip/XML surgery only, never
    openpyxl.load_workbook()+save()."""
    site_ids = [r.get("Node to be built as") for r in mm_objs if r.get("Node to be built as")]

    with zipfile.ZipFile(fpath) as zin:
        wb_xml = zin.read("xl/workbook.xml").decode("utf-8")
        rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        shared_strings_xml = zin.read("xl/sharedStrings.xml").decode("utf-8")
        infos = {i.filename: i for i in zin.infolist()}
        all_files = {n: zin.read(n) for n in zin.namelist()}

    shared_strings_list = re.findall(r'<si>(.*?)</si>', shared_strings_xml, re.S)
    sheet_name_to_rid = dict(re.findall(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"', wb_xml))
    rid_to_target = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels_xml))

    for i, sheet_name in enumerate(["Node_1", "Node_2", "Node_3"]):
        if sheet_name not in sheet_name_to_rid:
            continue
        target = "xl/" + rid_to_target[sheet_name_to_rid[sheet_name]]
        if i >= len(site_ids):
            log(f"· Global Local Script Order · {sheet_name} -> no corresponding CIQ node, left blank")
            continue
        node_id = site_ids[i]
        sheet_xml = all_files[target].decode("utf-8")

        sheet_xml = _patch_or_insert_cell(sheet_xml, "B1", str(node_id))

        for ref in re.findall(r'<c r="(\w+\d+)"', sheet_xml):
            if ref == "B1":
                continue
            txt, style_attr2, _ = _resolve_cell_text(sheet_xml, shared_strings_list, ref)
            if txt and "XXSITEIDXX" in txt:
                sheet_xml = _patch_text_cell(sheet_xml, ref, style_attr2, txt.replace("XXSITEIDXX", str(node_id)))

        all_files[target] = sheet_xml.encode("utf-8")
        log(f"✓ Global Local Script Order · {sheet_name} -> {node_id}")

    out_buf = io.BytesIO()
    with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in all_files.items():
            zout.writestr(infos[name], data)
    return out_buf.getvalue()


def detect_board_swap_nodes(ciq_wb, mm_objs, precheck_text):
    """Same board-swap detection as report_detect.detect_node_board_changes (the MCA Integration
    Report pipeline), reimplemented locally so the file-generation pipeline (generate_mca/
    generate_cenm/generate_cran) doesn't need to self-import app.py as a module. Returns just the
    node names — [(node, pre_board, post_board), ...] trimmed to node names only."""
    _, pre_nodes_set = extract_precheck_sectors(precheck_text)
    swapped = []
    for row in mm_objs:
        primary = row.get("Node to be built as")
        if primary not in pre_nodes_set:
            continue  # new node, not a swap
        e_name, g_name = row.get("eNodeB Name"), row.get("gNodeB Name")
        is_lte_primary = str(primary).strip().upper() == str(e_name or "").strip().upper()
        r = find_row_by_name(ciq_wb, "eNB Info", "eNodeB Name", e_name) if is_lte_primary else find_row_by_name(ciq_wb, "gNB Info", "gNodeB Name", g_name)
        if not r:
            r = find_row_by_name(ciq_wb, "eNB Info", "eNodeB Name", e_name) or find_row_by_name(ciq_wb, "gNB Info", "gNodeB Name", g_name)
        post_board = hw_string(r)
        pre_board = pre_hw_string(precheck_text, primary)
        if pre_board and post_board and pre_board.strip() != post_board.strip():
            swapped.append(primary)
    return swapped


def generate_node_deletion_templates(ciq_wb, mm_objs, edp_index, classification, precheck_text, user_id, date_str, log):
    """Site_Install_Generation_and_Node_Deletion_commands.txt — triggered for MCA/CENM/CRAN
    whenever a node is being deleted from ENM (classification['deleted_nodes'], computed before
    any scope-specific zeroing) OR has a board swap (detect_board_swap_nodes). One filled file
    per triggered node. xxxxIP_ADDDESS_SITExxxx = EDP's IPV6_ENODEB_OAM_IP for that node — if the
    node isn't found in EDP (expected for a deleted node no longer tracked there), the token is
    left in place rather than guessed, per confirmed behavior."""
    if TPL_NODE_DELETION is None or not TPL_NODE_DELETION.exists():
        return [], []
    base_tpl = TPL_NODE_DELETION.read_text(encoding="utf-8")
    deleted_nodes = set(classification.get("deleted_nodes") or [])
    board_swap_nodes = set(detect_board_swap_nodes(ciq_wb, mm_objs, precheck_text))
    trigger_nodes = sorted(deleted_nodes | board_swap_nodes)
    outputs, summary_rows = [], []
    for node in trigger_nodes:
        reason = []
        if node in deleted_nodes:
            reason.append("deleted node")
        if node in board_swap_nodes:
            reason.append("board swap")
        row = edp_row_for(edp_index, node)
        ip = edp_get(edp_index, row, "IPV6_ENODEB_OAM_IP")
        tpl = base_tpl.replace("xxSite_IDxx", str(node))
        if ip:
            tpl = tpl.replace("xxxxIP_ADDDESS_SITExxxx", str(ip))
            summary_rows.append({"Item": f"{node} · IP Address", "Source": "EDP · IPV6_ENODEB_OAM_IP", "Value": ip, "Note": " & ".join(reason)})
            log(f"✓ {node} · Node Deletion template · IP -> {ip} ({' & '.join(reason)})")
        else:
            summary_rows.append({"Item": f"{node} · IP Address", "Source": "EDP · IPV6_ENODEB_OAM_IP", "Value": "NOT FOUND", "Note": f"{' & '.join(reason)} — placeholder left in output"})
            log(f"✗ {node} · Node Deletion template · IP NOT FOUND in EDP, placeholder left in output ({' & '.join(reason)})")
        outputs.append((f"{node}_Site_Install_and_Node_Deletion_Filled.txt", tpl))
    return outputs, summary_rows


def get_universal_static_outputs(ciq_wb, mm_objs, log):
    """Returns a list of (filename, bytes) for the static reference files that ship alongside
    Final Connections / Pre Fibers for every scope. Integration_Checklist_v3.xlsx and
    Global Local Script Execution Order.xlsx now get Site ID/FA Code/Node ID auto-filled from
    the CIQ (see fill_integration_checklist / fill_global_local_script_order) — everything else
    in STATIC_OUTPUT_FILES stays a pure, unmodified passthrough."""
    outputs = []
    for fname in STATIC_OUTPUT_FILES:
        fpath = TDIR_STATIC / fname
        if not fpath.exists():
            log(f"\u2717 Static output not found: templates/Static/{fname}")
            continue
        if fname == "Integration_Checklist_v3.xlsx":
            outputs.append((fname, fill_integration_checklist(fpath, ciq_wb, mm_objs, log)))
        elif fname == "Global Local Script Execution Order.xlsx":
            outputs.append((fname, fill_global_local_script_order(fpath, mm_objs, log)))
        else:
            outputs.append((fname, fpath.read_bytes()))
        log(f"\u2713 Static output attached: {fname}")
    return outputs


def generate_pre_fibers(precheck_text):
    """One Excel file per CIQ: Cells + DUS/XMU (S.No) - RRU from Pre-checks' DL/UL Loss table,
    plus a blank 'Pre fibers' column for manual fill-in."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    rows = extract_dl_ul_loss_rows(precheck_text)
    if not rows:
        return None
    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = "Sheet1"
    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="000000")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.append(["Cells", "DUS/XMU (S.No) - RRU", "Pre fibers"])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
    for r in rows:
        ws.append([r["Cells"], r["DUS/XMU (S.No) - RRU"], None])
        for cell in ws[ws.max_row]:
            cell.border = BORDER
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 90
    ws.column_dimensions["C"].width = 14
    buf = io.BytesIO()
    out_wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============================================================
# RADIO SWAP (universal — compares Pre-checks' DL/UL Loss radio type against the CIQ's RRU Type)
# ============================================================

RADIO_TYPE_RE = re.compile(r'RRU[-\w]*\s*\(([^,)]+),')

def extract_precheck_radio_types(precheck_text):
    """Reuses the same DL/UL Loss row match as Pre Fibers, but pulls the radio product type
    out of the captured DUS/XMU description instead. A cell can legitimately appear more than
    once (a dual-band radio unit serving one sector through two physical radios/ports) — every
    distinct radio type seen for that cell is kept, not just the first or last."""
    out = {}
    if not precheck_text:
        return out
    for m in DL_UL_LOSS_ROW_RE.finditer(precheck_text):
        cell, desc = m.group(1), m.group(2)
        rm = RADIO_TYPE_RE.search(desc)
        if rm:
            radio = rm.group(1).strip()
            out.setdefault(cell, [])
            if radio not in out[cell]:
                out[cell].append(radio)
    return out


def ciq_radio_types(ciq_wb):
    """CIQ-side radio type per cell: 5G Info's 'RRU Type' for 5G cells, eUtran Parameters'
    'RRU type' for LTE cells."""
    out = {}
    if "5G Info" in ciq_wb.sheetnames:
        for r in sheet_objs(ciq_wb["5G Info"]):
            cell, rru = r.get("NRCellDU"), r.get("RRU Type")
            if cell and is_populated(rru):
                out[cell] = str(rru).strip()
    if "eUtran Parameters" in ciq_wb.sheetnames:
        for r in sheet_objs(ciq_wb["eUtran Parameters"]):
            cell, rru = r.get("EutranCellFDDId"), r.get("RRU type")
            if cell and is_populated(rru):
                out[cell] = str(rru).strip()
    return out


def radio_family(radio_string):
    """Extract just the RRU type — the token right after 'RRUS'/'Radio', ignoring the band suffix
    entirely (e.g. 'RRUS A2 B4' -> 'A2', 'RRUS 12 B4' -> '12', 'Radio 4890HP 48B2/B25 48B66 M01' ->
    '4890', 'RRUS 4890' -> '4890'). Handles both digit-leading types (strips trailing letters like
    'HP') and letter-leading alphanumeric types (kept as-is, e.g. 'A2')."""
    s = str(radio_string or "").strip()
    tokens = s.split()
    if len(tokens) >= 2 and tokens[0].upper() in ("RRUS", "RADIO"):
        type_token = tokens[1]
        m = re.match(r"^(\d+)", type_token)
        return m.group(1) if m else type_token.upper()
    m = re.search(r"\d{2,5}", s)
    return m.group(0) if m else s.upper()


def build_colocation_groups(ciq_wb):
    """Cell -> canonical co-location group key, from eUtran Parameters' 'Co-Located Technology
    Cell' column (lists peer cell names sharing the same physical radio, spans LTE+5G together).
    'NA' or blank means the cell isn't co-located with anything else. The group is registered
    under EVERY member's name (the LTE cell itself plus all its listed peers, including 5G
    cells) — a 5G cell only ever appears as a peer inside an LTE row, never as its own row, so
    without this it could never find the group it's actually listed in."""
    groups = {}
    if "eUtran Parameters" not in ciq_wb.sheetnames:
        return groups
    for r in sheet_objs(ciq_wb["eUtran Parameters"]):
        cell = r.get("EutranCellFDDId")
        colo_raw = r.get("Co-Located Technology Cell")
        if not cell:
            continue
        if colo_raw and str(colo_raw).strip().upper() != "NA":
            peers = {p.strip() for p in str(colo_raw).split(",") if p.strip()}
            peers.add(cell)
            group_key = tuple(sorted(peers))
            for member in peers:
                groups[member] = group_key
        else:
            groups[cell] = (cell,)
    return groups


def classify_radio_swaps(precheck_text, ciq_wb):
    """Cells present in both Pre-checks and the CIQ where the radio family genuinely differs
    (compared by RRU type only — ignoring band suffix — to avoid false positives from naming-
    format differences between the two sources). Follows Sector Del_Movement's rename mapping —
    a moved cell can be renamed, so the Pre-checks value must be compared against the CIQ using
    the cell's NEW name. A cell with multiple distinct Pre-checks radios (dual-band radio unit)
    shows all of them combined with '+' in the From field."""
    pre_radios = extract_precheck_radio_types(precheck_text)
    post_radios = ciq_radio_types(ciq_wb)
    colo_groups = build_colocation_groups(ciq_wb)

    rename_map = {}
    if "Sector Del_Movement" in ciq_wb.sheetnames:
        for r in sheet_objs(ciq_wb["Sector Del_Movement"]):
            src_sector, tgt_sector = r.get("Source Sector"), r.get("Target Sector")
            tgt_node = r.get("Target Node name")
            if src_sector and tgt_sector and str(tgt_node).strip().upper() != "DELETE":
                rename_map[src_sector] = tgt_sector

    swaps = []
    for cell, pre_radio_list in pre_radios.items():
        ciq_cell = rename_map.get(cell, cell)
        post_radio = post_radios.get(ciq_cell)
        if not post_radio:
            continue
        post_type = radio_family(post_radio)
        pre_types = []
        for r in pre_radio_list:
            t = radio_family(r)
            if t not in pre_types:
                pre_types.append(t)
        if post_type not in pre_types:
            label, sector = band_label(cell)
            pre_types_sorted = sorted(pre_types, key=lambda t: (not t[0].isdigit(), t))
            from_str = "RRU " + "+".join(pre_types_sorted)
            to_str = f"RRU {post_type}"
            group_key = colo_groups.get(ciq_cell, (ciq_cell,))
            swaps.append({"label": label, "sector": sector, "from": from_str, "to": to_str, "group_key": group_key})
    return swaps


# ============================================================
# GENERIC PRE/POST CONFIGURATION (MCA / CENM — any node set, not CRAN's fixed roles)
# ============================================================

def pre_node_label(precheck_text, node_name):
    """Determine a node's (P)/(S) identity pairing AND its BBU mode tag as they existed in
    Pre-checks — independent from the CIQ, since a node can genuinely convert LTE-only <->
    MMBB/TMBB between Pre and Post (5G sectors moving onto or off of it as part of the same
    scope). Confirmed real rule, derived purely from which cells are actually present in
    Pre-checks (not from the CIQ's BBU Mode column, which only reflects the Post-side state):
    LTE cells only -> SMBB. 5G cells only -> SMBB. LTE + 5G (no CBAND/DOD) -> MMBB.
    LTE + 5G + CBAND/DOD -> TMBB. CBAND/DOD 5G cells share band code 077 (_N077[A-F]_n,
    covering CBAND/DOD/DOD_BWE alike), same pattern nr_band_label() already uses."""
    pre_pairs, _ = extract_precheck_sectors(precheck_text)
    node_cells = [cell for (node, cell) in pre_pairs if node == node_name]
    if not node_cells:
        return node_name
    fiveg_cells = [c for c in node_cells if is_5g_cell(c)]
    lte_cells = [c for c in node_cells if not is_5g_cell(c)]
    has_cband_dod = any(re.search(r'_N077[A-F]_\d+$', c) for c in fiveg_cells)

    if lte_cells and fiveg_cells:
        mode_tag = "(TMBB)" if has_cband_dod else "(MMBB)"
    elif lte_cells or fiveg_cells:
        mode_tag = "(SMBB)"
    else:
        mode_tag = ""

    if fiveg_cells and lte_cells:
        m = re.match(r"^(.+?)_N\d{3}[A-F]_\d+$", fiveg_cells[0])
        secondary = m.group(1) if m else fiveg_cells[0]
        return f"{node_name}(P)/{secondary}(S){mode_tag}"
    return f"{node_name}{mode_tag}"

def generate_generic_pre_post(ciq_wb, mm_objs, precheck_text, precheck_node_names):
    """Pre = nodes actually found in Pre-checks. Post = nodes actually found in CIQ Mixed Mode Info.
    Each shown only on the side it's actually present — no 'vacated'/'new' padding.
    Dual-identity nodes (MMBB/TMBB) get the (P)/(S) pairing, tagged with their real BBU Mode."""
    def node_label(row):
        primary = row.get('Node to be built as')
        e_name, g_name = row.get('eNodeB Name'), row.get('gNodeB Name')
        bbu_mode = row.get('BBU Mode')
        if is_populated(e_name) and is_populated(g_name):
            is_lte_primary = str(primary).strip().upper() == str(e_name).strip().upper()
            secondary = g_name if is_lte_primary else e_name
            return f"{primary}(P)/{secondary}(S)({bbu_mode})"
        return str(primary)

    post_nodes, labels = {}, {}
    ciq_order = []
    for row in mm_objs:
        primary = row.get('Node to be built as')
        if is_wll_node_name(primary):
            continue  # WLL node — not a real node, excluded from Post Configuration entirely
        ciq_order.append(primary)
        labels[primary] = node_label(row)
        e_name, g_name = row.get('eNodeB Name'), row.get('gNodeB Name')
        is_lte_primary = str(primary).strip().upper() == str(e_name or '').strip().upper()
        r = find_row_by_name(ciq_wb, 'eNB Info', 'eNodeB Name', e_name) if is_lte_primary else find_row_by_name(ciq_wb, 'gNB Info', 'gNodeB Name', g_name)
        if not r:
            r = find_row_by_name(ciq_wb, 'eNB Info', 'eNodeB Name', e_name) or find_row_by_name(ciq_wb, 'gNB Info', 'gNodeB Name', g_name)
        post_nodes[primary] = hw_string(r) or 'NOT FOUND'

    # order: CIQ order first (so Pre and Post always list shared nodes in the same sequence),
    # then any Pre-only nodes (e.g. a fully vacated node) appended after
    _, pre_nodes_set = extract_precheck_sectors(precheck_text)
    ordered_names = list(ciq_order) + [n for n in precheck_node_names if n not in ciq_order and not is_wll_node_name(n)]

    pre_nodes = {}
    for name in ordered_names:
        if name in pre_nodes_set:  # only include nodes actually confirmed present in Pre-checks
            pre_nodes[name] = pre_hw_string(precheck_text, name) or "NOT FOUND"

    def lbl(n):
        return labels.get(n, n)

    pre_parts = [f"{pre_node_label(precheck_text, n)}({hw})" for n, hw in pre_nodes.items()]
    post_parts = [f"{lbl(n)}({hw})" for n, hw in post_nodes.items()]
    return " + ".join(pre_parts), " + ".join(post_parts)


def push_all_controller_siad_rows(siad_rows, edp_index, controller_objs):
    """For every 6610 controller in the CIQ's Controller Info, add its SIAD row (ANCEQ_* columns)
    and track whether EDP actually has it published. Returns {controller_id: bool}."""
    found_status = {}
    for r in controller_objs:
        if str(r.get("Controller", "")).strip() == "6610":
            ctrl_id = r.get("Controller ID")
            found_status[ctrl_id] = push_controller_siad_row(siad_rows, edp_index, ctrl_id)
    return found_status


# ============================================================
# GENERATOR: N2E (Nokia-to-Ericsson) — no Pre-checks at all, greenfield-style build.
# Pre Configuration is always the fixed string "Nokia"; Post derived from CIQ as usual.
# ============================================================

def n2e_node_type(row):
    """LTE-only (no gNBId) / 5G-only (no eNBId) / MMBB / TMBB (TRIMODE), from Mixed Mode Info."""
    has_lte = is_populated(row.get("eNBId")) or is_populated(row.get("eNodeB Name"))
    has_5g = is_populated(row.get("gNBId")) or is_populated(row.get("gNodeB Name"))
    if has_lte and not has_5g:
        return "LTE"
    if has_5g and not has_lte:
        return "5G"
    bbu_mode = str(row.get("BBU Mode", "")).strip().upper()
    if bbu_mode == "MMBB":
        return "MMBB"
    if bbu_mode == "TMBB":
        return "TRIMODE"
    return None


def fill_node_template_n2e(template_text, row, edp_index, user_id, date_str, controller_objs, summary_rows, log):
    """N2E placeholder fill — confirmed mapping, distinct from MCA's fill_node_template.
    xxOAMIPAddressxx always from EDP ipv6_enodeb_oam_ip matched by Primary ID, regardless of node type.
    5G-side EDP values (bearer/SIAD/Vlan) always matched by gNodeB Name, not by whichever identity
    is Primary — confirmed explicitly for MMBB/TRIMODE, and holds trivially for 5G-only too."""
    node_type = n2e_node_type(row)
    primary = row.get("Node to be built as")
    lte_name, gnb_name = row.get("eNodeB Name"), row.get("gNodeB Name")
    tpl = template_text

    def sub(placeholder, value, note=""):
        nonlocal tpl
        if is_populated(value):
            tpl = tpl.replace(placeholder, str(value))
            summary_rows.append({"Item": f"{primary} · {placeholder}", "Source": note, "Value": value, "Note": ""})
            log(f"✓ {primary} · {placeholder} -> {value}")
        else:
            summary_rows.append({"Item": f"{primary} · {placeholder}", "Source": note, "Value": "NOT FOUND", "Note": ""})
            log(f"✗ {primary} · {placeholder} -> NOT FOUND")

    for ph in ("xSite_IDx", "xSITE_IDx", "xxSiteIdxx"):
        sub(ph, primary, "Primary ID")
    sub("xxUserIDxx", user_id, "manual")
    sub("xDatex", date_str, "manual")

    oam_row = edp_row_for(edp_index, primary)
    sub("xxOAMIPAddressxx", edp_get(edp_index, oam_row, "ipv6_enodeb_oam_ip"), "EDP · ipv6_enodeb_oam_ip (by Primary ID)")

    if node_type == "LTE":
        lte_row = edp_row_for(edp_index, primary)  # LTE-only: Primary == eNodeB Name
        sub("xsecondary_IPV6_ENODEB_BEARER_IPx", edp_get(edp_index, lte_row, "ipv6_enodeb_bearer_ip"), "EDP · ipv6_enodeb_bearer_ip (LTE site ID)")
        sub("xLTE_IPV6_SIAD_BEARER_IPx", edp_get(edp_index, lte_row, "ipv6_siad_bearer_ip_def_router"), "EDP · ipv6_siad_bearer_ip_def_router")
        sub("xLTE_Vlan_IDx", edp_get(edp_index, lte_row, "bearer_enodeb_sb_vlan_id"), "EDP · bearer_enodeb_sb_vlan_id")

    elif node_type == "5G":
        sub("xgNBIdx", row.get("gNBId"), "CIQ · Mixed Mode Info")
        sub("xgNB_Namex", gnb_name, "CIQ · Mixed Mode Info")
        gnb_row = edp_row_for(edp_index, primary)  # 5G-only: Primary == gNodeB Name
        bearer_ip = edp_get(edp_index, gnb_row, "ipv6_enodeb_bearer_ip")
        sub("xSecondary_IPV6_ENODEB_BEARER_IPx", bearer_ip, "EDP · ipv6_enodeb_bearer_ip (5G Primary ID)")
        sub("xsecondary_IPV6_ENODEB_BEARER_IPx", bearer_ip, "EDP · ipv6_enodeb_bearer_ip (5G Primary ID)")
        sub("x5G_IPV6_SIAD_BEARER_IPx", edp_get(edp_index, gnb_row, "ipv6_siad_bearer_ip_def_router"), "EDP · ipv6_siad_bearer_ip_def_router")
        sub("x5G_Vlan_IDx", edp_get(edp_index, gnb_row, "bearer_enodeb_sb_vlan_id"), "EDP · bearer_enodeb_sb_vlan_id")

    elif node_type in ("MMBB", "TRIMODE"):
        sub("xgNBIdx", row.get("gNBId"), "CIQ · Mixed Mode Info")
        sub("xgNB_Namex", gnb_name, "CIQ · Mixed Mode Info")
        gnb_row = edp_row_for(edp_index, gnb_name)  # always matched by gNodeB Name, regardless of Primary/Secondary
        bearer_ip = edp_get(edp_index, gnb_row, "ipv6_enodeb_bearer_ip")
        sub("xSecondary_IPV6_ENODEB_BEARER_IPx", bearer_ip, "EDP · ipv6_enodeb_bearer_ip (by gNodeB Name)")
        sub("xsecondary_IPV6_ENODEB_BEARER_IPx", bearer_ip, "EDP · ipv6_enodeb_bearer_ip (by gNodeB Name)")
        sub("x5G_IPV6_SIAD_BEARER_IPx", edp_get(edp_index, gnb_row, "ipv6_siad_bearer_ip_def_router"), "EDP · ipv6_siad_bearer_ip_def_router")
        sub("x5G_Vlan_IDx", edp_get(edp_index, gnb_row, "bearer_enodeb_sb_vlan_id"), "EDP · bearer_enodeb_sb_vlan_id")
        ctrl_rows = [r for r in controller_objs if str(r.get("Controller", "")).strip() == "6610"]
        if ctrl_rows:
            sub("xController_IDX", ctrl_rows[0].get("Controller ID"), "CIQ · Controller Info")

    return tpl


def check_sa_conversion(ciq_wb, node_id):
    """SA Conversion: is this node's ID present anywhere in the CIQ's NR_SA tab?"""
    if "NR_SA" not in ciq_wb.sheetnames or not node_id:
        return False
    for row in ciq_wb["NR_SA"].iter_rows(values_only=True):
        if any(str(c).strip() == str(node_id).strip() for c in row if c is not None):
            return True
    return False


def generate_n2e(ciq_wb, edp_index, controller_objs, mm_objs, user_id, date_str, log):
    summary_rows, siad_rows, outputs = [], [], []
    tpl_paths = {"LTE": TPL_N2E_LTE, "5G": TPL_N2E_5G, "MMBB": TPL_N2E_MMBB, "TRIMODE": TPL_N2E_TRIMODE}
    sa_conversion_nodes = []

    for row in mm_objs:
        node_type = n2e_node_type(row)
        primary = row.get("Node to be built as")
        if node_type is None:
            summary_rows.append({"Item": f"Node: {primary}", "Source": "node type detection", "Value": "skipped", "Note": "couldn't determine LTE/5G/MMBB/TRIMODE"})
            log(f"· {primary}: could not determine node type, skipped")
            continue
        tpl_path = tpl_paths[node_type]
        if not tpl_path.exists():
            summary_rows.append({"Item": f"Node: {primary}", "Source": f"N2E {node_type} template", "Value": "NOT FOUND", "Note": f"expected file not in templates/N2E/: {tpl_path.name} — this node type's template hasn't been uploaded yet"})
            log(f"✗ {primary}: N2E {node_type} template file not found, skipped")
            continue
        tpl_text = tpl_path.read_text(encoding="utf-8")
        tpl = fill_node_template_n2e(tpl_text, row, edp_index, user_id, date_str, controller_objs, summary_rows, log)
        outputs.append((f"{primary}_N2E_{node_type}_Integration_Filled.txt", tpl))
        push_siad_row(siad_rows, edp_index, primary)
        if check_sa_conversion(ciq_wb, primary):
            sa_conversion_nodes.append(primary)

    controller_edp_found = push_all_controller_siad_rows(siad_rows, edp_index, controller_objs)
    add_outputs, add_summary = generate_6610(controller_objs, user_id, date_str, log, controller_edp_found)
    outputs += add_outputs
    summary_rows += add_summary
    dss_outputs, dss_summary, dss_labels = generate_dss(ciq_wb, mm_objs, user_id, date_str, log)
    outputs += dss_outputs
    summary_rows += dss_summary
    idl_outputs, idl_summary, idl_scope_lines = generate_idl_connections(
        ciq_wb, mm_objs, user_id, date_str, log, template_dir=TDIR_N2E_IDL, registry=N2E_IDL_TEMPLATE_REGISTRY)
    outputs += idl_outputs
    summary_rows += idl_summary

    binary_outputs = [(f"Final_Connections_{mm_objs[0].get('Node to be built as','site')}.xlsx", generate_final_connections(ciq_wb, mm_objs))] if mm_objs else []

    pre_line = "Nokia"
    post_parts = []
    for row in mm_objs:
        primary = row.get("Node to be built as")
        e_name, g_name = row.get("eNodeB Name"), row.get("gNodeB Name")
        is_lte_primary = str(primary).strip().upper() == str(e_name or "").strip().upper()
        r = find_row_by_name(ciq_wb, "eNB Info", "eNodeB Name", e_name) if is_lte_primary else find_row_by_name(ciq_wb, "gNB Info", "gNodeB Name", g_name)
        if not r:
            r = find_row_by_name(ciq_wb, "eNB Info", "eNodeB Name", e_name) or find_row_by_name(ciq_wb, "gNB Info", "gNodeB Name", g_name)
        hw = hw_string(r) or "NOT FOUND"
        if is_populated(e_name) and is_populated(g_name):
            secondary = g_name if is_lte_primary else e_name
            bbu_mode = row.get("BBU Mode")
            post_parts.append(f"{primary}(P)/{secondary}(S)({bbu_mode})({hw})")
        else:
            post_parts.append(f"{primary}({hw})")
    post_line = " + ".join(post_parts)

    # Carrier ADD — no Pre-checks for N2E, so every cell in the CIQ counts as an addition
    added = {}
    eutran_objs = sheet_objs(ciq_wb["eUtran Parameters"]) if "eUtran Parameters" in ciq_wb.sheetnames else []
    fiveg_objs = sheet_objs(ciq_wb["5G Info"]) if "5G Info" in ciq_wb.sheetnames else []
    for row in mm_objs:
        node = row.get("Node to be built as")
        e_name, g_name = row.get("eNodeB Name"), row.get("gNodeB Name")
        cells = []
        for r in eutran_objs:
            c = r.get("EutranCellFDDId")
            if c and e_name and str(c).startswith(str(e_name)):
                cells.append(c)
        for r in fiveg_objs:
            c = r.get("NRCellDU")
            if c and g_name and str(c).startswith(str(g_name)):
                cells.append(c)
        if cells:
            added[node] = cells

    classification = {"added": added, "moved": [], "deleted_sectors": {}, "deleted_nodes": [], "retuned": []}
    scope_of_work_lines = format_scope_of_work(classification, controller_objs, dss_labels, controller_edp_found)
    for node in sa_conversion_nodes:
        scope_of_work_lines.append(f"SA conversion.\t{node}")
    scope_of_work_lines += idl_scope_lines
    ngs_summary, ngs_scope_lines = generate_ngs_checks(ciq_wb, mm_objs, log)
    summary_rows += ngs_summary
    scope_of_work_lines += ngs_scope_lines
    ngs_tpl_outputs, ngs_tpl_summary = generate_ngs_template_output(ciq_wb, mm_objs, user_id, date_str, log)
    outputs += ngs_tpl_outputs
    summary_rows += ngs_tpl_summary

    return summary_rows, pre_line, post_line, siad_rows, outputs, binary_outputs, scope_of_work_lines


# ============================================================
# GENERATOR: NSB — no Pre-checks (Pre Configuration is always the fixed string "NA").
# Only 2 templates (MMBB, TRIMODE) — no LTE-only/5G-only variants, per the blueprint.
# Same confirmed placeholder mapping as N2E's MMBB/TRIMODE, minus the controller ID field
# (NSB templates don't fill xController_IDX directly — 6610 is purely the universal add-on here too).
# ============================================================

def nsb_node_type(row):
    bbu_mode = str(row.get("BBU Mode", "")).strip().upper()
    if bbu_mode == "MMBB":
        return "MMBB"
    if bbu_mode == "TMBB":
        return "TRIMODE"
    if is_smbb_lte_primary(row):
        return "SMBB"
    return None


def fill_node_template_nsb(template_text, row, edp_index, user_id, date_str, summary_rows, log):
    """NSB placeholder fill — confirmed identical to N2E's MMBB/TRIMODE mapping, minus xController_IDX."""
    primary = row.get("Node to be built as")
    gnb_name = row.get("gNodeB Name")
    tpl = template_text

    def sub(placeholder, value, note=""):
        nonlocal tpl
        if is_populated(value):
            tpl = tpl.replace(placeholder, str(value))
            summary_rows.append({"Item": f"{primary} · {placeholder}", "Source": note, "Value": value, "Note": ""})
            log(f"✓ {primary} · {placeholder} -> {value}")
        else:
            summary_rows.append({"Item": f"{primary} · {placeholder}", "Source": note, "Value": "NOT FOUND", "Note": ""})
            log(f"✗ {primary} · {placeholder} -> NOT FOUND")

    sub("xxSiteIdxx", primary, "Primary ID")
    sub("xxUserIDxx", user_id, "manual")
    sub("xDatex", date_str, "manual")
    sub("xgNBIdx", row.get("gNBId"), "CIQ · Mixed Mode Info")
    sub("xgNB_Namex", gnb_name, "CIQ · Mixed Mode Info")

    gnb_row = edp_row_for(edp_index, gnb_name)  # always matched by gNodeB Name, regardless of Primary/Secondary
    bearer_ip = edp_get(edp_index, gnb_row, "ipv6_enodeb_bearer_ip")
    sub("xsecondary_IPV6_ENODEB_BEARER_IPx", bearer_ip, "EDP · ipv6_enodeb_bearer_ip (by gNodeB Name)")
    sub("x5G_IPV6_SIAD_BEARER_IPx", edp_get(edp_index, gnb_row, "ipv6_siad_bearer_ip_def_router"), "EDP · ipv6_siad_bearer_ip_def_router")
    sub("x5G_Vlan_IDx", edp_get(edp_index, gnb_row, "bearer_enodeb_sb_vlan_id"), "EDP · bearer_enodeb_sb_vlan_id")

    return tpl


def generate_nsb(ciq_wb, edp_index, controller_objs, mm_objs, user_id, date_str, log):
    summary_rows, siad_rows, outputs = [], [], []
    tpl_paths = {"MMBB": TPL_NSB_MMBB, "TRIMODE": TPL_NSB_TRIMODE, "SMBB": TPL_NSB_SMBB_LTE}
    sa_conversion_nodes = []

    for row in mm_objs:
        node_type = nsb_node_type(row)
        primary = row.get("Node to be built as")
        if node_type is None:
            summary_rows.append({"Item": f"Node: {primary}", "Source": "node type detection", "Value": "skipped", "Note": "NSB only supports MMBB/TMBB/SMBB-LTE-primary — not 5G-only"})
            log(f"· {primary}: BBU Mode not MMBB/TMBB/SMBB-LTE-primary, skipped")
            continue
        tpl_path = tpl_paths[node_type]
        if not tpl_path.exists():
            summary_rows.append({"Item": f"Node: {primary}", "Source": f"NSB {node_type} template", "Value": "NOT FOUND", "Note": f"expected file not in templates/NSB/: {tpl_path.name}"})
            log(f"✗ {primary}: NSB {node_type} template file not found, skipped")
            continue
        tpl_text = tpl_path.read_text(encoding="utf-8")
        if node_type == "SMBB":
            tpl = fill_node_template_smbb_lte(tpl_text, row, mm_objs, edp_index, user_id, date_str, summary_rows, log)
        else:
            tpl = fill_node_template_nsb(tpl_text, row, edp_index, user_id, date_str, summary_rows, log)
        outputs.append((f"{primary}_NSB_{node_type}_Integration_Filled.txt", tpl))
        push_siad_row(siad_rows, edp_index, primary)
        if check_sa_conversion(ciq_wb, primary):
            sa_conversion_nodes.append(primary)

    controller_edp_found = push_all_controller_siad_rows(siad_rows, edp_index, controller_objs)
    add_outputs, add_summary = generate_6610(controller_objs, user_id, date_str, log, controller_edp_found)
    outputs += add_outputs
    summary_rows += add_summary
    dss_outputs, dss_summary, dss_labels = generate_dss(ciq_wb, mm_objs, user_id, date_str, log)
    outputs += dss_outputs
    summary_rows += dss_summary
    idl_outputs, idl_summary, idl_scope_lines = generate_idl_connections(ciq_wb, mm_objs, user_id, date_str, log)
    outputs += idl_outputs
    summary_rows += idl_summary
    ngs_summary, ngs_scope_lines = generate_ngs_checks(ciq_wb, mm_objs, log)
    summary_rows += ngs_summary
    ngs_tpl_outputs, ngs_tpl_summary = generate_ngs_template_output(ciq_wb, mm_objs, user_id, date_str, log)
    outputs += ngs_tpl_outputs
    summary_rows += ngs_tpl_summary

    binary_outputs = [(f"Final_Connections_{mm_objs[0].get('Node to be built as','site')}.xlsx", generate_final_connections(ciq_wb, mm_objs))] if mm_objs else []

    pre_line = "NA"
    post_parts = []
    for row in mm_objs:
        primary = row.get("Node to be built as")
        e_name, g_name = row.get("eNodeB Name"), row.get("gNodeB Name")
        is_lte_primary = str(primary).strip().upper() == str(e_name or "").strip().upper()
        r = find_row_by_name(ciq_wb, "eNB Info", "eNodeB Name", e_name) if is_lte_primary else find_row_by_name(ciq_wb, "gNB Info", "gNodeB Name", g_name)
        if not r:
            r = find_row_by_name(ciq_wb, "eNB Info", "eNodeB Name", e_name) or find_row_by_name(ciq_wb, "gNB Info", "gNodeB Name", g_name)
        hw = hw_string(r) or "NOT FOUND"
        if is_populated(e_name) and is_populated(g_name):
            secondary = g_name if is_lte_primary else e_name
            bbu_mode = row.get("BBU Mode")
            post_parts.append(f"{primary}(P)/{secondary}(S)({bbu_mode})({hw})")
        else:
            post_parts.append(f"{primary}({hw})")
    post_line = " + ".join(post_parts)

    # Carrier ADD — no Pre-checks for NSB, so every cell in the CIQ counts as an addition (same rule as N2E)
    added = {}
    eutran_objs = sheet_objs(ciq_wb["eUtran Parameters"]) if "eUtran Parameters" in ciq_wb.sheetnames else []
    fiveg_objs = sheet_objs(ciq_wb["5G Info"]) if "5G Info" in ciq_wb.sheetnames else []
    for row in mm_objs:
        node = row.get("Node to be built as")
        e_name, g_name = row.get("eNodeB Name"), row.get("gNodeB Name")
        cells = []
        for r in eutran_objs:
            c = r.get("EutranCellFDDId")
            if c and e_name and str(c).startswith(str(e_name)):
                cells.append(c)
        for r in fiveg_objs:
            c = r.get("NRCellDU")
            if c and g_name and str(c).startswith(str(g_name)):
                cells.append(c)
        if cells:
            added[node] = cells

    classification = {"added": added, "moved": [], "deleted_sectors": {}, "deleted_nodes": [], "retuned": []}
    scope_of_work_lines = format_scope_of_work(classification, controller_objs, dss_labels, controller_edp_found)
    for node in sa_conversion_nodes:
        scope_of_work_lines.append(f"SA conversion.\t{node}")
    scope_of_work_lines += idl_scope_lines
    scope_of_work_lines += ngs_scope_lines

    return summary_rows, pre_line, post_line, siad_rows, outputs, binary_outputs, scope_of_work_lines


def generate_mca(ciq_wb, edp_index, controller_objs, mm_objs, user_id, date_str, precheck_text, log):
    summary_rows, siad_rows, outputs = [], [], []
    tpl_mmbb = TPL_MMBB.read_text(encoding="utf-8")
    tpl_tmbb = TPL_TMBB.read_text(encoding="utf-8")
    tpl_smbb_lte = TPL_SMBB_LTE.read_text(encoding="utf-8") if TPL_SMBB_LTE.exists() else None

    for row in mm_objs:
        bbu_mode = str(row.get("BBU Mode", "")).strip()
        site_id = row.get("Node to be built as")
        if bbu_mode == "MMBB":
            tpl = fill_node_template(tpl_mmbb, row, edp_index, user_id, date_str, summary_rows, log)
            outputs.append((f"{site_id}_MMBB_Integration_Filled.txt", tpl))
            push_siad_row(siad_rows, edp_index, site_id)
        elif bbu_mode == "TMBB":
            tpl = fill_node_template(tpl_tmbb, row, edp_index, user_id, date_str, summary_rows, log)
            outputs.append((f"{site_id}_TMBB_Integration_Filled.txt", tpl))
            push_siad_row(siad_rows, edp_index, site_id)
        elif is_smbb_lte_primary(row):
            if tpl_smbb_lte is None:
                summary_rows.append({"Item": f"Node: {site_id}", "Source": "SMBB LTE template", "Value": "NOT FOUND", "Note": f"expected file not in templates/MCA/: {TPL_SMBB_LTE.name}"})
                log(f"✗ {site_id}: SMBB LTE template file not found, skipped")
                push_siad_row(siad_rows, edp_index, site_id)
                continue
            tpl = fill_node_template_smbb_lte(tpl_smbb_lte, row, mm_objs, edp_index, user_id, date_str, summary_rows, log)
            outputs.append((f"{site_id}_SMBB_LTE_Integration_Filled.txt", tpl))
            push_siad_row(siad_rows, edp_index, site_id)
        else:
            note = "SMBB but 5G-primary — no template yet" if bbu_mode == "SMBB" else "not MMBB, TMBB, or SMBB-LTE-primary"
            summary_rows.append({"Item": f"Node: {site_id}", "Source": f"BBU Mode = {bbu_mode}", "Value": "skipped", "Note": note})
            log(f"· {site_id}: BBU Mode = {bbu_mode}, skipped ({note})")
            push_siad_row(siad_rows, edp_index, site_id)

    controller_edp_found = push_all_controller_siad_rows(siad_rows, edp_index, controller_objs)
    add_outputs, add_summary = generate_6610(controller_objs, user_id, date_str, log, controller_edp_found)
    outputs += add_outputs
    summary_rows += add_summary
    dss_outputs, dss_summary, dss_labels = generate_dss(ciq_wb, mm_objs, user_id, date_str, log)
    outputs += dss_outputs
    summary_rows += dss_summary
    has_cran_node = any(str(r.get("Node to be built as", "")).strip().upper().endswith("F") for r in mm_objs)
    if has_cran_node:
        idl_outputs, idl_summary, idl_scope_lines = generate_idl_connections(
            ciq_wb, mm_objs, user_id, date_str, log, template_dir=TDIR_MCA_IDL_CRAN, registry=MCA_CRAN_IDL_REGISTRY)
    else:
        idl_outputs, idl_summary, idl_scope_lines = generate_idl_connections(ciq_wb, mm_objs, user_id, date_str, log)
    outputs += idl_outputs
    summary_rows += idl_summary
    ngs_summary, ngs_scope_lines = generate_ngs_checks(ciq_wb, mm_objs, log)
    summary_rows += ngs_summary
    ngs_tpl_outputs, ngs_tpl_summary = generate_ngs_template_output(ciq_wb, mm_objs, user_id, date_str, log)
    outputs += ngs_tpl_outputs
    summary_rows += ngs_tpl_summary

    binary_outputs = [(f"Final_Connections_{mm_objs[0].get('Node to be built as','site')}.xlsx", generate_final_connections(ciq_wb, mm_objs))] if mm_objs else []
    pre_fibers_bytes = generate_pre_fibers(precheck_text)
    if pre_fibers_bytes and mm_objs:
        binary_outputs.append((f"Pre_Fibers_{mm_objs[0].get('Node to be built as','site')}.xlsx", pre_fibers_bytes))

    _, pre_nodes_found = extract_precheck_sectors(precheck_text)
    ciq_node_names = {r.get("Node to be built as") for r in mm_objs if r.get("Node to be built as")}
    pre_line, post_line = generate_generic_pre_post(ciq_wb, mm_objs, precheck_text, pre_nodes_found | ciq_node_names)

    classification = classify_carriers(ciq_wb, mm_objs, precheck_text)
    radio_swaps = classify_radio_swaps(precheck_text, ciq_wb)
    scope_of_work_lines = format_scope_of_work(classification, controller_objs, dss_labels, controller_edp_found, radio_swaps)
    scope_of_work_lines += idl_scope_lines
    scope_of_work_lines += ngs_scope_lines
    pc_outputs, pc_summary, pc_scope_lines = generate_port_conversion_checks(ciq_wb, mm_objs, edp_index, precheck_text, log)
    outputs += pc_outputs
    summary_rows += pc_summary
    scope_of_work_lines += pc_scope_lines
    data2_scope_lines = generate_data2_testing_checks(ciq_wb, mm_objs, precheck_text, log)
    scope_of_work_lines += data2_scope_lines

    del_outputs, del_summary = generate_node_deletion_templates(ciq_wb, mm_objs, edp_index, classification, precheck_text, user_id, date_str, log)
    outputs += del_outputs
    summary_rows += del_summary

    return summary_rows, pre_line, post_line, siad_rows, outputs, binary_outputs, scope_of_work_lines


# ============================================================
# GENERATOR: CENM (always cENM_TRIMODE template, for TMBB-mode nodes)
# ============================================================

def generate_cenm(ciq_wb, edp_index, controller_objs, mm_objs, user_id, date_str, precheck_text, log):
    summary_rows, siad_rows, outputs = [], [], []
    tpl_cenm_tmbb = TPL_CENM.read_text(encoding="utf-8")
    tpl_cenm_mmbb = TPL_CENM_MMBB.read_text(encoding="utf-8") if TPL_CENM_MMBB.exists() else None
    # SMBB-LTE-primary shares the exact same template file as MCA (confirmed) — TPL_SMBB_LTE.
    tpl_cenm_smbb_lte = TPL_SMBB_LTE.read_text(encoding="utf-8") if TPL_SMBB_LTE.exists() else None

    tmbb_rows = [r for r in mm_objs if str(r.get("BBU Mode", "")).strip() == "TMBB"]
    mmbb_rows = [r for r in mm_objs if str(r.get("BBU Mode", "")).strip() == "MMBB"]
    smbb_lte_rows = [r for r in mm_objs if is_smbb_lte_primary(r)]
    if not tmbb_rows and not mmbb_rows and not smbb_lte_rows:
        summary_rows.append({"Item": "Node identification", "Source": "CIQ · Mixed Mode Info", "Value": "NOT FOUND", "Note": "CENM expects a BBU Mode = TMBB, MMBB, or SMBB-LTE-primary row"})
        return summary_rows, None, None, siad_rows, outputs, [], []

    for row in tmbb_rows:
        site_id = row.get("Node to be built as")
        tpl = fill_node_template(tpl_cenm_tmbb, row, edp_index, user_id, date_str, summary_rows, log)
        outputs.append((f"{site_id}_cENM_TMBB_Integration_Filled.txt", tpl))
        push_siad_row(siad_rows, edp_index, site_id)

    for row in mmbb_rows:
        site_id = row.get("Node to be built as")
        if tpl_cenm_mmbb is None:
            summary_rows.append({"Item": f"Node: {site_id}", "Source": "CENM MMBB template", "Value": "NOT FOUND", "Note": f"expected file not in templates/MCA/: {TPL_CENM_MMBB.name}"})
            log(f"✗ {site_id}: CENM MMBB template file not found, skipped")
            push_siad_row(siad_rows, edp_index, site_id)
            continue
        tpl = fill_node_template(tpl_cenm_mmbb, row, edp_index, user_id, date_str, summary_rows, log)
        outputs.append((f"{site_id}_cENM_MMBB_Integration_Filled.txt", tpl))
        push_siad_row(siad_rows, edp_index, site_id)

    for row in smbb_lte_rows:
        site_id = row.get("Node to be built as")
        if tpl_cenm_smbb_lte is None:
            summary_rows.append({"Item": f"Node: {site_id}", "Source": "CENM SMBB-LTE template", "Value": "NOT FOUND", "Note": f"expected file not in templates/MCA/: {TPL_SMBB_LTE.name}"})
            log(f"✗ {site_id}: CENM SMBB-LTE template file not found, skipped")
            push_siad_row(siad_rows, edp_index, site_id)
            continue
        tpl = fill_node_template_smbb_lte(tpl_cenm_smbb_lte, row, mm_objs, edp_index, user_id, date_str, summary_rows, log)
        outputs.append((f"{site_id}_cENM_SMBB_LTE_Integration_Filled.txt", tpl))
        push_siad_row(siad_rows, edp_index, site_id)

    for row in mm_objs:
        if str(row.get("BBU Mode", "")).strip() not in ("TMBB", "MMBB") and not is_smbb_lte_primary(row):
            site_id = row.get("Node to be built as")
            summary_rows.append({"Item": f"Node: {site_id}", "Source": f"BBU Mode = {row.get('BBU Mode')}", "Value": "skipped for template", "Note": "not TMBB/MMBB/SMBB-LTE-primary — still included in Pre/Post and SIAD"})
            push_siad_row(siad_rows, edp_index, site_id)

    controller_edp_found = push_all_controller_siad_rows(siad_rows, edp_index, controller_objs)
    add_outputs, add_summary = generate_6610(controller_objs, user_id, date_str, log, controller_edp_found)
    outputs += add_outputs
    summary_rows += add_summary
    dss_outputs, dss_summary, dss_labels = generate_dss(ciq_wb, mm_objs, user_id, date_str, log)
    outputs += dss_outputs
    summary_rows += dss_summary
    has_cran_node = any(str(r.get("Node to be built as", "")).strip().upper().endswith("F") for r in mm_objs)
    if has_cran_node:
        idl_outputs, idl_summary, idl_scope_lines = generate_idl_connections(
            ciq_wb, mm_objs, user_id, date_str, log, template_dir=TDIR_MCA_IDL_CRAN, registry=MCA_CRAN_IDL_REGISTRY)
    else:
        idl_outputs, idl_summary, idl_scope_lines = generate_idl_connections(ciq_wb, mm_objs, user_id, date_str, log)
    outputs += idl_outputs
    summary_rows += idl_summary
    ngs_summary, ngs_scope_lines = generate_ngs_checks(ciq_wb, mm_objs, log)
    summary_rows += ngs_summary
    ngs_tpl_outputs, ngs_tpl_summary = generate_ngs_template_output(ciq_wb, mm_objs, user_id, date_str, log)
    outputs += ngs_tpl_outputs
    summary_rows += ngs_tpl_summary

    binary_outputs = [(f"Final_Connections_{mm_objs[0].get('Node to be built as','site')}.xlsx", generate_final_connections(ciq_wb, mm_objs))] if mm_objs else []
    pre_fibers_bytes = generate_pre_fibers(precheck_text)
    if pre_fibers_bytes and mm_objs:
        binary_outputs.append((f"Pre_Fibers_{mm_objs[0].get('Node to be built as','site')}.xlsx", pre_fibers_bytes))

    _, pre_nodes_found = extract_precheck_sectors(precheck_text)
    ciq_node_names = {r.get("Node to be built as") for r in mm_objs if r.get("Node to be built as")}
    pre_line, post_line = generate_generic_pre_post(ciq_wb, mm_objs, precheck_text, pre_nodes_found | ciq_node_names)

    classification = classify_carriers(ciq_wb, mm_objs, precheck_text)
    radio_swaps = classify_radio_swaps(precheck_text, ciq_wb)
    scope_of_work_lines = format_scope_of_work(classification, controller_objs, dss_labels, controller_edp_found, radio_swaps)
    scope_of_work_lines += idl_scope_lines
    scope_of_work_lines += ngs_scope_lines
    pc_outputs, pc_summary, pc_scope_lines = generate_port_conversion_checks(ciq_wb, mm_objs, edp_index, precheck_text, log)
    outputs += pc_outputs
    summary_rows += pc_summary
    scope_of_work_lines += pc_scope_lines
    data2_scope_lines = generate_data2_testing_checks(ciq_wb, mm_objs, precheck_text, log)
    scope_of_work_lines += data2_scope_lines

    del_outputs, del_summary = generate_node_deletion_templates(ciq_wb, mm_objs, edp_index, classification, precheck_text, user_id, date_str, log)
    outputs += del_outputs
    summary_rows += del_summary

    return summary_rows, pre_line, post_line, siad_rows, outputs, binary_outputs, scope_of_work_lines


# ============================================================
# GENERATOR: CRAN (Trip 1 / Trip 2 / NSA — shared logic, per-variant options)
# ============================================================

def generate_cran(ciq_wb, edp_index, controller_objs, mm_objs, user_id, date_str, precheck_text, log, tpl_path, include_source_poles, needs_6673, out_name):
    summary_rows, siad_rows, outputs = [], [], []
    tpl = tpl_path.read_text(encoding="utf-8")

    macro = next((r for r in mm_objs if str(r.get("BBU Mode", "")).strip() == "MMBB"), None)
    lte = next((r for r in mm_objs if str(r.get("BBU Mode", "")).strip() == "SMBB"
                and is_populated(r.get("eNBId")) and not is_populated(r.get("gNBId"))), None)
    target = next((r for r in mm_objs if str(r.get("BBU Mode", "")).strip() == "SMBB"
                   and is_populated(r.get("gNBId")) and not is_populated(r.get("eNBId"))
                   and str(r.get("gNodeB Name", "")).strip().upper().endswith("F")), None)

    if not (macro and lte and target):
        summary_rows.append({"Item": "Node identification", "Source": "CIQ · Mixed Mode Info", "Value": "incomplete",
                              "Note": f"MMBB={bool(macro)} LTE={bool(lte)} CRAN={bool(target)}"})
        return summary_rows, None, None, siad_rows, outputs, [], []

    if "Sector Del_Movement" not in ciq_wb.sheetnames:
        summary_rows.append({"Item": "Sector Del_Movement tab", "Source": "CIQ", "Value": "NOT FOUND", "Note": "required for Source CRAN"})
        return summary_rows, None, None, siad_rows, outputs, [], []

    delmove = sheet_objs(ciq_wb["Sector Del_Movement"])
    target_name = target.get("Node to be built as")
    source_row = next((r for r in delmove if str(r.get("Target Node name", "")).strip().upper() == str(target_name).strip().upper()), None)
    source_id = source_row.get("Source Node name") if source_row else (delmove[0].get("Source Node name") if delmove else None)

    target_poles, source_poles = {}, {}
    if "5G Info" in ciq_wb.sheetnames:
        for r in sheet_objs(ciq_wb["5G Info"]):
            if str(r.get("gNB Name", "")).strip().upper() == str(target.get("gNodeB Name", "")).strip().upper():
                cell = r.get("NRCellDU")
                m = re.search(r"([A-C])_([12])$", str(cell or ""))
                if m:
                    target_poles[f"{m.group(1).upper()}_{m.group(2)}"] = cell
    for r in delmove:
        cell = r.get("Source Sector")
        m = re.search(r"([A-C])_([12])$", str(cell or ""))
        if m:
            source_poles[f"{m.group(1).upper()}_{m.group(2)}"] = cell

    for key in ["A_1", "A_2", "B_1", "B_2", "C_1", "C_2"]:
        t_token = f"xxTarget_SiteIdxx_Pole_N077{key}"
        t_val = target_poles.get(key)
        if t_val:
            tpl = tpl.replace(t_token, t_val)
            summary_rows.append({"Item": t_token, "Source": "CIQ · 5G Info · NRCellDU", "Value": t_val, "Note": ""})
        else:
            summary_rows.append({"Item": t_token, "Source": "CIQ · 5G Info · NRCellDU", "Value": "NOT PRESENT", "Note": ""})
        log(f"{'✓' if t_val else '·'} {t_token} -> {t_val or 'not present'}")

        if include_source_poles:
            s_token = f"xxSource_SiteIdxx_Pole_N077{key}"
            s_val = source_poles.get(key)
            if s_val:
                tpl = tpl.replace(s_token, s_val)
                summary_rows.append({"Item": s_token, "Source": "CIQ · Sector Del_Movement · Source Sector", "Value": s_val, "Note": ""})
            else:
                summary_rows.append({"Item": s_token, "Source": "CIQ · Sector Del_Movement · Source Sector", "Value": "NOT PRESENT", "Note": ""})
            log(f"{'✓' if s_val else '·'} {s_token} -> {s_val or 'not present'}")

    target_5g_row = edp_row_for(edp_index, target.get("gNodeB Name"))
    target_bearer = edp_get(edp_index, target_5g_row, "IPV6_ENODEB_BEARER_IP")

    fills = [
        ("xxMacro_MMBB_SiteIdxx", macro.get("Node to be built as"), "CIQ · Mixed Mode Info (MMBB)"),
        ("xxMacro_MMBB_gNB_Namexx", macro.get("gNodeB Name"), "CIQ · Mixed Mode Info (MMBB)"),
        # Confirmed real bug: the template's own legend and every actual cmedit command use
        # "xxMacro_5G_gNBIdxx" for the macro node's gNBId — "xxMacro_MMBB_gNBIdxx" (the token
        # this code filled instead) doesn't exist anywhere in any of the 3 CRAN templates, so
        # the macro's gNBId never populated in any generated CRAN rehome output.
        ("xxMacro_5G_gNBIdxx", macro.get("gNBId"), "CIQ · Mixed Mode Info (MMBB)"),
        ("xxTarget_5G_IPV6_ENODEB_BEARER_IPxx", target_bearer, "EDP · IPV6_ENODEB_BEARER_IP"),
        ("xxTarget_5G_gNBIdxx", target.get("gNBId"), "CIQ · Mixed Mode Info (Target CRAN)"),
        ("xxTarget_SiteIdxx", target.get("Node to be built as"), "CIQ · Mixed Mode Info (Target CRAN)"),
        ("xxLTE_SiteIDxx", lte.get("Node to be built as"), "CIQ · Mixed Mode Info (LTE)"),
        ("xxLTE_SiteIdxx", lte.get("Node to be built as"), "CIQ · Mixed Mode Info (LTE)"),
        ("xxSiteIDxx", target.get("Node to be built as"), "ambiguous — defaulted to Target, VERIFY"),
        # Confirmed real bug: a SEPARATE lowercase-"d" form of this token exists in Trip-1 and
        # NSA (never in Trip-2), always in a CommonBeamforming/NRSectorCarrier (5G radio tilt)
        # context — never filled at all before. Same ambiguity as xxSiteIDxx above, so defaulted
        # to Target the same way, pending confirmation.
        ("xxSiteIdxx", target.get("Node to be built as"), "ambiguous — defaulted to Target, VERIFY"),
        ("xxSource_SiteIdxx", source_id, "CIQ · Sector Del_Movement · Source Node name"),
        ("xxUserIDxx", user_id, "manual input"),
        ("xxDATExx", date_str, "manual input"),
        ("xxDatexx", date_str, "manual input"),
        ("xxdatexx", date_str, "manual input"),
    ]
    if needs_6673:
        switch_id = None
        if "Sidehaul Info" in ciq_wb.sheetnames:
            for r in sheet_objs(ciq_wb["Sidehaul Info"]):
                if str(r.get("Switch", "")).strip() == "6673":
                    switch_id = r.get("SH Switch ID")
                    break
        fills.append(("xx6673_switch_idxx", switch_id, "CIQ · Sidehaul Info (Switch = 6673 -> SH Switch ID)"))

    for token, val, src in fills:
        if val:
            tpl = tpl.replace(token, str(val))
            summary_rows.append({"Item": token, "Source": src, "Value": val, "Note": ""})
        else:
            summary_rows.append({"Item": token, "Source": src, "Value": "NOT FOUND", "Note": "left as placeholder"})
        log(f"{'✓' if val else '✗'} {token} -> {val or 'NOT FOUND'}")
    tpl = tpl.replace("xDatex", date_str)

    if "NSA" in out_name:
        summary_rows.append({"Item": "xx5G_Cell_namexx / xxFDD_namexx", "Source": "not in header legend", "Value": "n/a", "Note": "confirmed manual RF-judgment field — left untouched"})

    # Pre/Post configuration — compact line format
    lte_row = find_row_by_name(ciq_wb, "eNB Info", "eNodeB Name", lte.get("eNodeB Name"))
    lte_hw_post = hw_string(lte_row) or "NOT FOUND"
    lte_hw_pre = pre_hw_string(precheck_text, lte.get("eNodeB Name")) or "NOT FOUND"

    macro_primary = macro.get("Node to be built as")
    macro_is_primary_lte = str(macro_primary).strip().upper() == str(macro.get("eNodeB Name", "")).strip().upper()
    macro_secondary = macro.get("gNodeB Name") if macro_is_primary_lte else macro.get("eNodeB Name")
    macro_row = find_row_by_name(ciq_wb, "eNB Info", "eNodeB Name", macro.get("eNodeB Name")) or \
                find_row_by_name(ciq_wb, "gNB Info", "gNodeB Name", macro.get("gNodeB Name"))
    macro_hw_post = hw_string(macro_row) or "NOT FOUND"
    macro_hw_pre = pre_hw_string(precheck_text, macro.get("eNodeB Name")) or "NOT FOUND"

    target_row = find_row_by_name(ciq_wb, "gNB Info", "gNodeB Name", target.get("gNodeB Name"))
    target_hw = hw_string(target_row) or "NOT FOUND"
    source_hw = pre_hw_string(precheck_text, source_id) or "NOT FOUND (no Pre-checks match)"

    lte_node = lte.get("Node to be built as")
    already_listed = {str(lte_node).strip().upper(), str(macro_primary).strip().upper()}
    source_is_distinct = is_populated(source_id) and str(source_id).strip().upper() not in already_listed

    pre_line = f"{lte_node}({lte_hw_pre}) + {macro_primary}(P)/{macro_secondary}(S)(MMBB)({macro_hw_pre})"
    if source_is_distinct:
        pre_line += f" + {source_id}({source_hw})"

    post_line = f"{lte_node}({lte_hw_post}) + {macro_primary}(P)/{macro_secondary}(S)(MMBB)({macro_hw_post}) + {target.get('Node to be built as')}({target_hw})"

    push_siad_row(siad_rows, edp_index, macro.get("Node to be built as"))
    push_siad_row(siad_rows, edp_index, lte.get("Node to be built as"))
    push_siad_row(siad_rows, edp_index, target.get("Node to be built as"))

    outputs.append((f"{target.get('Node to be built as')}_{out_name}_Filled.txt", tpl))

    controller_edp_found = push_all_controller_siad_rows(siad_rows, edp_index, controller_objs)
    add_outputs, add_summary = generate_6610(controller_objs, user_id, date_str, log, controller_edp_found)
    outputs += add_outputs
    summary_rows += add_summary
    dss_outputs, dss_summary, dss_labels = generate_dss(ciq_wb, mm_objs, user_id, date_str, log)
    outputs += dss_outputs
    summary_rows += dss_summary

    # Confirmed: CRAN sites never had any IDL Connections generation at all — reuses the
    # SAME registry/templates already built for MCA sites with a CRAN-styled node
    # (MCA_CRAN_IDL_REGISTRY / templates/MCA/IDL_CRAN), since CRAN's own reference sheet
    # (uploaded IDL_Connections.xlsx, "CRAN" tab) confirms the exact same combos/templates
    # apply here — this covers Build Types L-1, L-2, L-2B, L-3B, L-10, L-11, L-12. Six other
    # documented Build Types (L-4, L-5, L-5B, L-6, L-8, L-9) have no template file in the
    # repo yet, so those combos still fall through to "IDL Template not found" until the
    # actual template content is available.
    idl_outputs, idl_summary, idl_scope_lines = generate_idl_connections(
        ciq_wb, mm_objs, user_id, date_str, log, template_dir=TDIR_MCA_IDL_CRAN, registry=MCA_CRAN_IDL_REGISTRY)
    outputs += idl_outputs
    summary_rows += idl_summary

    binary_outputs = [(f"Final_Connections_{target.get('Node to be built as','site')}.xlsx", generate_final_connections(ciq_wb, mm_objs))]
    pre_fibers_bytes = generate_pre_fibers(precheck_text)
    if pre_fibers_bytes:
        binary_outputs.append((f"Pre_Fibers_{target.get('Node to be built as','site')}.xlsx", pre_fibers_bytes))

    # CRAN's checklist now matches MCA/CENM (Carrier ADD/Delete/Move, Retune) — run the same
    # classification off the same Sector Del_Movement tab CRAN already reads for role detection.
    # Pre/Post Configuration stays CRAN's own distinct role-based format, untouched above.
    radio_swaps = classify_radio_swaps(precheck_text, ciq_wb)
    classification = classify_carriers(ciq_wb, mm_objs, precheck_text)
    raw_deleted_nodes = classification.get("deleted_nodes")  # captured before the CRAN-specific zeroing below
    classification["deleted_nodes"] = []  # every CRAN rehome vacates a source node — not a noteworthy anomaly here, unlike MCA/CENM
    scope_of_work_lines = format_scope_of_work(classification, controller_objs, dss_labels, controller_edp_found, radio_swaps)
    scope_of_work_lines += idl_scope_lines
    ngs_summary, ngs_scope_lines = generate_ngs_checks(ciq_wb, mm_objs, log)
    summary_rows += ngs_summary
    scope_of_work_lines += ngs_scope_lines
    ngs_tpl_outputs, ngs_tpl_summary = generate_ngs_template_output(ciq_wb, mm_objs, user_id, date_str, log)
    outputs += ngs_tpl_outputs
    summary_rows += ngs_tpl_summary
    pc_outputs, pc_summary, pc_scope_lines = generate_port_conversion_checks(ciq_wb, mm_objs, edp_index, precheck_text, log)
    outputs += pc_outputs
    summary_rows += pc_summary
    scope_of_work_lines += pc_scope_lines
    data2_scope_lines = generate_data2_testing_checks(ciq_wb, mm_objs, precheck_text, log)
    scope_of_work_lines += data2_scope_lines

    del_outputs, del_summary = generate_node_deletion_templates(
        ciq_wb, mm_objs, edp_index, {"deleted_nodes": raw_deleted_nodes}, precheck_text, user_id, date_str, log)
    outputs += del_outputs
    summary_rows += del_summary

    return summary_rows, pre_line, post_line, siad_rows, outputs, binary_outputs, scope_of_work_lines


# ============================================================
# CHECKS-PERFORMED PANEL (per-scope checklist, matched against the blueprint) —
# derives pass/fail per check from the already-computed scope_lines rather than
# threading new return values through every generate_* function.
# ============================================================

SCOPE_CHECKLIST = {
    "CRAN": ["Carrier ADD", "Carrier delete", "Carrier moving", "IDL Connections", "DSS checks", "Radio swap", "Retune", "6610 Present", "NGS Checks", "Port Conversion"],
    "MCA": ["Carrier ADD", "Carrier delete", "Carrier moving", "IDL Connections", "DSS checks", "Radio swap", "Retune", "6610 Present", "NGS Checks", "Port Conversion"],
    "CENM": ["Carrier ADD", "Carrier delete", "Carrier moving", "IDL Connections", "DSS checks", "Radio swap", "Retune", "6610 Present", "NGS Checks", "Port Conversion"],
    "N2E": ["Carrier ADD", "IDL Connections", "DSS checks", "6610 Present", "SA Conversion", "NGS Checks"],
    "NSB": ["Carrier ADD", "IDL Connections", "DSS checks", "NGS Checks", "6610 Present", "SA Conversion"],
}

# label -> function(line) -> True if that scope_lines entry counts as a "found/applicable" hit for this check
CHECK_MATCHERS = {
    "Carrier ADD": lambda l: l.startswith("Integration:"),
    "Carrier delete": lambda l: l.startswith("Deleted Node from ENM:") or l.startswith("Deleted Sector:"),
    "Carrier moving": lambda l: l.startswith("Moved Sectors:") and "CHECK CIQ" not in l,
    "IDL Connections": lambda l: l.startswith("IDL Connections:") and "not found" not in l.lower() and "could not determine" not in l.lower() and "missing" not in l.lower(),
    "DSS checks": lambda l: l.startswith("DSS Activation:"),
    "Radio swap": lambda l: l.startswith("Radio Swap on:"),
    "Retune": lambda l: l.startswith("Retune on:"),
    "6610 Present": lambda l: l.startswith("6610 Controller Integration:") or l.startswith("EDP is not published for the controller"),
    "SA Conversion": lambda l: l.startswith("SA conversion."),
    "NGS Checks": lambda l: l.startswith("NGS Activation on :"),
    "Port Conversion": lambda l: l.startswith("Port speed 1G to 10G conversion with MPST:"),
}

# (scope, check label) pairs that aren't wired into the tool yet — shown as not-run rather than a
# misleading "fail", since a fail here would otherwise look identical to "this site has none of these"
NOT_BUILT_YET = set()


def derive_check_status(top_scope, scope_lines):
    checklist = SCOPE_CHECKLIST.get(top_scope, [])
    lines = scope_lines or []
    results = []
    for label in checklist:
        matcher = CHECK_MATCHERS.get(label, lambda l: False)
        found = any(matcher(line) for line in lines)
        results.append({"label": label, "found": found, "not_built": (top_scope, label) in NOT_BUILT_YET})
    return results


def _build_checks_html_rows(top_scope, scope_lines):
    statuses = derive_check_status(top_scope, scope_lines)
    html_rows = []
    for s in statuses:
        if s["not_built"]:
            icon_cls, icon_char, label_cls = "fail", "\u2717", "dim"
        elif s["found"]:
            icon_cls, icon_char, label_cls = "pass", "\u2713", ""
        else:
            icon_cls, icon_char, label_cls = "fail", "\u2717", "dim"
        suffix = " (not built yet)" if s["not_built"] else ""
        html_rows.append(
            f'<div class="qkx-check-row"><div class="qkx-check-icon {icon_cls}">{icon_char}</div>'
            f'<div class="qkx-check-label {label_cls}">{s["label"]}{suffix}</div></div>'
        )
    return html_rows


def render_checks_panel_animated(container, top_scope, scope_lines):
    """Plays the reveal-row-by-row-then-fade-out animation in place at the TOP slot (where
    Pre/Post configuration will render once this finishes), then fully empties the container.
    `container` must be an st.empty() placeholder so it can be cleared afterward."""
    html_rows = _build_checks_html_rows(top_scope, scope_lines)
    with container.container():
        with st.container(border=True):
            st.subheader("Checks Performed")
            rows_ph = st.empty()
            shown = []
            for row in html_rows:
                shown.append(row)
                rows_ph.markdown('<div class="qkx-checklist">' + "".join(shown) + "</div>", unsafe_allow_html=True)
                time.sleep(0.18)
            time.sleep(0.6)  # let the completed checklist register before it fades away
            rows_ph.markdown(
                '<div class="qkx-checklist qkx-checks-fadeout">' + "".join(shown) + "</div>",
                unsafe_allow_html=True,
            )
        time.sleep(0.55)  # match the fade-out animation duration
    container.empty()  # fully removed — Pre/Post configuration then takes this spot


def render_checks_panel_static(container, top_scope, scope_lines):
    """The settled, final version — no animation — shown permanently at the bottom of the
    left column (under the processing log) once the top-slot animation has finished."""
    html_rows = _build_checks_html_rows(top_scope, scope_lines)
    with container.container():
        with st.container(border=True):
            st.subheader("Checks Performed")
            st.markdown('<div class="qkx-checklist">' + "".join(html_rows) + "</div>", unsafe_allow_html=True)








# ============================================================
# PARAMETER VERIFICATION (MCA / N2E / NSB) — CIQ vs Pre logs vs Onsite logs
# Confirmed against the blueprint (temp_blueprint.xlsx) DIRECTLY by inspecting the raw sheet
# grid (merged cells, column layout) — NOT just header text. Real layout is WIDE: one ROW PER
# SECTOR, each parameter as its own 3-column group (Pre/CIQ/On-site) placed side by side across
# the sheet, with ONE shared Comments column per row combining whatever mismatch groups
# triggered — never one row per (sector, parameter) with its own comment.
# - MCA sheet has TWO separate tables per generation: one for rachRootSequence/PCI/Cellrange/
#   TAC (4G) or .../nRPCI/.../NRTAC/nCI (5G) — the "pre-existing vs newly-added" rule; another
#   for EarfcnDL/UL/TX/RX/Bandwidth/ConfiguredOutputPower/CellID (4G) or arfcnDL/UL/bSChannelBw/
#   ConfiguredOutputPower/cellLocalId/ssbFrequency (5G) — expected = CIQ always.
# - N2E_NSB sheet: EVERY parameter has ONLY CIQ/On-site columns, no Pre column anywhere at all.
# - Row 1 (both sheets): FDD naming / Rilinks / sector carrier / ISDLONLY / NRCELL CU naming —
#   CIQ vs On-site only. Confirmed real onsite sources: FDD naming = EUtranCellFDD exists
#   onsite with CIQ's naming; Rilinks = RiLink MO scripted+ENABLED for this sector's radio;
#   sector carrier = that RiLink's DATA-port (riPortRef2) matches CIQ's Radio Port; ISDLONLY =
#   onsite isDlOnly consistent with onsite TX=0 (LTE only); NRCELL CU naming = NRCellCU exists
#   onsite matching CIQ (NR only). LTE radios (RRU-n) carry no sector name of their own —
#   resolved via RetSubUnit.userLabel (lists cells sharing an AntennaUnitGroup) ->
#   AntennaNearUnit.rfPortRef (that group's RRU-n) -> RiLink (matching riPortRef2), with a
#   cross-node fallback for co-located sites sharing one antenna.
# NOTE: reportlab is imported lazily inside build_parameter_verification_pdf() only — a new
# feature dependency must never crash the whole app on startup if it isn't installed yet.
# ============================================================



# ============================================================
# AMOS/moshell hget & get command output parser
# (validated against real .log / _onsite.txt captures)
# ============================================================

RULE_RE = re.compile(r'^=+$')
CMD_RE = re.compile(r'^\S+>\s*h?get\s+(\S+)\s+(.+)$')
ANY_PROMPT_RE = re.compile(r'^\S+>\s*\S')


def _col_bounds(header_line):
    return [(m.start(), m.group()) for m in re.finditer(r'\S+', header_line)]


def _clean_value(raw):
    raw = raw.strip()
    if raw == '':
        return None
    m = re.match(r'^-?\d+\s+\(([A-Z_]+)\)$', raw)
    if m:
        return m.group(1)
    m = re.match(r'^\[\d+\]\s*=\s*(.+)$', raw)
    if m:
        return m.group(1).strip()
    if re.match(r'^i\[\d+\]\s*=$', raw):
        return None
    if raw.lower() == 'true':
        return True
    if raw.lower() == 'false':
        return False
    return raw


def _rebuild_header(lines, start, n):
    parts = []
    j = start
    while j < n and not RULE_RE.match(lines[j].strip()):
        if lines[j].strip() == '':
            break
        parts.append(lines[j])
        j += 1
    if not parts or j >= n or not RULE_RE.match(lines[j].strip()):
        return None, None, j
    wrap_width = len(parts[0]) if len(parts) > 1 else None
    return ''.join(parts), wrap_width, j


def parse_hget_blocks(text):
    lines = text.replace('\r\n', '\n').replace('\r', '\n').split('\n')
    i = 0
    n = len(lines)
    while i < n:
        m = CMD_RE.match(lines[i].strip())
        if not m:
            i += 1
            continue
        mo_pattern, attrs_str = m.group(1), m.group(2)
        block_start = i
        i += 1
        header_line = None
        wrap_width = None
        j = i
        limit = min(i + 40, n)
        while j < limit:
            stripped = lines[j].strip()
            if ANY_PROMPT_RE.match(stripped):
                break
            if RULE_RE.match(stripped):
                header_line, wrap_width, j = _rebuild_header(lines, j + 1, n)
                break
            j += 1
        if header_line is None:
            i = block_start + 1
            continue
        j += 1

        col_positions = _col_bounds(header_line)
        # Confirmed real format: some commands (e.g. onsite's "get . cellrange") render as a
        # PIVOTED "MO / Attribute / Value" table instead of the normal wide table — one row per
        # (instance, single-attribute) pair rather than one row per instance with all attributes
        # as columns. Detect and pivot it into the same {instance: {attr: value}} shape.
        is_pivot = [c for _, c in col_positions] == ['MO', 'Attribute', 'Value']

        # Confirmed real corruption source: some hget queries print columns that were never
        # asked for, and those columns describe a DIFFERENT MO. 'hget ^EUtranCellFDD pciConflict'
        # renders one row per conflicting neighbour, carrying that NEIGHBOUR's cellId/enbId/
        # mcc/mnc alongside this cell's pciConflict. Merging that blindly overwrote each cell's
        # real cellId with a neighbour's (e.g. DXL04017_3C_1 took 3A_1's 149 instead of its own
        # 151) or blanked it entirely. Only columns the command actually requested are kept, so
        # a query can never contribute a value it wasn't asked for. Pivot (MO/Attribute/Value)
        # tables are exempt — their column names are structural, not attribute names.
        keep_cols = None
        if not is_pivot:
            _pats = [a for a in attrs_str.split('|') if a.strip()]
            _matched = set()
            for _, _cn in col_positions[1:]:
                for _pat in _pats:
                    try:
                        if re.search(_pat, _cn, re.I):
                            _matched.add(_cn)
                            break
                    except re.error:
                        if _pat.lower() in _cn.lower():
                            _matched.add(_cn)
                            break
            # If nothing matched, the command/column naming isn't understood — keep everything
            # rather than silently discarding the whole block.
            if _matched:
                keep_cols = _matched

        rows = {}
        k = j
        while k < n and not RULE_RE.match(lines[k].strip()):
            line = lines[k]
            if line.strip() == '':
                k += 1
                continue
            if ANY_PROMPT_RE.match(line.strip()):
                break
            row_text = line
            k += 1
            if wrap_width is not None:
                while len(row_text) >= wrap_width and k < n:
                    nxt = lines[k]
                    if nxt.strip() == '' or RULE_RE.match(nxt.strip()) or ANY_PROMPT_RE.match(nxt.strip()):
                        break
                    row_text += nxt
                    k += 1
            bounds = [p[0] for p in col_positions] + [len(row_text)]
            values = {}
            for idx, (pos, colname) in enumerate(col_positions):
                raw = row_text[bounds[idx]:bounds[idx + 1]]
                values[colname] = _clean_value(raw)
            if is_pivot:
                instance = values.get('MO')
                attr_name = values.get('Attribute')
                attr_val = values.get('Value')
                if instance and attr_name:
                    rows.setdefault(instance, {})[attr_name] = attr_val
            else:
                instance = values.get(col_positions[0][1])
                if instance:
                    if keep_cols is not None:
                        values = {kk: vv for kk, vv in values.items()
                                   if kk in keep_cols or kk == col_positions[0][1]}
                    # The same MO can appear on several rows of one table (again, the
                    # pciConflict output). Merge instead of replacing, and never let a later
                    # row blank out a value an earlier row already supplied.
                    tgt = rows.setdefault(instance, {})
                    for kk, vv in values.items():
                        if vv is not None or kk not in tgt:
                            tgt[kk] = vv

        yield {
            "mo_type": mo_pattern,
            "attrs_requested": attrs_str.split('|'),
            "columns": [c for _, c in col_positions],
            "rows": rows,
        }
        i = k


def parse_hget_text(text):
    return list(parse_hget_blocks(text))


def merge_by_instance(blocks, mo_type_filter):
    merged = {}
    for b in blocks:
        if mo_type_filter not in b["mo_type"]:
            continue
        for instance, vals in b["rows"].items():
            tgt = merged.setdefault(instance, {})
            for k, v in vals.items():
                if v is not None or k not in tgt:
                    tgt[k] = v
    return merged


def merge_by_instance_prefix(blocks, instance_prefix):
    merged = {}
    for b in blocks:
        for instance, vals in b["rows"].items():
            if instance.startswith(instance_prefix):
                tgt = merged.setdefault(instance, {})
                for k, v in vals.items():
                    if v is not None or k not in tgt:
                        tgt[k] = v
    return merged


def top_level(instance_ldn):
    return ',' not in instance_ldn


def pv_load_node_tables(text):
    """Parse one node's log text into {'lte_cell', 'nr_cell', 'nr_sector', 'lte_sector',
    'rilink', 'nrcellcu', 'retsubunit', 'antennanearunit'} tables."""
    blocks = parse_hget_text(text)
    lte = {k: v for k, v in merge_by_instance_prefix(blocks, 'EUtranCellFDD=').items() if top_level(k)}
    nr_cell = {k: v for k, v in merge_by_instance_prefix(blocks, 'NRCellDU=').items() if top_level(k)}
    nr_sector = merge_by_instance(blocks, 'Sector')
    lte_sector = merge_by_instance(blocks, '^SectorCarrier')
    rilink = merge_by_instance(blocks, 'RiLink')
    nrcellcu = {k: v for k, v in merge_by_instance_prefix(blocks, 'NRCellCU=').items() if top_level(k)}
    # Confirmed real chain (LTE RRU-to-sector correlation, since RiLink's own naming carries
    # no sector info for LTE): RetSubUnit.userLabel lists "<label>__<cell1>;<cell2>;..." for its
    # AntennaUnitGroup=N; AntennaNearUnit.rfPortRef gives that same AntennaUnitGroup=N's
    # FieldReplaceableUnit=RRU-n. Chaining cell -> AntennaUnitGroup=N -> RRU-n -> RiLink
    # (matching riPortRef1) resolves the mapping RiLink's own naming can't provide for LTE.
    retsubunit = merge_by_instance_prefix(blocks, 'AntennaUnitGroup=')
    retsubunit = {k: v for k, v in retsubunit.items() if ',RetSubUnit=' in k}
    antennanearunit = merge_by_instance_prefix(blocks, 'AntennaUnitGroup=')
    antennanearunit = {k: v for k, v in antennanearunit.items() if ',AntennaNearUnit=' in k and ',RetSubUnit=' not in k}
    # AntennaUnitGroup=<g>,RfBranch=<k>.rfPortRef names EVERY radio feeding that antenna group,
    # not just the RET-carrying one that AntennaNearUnit reports. A shared alpha antenna group
    # routinely carries two or three radios on different bands (confirmed real: group 1 carries
    # both RRU-1 and RRU-4), so AntennaNearUnit alone cannot identify a cell's radio.
    rfbranch = merge_by_instance_prefix(blocks, 'AntennaUnitGroup=')
    rfbranch = {k: v for k, v in rfbranch.items() if ',RfBranch=' in k}
    return {"lte_cell": lte, "nr_cell": nr_cell, "nr_sector": nr_sector, "lte_sector": lte_sector,
            "rilink": rilink, "nrcellcu": nrcellcu, "retsubunit": retsubunit,
            "antennanearunit": antennanearunit, "rfbranch": rfbranch}


def _lte_cell_antenna_group(all_tables_list, cell_id):
    """RetSubUnit.userLabel is '<antenna>__<cell1>;<cell2>;...' under AntennaUnitGroup=<g>.
    Returns (group_id, antenna_label) or (None, None). Antenna wiring is static, so this is
    often only present in the Pre log — hence the list of table-dicts to search in order."""
    for tables in all_tables_list:
        for mo_name, attrs in (tables.get("retsubunit") or {}).items():
            label = attrs.get("userLabel") or ""
            antenna, _, cells_part = label.partition("__")
            cells = [c.strip() for c in cells_part.split(";") if c.strip()]
            if cell_id in cells:
                m = re.match(r'AntennaUnitGroup=(\d+)', mo_name)
                if m:
                    return m.group(1), antenna.strip()
    return None, None


def _rrus_in_antenna_group(all_tables_list, group_id):
    """Every RRU feeding this antenna group, read from its RfBranch entries (with the
    AntennaNearUnit's own RRU as a fallback for captures that omit RfBranch)."""
    rrus = []
    for tables in all_tables_list:
        for mo_name, attrs in (tables.get("rfbranch") or {}).items():
            if not mo_name.startswith(f"AntennaUnitGroup={group_id},"):
                continue
            m = re.search(r'FieldReplaceableUnit=(RRU-\d+)', attrs.get("rfPortRef") or "")
            if m and m.group(1) not in rrus:
                rrus.append(m.group(1))
        if rrus:
            return rrus
    for tables in all_tables_list:
        for mo_name, attrs in (tables.get("antennanearunit") or {}).items():
            if not mo_name.startswith(f"AntennaUnitGroup={group_id},"):
                continue
            m = re.search(r'FieldReplaceableUnit=(RRU-\d+)', attrs.get("rfPortRef") or "")
            if m and m.group(1) not in rrus:
                rrus.append(m.group(1))
    return rrus


def _lte_cell_to_rru(all_tables_list, cell_id):
    """Backwards-compatible single-RRU resolution (first radio of the cell's antenna group)."""
    group_id, _ = _lte_cell_antenna_group(all_tables_list, cell_id)
    if group_id is None:
        return None
    rrus = _rrus_in_antenna_group(all_tables_list, group_id)
    return rrus[0] if rrus else None




# ============================================================
# CIQ-linked comparison logic
# ============================================================

# ============================================================
# Confirmed against the blueprint (temp_blueprint.xlsx) directly, twice:
# - MCA sheet: rachRootSequence/PCI/Cellrange/TAC (4G) and rachRootSequence/nRPCI/Cellrange/
#   NRTAC/nCI (5G) all sit under the "pre/CIQ/On site" 3-way table with the explicit rule
#   "Matching as per pre for pre-existing / CIQ for newly added" — this is Category A.
#   EarfcnDL/UL, Bandwidth, CellID, TX/RX, ConfiguredOutputPower (4G) and arfcnDL/UL,
#   bSChannelBwDL/UL, cellLocalId, ssbFrequency, ConfiguredOutputPower (5G) also have a "pre"
#   column but no such pre-vs-new rule — Category B, expected = CIQ always.
# - N2E_NSB sheet: EVERY parameter, both tables, has ONLY "CIQ"/"On site" columns — no "pre"
#   column anywhere at all. Confirmed real-world reason: N2E is Nokia-to-Ericsson (no prior
#   Ericsson state exists) and NSB is a brand-new site (nothing existed before either way).
#   So for these two scopes, every parameter — Category A concepts included — is simply
#   "Post must match CIQ", with no Pre lookup, no is_new/move_map logic, no Pre-log upload
#   needed at all.
# ============================================================

NO_PRE_SCOPES = {"N2E", "NSB"}

# Marker written into the "Pre" column for a newly-added sector — there is no Pre state to
# show, and a blank would be indistinguishable from "log didn't report it". Rendered amber.
PRE_NA = "NA"

# Category A = "expected is the PRE value for a pre-existing (commercial) sector, and the CIQ
# value for a newly-added one". Confirmed scope: rachRootSequence, PCI / nRPCI and Cellrange
# ONLY. TAC, NRTAC and nCI are Category B — Pre is shown for reference but the on-site value is
# always verified against the CIQ, and any difference is flagged red.
CATEGORY_A_LTE = {"rachRootSequence": "rachRootSequence", "PCI": "physicalLayerCellId",
                   "Cellrange": "cellRange"}
# Confirmed against the real CIQ: 'eUtran Parameters' genuinely has no tac column — the LTE
# TAC lives on the 'eNB Info' tab ('tac', one row per eNodeB). run_parameter_verification()
# resolves it per node and injects it into each eUtran row under this synthetic key, so the
# comparison below can treat it like any other CIQ-sourced value.
CIQ_TAC_KEY = "__ciq_tac__"
CATEGORY_A_LTE_CIQ_KEYS = {"rachRootSequence": "rachRootSequence", "PCI": "PCI",
                            "Cellrange": "cellRange"}
CATEGORY_A_NR = {"rachRootSequence": "rachRootSequence", "nRPCI": "nRPCI",
                  "Cellrange": "cellRange"}
CATEGORY_A_NR_CIQ_KEYS = {"rachRootSequence": "rachRootSequence", "nRPCI": "nRPCI",
                           "Cellrange": "CellRange"}  # confirmed: 5G Info uses "CellRange", not "cellRange"
CATEGORY_B_LTE = {
    "EarfcnDL": ("earfcndl", "earfcnDl"), "EarfcnUL": ("earfcnul", "earfcnUl"),
    "Bandwidth": ("dlChannelBandwidth", "dlChannelBandwidth"),
    "CellID": ("cellId", "cellId"),
    # LTE tac is a node-level value on the CIQ's 'eNB Info' tab, injected per row under
    # CIQ_TAC_KEY by run_parameter_verification().
    "TAC": ("tac", CIQ_TAC_KEY),
}
CATEGORY_B_NR = {
    "arfcnDL": ("arfcnDL", "arfcnDL"), "arfcnUL": ("arfcnUL", "arfcnUL"),
    "bSChannelBwDL": ("bSChannelBwDL", "bSChannelBwDL"), "bSChannelBwUL": ("bSChannelBwUL", "bSChannelBwUL"),
    "cellLocalId": ("cellLocalId", "cellLocalId"), "ssbFrequency": ("ssbFrequency", "ssbFrequency"),
    "NRTAC": ("nRTAC", "nRTAC"), "nCI": ("nCI", "nCI"),
}



def norm(v):
    if v is None:
        return None
    return str(v).strip()


def verdict(post_val, expected_val, category):
    """Returns (color, note). category: 'A' or 'B'."""
    p, e = norm(post_val), norm(expected_val)
    if p is None and e is None:
        return "gray", "No data available"
    if p is None:
        # The parameter simply wasn't returned by the onsite capture (e.g. the NRCellDU hget
        # didn't request rachRootSequence). That is a gap in the log, not a configuration
        # mismatch — it renders as NA in grey and is deliberately kept out of the mismatch
        # comments, so a missing column can't masquerade as a site fault. A cell that genuinely
        # doesn't exist onsite is already caught by the FDD naming / NRCELL CU naming checks.
        return "gray", "Not captured in the onsite log"
    if e is None:
        return "amber", "No baseline to verify against"
    if p == e:
        return "green", ""
    return "red", None  # note filled in by blueprint_comment() below


# ============================================================
# Blueprint-exact comment text (temp_blueprint.xlsx) — no generic "match"/"expected X got Y".
# ============================================================

CATEGORY_A_PARAMS = ("rachRootSequence", "PCI", "nRPCI", "Cellrange")


def blueprint_comment(param, color, is_new, gen):
    """gen: 'lte' or 'nr'. Returns the blueprint's own comment text for this scenario.

    Category A (PCI/nRPCI, Cellrange, TAC/NRTAC/nCI, rachRootSequence) wording is fixed by the
    user's spec and is DIFFERENT for pre-existing vs newly-adding sectors:
      match   + pre-existing  -> "Matching as per pre"
      match   + newly adding  -> no comment at all
      mismatch                -> "<PARAM> Mismatch found on the Pre existing sectors"
                              /  "<PARAM> Mismatch found on the newly adding sectors"
    """
    if param in CATEGORY_A_PARAMS:
        if color == "green":
            return "" if is_new else "Matching as per pre"
        if color in ("amber", "gray"):
            return ""
        tail = "newly adding" if is_new else "Pre existing"
        return f"{param} Mismatch found on the {tail} sectors"
    if color == "green":
        return "Matching"
    if color in ("amber", "gray"):
        return ""
    for params, tail in (COMMENT_GROUPS_LTE if gen == "lte" else COMMENT_GROUPS_NR):
        if tail is None:
            continue
        if param in params:
            return group_mismatch_comment([param], tail)
    return group_mismatch_comment([param], "")


# Category-B comment groups. The parameter list printed inside [Parameter - ...] is built from
# the parameters that ACTUALLY mismatched on that row, not from a fixed string naming the whole
# group — so a row where only EarfcnDL is wrong reads "[Parameter - EarfcnDL]", not the full
# "BW/Erfcndl/erfcnul/cell id". The second element is the fixed advice appended after it.
_RETUNE_TAIL = "Verify the Revision history for retune."
_SWAP_TAIL = "Verify for the Radio swap"

COMMENT_GROUPS_LTE = [
    (("rachRootSequence", "PCI", "Cellrange"), None),  # Category A — is_new-aware wording
    (("EarfcnDL", "EarfcnUL", "Bandwidth", "CellID"), _RETUNE_TAIL),
    (("TX", "RX", "ConfiguredOutputPower"), _SWAP_TAIL),
    (("TAC",), ""),
]
COMMENT_GROUPS_NR = [
    (("rachRootSequence", "nRPCI", "Cellrange"), None),  # Category A
    (("arfcnDL", "arfcnUL", "bSChannelBwDL", "bSChannelBwUL", "cellLocalId", "ssbFrequency"),
     _RETUNE_TAIL),
    (("TX", "RX", "ConfiguredOutputPower"), _SWAP_TAIL),
    (("NRTAC", "nCI"), ""),
]


def group_mismatch_comment(triggered_params, tail):
    """Category-B mismatch comment. The parameter name is deliberately NOT printed — the cell
    that mismatched is already highlighted red in its own column, so naming it again in the
    comment is redundant. Only the advice that belongs to the group is printed."""
    if not triggered_params:
        return ""
    return f"Mismatch found, {tail}" if tail else "Mismatch found"


def pivot_param_results(cells, gen, has_pre):
    """Confirmed against the blueprint directly (temp_blueprint.xlsx): the actual layout is
    WIDE, one ROW PER SECTOR, with each parameter as its own group of 3 columns (Pre/CIQ/On
    site) placed side by side, and ONE shared Comments cell per row combining whatever mismatch
    groups were triggered — not one row per (sector, parameter) with a comment each. Returns
    list of {'sector', <param>_pre/ciq/post/color per param, 'comment'}."""
    groups = COMMENT_GROUPS_LTE if gen == "lte" else COMMENT_GROUPS_NR
    rows = []
    for cell_id, results in cells:
        by_param = {r["parameter"]: r for r in results}
        row = {"sector": cell_id, "is_new": any(r.get("is_new") for r in results)}
        for param, r in by_param.items():
            row[f"{param}_pre"] = r.get("pre")
            row[f"{param}_ciq"] = r.get("ciq")
            row[f"{param}_post"] = r.get("post")
            row[f"{param}_color"] = r.get("color")
            row[f"{param}_note"] = r.get("note")
        comments = []
        for params, fixed_text in groups:
            triggered = [p for p in params if p in by_param and by_param[p]["color"] == "red"]
            if not triggered:
                continue
            if fixed_text is not None:
                comments.append(group_mismatch_comment(triggered, fixed_text))
            else:
                # Category A group: use whichever triggered param's own note (already carries
                # the correct pre-existing-vs-newly-added blueprint text for this sector).
                comments.append(by_param[triggered[0]]["note"])
        row["comment"] = " / ".join(dict.fromkeys(comments))  # dedupe, preserve order
        rows.append(row)
    return rows


def pivot_naming_results(cells):
    """Same wide-row pivot as pivot_param_results, for the FDD naming/Rilinks/sector carrier/
    ISDLONLY/NRCELL CU naming table (blueprint row 1) — one row per sector, each check as its
    own CIQ/On-site column pair, one combined comment."""
    rows = []
    for cell_id, results in cells:
        row = {"sector": cell_id}
        comments = []
        for r in results:
            key = r["parameter"].replace(" ", "_")
            row[f"{key}_ciq"] = r.get("ciq")
            row[f"{key}_post"] = r.get("post")
            row[f"{key}_color"] = r.get("color")
            if r["color"] in ("red", "amber"):
                comments.append(f"{r['parameter']}: {r['note']}")
        row["comment"] = " / ".join(dict.fromkeys(comments))
        rows.append(row)
    return rows


def build_enb_tac_map(ciq_wb):
    """{eNodeB Name -> tac, str(eNBId) -> tac} from the CIQ's 'eNB Info' tab. Confirmed real
    column names: 'eNodeB Name', 'eNBId', 'tac'. Both keys are kept because eUtran Parameters
    identifies its node by eNBId while everything else in the app keys off the node name."""
    out = {}
    if "eNB Info" not in ciq_wb.sheetnames:
        return out
    ws = ciq_wb["eNB Info"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return out
    hdr = [str(h).strip() if h is not None else "" for h in rows[0]]
    for r in rows[1:]:
        d = dict(zip(hdr, r))
        tac = d.get("tac")
        if tac is None:
            continue
        name = d.get("eNodeB Name")
        enbid = d.get("eNBId")
        if name:
            out[str(name).strip().upper()] = tac
        if enbid is not None:
            out[str(enbid).strip()] = tac
    return out


def build_sector_move_map(ciq_wb):
    """{target_sector_name: {'source_sector': str, 'source_node': str}} from Sector Del_Movement."""
    move_map = {}
    if "Sector Del_Movement" not in ciq_wb.sheetnames:
        return move_map
    ws = ciq_wb["Sector Del_Movement"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return move_map
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    for r in rows[1:]:
        d = dict(zip(headers, r))
        target_sector = d.get("Target Sector")
        if target_sector:
            move_map[str(target_sector).strip()] = {
                "source_sector": d.get("Source Sector"),
                "source_node": d.get("Source Node name"),
            }
    return move_map


def compare_lte_cell(cell_id, ciq_row, pre_tables, onsite_tables, move_map, source_pre_tables_by_node, scope="MCA"):
    """Returns list of {'parameter','category','pre','ciq','post','color','note'}."""
    has_pre = scope not in NO_PRE_SCOPES
    ldn = f"EUtranCellFDD={cell_id}"
    on_cell = onsite_tables["lte_cell"].get(ldn, {})
    pre_cell = pre_tables["lte_cell"].get(ldn, {}) if has_pre else {}

    is_new = True if not has_pre else ("existing" not in str(ciq_row.get("Carrier Cell Intention", "")).lower())
    if has_pre:
        move_info = move_map.get(cell_id)
        if move_info and move_info.get("source_sector"):
            src_node = move_info.get("source_node")
            src_tables = source_pre_tables_by_node.get(src_node)
            if src_tables:
                pre_cell = src_tables["lte_cell"].get(f"EUtranCellFDD={move_info['source_sector']}", pre_cell)
        if not pre_cell:
            # C-band / DoD movement: the sector physically moves from one node's pole to
            # another's. In Pre it is still on the OTHER node; in Post it is on the target node
            # under the SAME pole ID (same cell name). Sector Del_Movement covers the case where
            # the move was declared; when it wasn't, fall back to finding the identical cell
            # name in any other node's Pre log rather than reporting "no Pre data".
            for _other_node, _other in (source_pre_tables_by_node or {}).items():
                _found = (_other.get("lte_cell") or {}).get(ldn)
                if _found:
                    pre_cell = _found
                    break

    def sector_lookup(tables, cell):
        ref = cell.get("sectorCarrierRef")
        return tables["lte_sector"].get(ref, {}) if ref else {}

    on_sec = sector_lookup(onsite_tables, on_cell)
    pre_sec = sector_lookup(pre_tables, pre_cell) if has_pre else {}

    results = []
    for param, cell_key in CATEGORY_A_LTE.items():
        pre_v = pre_cell.get(cell_key) if has_pre else None
        ciq_v = ciq_row.get(CATEGORY_A_LTE_CIQ_KEYS[param])
        expected = ciq_v if (is_new or not has_pre) else pre_v
        color, _ = verdict(on_cell.get(cell_key), expected, "A")
        note = blueprint_comment(param, color, is_new, "lte")
        # A newly-added cell has no Pre state at all: show the explicit NA marker so the Pre
        # cell reads as "deliberately absent" (amber in the PDF) instead of a blank that looks
        # like missing data.
        if has_pre and is_new:
            pre_v = PRE_NA
        results.append({"parameter": param, "pre": pre_v, "ciq": ciq_v, "is_new": is_new,
                         "post": on_cell.get(cell_key), "color": color, "note": note})

    for param, (cell_key, ciq_key) in CATEGORY_B_LTE.items():
        ciq_v = ciq_row.get(ciq_key) if ciq_key else None
        pre_v = pre_cell.get(cell_key) if has_pre else None
        post_v = on_cell.get(cell_key)
        # Category B is strictly "on site must equal CIQ" — the Pre column is shown for
        # reference only. A value that still matches Pre but differs from the CIQ is a real
        # mismatch (the change hasn't been applied) and is flagged red, not softened to amber.
        color, _ = verdict(post_v, ciq_v, "B")
        note = blueprint_comment(param, color, is_new, "lte")
        if has_pre and is_new:
            pre_v = PRE_NA
        results.append({"parameter": param, "pre": pre_v, "ciq": ciq_v, "is_new": is_new,
                         "post": post_v, "color": color, "note": note})

    for label, cell_key in [("TX", "noOfTxAntennas"), ("RX", "noOfRxAntennas")]:
        ciq_v = ciq_row.get(cell_key)
        pre_v = pre_sec.get(cell_key) if has_pre else None
        post_v = on_sec.get(cell_key)
        color, _ = verdict(post_v, ciq_v, "B")
        note = blueprint_comment(label, color, is_new, "lte")
        if has_pre and is_new:
            pre_v = PRE_NA
        results.append({"parameter": label, "pre": pre_v, "ciq": ciq_v, "is_new": is_new,
                         "post": post_v, "color": color, "note": note})

    ciq_v = ciq_row.get("configuredOutputPower")
    pre_v = pre_sec.get("configuredMaxTxPower") if has_pre else None
    post_v = on_sec.get("configuredMaxTxPower")
    color, _ = verdict(post_v, ciq_v, "B")
    note = blueprint_comment("ConfiguredOutputPower", color, is_new, "lte")
    if has_pre and is_new:
        pre_v = PRE_NA
    results.append({"parameter": "ConfiguredOutputPower", "pre": pre_v, "ciq": ciq_v, "is_new": is_new,
                     "post": post_v, "color": color, "note": note})

    return results


def _rilink_for_sector(onsite_tables, sector_suffix):
    """Find the RiLink whose riPortRef2 (radio side) names this sector. Confirmed real naming:
    5G radios are 'AAS-<band>_<sector>' (e.g. 'AAS-055619_N077A_1') — hyphen before the numeric
    part, not underscore, so match without assuming a leading separator. LTE radios are just
    'RRU-<n>' with no sector info in the name at all — cannot be matched this way; returns None,
    which correctly surfaces as "not found" rather than a false match."""
    rilinks = onsite_tables.get("rilink", {})
    for rl in rilinks.values():
        ref2 = rl.get("riPortRef2") or ""
        if sector_suffix and sector_suffix in ref2:
            return rl
    return None


_PORT_BLANKS = {"", "N/A", "NA", "NONE", "NOT USED", "-", "--"}


def _ciq_bbu_ports(ciq_row, gen):
    """CIQ's BBU-side (DUS/XMU) letter port(s) for this sector — the port the sector is
    ACTUALLY meant to be scripted on (e.g. 'A' for the 700 Alpha sector). LTE carries it in
    'DUS / XMU Port' (+ its expansion column); 5G Info carries it in 'Port 1'..'Port 4'
    (multiple ports = one radio cabled to several BBU ports, any of them is a valid match).
    'Radio Port' is a DIFFERENT thing (the DATA-n port on the radio) and stays with the
    'sector carrier' check — never used here."""
    keys = (["DUS / XMU Port", "DUS / XMU Port Expansion"] if gen == "lte"
            else ["Port 1", "Port 2", "Port 3", "Port 4"])
    ports = []
    for k in keys:
        v = ciq_row.get(k)
        if v is None:
            continue
        whole = str(v).strip().upper()
        # Check the whole value against the blank set BEFORE splitting — otherwise "N/A"
        # splits into the two bogus ports "N" and "A".
        if whole in _PORT_BLANKS:
            continue
        for part in re.split(r'[,/;]', whole):
            s = part.strip().upper().replace("_", "")
            if s and s not in _PORT_BLANKS and s not in ports:
                ports.append(s)
    return ports


def _bbu_port_from_rilink(rl):
    """riPortRef1 is the BBU/DU side of the RiLink and carries the port id. Confirmed real:
    the RiPort can be ALPHABETICAL when the sector is scripted on the BBU/DUS
    (e.g. 'FieldReplaceableUnit=DUS-1,RiPort=A') or NUMERICAL when it is scripted on an XMU
    (e.g. 'FieldReplaceableUnit=XMU-1,RiPort=3') — both forms must be read, so the pattern
    accepts letters and digits. riPortRef2 is the radio side (DATA-n) and is never used here."""
    m = re.search(r'RiPort=([A-Za-z0-9]+)', (rl.get("riPortRef1") or ""))
    return m.group(1).strip().upper().replace("_", "") if m else None


def _rilinks_for_sector(onsite_tables, sector_suffix):
    """ALL RiLinks whose riPortRef2 (radio side) names this sector — a sector cabled on both
    DATA1 and DATA2 has TWO RiLinks, and therefore two BBU RiPorts, both of which must be
    scripted. Returning only the first silently hid the second port."""
    out = []
    for rl in (onsite_tables.get("rilink") or {}).values():
        ref2 = rl.get("riPortRef2") or ""
        if sector_suffix and sector_suffix in ref2:
            out.append(rl)
    return out


def _radio_data_port(rl):
    """riPortRef2's radio-side port, normalised: 'RiPort=DATA_1' -> 'DATA1'."""
    m = re.search(r'RiPort=([A-Za-z0-9_]+)', (rl.get("riPortRef2") or ""))
    return m.group(1).strip().upper().replace("_", "") if m else None


def _radio_fru(rl):
    """riPortRef2's radio, e.g. 'RRU-16' or 'AAS-055619_N077A_1'."""
    m = re.search(r'FieldReplaceableUnit=([^,\s]+)', (rl.get("riPortRef2") or ""))
    return m.group(1).strip() if m else None


def _rilinks_on_bbu_port(onsite_tables, port):
    """Every RiLink on this node scripted on BBU/XMU RiPort <port>."""
    want = str(port).strip().upper().replace("_", "")
    return [rl for rl in (onsite_tables.get("rilink") or {}).values()
            if _bbu_port_from_rilink(rl) == want]


def _ciq_radio_ports(ciq_row):
    """CIQ 'Radio Port' — the radio-side DATA port(s). 'DATA1/DATA2' means both."""
    v = ciq_row.get("Radio Port")
    if v is None:
        return set()
    out = set()
    for part in re.split(r'[,/;|]', str(v)):
        t = part.strip().upper().replace("_", "").replace(" ", "")
        if t and t not in _PORT_BLANKS:
            out.add(t)
    return out


def check_rilinks(cell_id, ciq_row, gen, node, onsite_by_node, onsite_tables):
    """Rilinks verification, driven by the CIQ's RADIO MAPPING rather than by the antenna
    (RetSubUnit) chain.

    Why: the antenna chain lives in the Pre log and identifies an AntennaUnitGroup, not a
    radio — and a group routinely feeds several radios, so it can't say which one is this
    sector's. Anything derived from it was either unavailable (nodes whose Pre log wasn't
    uploaded produced no Rilinks result at all) or ambiguous.

    The CIQ already states the radio mapping directly: which BBU/XMU RiPort this sector is
    cabled on ('DUS / XMU Port' + its expansion for LTE, 'Port 1'..'Port 4' for 5G) and which
    radio-side DATA port it uses ('Radio Port'). The onsite RiLink table states, for each
    scripted RiPort, which radio and which DATA port it reaches. So the check is:

        for each RiPort the CIQ specifies for this sector, on THIS node:
          - is a RiLink scripted on it at all?
          - does that RiLink reach the radio-side DATA port the CIQ specifies?

    RiPorts are alphabetical on a BBU/DUS and numerical on an XMU; both are handled. More than
    one RiPort is joined with ' | '. Sectors that share a radio (the CIQ's 'Co-Located
    Technology Cell' set — e.g. a 4G carrier, its second carrier and the 5G cell on the same
    RRU) all carry the same CIQ RiPort and therefore all resolve to that same single onsite
    RiPort, which is exactly what the node shows. Operational state is deliberately not
    considered: a link can be DISABLED at capture time and still be scripted on the right port.

    Returns {'ciq','post','color','note','radios'}."""
    own = onsite_by_node.get(node) if (onsite_by_node and node in onsite_by_node) else onsite_tables
    own = own or {}
    ciq_ports = _ciq_bbu_ports(ciq_row, gen)
    want_data = _ciq_radio_ports(ciq_row)

    found, missing, wrong_data, radios = [], [], [], []
    for port in ciq_ports:
        rls = _rilinks_on_bbu_port(own, port)
        if not rls:
            missing.append(port)
            continue
        found.append(port)
        for rl in rls:
            fru = _radio_fru(rl)
            if fru and fru not in radios:
                radios.append(fru)
            dp = _radio_data_port(rl)
            if want_data and dp and dp not in want_data:
                wrong_data.append(f"{port}\u2192{dp}")

    ciq_text = " | ".join(ciq_ports) if ciq_ports else None
    post_text = " | ".join(found) if found else None

    if not ciq_ports:
        color, note = "amber", "CIQ has no BBU/XMU port for this sector"
    elif not (own.get("rilink") or {}):
        color, note = "gray", "No RiLink data in this node's onsite log"
    elif missing and not found:
        color = "red"
        note = f"CIQ Port {' | '.join(missing)} is not scripted on {node}"
    elif missing:
        color = "red"
        note = (f"CIQ Port {' | '.join(missing)} is not scripted on {node} "
                f"(Port {post_text} is)")
    elif wrong_data:
        color = "red"
        note = (f"Scripted on Port {post_text}, but cabled to {', '.join(wrong_data)} \u2014 "
                f"CIQ Radio Port is {'/'.join(sorted(want_data))}")
    else:
        color = "green"
        note = f"Matching \u2014 scripted on Port {post_text}"
    return {"ciq": ciq_text, "post": post_text, "color": color, "note": note, "radios": radios}


def verify_naming_and_config(cell_id, ciq_row, onsite_tables, pre_tables, gen, all_onsite_tables=None,
                              all_pre_tables=None, node=None, onsite_by_node=None, pre_by_node=None):
    """FDD naming / Rilinks / sector carrier / ISDLONLY / NRCELL CU naming — confirmed against
    blueprint row 1 (temp_blueprint.xlsx, both sheets): CIQ vs On-site only, no Pre column at
    all for this table in either MCA or N2E_NSB. Columns generated per gen ('lte' or 'nr') —
    FDD naming is LTE-only, NRCELL CU naming is NR-only; Rilinks/sector carrier apply to both.
    pre_tables is used ONLY as a fallback source for the RetSubUnit/AntennaNearUnit RRU-mapping
    chain (confirmed real: antenna wiring is static, so it's often only queried in Pre logs,
    not re-queried onsite) — never for the Pre-vs-CIQ-vs-Post comparison itself, since this
    table has no Pre column at all, in any scope.
    Returns list of {'parameter','ciq','post','color','note'}."""
    results = []
    sector_suffix = cell_id.split("_", 1)[1] if "_" in cell_id else cell_id

    if gen == "lte":
        # FDD naming: does the EUtranCellFDD instance exist onsite exactly as CIQ names it —
        # i.e. is the cell built with the naming CIQ specifies.
        exists = f"EUtranCellFDD={cell_id}" in onsite_tables.get("lte_cell", {})
        color = "green" if exists else "red"
        note = "Matching" if exists else "Cell not found onsite with CIQ naming"
        results.append({"parameter": "FDD naming", "ciq": cell_id,
                         "post": cell_id if exists else None, "color": color, "note": note})

    # Rilinks — verified from the CIQ's radio mapping (see check_rilinks). The CIQ states the
    # BBU/XMU RiPort(s) this sector is cabled on and its radio-side DATA port; the onsite RiLink
    # table states what is actually scripted on those ports. Multiple ports join with " | ".
    _rl = check_rilinks(cell_id, ciq_row, gen, node, onsite_by_node, onsite_tables)
    results.append({"parameter": "Rilinks", "ciq": _rl["ciq"], "post": _rl["post"],
                     "color": _rl["color"], "note": _rl["note"]})

    # Sector carrier = the sector carrier ID, not the radio DATA port.
    #   LTE -> CIQ 'sectorId' vs the SectorCarrier the cell actually points at onsite
    #          (EUtranCellFDD.sectorCarrierRef = 'SectorCarrier=<n>').
    #   5G  -> "sector carrier" means NR SECTOR CARRIER: CIQ 'NRSectorCarrier' vs the
    #          NRSectorCarrier instance that exists onsite for this cell.
    if gen == "lte":
        ciq_sc = ciq_row.get("sectorId")
        on_cell_sc = onsite_tables.get("lte_cell", {}).get(f"EUtranCellFDD={cell_id}", {}).get("sectorCarrierRef") or ""
        m = re.search(r'SectorCarrier=([A-Za-z0-9_\-]+)', str(on_cell_sc))
        onsite_sc = m.group(1) if m else None
    else:
        ciq_sc = ciq_row.get("NRSectorCarrier")
        on_cell = onsite_tables.get("nr_cell", {}).get(f"NRCellDU={cell_id}", {})
        ref = on_cell.get("nRSectorCarrierRef") or on_cell.get("nRSectorCarrierId") or ""
        m = re.search(r'NRSectorCarrier=([A-Za-z0-9_\-]+)', str(ref))
        if m:
            onsite_sc = m.group(1)
        elif ciq_sc and f"NRSectorCarrier={str(ciq_sc).strip()}" in (onsite_tables.get("nr_sector") or {}):
            onsite_sc = str(ciq_sc).strip()
        else:
            onsite_sc = None

    label_sc = "Sector carrier" if gen == "lte" else "NR sector carrier"
    if onsite_sc is None and ciq_sc is None:
        color, note = "gray", "No data available"
    elif onsite_sc is None:
        color, note = "red", f"{label_sc} not found onsite"
    elif ciq_sc is None:
        color, note = "amber", "CIQ has no sector carrier to verify against"
    elif norm(onsite_sc) == norm(ciq_sc):
        color, note = "green", "Matching"
    else:
        color, note = "red", f"{label_sc} mismatch — CIQ {ciq_sc}, onsite {onsite_sc}"
    results.append({"parameter": "sector carrier", "ciq": ciq_sc,
                     "post": onsite_sc, "color": color, "note": note})

    if gen == "lte":
        # ISDLONLY is compared CIQ vs on site, like every other naming/config row. The CIQ's
        # own 'ISDLONLY' column is the expected value (TRUE/FALSE as text); onsite it is
        # EUtranCellFDD.isDlOnly (true/false). Both sides are normalised to a boolean so the
        # text-vs-boolean difference can't cause a false mismatch.
        def _as_bool(v):
            if v is None:
                return None
            t = str(v).strip().lower()
            if t in ("true", "yes", "y", "1"):
                return True
            if t in ("false", "no", "n", "0"):
                return False
            return None

        onsite_cell = onsite_tables.get("lte_cell", {}).get(f"EUtranCellFDD={cell_id}", {})
        ciq_isdl = _as_bool(ciq_row.get("ISDLONLY"))
        onsite_isdl = _as_bool(onsite_cell.get("isDlOnly"))
        if ciq_isdl is None and onsite_isdl is None:
            color, note = "gray", "No data available"
        elif onsite_isdl is None:
            color, note = "red", "isDlOnly not found onsite"
        elif ciq_isdl is None:
            color, note = "amber", "CIQ has no ISDLONLY value to verify against"
        elif ciq_isdl == onsite_isdl:
            color, note = "green", "Matching"
        else:
            color, note = "red", f"ISDLONLY mismatch — CIQ {ciq_isdl}, onsite {onsite_isdl}"
        results.append({"parameter": "ISDLONLY", "ciq": ciq_isdl, "post": onsite_isdl,
                         "color": color, "note": note})
    # ISDLONLY is an FDD(LTE)-specific concept (EUtranCellFDD.isDlOnly) — no equivalent
    # NRCellDU attribute exists, so it's skipped entirely for NR rather than shown as "no data".

    if gen == "nr":
        ciq_nrcellcu = ciq_row.get("NRCellCU")
        exists = bool(ciq_nrcellcu) and f"NRCellCU={ciq_nrcellcu}" in onsite_tables.get("nrcellcu", {})
        color = "green" if exists else "red"
        note = "Matching" if exists else "NRCellCU naming not found onsite matching CIQ"
        results.append({"parameter": "NRCELL CU naming", "ciq": ciq_nrcellcu,
                         "post": ciq_nrcellcu if exists else None, "color": color, "note": note})

    return results


def compare_nr_cell(cell_id, ciq_row, pre_tables, onsite_tables, move_map, source_pre_tables_by_node,
                     scope="MCA", node_has_pre=False):
    has_pre = scope not in NO_PRE_SCOPES
    ldn = f"NRCellDU={cell_id}"
    on_cell = onsite_tables["nr_cell"].get(ldn, {})
    pre_cell = pre_tables["nr_cell"].get(ldn, {}) if has_pre else {}

    # KNOWN LIMITATION (MCA only — moot for N2E/NSB, which never use Pre at all): unlike
    # eUtran Parameters' "Carrier Cell Intention" column, 5G Info has no equivalent new-vs-
    # existing classification column at all — confirmed by checking its full header list. A
    # cell explicitly listed in Sector Del_Movement is reliably "existing/moved", so that case
    # is handled correctly below. For everything else, is_new can't be determined from 5G Info
    # alone; defaulting to False (treat as existing, use Pre as expected) is the safer
    # assumption, since PCI/nRPCI errors are more consequential than a newly-added cell being
    # compared against a nonexistent Pre baseline — but this needs a real classification source
    # to be fully reliable. Flagging rather than guessing.
    # 5G Info has no "new vs existing" column (unlike eUtran Parameters' Carrier Cell
    # Intention) — confirmed by checking its full header list. The one reliable signal is the
    # Pre log itself: if this node's Pre log WAS uploaded and the cell simply isn't in it, the
    # cell is newly added. Without a Pre log for the node, absence proves nothing, so the
    # safer "treat as existing" assumption is kept. Resolved after the move lookup below.
    is_new = True if not has_pre else False
    pre_sector_source_tables = pre_tables
    pre_sector_id = cell_id
    if has_pre:
        move_info = move_map.get(cell_id)
        if move_info and move_info.get("source_sector"):
            src_node = move_info.get("source_node")
            src_tables = source_pre_tables_by_node.get(src_node)
            if src_tables:
                pre_cell = src_tables["nr_cell"].get(f"NRCellDU={move_info['source_sector']}", pre_cell)
                pre_sector_source_tables = src_tables
                pre_sector_id = move_info["source_sector"]
        if not pre_cell:
            # Same C-band / DoD movement fallback as the LTE path — Pre lives on the other
            # node's log, Post on the target node, under the same pole ID (same cell name).
            for _other_node, _other in (source_pre_tables_by_node or {}).items():
                _found = (_other.get("nr_cell") or {}).get(ldn)
                if _found:
                    pre_cell = _found
                    pre_sector_source_tables = _other
                    break
        if node_has_pre and not pre_cell:
            is_new = True

    def sector_lookup(tables, cell_id_key):
        # Confirmed real key format: this block's instances are "NRSectorCarrier=<cell_id>",
        # not the bare cell name.
        return tables["nr_sector"].get(f"NRSectorCarrier={cell_id_key}", {})

    on_sec = sector_lookup(onsite_tables, cell_id)
    pre_sec = sector_lookup(pre_sector_source_tables, pre_sector_id) if has_pre else {}

    results = []
    for param, cell_key in CATEGORY_A_NR.items():
        pre_v = pre_cell.get(cell_key) if has_pre else None
        ciq_v = ciq_row.get(CATEGORY_A_NR_CIQ_KEYS[param]) if ciq_row else None
        expected = ciq_v if (is_new or not has_pre) else pre_v
        color, _ = verdict(on_cell.get(cell_key), expected, "A")
        note = blueprint_comment(param, color, is_new, "nr")
        if has_pre and is_new:
            pre_v = PRE_NA
        results.append({"parameter": param, "pre": pre_v, "ciq": ciq_v, "is_new": is_new,
                         "post": on_cell.get(cell_key), "color": color, "note": note})

    # Confirmed real bug: arfcnDL/arfcnUL/bSChannelBwDL/bSChannelBwUL only exist in the
    # NRSectorCarrier ("Sector") table, not NRCellDU — reading them from the cell table
    # silently returned None always. cellLocalId/ssbFrequency live on the cell itself.
    SECTOR_LEVEL_B_PARAMS = {"arfcnDL", "arfcnUL", "bSChannelBwDL", "bSChannelBwUL"}
    for param, (cell_key, ciq_key) in CATEGORY_B_NR.items():
        ciq_v = ciq_row.get(ciq_key) if (ciq_row and ciq_key) else None
        if param in SECTOR_LEVEL_B_PARAMS:
            pre_v = pre_sec.get(cell_key) if has_pre else None
            post_v = on_sec.get(cell_key)
        else:
            pre_v = pre_cell.get(cell_key) if has_pre else None
            post_v = on_cell.get(cell_key)
        # Category B is strictly "on site must equal CIQ" — the Pre column is shown for
        # reference only. A value that still matches Pre but differs from the CIQ is a real
        # mismatch (the change hasn't been applied) and is flagged red, not softened to amber.
        color, _ = verdict(post_v, ciq_v, "B")
        note = blueprint_comment(param, color, is_new, "nr")
        if has_pre and is_new:
            pre_v = PRE_NA
        results.append({"parameter": param, "pre": pre_v, "ciq": ciq_v, "is_new": is_new,
                         "post": post_v, "color": color, "note": note})

    for label, cell_key in [("TX", "noOfTxAntennas"), ("RX", "noOfRxAntennas")]:
        pre_v = pre_sec.get(cell_key)
        post_v = on_sec.get(cell_key)
        color, _ = verdict(post_v, pre_v, "B")  # no CIQ column confirmed for NR TX/RX yet
        note = blueprint_comment(label, color, is_new, "nr")
        if has_pre and is_new:
            pre_v = PRE_NA
        results.append({"parameter": label, "pre": pre_v, "ciq": None, "is_new": is_new,
                         "post": post_v, "color": color, "note": note})

    ciq_v = ciq_row.get("configuredMaxTxPower") if ciq_row else None
    pre_v = pre_sec.get("configuredMaxTxPower")
    post_v = on_sec.get("configuredMaxTxPower")
    color, _ = verdict(post_v, ciq_v, "B")
    note = blueprint_comment("ConfiguredOutputPower", color, is_new, "nr")
    if has_pre and is_new:
        pre_v = PRE_NA
    results.append({"parameter": "ConfiguredOutputPower", "pre": pre_v, "ciq": ciq_v, "is_new": is_new,
                     "post": post_v, "color": color, "note": note})

    return results


# ============================================================
# File-to-node matching + orchestration
# ============================================================

def match_file_to_node(filename, node_names):
    """Match an uploaded log filename to a CIQ node name. Confirmed real naming:
    pre logs are '<NODE>.log', onsite logs are '<NODE>_onsite.txt' — but matches any
    '<NODE>' prefix followed by '.', '_', or end-of-basename, to tolerate minor variations.
    Longest node name wins first, to avoid a shorter node name being a false-positive
    prefix of a longer one (e.g. 'DXL0233' vs 'DXL02330')."""
    base = re.sub(r'\.(log|txt)$', '', filename, flags=re.IGNORECASE)
    for node in sorted(node_names, key=len, reverse=True):
        node_u = str(node).strip()
        if not node_u:
            continue
        if base == node_u or base.startswith(node_u + '_') or base.startswith(node_u + '.'):
            return node
    return None


def run_parameter_verification(ciq_wb, mm_objs, pre_files, onsite_files, scope="MCA"):
    """pre_files / onsite_files: list of (filename, text_content) tuples. pre_files may be
    empty for N2E/NSB, which never require a Pre log at all (confirmed against the blueprint —
    no "pre" column exists anywhere in the N2E_NSB sheet).
    Returns {
        'node_results': {node: {'lte': [(cell_id, [param_result,...]), ...], 'nr': [...]}},
        'unmatched_pre': [filename,...], 'unmatched_onsite': [filename,...],
        'nodes_missing_pre': [node,...], 'nodes_missing_onsite': [node,...],
    }"""
    has_pre = scope not in NO_PRE_SCOPES
    node_names = [r.get("Node to be built as") for r in mm_objs if r.get("Node to be built as")]

    # Confirmed real gap: a node's 5G cells are named after its gNodeB Name, which for an MMBB
    # node is NOT the same string as "Node to be built as" (e.g. node DXL02330 carries gNodeB
    # DXFN014017, and every one of its NRCellDU instances starts with DXFN014017_). Matching
    # cells to nodes on "Node to be built as" alone silently dropped all of them. Build an
    # alias map so a cell (or a log filename) can be resolved by any of the node's names.
    node_alias = {}
    for r in mm_objs:
        built = r.get("Node to be built as")
        if not built:
            continue
        for key in ("Node to be built as", "eNodeB Name", "gNodeB Name"):
            v = r.get(key)
            if v and str(v).strip() and str(v).strip().upper() != "N/A":
                node_alias[str(v).strip()] = str(built).strip()

    # Sector Del_Movement's SOURCE nodes are real nodes that are NOT part of this build (the
    # C-band / DoD sectors are moving off them). Their Pre logs are still needed — that's where
    # the moved sector's Pre state lives — so they're matched and loaded, but they are never
    # reported as build nodes and never get result tables of their own.
    move_map_early = build_sector_move_map(ciq_wb) if has_pre else {}
    external_pre_nodes = {str(v.get("source_node")).strip() for v in move_map_early.values()
                          if v.get("source_node")} - set(node_alias)

    match_names = list(node_alias) + sorted(external_pre_nodes)

    pre_by_node, onsite_by_node = {}, {}
    unmatched_pre, unmatched_onsite = [], []
    if has_pre:
        for fname, text in pre_files:
            hit = match_file_to_node(fname, match_names)
            if hit:
                pre_by_node[node_alias.get(hit, hit)] = pv_load_node_tables(text)
            else:
                unmatched_pre.append(fname)
    for fname, text in onsite_files:
        hit = match_file_to_node(fname, match_names)
        if hit:
            onsite_by_node[node_alias.get(hit, hit)] = pv_load_node_tables(text)
        else:
            unmatched_onsite.append(fname)

    nodes_missing_pre = [n for n in node_names if n not in pre_by_node] if has_pre else []
    nodes_missing_onsite = [n for n in node_names if n not in onsite_by_node]

    def _owner_node(cell_id):
        """Longest alias wins so a shorter node name can't claim a longer one's cell."""
        for a in sorted(node_alias, key=len, reverse=True):
            if str(cell_id).startswith(a):
                return node_alias[a]
        return None

    empty_tables = {"lte_cell": {}, "nr_cell": {}, "nr_sector": {}, "lte_sector": {}, "rilink": {}, "nrcellcu": {}}
    source_pre_by_node = pre_by_node  # for Sector Del_Movement lookups across nodes (MCA only)

    lte_ciq, nr_ciq = {}, {}
    enb_tac = build_enb_tac_map(ciq_wb)
    if "eUtran Parameters" in ciq_wb.sheetnames:
        ws = ciq_wb["eUtran Parameters"]
        hdr = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(hdr, row))
            if d.get("EutranCellFDDId"):
                # LTE tac is on the 'eNB Info' tab, not this one — resolve it per node here so
                # the comparison sees it as a normal CIQ value. Match on eNBId first (this
                # sheet's own node key), then on the node name the cell belongs to.
                _tac = None
                if d.get("eNBId") is not None:
                    _tac = enb_tac.get(str(d.get("eNBId")).strip())
                if _tac is None:
                    _cell = str(d["EutranCellFDDId"])
                    _owner = _owner_node(_cell)
                    if _owner:
                        _tac = enb_tac.get(str(_owner).strip().upper())
                d[CIQ_TAC_KEY] = _tac
                lte_ciq[d["EutranCellFDDId"]] = d
    if "5G Info" in ciq_wb.sheetnames:
        ws = ciq_wb["5G Info"]
        hdr = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(hdr, row))
            if d.get("NRCellDU"):
                nr_ciq[d["NRCellDU"]] = d

    move_map = move_map_early

    node_results = {n: {"lte": [], "nr": [], "naming_lte": [], "naming_nr": []} for n in node_names}
    for cell_id, ciq_row in lte_ciq.items():
        node = _owner_node(cell_id)
        if not node:
            continue
        pre_t = pre_by_node.get(node, empty_tables)
        on_t = onsite_by_node.get(node, empty_tables)
        results = compare_lte_cell(cell_id, ciq_row, pre_t, on_t, move_map, source_pre_by_node, scope=scope)
        node_results[node]["lte"].append((cell_id, results))
        node_results[node]["naming_lte"].append((cell_id, verify_naming_and_config(
            cell_id, ciq_row, on_t, pre_t, "lte",
            all_onsite_tables=list(onsite_by_node.values()), all_pre_tables=list(pre_by_node.values()),
            node=node, onsite_by_node=onsite_by_node, pre_by_node=pre_by_node)))

    for cell_id, ciq_row in nr_ciq.items():
        node = _owner_node(cell_id)
        if not node:
            continue
        pre_t = pre_by_node.get(node, empty_tables)
        on_t = onsite_by_node.get(node, empty_tables)
        results = compare_nr_cell(cell_id, ciq_row, pre_t, on_t, move_map, source_pre_by_node,
                                   scope=scope, node_has_pre=(node in pre_by_node))
        node_results[node]["nr"].append((cell_id, results))
        node_results[node]["naming_nr"].append((cell_id, verify_naming_and_config(
            cell_id, ciq_row, on_t, pre_t, "nr",
            all_onsite_tables=list(onsite_by_node.values()), all_pre_tables=list(pre_by_node.values()),
            node=node, onsite_by_node=onsite_by_node, pre_by_node=pre_by_node)))

    for node, res in node_results.items():
        res["_wide_lte"] = pivot_param_results(res["lte"], "lte", has_pre)
        res["_wide_nr"] = pivot_param_results(res["nr"], "nr", has_pre)
        res["_wide_naming_lte"] = pivot_naming_results(res["naming_lte"])
        res["_wide_naming_nr"] = pivot_naming_results(res["naming_nr"])

    return {
        "node_results": node_results,
        "unmatched_pre": unmatched_pre,
        "unmatched_onsite": unmatched_onsite,
        "nodes_missing_pre": nodes_missing_pre,
        "nodes_missing_onsite": nodes_missing_onsite,
    }


import io

# COMMENT_GROUPS_LTE / COMMENT_GROUPS_NR are defined once, above, and reused here.

LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAlkAAADSCAYAAAB5ENV1AAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAADsMAAA7DAcdvqGQAAKs4SURBVHhe7J13gCRHdfB/VdU9M5v3knI8ZSQhCQWyQCKKZMAIRDBgssnJNjgCFphsDAYTjEifkMgokhEiSASJoJzDnXS6HDZN6K563x9VPdMzO5vudu/2Tv07lXamp0N1hVevXlW9UiIiFBQUFBQUFBQUzCu680BBQUFBQUFBQcGOUyhZBQUFBQUFBQULQKFkFRQUFBQUFBQsAIWSVVBQUFBQUFCwABRKVkFBQUFBQUHBAlAoWQUFBQUFBQUFC0ChZBUUFBQUFBQULACFklVQUFBQUFBQsAAUSlZBQUFBQUFBwQJQKFkFBQUFBQUFBQtAoWQVFBQUFBQUFCwAhZJVUFBQUFBQULAAFEpWQUFBQUFBQcECUChZBQUFBQUFBQULQKFkFRQUFBQUFBQsAIWSVVBQUFBQUFCwABRKVkFBQUFBQUHBAlAoWQUFBQUFBQUFC0ChZBUUFBQUFBQULACFklVQUFBQUFBQsAAUSlZBQUFBQUFBwQJQKFkFBQUFBQUFBQtAoWQVFBQUFBQUFCwAhZJVUFBQUFBQULAAFEpWQUFBQUFBQcECUChZBQUFBQUFBQULQKFkFRQUFBQUFBQsAIWSVVBQUFBQUFCwABRKVkFBQUFBQUHBAlAoWQUFBQUFBQUFC0ChZBUUFBQUFBQULACFklVQUFBQUFBQsAAUSlZBQUFBQUFBwQJQKFkFBQUFBQUFBQtAoWQVFBQUFBQUFCwAhZJVUFBQUFBQULAAFEpWQUFBQUFBQcECUChZBQUFBQUFBQULQKFkFRQUFBQUFBQsAIWSVVBQUFBQUFCwABRKVkFBQUFBQUHBAlAoWQUFBQUFBQUFC0ChZBUUFBQUFBQULACFklVQUFBQUFBQsAAUSlZBQUFBQUFBwQJQKFkFBQUFBQUFBQtAoWQVFBQUFBQUFCwAhZJVUFBQUFBQULAAFEpWQUFBQUFBQcECUChZBQUFBQUFBQULQKFkFRQUFBQUFBQsAIWSVVBQUFBQUFCwABRKVkFBQUFBQUHBAlAoWQUFBQUFBQUFC0ChZBUUFBQUFBQULACFklVQUFBQUFBQsAAUSlZBQUFBQUFBwQJQKFkFBQUFBQUFBQtAoWQVFBQUFBQUFCwAhZJVUFBQUFBQULAAFEpWQUFBQUFBQcECUChZBQUFBQUFBQULQKFk7a5IR1hIFvr+BQUFBQUFeyBKRB6kTagF1M7RM7ulsOo8MEfy91Tgcl/n640Ei0KDKETteJQLCvZ88jVRzUNF3xPp7BkW6VSw5/IgVbIs4EiqiigyKCOkTvmqr3xy+L/tgkCJFwQtcdB5Di0VRxQKhaQOrRRKA0qhTCaEFUrtgGDJP1o7GtaSigPEi6zOaM2RehyhbYMBUwHRiI9+QUHBlPj6B4LgELF4KWBANOipK1BndZ36zOnpvA/bea/5ug+5e7VL03bUvHUNCwoWFw9SJSsBhA1rDVdecR0NV6bm+gDBNZWsfI/Uo8QLAk2QGF3OIZwDCi2KcgQHHhDxmNP3x6YWbWqIVFBK77iS5by0ctriMPz22lXccbtlPrLURvCY05Zw1EH9KGIc07YRBQUFuGYQEZxTYBU4hVIa3VHfBQj9tklKTXamkskKSSfT3YfcvXS3Hztws7jP9sRJVBCZStC6vcPmPxZKVsGeyYNUyUoRHLffEXHOOe9h09YSieoF8pasyQrUZCVLJoujnJKlUPSWNCecMMB5X34FlahBpKogfaDMjitZ4dFWW7aOC29945e44uptOJGcSNwOFAz3DfPBc8/gWU89CCQmVWDCA3co3gUFeyyCIEAD0DSqJW66dRMbtnlZoglj7qHeZopH9jmPwp+nMjEzDdPdh9z1c1KyprnPnOIU/ko4ZozitOOWMDxQKFUFDw4enEqWOESl3HR3iac/671s2dJLRMkLyOZon2tTn7xg8T82BUwXRSyvZPmvKQccWOK733oTB+6dUCJBUUHUfAwX+og0JGH9Fnje0z7HHetSnIi/9w7cfrh3GR/54KN47rMOAWtIjCUSA4WSVVAwJV6hqKIosW2L4c1vP48r/7gR5xzWtluZM8WDjplcZPImyJqZalv+PplCk5HdJyfapsXRrhhlNO8zxzjl7yMKKj0Vfnzh63nI4b5TW1CwpzOberfnIRqFAuWwkcMaQSMYQIuEAEakGbTkfiOE/LFJv/mgTMrGreNc8/s6DoVyPbCjQ4UZSkAJSQr33LuNNetSlAMtfqhSi0I7hZkh6C5BqRqoJBQRhSJB7ahiWFCwxyOAopEmiIPRbZqxrT1MjPRRGx+kMTHUDEl1iGTCh7RLSCaGaFRb508V8vfpvFfz+Czu03mvrveZ5b2y++TvYSeGcBNLcGkhQwoePDw4lSwAMZRSjXERVgv1COoRJFqRaIXTBqcjnI4QHeG0ITWaJNI0msHQiKKOEH6LFY1YgYowaZlvf/sy0BGWHbMwtSEasNRrJb72ld8zojS1WFOPNbVIUTOKxChSPXVIsr9G0cgFpzRIlLPMlToeXlBQkMdPe7cIESYqoSIQAy4CawSrHWkuWOX8BPncXC6VzenC4ZQ/J39N1zDNfZjLfbTDddynM05W+ed1XtcZ8u+mcChxGOeDX2JUUPDg4MGrZDntbTTBOjMrK80MP3fDWYVNhVtvvo31ax34Ebf5QUCsYWQk4be/vQ7nOgcddhTTHCdQ2/PyBQUPOqRp/d0tWahoL9R9CwoWOQ9eJStHGHVrm8yZzUHYcdkQkSZCvWa545YaeqYZo7NFASI0Gor7Vo9SnVAg3n3DjjDVO3c7VlBQMHeasqWLvCkoKNizeHAqWUGiNSdnzuMI3iTE4NDUEseVV95Cw050nrHdiIJ6Q/jd725gdDSlr7dvGlGdTUGdXsnLlM0sbfzBma4qKCiYC/PV19qdyE+GLyh4sPDgVLIC+Qqft1ypDpWkeU62pHoOAQwiFZK0wo9/cg1bXWnSSqLtRitG6ppvXf5bGrpC0qhihLagpPUWrX+T1a3OFUOS8yJfCMaCgpnprPtTWcfzlSk71q1O5n+fKXTSdp+5yK3OCHTKiVmGPM13y3xl5V+2y/MKCvYkHrxKVhfFoduwYSf5c2YTAASDtSXWbRzlD395gNQSlnJP86BZkFjLbXdv4v71QqpKQB3VXAnpVz4GP/b+grwk7ZSEmfALZELRT+advMS8oKCgnc4qNel7Jhc6jk9FpyyZLkxHs8rPIsxE5/nThU66HPJM+UNBwe7Pg1fJCkyq3926YjuAKIcoi0VIgeuu3YxNwVmLiFe2ttdVWbXW4Jab7qc+UQkuKZLOU5qIalmwmmrTPL5nQUGBR6G8otH5Qwcz/b49ZA5Hs3s3h/y7MN1vC8WueGZBwa7kwatkBSHYWeezhcvdftsulEVUigMaieO3v17rt5UArLNkLk/npmh5dalWU/zp2rvAlVAq8asBu53ZMf9suidl790tFBQUTI/f5yHUlykUrW71qdux7UFncm2KZxOes0uVnc5nd34vKNiDePAqWYEFrd9K/BY+yiECSeK46bqNrFo9Gra+yfujmYlwvoBgcc4yMVbid7+9BZQFVQfXN0tXCzOL9KyRKCgomAu+XmWdJqXSlq8pFULTppwPQQ5k5yg3efyt7bfu5yisj4boICu6VeKs/vu4Kuyk+0x+zmzDDPfB+ji1R6GgYI/lwbmtTpBzd94NT3zeuazfVqIkZS/uQnJ0bua6PYgSvxeiaO9eQeqU4yH+4V/P5NV/fST9vQ1EgffYNdNehl5hQ2JQCUla5ysXbuXfzr2Y1DrSNJ2kFGXCTDp6rko0aoZcH+7v4yP/eSbPe+ZKEEiVIyp08oKCGfB2cBGo1Wp8+BOX8KdbBkEpvy2XYoq+bWdHa/I5IkGBapLTVEIFr47exw23jDCxrYIYhzMTmA5NRkkJqxM0PSzv0TzshAHGzZLwa4iH6lQCpyOvLXnlrvVT/npHXCrxif84k4P3m2x1LyjYEymUrJ2lZAFaaijp5fQnruBbnzuHyFicSrySJfEslCznXUirGtUJw2vefBE/uuIBrDicuEmKkxDkb6FkFRTsJFxTQbJpg2o1JUkMSvntrrrX8WD1aaNDWQEks1I1yZQrHwS44vdV3vL3X2Fia5kUhzUNTIcCpxCsGcdoeOTJe/PFT7+UfhN2dOimXKlOBbCDZhzoqhzmFUilNaWSJS5HuWsKCvZcutWIBw2ZRXuu5EXDbMWEorWx8prVo2zdCDgTfplNJJRXsADrhE3rHXfduRYnLljJJp3t4zabW3dBZfI2972goGAmvEgVBGUi+gZ6GF4SMzgc0zcc0TNk6BmKmqF3KKJ3uETvUEzfcCkXYvqW5MJw5+8d5w5H9A9HDA4lxCUJm1SnTZmRR0mMcv0gvRgtDA5B73AIQ4reIU3vkPFxG4roHZwhZOcNReHazuB/6xsy9A4oTDlqbtaTU+UKCvZIHtRKljAX7SGoLaql0HhFRIMYvy10p6YzCf+wtWvHuPGGdaRJ1guc6boWogRr4c7bN7Bp4xiEFYr53wUfJ7WDs1ubV4cpFAUFBbMk1EnBtXosKsxRarMUeTcrSoW/zdDqKClAKfw5XYP/HQVGj6FUitINlLbhh46oqUZ4dhkh7jJUSfg9zBObU+hG9lvrfVt/C9lSsGfz4FSygkBq+oEKtTwv1PK0VuVpwKJRGJsSiwZnkODV3SGTV+3kFbIga8dGHV/9/s8ZT/HOSueQDYJj3MIPfnkTo1UVJtCH31SmZBmcVShiNDHiFFpplCgf2u5YUFAw76hslWFrtaEKNV37Lhk6U65UF6WqTeGaQvFqnuc3XdYIxkU4saD9NhZaSaah5QIo5fykfA2gm/fLx60VV9UMKve5FfxtsvOzuLTfJx/ndhk0+y5mQcHux+xb9z2QTHkiq/y5RTGdZKJDkaKAgbKg3Dg9lRgJHTiX+aLK3zMEcvdNUsW1d2xh7RYQMUA0xVyNThyCsHY04UdX3cJ4YtpElAtKFijiuExkQJNSjiKSRhr89+hWBAsKChYEhUKplgqiMK2gdCs0VRXTpqp0qlL+jrl7YPx1zXv5+V7G9Xr/eyoCpYMy5X9rBsperdIpoixK6aZdrfk80SGYVqBLyH5z4fxwfZtKmN2ryzldRG1BwR7Fg1fJyiaEd/akmlpRN7wCU44jzjjzYbzhTX9FYtehtGua7aei1a8FHQkjm2D1vZu88jWrbBBQFgesvnecrZtqaLI5Xa1Ie33PkjZq7LOf5l3//BJMVEXrbFhgmkgWFBTs1qhMpinVtJJ1kh1R2f9yHhW2m9lcPJtzCgr2MGbTuu+ZTOGqJZs90HncX+BAIrRyHH3MCp79vGNYeUQfRqUo1fCr9pT2Q3Pa9y610i1TeziulaE6Adf98U5so/M5UyFASkMcN1+/iYkxENH4ee8+tkpCz1nViY3jsY9byZPPWs7AgKMUV9Aq7W6mK9glSG6mSsFcHfLufkyyKIUwW7LdIWYMZJ+nLlxNMdApB6e5pqCgYO48eJWsOZM1iQqlII7rDC+3nH7GcWiTopVFqckredoJAlVKJI0qV/zkGhr1znOmQgDLRK3K9773C+/lvZt6JgaFoq+3xKNPP5KokiCMkopFxD3ohgrzjQ9OILeVkcyhUZ/LuVORPdOJPCiVq+b06A6lgM58KtgxmgqUdA/ZCUVaF+xhLMY+woNTyQq5kG2krHzbi8vpH22Z1VwYJChJqUYGpx3LY8OzH/1wVlRiIhlCSefAX+gfKpWbr+VQYhBJuGOL5Xe3jpG6qVbltGOd4c47x1l9dwOXZn50whyrMH/CEYNE9PVVOXr/AXoTjXblMJ/LT9BfbIVwwRDfkIikQIIowVlAVRE1imWElG1YqWMTqNcs1eoEtfoomzamrFsDD9wv3Ht3ndtu3cx996/lgXUbeWD9JtZv2sLW8VHGahOM1xJqCaQOHAmOBkIDoYowgVANxwWnUkSlYSfL/KqrPYus/uRD9ougcKKQEJwlLMoQsA3A4mwDxAJC4vyenAI0JKFuGySSYklx3hsUDhv+OfwvPrhsxdvkiOx2dFrAuoW8stqyaE0RCJ2H0PloS5tOM3/H9Irm6R3ndB6adCB3Tn72WUFBG23lsQHUgKQpMyUE16z9XgL47mvoT7gQOsvsTqZwRjpLZ6QS/meoYQaW8u7Xn8gbX34qI1vhdW/5Oj+9agtSSdHB8ehkBCXOryaUGKVHKQ1q3vTqx/KPf3caZqrLAiKORmr53Fd/x4c+9HsaVYVTGhckmFLglCJVMRUSTj4u5utffS2NOpzxtHN5YGQJytYR6UGRzDjldEl/Hx/9wJk891krwYHTLnim371wkuAE7ykfQDUQZ0kThU1L1CZg81bNzXdNcMONN3PzLbexfsMmanWDc4bUORpJgpWUUlkQ60ApyqUSSMrw0AD7778vxxxzJEcfeSjHHRjR3w/lMpTKEMXiV3o5ELGIToLvRj9xmbDWqqtVcjdmKoGmcIjzypNCI+KbWucgtZbUNkjSBsYNoawXli4ITRtuqBQ4B8ZYyj2WKE4xUR1tLIpKmJQdIUQ+ZTsjsmcldRtX/WozL3zT/zGyrYSSqWq5QdOAWHjEw4f4znkvoxKH+pGRTyPx3zu7gm3KVDhnWmZzTkFBs9Q6vy0dafheCnU7tHn50i00t20ChYSFJN3Kf1u53Qk8qJWsu+7ySta6kdkoWQpEMKpG1L8sKFmnUE/gW5fcwr+e+2PGEouSqO26FqHZCY2KwqLKDZ74mH34/MdfwsBgh8CivSQ4J0xMCG/+h+9y+Q9X41KLI8IRhfMUosBqQ6+xvOvtj+FVLzuBsVE482nvZ83oElRSwz1IlKyQ2lhJ0EpRSy1JElEdh3vWJVx19Q3cdP0q7r5jA6tWbaKaRjhrwJWxFu/kFd/JB++pWiGk1mK0QWnjK7+yaGPRJsFEjuGohwMOWs5BBw9xzLEHc8YjTuTwfSHqqVPpTbASo40OS/x9emaLIvznPYMs/TsRSVE4rBVwMY26YqKmuH8T3HbrGm67dRX3r9nE+Fid0dEJavUa1gnj42OYOKavrw+loFyp0NdbYWiwzD77D3LIoUs49LB9WblikEoZKj0QlxxGa3ARiF8bDCC6FbO5zInaHZg3JYvJhXFaJaugYN6QDiu/xjnxTnTzTVA2+h08k3iZY4NSJiilcythW+zsclsoWdupZP3T353Im15xCs7A9Xds469f+DG2ji6fJvvympOgbIlEj3HMIXDhF9/BwStDU5uVmgzlr02tY82aGk9/7hdYu0GjpIYjxhG2w8Cfl2hYPlDiom+9hiNXwtZN8ISnv5/7R5egkzpWelE0gr/lqVnS38dHPnAmf/0sv62OVQ6zGylZmVG5IZZ6Ahs3WX7zq3v4waV/4vc33MNEzVGrCkgJbExJV1CARgdTtN/YO1+RwfscUnhfQ4L1m3+Hk5wCXC+oFB2lOFdlRX+Fhx21P09+2nE84rGHsM9+/fSWUmLTsl/tWUqWT6xOJcsPTimcpGzZFrF+vePWW+7juj/fxW9/fzN33jfB+EQNa71nJSe+x+r903law/n+g1gFTohiR7kEpRIcuKTCCSeu5LiHHsAxxx7IMUcuYahcR5xQLpd9/9eEOg5dnXXuzhRK1vySH16dju1ayJD7PPurHgwI1lqsdThrcKlBnB8wnKhCbQKs8+XROi+TeyrQ1wcVA0Y30CYlihVRZNAqbkvhyeW2o82dZx68ShZwx13wpNxwITklq7OydCpZ//x3J/LGV56CLgsbRoWXv/qT/O5qg5vKkNWGQ9kKiZpg7yWWd7/1qbzkb44jUqCVDb5n8qUhpZYI37/8Rt72tp/RsBW0qobGqF04Jspx6rF78cXPn82+ewub1sY8+Rn/wX3jS4jTlFR6gfqMStbuuHehiOCcQxlFKlC3mnUb4OKLr+G8837Iugc0MISohNSmaKN9fktERD9KWZRKEe1n9njFwGeFn0nkK7Q3hEirIIkBiYLrRY/SXoVy1lKKYpxMMLCkxnPOPpW/fcHjOeSAEnEERol3GCl4K6fyw5G75/ChhE2MfTkRsSitsKJJUs3WUbjxxtV85et/5NdX3Ux1QpEkMfW6UKlUsFIF472V43pbQ7zNu2f4imGtpVyqkCYWsT7vjTJAgo4alMtwyP6GV774ETzhiY9k6XJNuSJo7fx8IJGQzt6yKMiker+7MW9KVpdk6JQYkxurxYf4yuvRvs5K021OK32UEpxrTTxLBBoWqhNQb0B1ok6a+mtbeHlQLpcolQw9PTBQAqNBaTAG/Jg3aGVCgkkYAms5oc6n4WJPz/mhlSl+cC/ITfF5YBMYHYfV91vuW72OG6+/jVtvu5trbt3MRK2BuND5RRDnEBGM1pRKMYftbzj+uEM54cSHcORRh7JieQ8rhiAqgTbZPMDMaWWW2tnnhUn9B6eSFfL49jAna8O2EhVX9hmXJYeeWcl6w6tOwZQTxhPNVVdt5jUv/xrbVJce4ST8Kj9BiEslHv/offny555NTwyKBrhss1YfhJTRuuLN77mEH3zzARJr0KrqG7UOxUci4YVn7cuHP/xXlCuK9feXeNJZ7+O+8SWUnMVKL0K9i8hsZ2igjw//55mc/YyViJ+KTLwbKFkoqKU1tk0YrrhyFZ/6759xw3Xb0JEhimMEv7qveQ1hzYDyVhPdZo/2qNwydxFBOT+pOLNodSU01l5Rc15FU95CundfynOf/TDe9MYzGBqE2CRo5Z1UoixKEUzdU918cSJYxFnEeeUyigRUTDWJuPCbf+TLX/oNt9xdI1ElP/HdVzgATHhXnwYdN56BTNAqFdyYKF9XlQNNhLgG/cMjPOr0fXjTW57HcUctoURK7NKw9VQM2m+L9WBUsr79pZfR06lkdaFTYixcszR/hK1dfTOqBFF1bx0VAxITab9gxTpoNGDjOvjNb67nF7+/gZvvXsOWzVXSJALKVMfT5vQBAG0UIhZjDKVSib6+Pvbv38axx+zPGWeeygknraSvF4Z6fDx0BJgUoYajAvgeeT4NJ0ufPQwJApU0LFspISiqNWjU4fKLf8ePLv8Dt9wFm0YrjI2NoZVCmzDkl5fdzc+hQScsnlf4Dli5xPIhzUMPE17ysqdz8imHsmQYYiMYnYT0V6FkqwVL/QenkoXPk9vCcOGGbSV6tkfJevUpmBhqVti4NeEFz/kCt9zvphBskxEErQ1HHhpz0bdew5JhPwCIC/OsQhSsTdg0onjC8z7B2rv6sKL8lhjNMevWHfuHYz7yL0/m2X91KCYS1t8feyVrbAmRS7HSB2pmJWuwi5JVWqBCOF+ICA1JuX/DNj7+0Sv48U/uYutoRCMxxFGEdT69vPUqXKOyiqlwc1GyQqM+ZSszScnyz1ECkS0zWKpw/HGat7398TzmUYeAAx35DX2Vzhr7qW6+OBGx4BxONCmGaiPlou//mfO/ejN33jHORDXG6pRE13Fagrz0GbHjSpZP8ryS5VcXeS/jPT1lrK0yPNTHs59zIK977cPZZ1mZiopRygYla7IFe3dje5WschxPSvbO750SY3cooUKCkwRxvX7BhLKIquNQ2DSmWlP84do7ufwH93DjTfex+r4q1YmYhq2DsVinwMQ06kLsSiinmvtPOpcSlyLStOGruzYoo1EGdJzQ0wtHHbgXJx+ynCc/4whOOGU5PX0ao8qQG1JXtEYJJkufPYxQIK2CuljG6oYrfrGaSy+9nT9du5r16yZ8mVJCalNcWHmvdLD6+R/DvXxn1+PPk0zuOu+ySKEoR2UqpZhlK+DxZxzFy152Eoce1E9fHBFmDixoQV60eeqbplboLiy2k1C6MyGRT9+s0e083g3fOAsmEqLyFk571FJUUM6y9mPSyqYcCoU4x/oNE9x8a52UulewHOGtLQKkqWbd2gnW3W99D0y53PYUeRy9PY7THnU4URQm/OlQMnVrefdsyKdNtyctNlJrqTYa3Hjbfbz9jT/hou/dx+atJVIXNYcFVRiGy8+fmOs8ChX+N62ClZXe5ua93kGtCcNSqaRMuFH+9JdR3vGWy7j4+/eSyARarwFlEacRN7f8Wgz4bWIsVle57tb1vOEtl/Iv77uGP942wla7FYk3ecGXuUQJ6ahVflbadqD86lpC3gK+U6Q1mAjiiGpao+ZqrN+2jm98/WZees43+c53b2NbbZxEworPBzFZHc+H3ZLO6iIxil6UhlRSrKpTbfRw000pn/nUn3nJ87/FW/7uSr73vXu5+aZxtmwep17bCkkdVbPEjZS4Ok6vm0DLOKKqCAmCRWuFtQ5FhHIRKlWouoaqQrZpqg8I11+zkfO/cyevft0PefnLLuJr/+8WbrtnDePpFhpsQ6hPGo3Yo1FQd7Bxm/Dd797C3736Qt79ru/zw5/9hfvW12i4EnXxc2lFK5QxoU77fYEl10b7XQ2yLaVa21SBRukIE8XoWEjMJkYbVVatSfn6N6/j+S/4Gh/48K+44bb11JKkTcFYCJm7m+du0JbmkTndUQASFHUG+iuc/cLT0EHjbp0ShkSmCUnD8O1v/zQMZWVX+hUSCNTrlu9++zLSel/o/Yc+ZZuipVEKjnvoYSxZ2tFoZQ2PMJe3W9yCNugx4rxzy4ZNufueNbzjdb/ht79fTc0ZrDJYCdarzpwNCZQ1yahsfk64fW7CtT+jI912JHEiRSJCnRprN9X4wHt/xDfPW8NYbRnWRv5JYehqLgrgzkYkG44JZdnB2LZxfnnFZl7zyi/zsyvXMOJgXNWopyUS2wM69cVxitdS/tWnLqaZEGybkCzButBZ7h2YBinjuCjFarCU2Litzp13N3jfv/+Yj33092wZq5CI8553JOc3qmD3RAScQ5wvJ4oU1Chia6y6ZYyPfuiXvOKVX+Qzn/0VN90ywuYtKY1aDUkgUn24tIck7cPKACkVEh2Rai9PWk2m+PlWKgXdQHQD0TWcGUd0FasTUgV1cYy6Optq41z1x9W8+19/xAue/SW+8/XVrF5tqZFiqbaVWxHn3b1km+LudrjwDi2FRbynQuqNBjffWuMd7/gu5773Cn7zmzWMjU/gJmIiO0KsNhPTwGQdMQnWkGb6TCE4cmRpqURBWkLVlxLjiFwVqSo2bm3wxS//mte96itc/K3b2LK+TpqmrRYik2fzxC5UsrJXCiu5aAStvo6jHhw6pqHXkITfffDOyTLHZNNJ5CkIp6sOZ6T53uzs72rQGGJV5ujDD+Xkw/YhUmlIWuPnXk3Wq5rBAfVEc/VVdzMy1uOVLNXwq9ZE4VzKSDXiu5fd6xsWbJuhOUNJTKQNjzthCO3SUEb8i3kP4wLZ3KN89KegaW3IPneesBgQ38onVtgwEvHWd13Gzau2kVAidf6dldI+nbOQy1tf+rJ89g21CwHXi7O9KF3BKcGK30y3ZZ0CrRtolaIlRmNQOkEpF5Q375/JB9UKQRm2KBpiqKK5d0uVj3zmR/z4h5uYSCzOJJlevGhxwVTvAOsU1XqDrVssX/rmKG9957dZuyEiSVJo1Ims307KKvHuQkOFy5wB+4UECkQ366QWhXYK7TTG+hBZTeQMJRsRpRGxjYidI6bmnZmIn2MhBMuW8oqWiQBxfihYaeKeiDoNRhqGC759Ha/7u6/ywAZDPQkWROog6YL0ahcjTaU2H7rgbQStsHiLqJ/zqqxCOYuQsHWkhwsvWsM5r/0WXzz/j6xe6xipK8ZtAyk5LEJdOepicZFCxYLTKVZLmKRuQr3P9qn19dzXAoXTCqc1TmmsCp8NOCM443BakaoSVlXYuK3M+8/9Ge940+Vc9YttbNk2RiOpgQ3+d3PjN7tf+av69lkl3gl0mKTeEMf6Ucf/nn8zf/Oaz/HzX65iwzZLQkyaxqG+xgiRt0JnKDKTt//SUeh8cW1W9rZ/4C1hEkXUxdAgJlEaayOsHeT2e+Ef3nspb3jX5dyxJmKCEVL8avFJneodYBcqWXm8gM28lntHYlEo3BGEv1nwfi/yBu7toIviMOdkFbwFCUNERF9F8/jTD6JcikA1wgmmiwRrD6l1bBtJuOuOce9wUfl5LSKaNHXcd98Im7aWwvmZY7b8PXxBrFQcpz70QOIovFlwjNoqMpML6e6NI3Upm7dZ/ulfvsq1141RlxQXNJT2V+1IryYKJRHalTDOYJzDSEpMnZgEY1Mi0Qz0DBDrMsaUUKYEOkJU7N1oBI/+TllE22AFy57TEgCg0JkmrxROGaxWNEzKfZs3ce5/Xsyfrl9Nte4VhcWO7wQJVqBhy/zP//6Sj33iUtZtUjRSBThfc8V7WBPlmm4Y8qnicaBSlKojuhpcY2hEqZC+WdMT/Dxr660FJiHVDqsIeRDcPmRDw4TOVKipPh6C05aGOLZOwO+u3cB/fuByRiYgseBUilPpvPZmC3YGzluXQskSLPXEcMftFf7xH77H+/7jcm6/N2WkqkicJkVhtcNqGzoAWRmyYWcG7zTX3y94WcsKVFZwlQodVxVWF2chHGsW9PBBCykRY/UKf/rTCH//5m/y2U/8hc1bDbWG+PljXtNqqx27D8EJqHj54JSl7uCe+6r8y79cykc/cSX3PQDVhvZ1VQSIEZUps+2rtCEoWKpL29UmQNqlSfY98w0vOsapCKdBEYFEqKifiXSQn1y5ije/8Rv8+S81xmtCGvJvvth1SlY2sEprk2MRE1Z9GL/5cXPrjaCABVFpLUHxapbgzrvPih220ISL/VsoYgNnPvlAtGkAdd8tmYWcFizjtTo3XLeR1KWgYr/yQUHaMNxx+wYatdBgqEypElTw946AUGfZ3oZjjzvIu4FAecdCmcV5FvHIo0L65L8vNlIRxuuan/3idn52xX1Y3d/2mpklaTq8xURjXIxJDGVrGNAlBstj7L9inGOPiDjthH4efkIvT3r0vjzmxEEecXwvx68U9t1L6OmvU+qpYiKLkgqgfLWWvFIXRG4r65rpKUp5zw2VmFVrq3z4wz9hdCxqZttiJYu/wzFRh6+c/yf+72u/o1pLsOKaG6QDIK1ZJ1mKtOMwWtDaW2k1CqMdUWwplaHco4hLQqkMlR6D0ilRJEgEqS6RSg+pirHa4lT7whNvPWwNI2affePoZz3WEsMPfngrXzzvN4zVoZ7iO3qL3Zy4kEzOpEVObmjAaVKdMCqaK3+3jVe/9gJ++OPVjI3HGF1Ga+NdrGTFM2uKOpjvJMhkqiiF1Q2sStm0RfHlL/+Rf3vfN7jngTrVRKEooSTsVrCblUGRGOcibKoQiak1SlxzwzjveOc3+MHld7Bts8W5zIFQNomju1TopF2iTndFq7Y3JZVkdT/Uf7zDaScOK5o/X7+Od7zt69x86wR1J1i8JXs+2HVKVsBai8LgbasarEKsIixU8nNuQpBgTrWJIamDs/NQAHcgHTsv1QoecnwvJ592EFopxPnhvZnw+y8ZLvr+NaS27ocglUYpy0TVcvllV6NVb85sGgirXFAKbRLOeOLxxGUoRcHPVrZ8uW0Vxu5OeA8FTkWsWat4z3u/TuqW0pDww0yEU7IqaARiEQbKlkP2jXncIw/kwx/4G77/7Xfw7QtfzflfeTnf+OqL+drnns2FX3wxF/zfy7j4gjfy08vexrcvfCdveN2pnHhchSV9KS51xFHUUTqmSP+sd6YjGtZgjeJPf0z46teuoTbrjcN3Jq13UMpbn50Y/nTdFj7+iSsYTWMa4sujUqHcidcoO9++NefNF1RxdTQ1Kj0pK/YqcczRZZ5wxr685JwTeevrn8C//ONzeN0rT+dZTz2Cx5y2jJX7Ndh7aQMtY/TEZYwzaCx+o6nOp+HVqxCvzkUPohSJK/O1r9zAVX+4g0T6saK71PA9mG6v2u3YIkOaIXTGrZ+HWSXi+z+8mTe87Svcck/KhCuRSN13Up1FBR9pKiy6IFTHKdmOtMiUqsk4UAlWaqhYGE8U37/kHt72zi9z3ypDUlOI65D1uwlCHScpWpWQNOLmW2q86R1f5ddXbWOibolKEvaSDRIxmzfTvH5qJMiNTHA3kyffJIZpGUg2GuB/9J1cH7IS02wXjVCXUVbdl/KqV36S627cRCrGn9U8f/vZJS4cJBS+NE353R+u55vfuBLcCp8gknl59sKPZq9fglcFQVLLXnsv4aUvfxwHHdzfefuZCWaCO4MLh+11Rvp3rz6FuCStG2JoyEa+fvEm3v33P6BeL4GeXiP2V1sqccRQHPH97z6Nhxx6MJQUQsI9d6c88zkfZe2WvRDtW96WUHCIEsBgzARf+X+v4cmnLMUY8XNLLKx7AM54xnu9nyxrsczO4/vickaaTz8JKSBsHtX89/9cxf9+/momnIGSIkqnEmw5lK9+4gStFCUlLO/VvOxvn8ALXnQcQ0sbVMoao8LwkiRhTk8EypCkdaIoRlDU0wlQQr1e4r57U8674F4uvuQnbBuxWOetrVlZUuLdBBCGvfx/4hv5NKESlyhTYr8DFF//0t9y5Mo4G3TAheGKXSd2JQxVx6G8C6lTrF6jOfsFH+fuNYYkrmHq3t+SyjYrDm/QhvZ5qMUrRpGu0tczwfOe/yT++pzTWb43LO+H2Gez3/rROiolTZKCsyn1eo3N45rb76lx/pd/we+vXsXYaEJNKjgVhugl1E3t8rVmktAUcd4RsFQ45ZQ+/ud/X8r+yyEi8RaFIIc6X2Mxsj0uHGbrjHSx4fBZ3LRI4TCpYqyR8MUL/8RHPvkztmyNiU0ljIbUQSVoTGt7M6X9/MKw+rg1scK34gpCw+wb8GajrfzzsvTNXAc0y1awnKD8qjh/PfjZmM47MHYRmgjrhISIsh7jhCNK/O+n3sDhRyjvV2s3KXfg206ntiIuJqn3c8uNCW9428e5bWMf9eo44soYHYwOkq0GDDRX5Ss0Ouzz660Ekg1J6DrGaIwxzQVm4gSbgqQgolAuRvDDvt73oc9DwWVip5lFIScRBKUd0KAsSzh0pfCVL76alfv7+Xdov3oxf9VcMO95z3ve03lwoZFQ4OpJnR9fuZHPfOlu/nincN2ddW64o871d9S58Y4aN95R46bba9wUPt94R5Ub7hjjL7em3P7AJp521onsu5dXjraHTVvhK9/6JeN1QyyRrwSZUjepF+G/a5WiS72cfuo+nHryft6rb7OqaZTrpRxv5qLv3kmtXs6N63dHAVr7Si+NhBOOPIaHHK3QWpMkih/9dDUXX3onLgYrvmHO8lkpCXvnaZb0wlte/TCWDJf8zBMlaOWYGNd87fxfMpb04ItfKfjZnTpOAOVyiSc+8VAecuQSUF4pyHwZ7TxChvhq0/bdUeXW+xLO/fAlTIwOoJRBBb8oU6FyQZsUDZRL8OjHLOPDH348Zz3zaAaHGpRi3+h607JBqQil4rCps0LrUkgNTaRKRJSpmJi9hmMee+pSHnHqwdxww91sGrMkzgsNaM0p8CnvlS/vvkBhtFcMRBy1qoGBGiecdDBl7VAKnI9J+wvtdDIHfqOAZX21zEc/eQ1X/3IUp8cA0GK8uxDl65C3FOSsRlohegKDJqbC8uGIV73iofzXB87krCes5IAVwtIeoWKgpCHWglGO2N+WSAuRgXIpYqgv4pB9yzz9iQ/hrDNOpF8bblizgbERRxRFXonWkV8B1ky7VhqqIIeUaJAyqDE2rqswuMxw4kP3pmRSv1ChWeGaly5aVq+q8p0f/JF63ZfV7nhXGxg44IAKz3/2icQm80i+e7yn4BUsF7ZXUQosdSaqE3znOw/wsU/+nI3jMVaXUNahxPnZQi68Z0gdpRxGKSKtsDZF0pRICUocGotRKb0VRTmGnlKJSjmmr6dMKVaUIkscWRQpsdZosRixuLSOVl4++6TMcsLLAf/NlyuF9w4vOkUpy7qtCX+66xZOOekAVgwaP38orJxd7PmilMJZRyox994b8Q//+F3uXh3TSBqgomC4D+8d5AJZXoa5cNqUSW0JJylKxunrgYMOGOIZT3sUz3vBoZxzzmn8zYsfw9lnn8pzn3sSZ531UJ5w5okc95BD6StrRjdP4NI6jWQcFRmscxjlu3oE67pW5KzdPhZeCY+xYtm6FWrb7uVxjz6YUsX4zopKmnk214zYJaaJZvMu4KwlSROSNCENf5M0IUkS0qT1OQmf0zQhTVJfIfLud+dCSKd8E77jhGpkHPvuux+nn3EMiduKc9MPF4oCZ8ElEcot44pf3syog7oeYdSO8bs//5mGi8D2TcpeXzgVUaw46JCl7LfvkpwFbvKEJD9803ZoSvJpk6k3Ox8vjDLh1LIYJjRqMb/42V/YsH4CpQwawcywp1H2PspViFD0lzXPe/ZJfOK/XsZxxx1FrAQjGk2YExd2fM//m1zBwjEFymh6KjEnn7Iv//t/53DaSf1UdGugcCrbYfMtQ5e3kSRc/O1r2bChDjrxcxUnPXdnky8JmjStcMuNo3z/op+hyxNolWDSAT9RPZTr1pBgO8pWiGUJBx1Q5twPPIPX/t3p7LXfPvT09vheqsopNlOglFeUo7hOqW+Uwx7ieNs/P4KPfvQsjjysFyXjGGP8ApJQ9iF0UrrgO1iGWm2Cb3zjB2zcMOaVrxnisafQ3oWZMpkWFUK2XqpByhZSqXPFFeN8/L9+xOg2g7GKyPo3y6xRPjsV+CUbfj6KBWqD9LGEQT3AfgPLefxpB/KyFx7Jh//jSZz32WfxvQufx69/8Wp+84tX8sufvZyfXP4iLvrWCznv08/mP//1ibzupSfy1NP348gDljJcLtGjI0xq0C5CO4VquhxR3uGwhBKuACxGVUFXiXQvf7lmHZ/93C9Yt1mTEF5STd+OLBascmwbM3z0Y5dzy20jTNTHvbUxW8mfk6hkckITrEUK22hQ0YqDljjOfsqBfObjz+Wib7+Yf3n3QfzNc4/kqY9dwSOOi3jEMcKjj3E86WTN8x5f4e0vOYQvf+ypXHLJq3jr2x/PKQ9bSYSlZAgWx2AZ6xpaA4tKKRKb8r2LrueyH95NzTqsWBS13Fjl3NglSlYnPgPax0ubZqXss5LWCkSynNlOstu2msjWT6Fx6Dw+NSqnBACS0tfTy9+87DEML61DZh6dCgFwGOOopnWuueF+fvH7MX77px5+ebVw1R/ugXgczDZfMPNzSvC9cKMdz/7rxxFF3t8KobBkc2OaVoQ5zM3yk+p94fNvN7vrFgafG37INgJSRrcafvebe3FJP6n1flmaez7OQGInsA3Fk5+wkn999xkcsBRKJsUY8RYraSlYU5EvOyGpvSyM65SMZr+lPbznXc9l3+UTaNNeVn0Uu0Q0K0IijI3087vfrMLZCt5dTjaZflehEGLA0kgtY9sivnvh1dTGS1Qb434doQqOGZUOK/na5z5p5Y/1mmGWDG/j39/7JJ74xH0Z6I2IozDRNyRCfi5Edn0zJsqb7/0E4T60G8ToHkoVyxMefhCf+dRzOerwfhQNTOy3QvENXPuSeGnWd4XSJaxEWBpsXA83Xb8GZ32cmsMV3fJsDyGTYPmw2PGWfG/lcQi33NTgff/xbdZv9JsKR04RO4ORjvITZL92IEmKIaFU3sARRzV409uO49sXvYhPf/av+ed/egZnP/dEHvvwwzjh6APYZ0nEPksM+yw1rNx/gOOPWsHjH7OSF73gJN75ztP57089lwu+9XI+9t9/zcMfPszSYUclUn7zHBfat1w5bBZFiTFuH5zEJK6BoPjeZdfz9e9fx0jDe0f3VuTFjQMm6mU++d+/4Ypf3s54WqdmvTqrOmRBPogSVFTHSImBKOHEo6p87n9fwgc/+GKe+PjD2Guwh6WVJQxGg/TrPnp0L2XdTykaIIp7UaUSql8TDcGBBype9eoT+cxnz+FFL34oS5c4ylFM2kgQa336O++mobMzKF7KgtZU7VI+8j8/ZtUaTS2xQNhcfjtk8A5oKjtIW1yzV5wu5Amt2gKIgm5Pmx1ZXAStNfsdpDnksOHmFpiEIbzW6V5wK/zwR2LHkGiUVetHeONbvs7LX30er3/jeax5wAAlP94MOWuKH0vWaPr7enjcGYfh9ynvQhAwvt2a/dvtWsWqk/DeIUrbtgi33LAesZXgtM8L29nQNzDIgQeVeOMbn8byAX9M6bTprb+zEExXEie3SmW0sgyWKhx76FJe/fJn+zxz/mQ/1yAI2CB0u+VJow6//fUqqqPZvXe1kA1lGCGOSmzeWOe3V99Era6JzADYXlBJaPRaVzTLqwLEoFH09G7lXe9+Bo98xAH0ljRxUL7IqnVgcqp04v3xgO9sKKUY0MJRh0X8y788j96BKkla94J10t3yualopDW0jjGRojpe4hc//wO1qg1Fzvvym02MCnYmqZevErNxo+NjH7yCDZtSGjZBlA3+1nLzftrkg5+SoZyw7z69/Pv7zuR/P/+3vPb1j2a/A8cZXqrp7dPEFUFXGlCqIVEDa1KsTnE6QUwd4jq67Cj1aPr6S+yzj+PJTz6Iz3z65Xz1y6/lKU85EN1TIu6p4MSvZu2UI0qBdQ2MUYDBNYaZqO3NF778c+5bUyORkXDmIix/uZexNuUv16VcetlNTNQjGq5BqSdqGYCy3mizM6VBFFoi0nQpJt7AC154GF/8wut42ClD9C1tYHpTJLbB35j47bh08PmIbxuVioLTBUekGvSVEvbd2/KP73oC733vc1g+vJTecsVHN0yIn5SSyoUggGPcNbh7bZ0LvnkttUTh6uVuYnpW7Dolawqa7VXuhZofhaAT+37yjtLsoGZlpUPAzyZN/Xn5yuuHrFYMRZx+0kp6VN27pcAERSCzkDTfFGsdUVxCtGBFMzZWYfMmod7ooVoziJS8laZZSPHGbpeitOPA/XrZbymtjaU7yN4lr7HPhC+GOTPqrK5aaHwXtNqo8ds/3cGmkYY3+avU+1lRtjnZMQvNNMv1nAb0GP/07qezcmWMFYc4C2GOVbBc+0vCU/O51Z5z7TTPVwqloVSGJz/hIazoH8OP5xufjsFK08yXjvuA0Ehq/P66O9haA6XqMy5UWHhCTF1MrSZc8ZtbWLNhAhMZkjTxvmhEvI8fCcNyArhQX0VIE0Wkyhy90vGE0w+lrwxKLH5QxJ/nOyMKmltkdKZ6PoTfw1clBoWhr6x5+MOX8VdPP55YakFahHt7o4LfoLrpKNgRRfihfQFxEb/6/d1sHtHYpm472RK2R7EbvFZnfRFAxDA+Bl/63A387nfrqTbAlFRY1a3Qytc8pTI7qcbqMroEvb2jPOsZh3Pe51/LC84+kUMPjukrW3pLPWgRjMq8PjXtsqGo5T95JS4rjRpFJXbsvRROOamXD3/gr/jouU/jkP2qlEsWS6Ppyd37hsr869WD0qERZcDFbFln+OC5F7B5k8Pu6urfSVtmOKxLWLcm4txzv8P6zSkN57xsSDo6/hKses77DNTOUJYelkuNd7z+FN75tmeyzz4RUbYjHK3tyZoNdvNzkLWtrhwav31Zj1Es7XM87akrefPbTmHpkCLWBqPjsI1Rt5EK8Uq7skS6RKPey/nf+CVr1ilq9c5zZ8+OayrzTD4d83jtM2gJmRUr3+2dC6FwZBmTV7TmQr7Ctw76IavINnjWE09juE/8eLoQ/oYkb16sQBmcUz7jxRcsrQRx3vFotii9adoMyWCdI44dR6xcQn8ZYtPXFhU660K4x2xpKVZd33QX4ONhbcTv/3wndSt+UrPyFgf/t12xaipXTRGrOPKQPk47eW/iWHBGkWoLZH5pWgpWFmaN1w18JdZgSrDPXvCMJx+P0a2mniw1p7BigUIpy5aJGjffvSVM3M75ndpleOXJpmWu+M2N1G2Mk8SvoM05yVXZHBgJ3Y+s0IlBKctLXvh4lvRrIgUmq8o5hSpr0NpzoVvwc+d8nolP31gRmRI95QbPecbDGer1d1S0rKBtdSL8z7nMUqVxzrFpNGLthlpueGfOpWG3oWlg7xQWi4x89AS8zBTNfavgsktWMVZLEB35gSCVdRP9Z5VVTqVRJqFcGeX5LzyRD33kbI49OmKwbKgoRUxMRIVYm+AOW2Ek8rs6SBT29/A12RBhMGRdaKNAYzCho60VLB2C5z/rQD72obM55sgV9FdiTGYxp6VoKS2IiwGFU4nf9cMa/vDbtdx8o2W0Wt3l3axJZOVEpTRSxeWX3cPdd4/RSILLYOfXRPtVmS38zgugROgtx6hkE6998bG8+qWPYMVwhNGglV956dM3mFXEp3smG7w4DKUhdMy8HDdhYU1Mr3E8/+wjefazTqS/p4xyzq/yFu+Q1IdM9rQwYjBU2LQZfnHlXaQSytF2yOBFo2Q1RVinHG3+IM0vvjGc+8t2Y37u0iIrd1Gk2Xf/fk465VBEjSOqESxZk5N82jiExqqzLVZAqVShVBHOevqj0Zm5cx5pG93M/7BLyEbRFWnDcOvN9+FstvImlI1ZzJuJoohTH3kYy5b1eluSss1SNV/ky2Zcgmc/90mUyyFuzVYiF8/msdaBRiPFJoa//OkOcOVwyfTvtrPYujll1d2b/TYY4VVVU4fptD4RrHcKE2n22b/MIx99IqWyRqx4YddUlHagXue1Y/yw7HHHL+OwI/bpPLONpvxXQHBkKoBNSzywZmNzSy+fXzsQv4Idpr1UeWm6dRT+/h/PY82GKi5KsLjQCfXuUryLm9b1WimGKiO86uWP5u1vfyb9Aw3K5fb7klmlVK7TNals5mMzfZlQkeXUU/fjvf92Dgcs65tkQ/GfFSJ+31IBxAipThmrat7znm+wbcyGrawWHxbh7lUNLrjw14yNN4I7i1bHNo8ocEYj2qFUQn18HY9+zHL+5rUPo1QKZyv87g/4LyrrdnWpe5PzpYUGjCjKSvGG15/OyoOG0S71fjmltYOMEIXFRbkQ2lNnK3ztKz9i67Zt2y1/J7f4uwg/0TqbW9MteCEXqsqMBXs2zKJNnhMCoC1KwBjNwBI4/czjiMs1hGp4v5kr5WwQIEkdS5f2cNzxS7wVpzW2scPkFaz5TKPtxWeV9+g8PuYY26ZRzN19R6lU4vQzjkUEtFbozGdW0H/mFeWVrL32jegfiPyyefE2m84F9p2lwsQlqlXFvXdt9v50pF152zX45yf1Mo16hbTRio9AU7jmB5kJli0tBpGEAw+uMDCk0MYSxaFPOp9WIgVKGSq6QqkMx51wyJT3zqq/yw2jZ5bitBGzZfMYSiWIqgUZtAfTWfjnWTbOB/lSIkBKxF9u2MYtt49RSxWJMqC7tw0KqMQleqMST3zMUbzubx9Hjxkj1rHfFHxOL5wv39OU3aDkKeUwojnphCofft8rWDHUg8nt29kcSQllD4TEWlIlpLrMPastv//DeuqNYG2dbTQXED/M6UhUwkRa5stfuZo7V43hnA7KaSvkk0eUn12aqhQnNfZabnjnO57O8HJNpeInyPrsyPZ+9SrWlGk8AxqoaGGvZfC3L3s8w+USyuW3QmrJrmy0TFBoHeOcwxjNvXdt5cbrV5E0Mmv93Fg0Stbs6FS6doxWoW4xTZWZFd70CwpNT6/i1NOOZmg4plRiXuOulKKvb4AjjjiQZUv9cAl6FivQZvg5oy1tWiMtuxjvf2l0LGFi3PqhWciqCbRVk7Yq1NSoo9hwzHF7E8dZbcZXg2DRaiGzyKtZnKNSKmXQUvNDWl1M014Y50NmEI/ZtrWOs2BUh7PIXYWyVGswOpJ63zKtH6AjJ8jlBAJiEpbtVcJEDpVfkj5N8s0Z5TekTfHz9Y4+JlOypg5ZaueDsxGjIxMo7TCk6Jnq1e5OR5FsJsUixNc6oVqHD370SzhdwZoEK+VQ7jov8MeqExMceOA+/Ps/vYC9BmIGSoOonKRYGBQiJTSWwYrh1JMNj3zUkZjIBEWFMEeo/RqlY4QSVoSJRp3vX3IL1fGadznhZ+TOajeRhURUnVQljIzBVb/eRCqxV1amnAbhlUo/V9YR6ZQXPf8sjj5iCWU1gJJSWI0fVoqTyfcdQVCSEgGPf/x+HLDXMozG7y2swsIZ7Ydn/cIdf9xZRWQMjlGiqIcb/ryeWn37jBidubvIyVYA7KCyEgSIyzVsKhSAHZUtiihcr9G6xAErenjGox5Kb2MIk+yP1mnQ8IPH28wevR30G+GU4/ciNmlovHW7BSq8S3MAZw7vF+wPOdVl1xSVVqOtgQqiLeVyPyv3Nxx1aMzKg/tZeXA/hx3kw+G5cNhB/Rx2cD+HHtzPoYf0c9jhwxxzzF5U4jCHQpRflRampLdPLs+nXJ5uFo1M2ZrcCYh0xNJhWLpkC6KE8dinaeddM5pXK42QsG3rFtLEi9Vdj5+EmiQwzgSUvRdlhZ93kfcw7pTCKoXDu63WZjMOS1xOQ3FXftWeCquF5oNcomoMPSTsPVTJbSwfhiadaQs4v2eqiMGFIKrOFruJmtFoO+Qbvu2sp4sdyYRCPiw2sooRFIxEUq7+7VZuu7VMLU1BHMZXnLaaIuDn0jjD0qEhXv3mA1m6V4rWDo13OKraOliZxJmqjekmEzrJ7uHxBt4I6GNwUPGv/3g6++9VR0hRNLwD39D+aL/LC5EojARnpLrMr379ANfeNcpovdHcY85Pn9g1KKVQVEjo5XNf/TOrttRJsbNaHBLpKjotsc/yCi992UPpLQk6diiNXx6kmg7gPe3JOScEP81CYVm2bCPPftZxVPpqUGogcR1rqlhdx+oEq1OsdjgtuKiKqDrGlUiV8Ou/NKjW9nhLluTm3+zg/KMgSLKq1JZ/O5Ch7fLJf+vpEf7qrx6PoooyIy3z0DwIs97elEec9hCv1int/Ri1lc4W2dHuv05GMiV0x5JkXsg/W6ixYr8JPvmF5/P1C8/hu988h+998xy+94328P1vnMPFF57DJRecw2VfP4fLzn8RP7jgRZz/mafRH6UYccE3W2aOzp7UlObTMJUAzuPv5UQwBg48YN9wdPI1WfqGNiIUEQUi1Go1khTsDE5tdw6td5bMVpUb7ugc9pDwGdFoF+MU4HqDHzIVeuWzSe85EKxrSjRGGSq9sHw/PaewYl/NiuWDDPXtTUQfqFJ+0tkeSb6eL8rXVD5iAtRtQr0Wc/FFv6RWtxit/Y/KhaX87Qh+ddoRR/fx6EceQ1QOvs8W9E2bNbr5DXyLu/+BhpNPOxqjK8EFQfsKPJVTuJT4TnitbvnxT64FXQrztBcy7rNDUFTH4cc/uppaYwJnvZI1I1YomRpnPfUUlizRlEsGwrCqz7t8Dnbm5lxR3iLoDFGpn6edfQynPaTMo4/o41FH9vHIo/p4RC488sg+HnFkHw8/qp9TjhrilKOHOPnwXkqNByiV/Zy5ubJL9i60gLFQrVf50tdv5N/f/1PSmebXSGbWTxFbYfm+lgvPezknHzfYeebMhDe+7S544tnnsnFriR7rn+/H52n6TMoSxw95CIYqZmA5//J3J/K6V59CVJqsL0lLJpAC99xtecFff4j7txic9GI7esSChAZnOpTfOV688mO05tQTSvy/r76ankqDsor9YmIVNtxWvpKvX2N4wlnnct/YMDpp4KQXTDJjwzY00MuH//MJnP2MlQhgccS7QCefLArTIEYNOND4Xp2aYqpCptP6PPG9XL8PlS9PSuHnTTSflOVNZ65mTJ9unVgLjUTzb+deylcuuJsJhNj6qt+Mb4c10ysmBq2Fow9OuPxbb2VgSbKLhwwFoY4SwzV/iHneKz/GaLWES7P08A2dU1nJ9++j0ESiiVzCqLGc8+yT+Mj7Hs9gf5jYGoZG9bwMDbQXFnGWraOOu1aP5jTAmRcQiAJj+lm2HPbby68n0077xUuLnO3Zu/DbX5q8d2G3kr9L8eIMDFRpcMctcM6L/4u1W3q99SScFlTs5mUO76Klv9zHP/zzSbzg+cezpCf2a4lbRQLdrH/Z3+yOU8mB6chkCa3h8txd6rbKd7+/ibf9/XewroRWE6R49zuKZi+rDaVSTn1kP5/52PM4eO8eCMOMWb3ZFVZWC1x7Q4Pnv/TTbB0VdL3Xbz+jXIhPiFNn1LRjqHcb3/jq2znhGEM5jhEd3iM7OZ/8Gd2OzYCEplUcoFOcM9Sryn9XbaJqMmHkxyhfsKJyQlTyax3nwtzOnjeyAugLYZtknBNTpc70ZE+T3JdJd5oiSqJAOe+eMH9KVq3yx0LesHwvw9OedQJKTFC7sl9zBXE2iI+A4IjjiEc++miiWIhM1Jz30nm/fIdnDk/KF/c5XrdQZC+SeWMPK3a1Aq0R7W3M2aYAzaBo/uaUP1/IrURrm5c5/2+qtcI5WLZ0WUhR37vtLF6ZRajZwQ6C1jmvxJhF1ML70QpHmubL8lT4NxV6UEpxz91bqDd87ZEs7TsTY55QWrNk2HDy8YOcfFwIxw9yygzh5OOHeegxFfbbKzgwxHvdXqBoLgry0mi63NylKD/xO7EJl158FeMjUZcVd52x91asKIZHnL435UriO1TbLRcDeYE/qWBMOtBEAbGJOPFhBzA0nJX/3HZgU1wqornlxvWsXbfFW7WDFX5X4ixc9L2fMj6qsC5pzUfJXmOKwqRVxEknH8rBK8tEJb9yN3R5O0/dYfxcZYuKUrTWREbR1+/oH0wZGLQMDFkGhlxHEPqGoHcYeoeFylCDykBKXOpcrjQ7dlBy53v1sxvOyFYOEObGiYT2Lmyrk/nXCSe2XxsGKURZlKoEPzkZdto4tOqDH8lOlSUFlEsxVrzlwBm05IIzmHAsP2HZOO0dGebSeyohpRHKlTovfemTGapYnPhC5UXDFDWqG7lxGSd1+uOUR5x8FL0lhfej3Twxd9EO0qoz83nXOdGepv79W3Hx6egUOB2CUrhOZ6ThmsylYLZBaHua5d+0MxfztH7vlLOdAVo9zKGhUvDl5CbfWrI3CeUilP1svzVfy+ZQVhaYJUtgaHAAJVHThxuEDkjzrOyoQ5TF4dBac+vt9/OX69dQSzXWae+1Oec6IdtnLF9D8uk5J9qsE74EgHf22D1oRGmUBNelApq4afEs2IUo8b6jECYmevj9tauo2+B8ORQO1fxf67DSmoZLOfb4Q9ln2SA9CpSkvlyG4HD4Eur/tbrQOTOItM4XJGxonF3tcOJwYn3IzpNgZQ8r51sSQ7NkOSxfEYdJ4j1t9b/1HB+yFxzZGvHLX95AI8mObVet2E46nyeMjsOPf35rcx5yM64ddaWzHvdEwsNPXUn/IGhV6j7BfToRPEeUEpRKcgXF+0trWRnbgz8rKwHZW+XPmRs7qGT5whp22PQv0ZmiHcFnhsfomDi2LFsiDPdOhFBjuM+Hod4ag701BnvrDPbVQmgw2D/OYMX4lQoQksL6+EhHHDpw4peeNqhTTycwtkEPJtulMviL9UpVPvjG2WHEEKfR7JJOwNHAmJQVS2KOPbxEbHSzIe0SvRnwza2OhT69lcMOWOFjkelfev7MxrlOSZs1bFfgi3argKtmc5k5BWwpxtm/LB+b+YlpLjaAkE5tSdV82xkqUq6yNScfdQ/egSVEBnr7VJgvkm2skSfXsWh2NGh+9x87r9l1lHur9JS9A11p2uZA8Bvhtiqfn9julEVMDSWKiZrj3997ITffNkridLg2r2Rlwq1diZ01uexpHWj/0S9z6BaCGq4ywdwqZ22+jeYcqcVNy3q6yN9N+SZvywbNHasnqOfyVuE7VEKrCqK8vC+V4GEn7UOv6qXsIiIxaNHNkLkIyMqcl7K5gpQrTy6nnglB0WoGP5WjqVIFtwCdxgNBMBEccfgBREbTSJLcb+0KlpBzhiuD/P63q6jVTJt83rn4FKo3atx6l2XNxl7qSc2P0ig7aVFaZ7ESoMw4Jxy3PwYFrqd9H+Lpyl5bvZ4tKiiBPShlwpC/+OPN9iAvNLxFTfslB6H1KIVh2Tk/HGanKUxHGJNxxg/jdM6UniI0glpWjiOe8bST+NYF7+QH3/7HjvDuXHhXK3zrXVz2zX/k659/JUccXAFxQb0q+XldEvaJCytBmwaukPcimgRNSoRSJbSLEOsQXcea7sHpOqLGMWoLxkWUbIJxfqn4zGg0EeUyPO4JD0WkNcTSnmWzKUECyqK14fAj92KvvcM9fJmZ8erdnfaq0PHbNE7p6Ph9uvMWAmNg+bJl4VuwS80QBSE7KbOYznDBTkEBjnJJqFQStE675kg+ptlnp4DQbKy6v8I/vfub3HTLBKkFoRw6bBJ6tTsolibRrcTMnu2/smBeEU1qhZ/85AY2bNjWsqFOlUFhpd5gBAftayjFY0DZbz9m/apSXBixsP6vCt+z+QbSXIPo//oV1/nQ4cQyH4KDSxeC7wX79gCgFGsUVaJy0mwep0OUY/XqbWzZUgudgp1JVocsoHFJhRuuuwtrfXsEBlF2xrdQCso9dQ4//ID2zub0l80rM7UV882cJ75nOmrDpfzuV+v41S/uZLxWRUUNoNdri9O9gAh9eyW8/mVPZKhHkWARZSm5MFlOmVBpWu4D8rXIa/V+eNEoR8NWuen2jZz/nWtQJUM8tqJ1bl7nC7dwCKmMYTSYdBlJTdMz0INTGjeNbFeqjjEbMOlyymI544z9ePij9kdFU4jwLFWV7424xHDv/SM89a++yOZtEU58j9mfmvXfswu7ZIkCpSxaHGnJ8On3P4Oz/+oITNRhddLhciWAY8P9hjOf5ie+m6SBneXE9yX9fXz0A2fy3GetBAdOu9ymMAVzwaXwi6vW8tJXfJNxN4GTCmS2uY6szvLRUiJSjiMOanDZ997GsmELYXLorkEQLIoG1XrM69/+fS7+4SrqtowK1h6UX7HZWgWMf89Q0L2i5VWoio457OAhXvD8E3nhOcfS0zNBOY6CspWi6G9TtrrWsVmT1amZym+rTmRDCV2f23q1Rcf2THz/znkvoxKFqRfd3qnbsZ2NCEjKaNXyqtdews9/fS9Ox7hsQVSXOKrM4acb573/8Tye8ux+euN+InGYME2FUC7zWapDcVHOO47J19GsDGdnT9XUqdx12V+jfGkaM5otE738x/su59JL7iQVjdKNdotOk/ZtafYZTvjal17GySctw1mLMjnL/EIjhO1+ItKa5rVv+x6X/PIBGo1Gs1K02qLWqEG40B/Vmqc/cZiPfezF9PcllK1CSeSrZvYKzVdZxBVtDsxZycqoN+p8+qur+PB//xRrExwTKOlHKZNbqdEN4YgDBrj4my9huOLAeOVGa9s01GWirlvEMnEpQOwa1GtVfnzFA7zxnRfT0DFKt+ZpCT5/svPxL0xsJ5CGZv+9Bvn8F17FyiMm0BIFj9WTkSD2HYJ2mpKq0xtFCDHK+PHo7ldmzxVwjtF6nbf+/V+49PI/Yp1rCv1geG67YhJhCAOVsHzfA7n0gnM4bH+/8kHIyrH4yhY+E1YXnnnWudw3PoxpNLD0ovTMStZwfx8f+c8zed4zV4JAqhzRjI3U4mJRVFERnFVBybrQK1l4E3leEGdkcc0rWZd+/20sH9rVShYIDkUDR8T3frCFN779q4xWy83hNCFFQ8fy+MlKFoBRNSqyhIoyHHao48UvfASPePhR7H9wiVJPDWPKbbbe6epYnnxyts7P4jNT+Z2lkrWI2W4lq2N14eJ7aZ+H967ZxF8/71vcc38d0YIjmlbBEFE4FzG8ZBk2WkOseohdSoTfV09oDX60WZddfs5m6/dOC/RUz1b4KKtcHdcojECSjqCjmM1bhYYrI6YKtjT55tCWEQ7FcAk+9V9P5RlPOQLEoSL/+1TxmFcciE6wSUR1m+JJz/oYt28okdqW/6jJSlYYTgpbBsWx4Z///hRe9bcPpxRZInEoidvLW/PzopDgO8xMUmdqBFx9HF2fgHFBjVeQaoKrpiQTCUl1ijCR0kgSdOTbDK1SlPLL8rONdQjJqkPwxkgfIiAWKAleBAqkacRErUKjapioOiaqjuqEozbhqI07GuOOZMyRjjmSMaE21ks9KWNVTKlsGCyV6I8jBmJDfwh9pVboL0UMlyKWlUosqcT0l/vQpoIJClZIjhCyOTft4l5h6C1VeNqZgjHr0ToJLq0ErTOR3l0kkt1bgSjNQ/eFfZaBCd5TlQpz+Tpb7Nx1LjRyeRvDdHRaAWdzzWLFBfuIpU7CJlImSKjSEP/X0sBKQj1JqNWFiQnYvBlWr65x33117r13grvvrnP7zXDbTXDLDXDz9T7cdF0I18ON18MNN8B1N8Kfb4Q/3QR/uMHxp9vhztUxjTQBN+wdYM5Iqwb4btBiEDQKwnzF0x4ac/iBMWnVYdBEGlTw7qvCHMGWwA1X546LK1FzNUbTCW64c5z3ffRKzn7Fp3ndP1zA9362kXtWO9avrzE6VqWebEMxBoBzWUn0HRM/ZdnXuVYd7CRLywzJms3c56k7Hfn7Tr737k8YlW4Piw6FtcLdD6SMpbXQ7Z05qgovJ7dtXc/oJmHzhiobNjVYu9nywOaUtZtT1m1KWb8pZcPGVli/yR/Pwobw+8ZZhg0bUzZuSlm/MWXdxoR1GxMe2Njg/k0NNmwrsW4jNNLYz6m3cVPBym8h4+W9L6eZHE5SxTXX3EO9CjpYsXYaCgSNEsUtq2DtqPjVz5lT52x+W3Dw7Ou/Q6m0Wc0UlsNW7k0pAo1faNI2bEi+os0mhxc/22/Jqtf59Bdu5OP//VPSehkRjdNJc9PYKW8qwuEr+/nxpa9gqJx6UakFgg+grPc6bfI25aGlVpvg8p+u51Vv+QEKRxLKnfYlonWfTOCLYGyM1Y6DDuzh/335FRx7eOKXpuVo61SoVry6xaxd+KbZ09vP9Zt4s2bNep7y/M/wwH1lhD7Q+En7wVQ8VbqJArTDxPD65x7Hu/71SfT1uFxcslgEvwXBkrVujeHxT28fLjQq6fBuPpnBgT4+/J9ncvYzVgKQ4ijtgE6+y8i1joJDVI3UOmxqSBoRaRIxVlXccNsot912N3ff/QCr71vHxIRmfCLBiTBenSC1lvHRsEJFhZ6j0CYgMmU2KzsKwbmUcqlCvQaNmqGeeP9XmbWnUy/OctNSJkI44uAGl3zvLawYlikdze4sHKBxWOuYmKjy4f++ms9/6WbqdQ3xGNotQVQtvEN4MeV7ASpXtrP66D8rkAjlYpROieIG5Ypl2bBhxfIeVq7cmxNPOpxHnHYsBw74vSDjEpTKQGTDNhxZJisgWASmTarpyz4+Z8JUhcnkj3Y/Y9exPZas3cFPliCkacLXL7uXf/3X7zMxokEZnJqhwyJkNbG5eEThy+V07yjghyinY4Z7MON9OkuSd8uQyQbfjvpSKIAoQ0WEMx89zGc/9VIGB0GF3RJ2hiXLxyrFNSK+94MNvPldX6Na951A3dZg+nrt42T9djWuQmJThpdEXPiVF/GIk5Y15UTblfn7LPwr7RR2qNXM0qNzufx0+Guy87INJGe+rg2VhWDCIZur5RutttGKLmSLNwU/v6vt5HBvlXtGu4I1E1nRaY9CaJNZvmIZT3rqw6iUTNgPLZvwOx3Kx9elRLHmCU9+GHEpuyr/pM6n5vKo66/TM5u3XdRIWH0tkLqURkMztrmXP97iOP/i2/mPj/+Ml7/+izz12efy6jd+kg/+1w/4+jev58pfb+CP141wy+01br+zwarVivvXGLaMlNk6WmHbaC/bRnvZOtrT/Oy/tz6PjPgwPjbI+vWa0dGINBG0mWsuLB58edAYE9Hb18szn/Fw9lmRUi4LViWkJvgkC0qoau7B2P7Gkt1M4bfVURbRDT9fslFibKTC3fcqrvnLON+66Db+7dxLePpz3s/ZL/0M//jeS/nUeb/lyj+OsGmroT6hcKlGxCL4vcVa1q6pyGI1Nc3q3+XM3TcHd18UkLqEu+/cSqNuUBIFtzozoGgbk8/nZV4mdgZ/cjZMMEXocl1nmJ7OszpLmo+7NE3C3m5776qNJA7sjOV8fvH1K8UJrFq1ljTJOiJd4t1Eo6SEkgiRlEoFBvp6mle0S4iO98m+dvlpd2IWpXR6mornNGntzZ75dAonem+SU184Fdmzsg6rtJSsbCnslLmSCX5FqwVGulQgr1y1CsFsmXxu84hAFBme8KQTMErQCkSsP0NyLzUp+LvEUcReey/l+BOXobzD88mm1nmiMwaT32oRIUCXPbOcOKoWttThrgcivv7tm3nNmy7geS/5BO9678V89qvXcOUftrFmU4VqTWjUQWyEooJYAbE4m6CcRTm/T58WhQ67OmmnmkFlf6X1V4lCrFCKSzjn/ByFKXu1i598GdDacORhQ/ztS5+A1hMo00MSJf6czvrT5ZWzGioIolwIfhGMKIfSCnGGNI2pVUuMjvZw7a11vn7RjXzwkz/lxa/8FGe/9JP8zxd+y89+s4l7HuhlW72fRuKX7U+bzG11bWpmd9aewe5Q19O0zi23rMcmmUPi2cd00nvl24pugVxHoVvI7tZ5XbcwC5rGgSkRwPveGhsTqjVQZpbb2MwLAtQBizhYu2Yj1mbtVovJZUiBxCgxaA3lCvT2zrC7S56d9XoLyA4pWS1BGdJCgtCUNhHaVICyeRPNhGsmYFuWTK4QGR0P9GV4ulzwJ+XjM/nh+WN0xCAfi64xgi5X+NCyf+VPVMCxRy9n6XAFrRQi3rGpH+bLlM7JQURjU8cpJxxKqdyK86RnzCcdc2oWLcpvEKvCpDMRSBxsq2quu63OG/7+25zxlP/kHf/8Q3746/vZuqGX6rZelB1ArHc9okT5VUWhPAjeQtvc6FirpsNTqyUENyk41R6ILPVkjLik/PcZZ9/laEv8WV2x01DAUD+85EUncvTRg2hKzfe3KO9kdDaWBjSI8V6vxfjhEZWiVIIiRYt4xRY/zwNXIq33MTHSx3XXOz76P9fyytefz5lPfT+vf+ul/Ohnd7N5VDPW8J77mlNys7zNOlZOfKOVW3IyHZ3SoGDnIvi5t3fcvhYbfE7PvUbMLReb7Va3sB1Pn5Z845GnaaFQQWpoklQxNmGbW7/tLLwvSkVqYe26rWG5i2/rJkc+a6D9GSiHMpqBHkVvxdfzVtsrret37ivtFGYjBadEEO/cM1OkWu7T/PGcdSn7zZ8b9g7KClBI2SBGp07nfF52nJT1MJqIoFyrNyEhPk6kY6PpUGGa9+x8SGcI1zUnzobnt4Uw8S+vaGVGMg379qT8zYvOJDYGRY+f8D+5U9COKMo64oyTBilFVQirYyYlxDyRn7TcOYF50ZA1nPgd7UmBBoxsc1z8k7t43iv+j+e88DP86EdrGa/3kWqFMxGmHBOVTHMXVqcNCX00dJmGUSTGkuoES4KoFGVSlPYBlfjPKvHKnUr8MZWgSTC54PdV1BjTh8s2SMZbbWZuInLlTGXfFxdKWYYHhc98+vUctK+hJ61QNwkWg3LlsI0USNNi3S2EOhXSUZTDobDKYLXBaoUzXrFNtcNqwRmF1X7qc7WhGRkrsXVkKT/60Wpe8caf8Min/A/Pevln+ciXf8lt6xpMTEAjxT9LnPepETanbkms6cnX7z2N5vSKfFhsiGLLxjLjE967urd8ziGiKuz9qphdLnZao6YK80K+dWzJBpVZzptBAwaLY6JaRcS3ezsLhwYiXArj4357LR0s9lmaCvgpALiwCMyFeVkp2pQYkDo9GlRzc3hf4DIHrtBR2Xbe6y0YO6Rk7Wwm9SYWqlexwJRLQzzl6ccyvFR7ua8SUPVpgzYJff1lTj7tKLSO94Syt4OEyiuClR4aSjGmxvjTbXfyr/92Ef/y95dy6/UTjI3WsVPuDr9zUrH1lJ3zvJ2C4MWurnPwAZZPf+YFnPiQQQaVEEsDRRWlxv0QQ9bZmpJWy97sd3Wgsk5u7reWTuBwkpLaOkk6wZaNCTf90fKZD93Ec5/6BV7zmgv43P/9lutu38Dm+hgNU8Npg5MYRwyku5kEeXAyMTGBFUFpQkflwYUSEBSJTRkdn0DNNOl/nhHCTicKRsfGmvuqTo94T/BBkVq2fHnLE82DpNLNm5IlYWVVJxK2GmhPz7wUnbu6KpPsSC2aCnC3HwPZ+LcOf+f2dJregLtfOYskFVi6osZDjh1CZBQrCUrZsC1B96C15eBDlrL/ASWM9j5H/L1yOyE3rQM7jiI45SPUk47fdzlhpwErMOHG2TCuufCiNbzqtb/gG9+/iU0jE1SrNYyKmyWlbTJ2WKw3aa5F6Bn6z9kUvc7f55Yakiv/LigLnWV07nfdtfiuTR0llshpHnb0Uj72gSfx2BMOZFlsKLsyab0PZ3vb+ujdQm5EJIdPEW8V9g5clFLoGYJSChEhSWo06qNs2FrjJ1dt4sOf+i3PfcEX+cd//hOX/WwL922tUlVVhAkcdB/AbWlxbWQ1bXfKrxnpfM/O74uAbdu24TJZtMjittMQ7zXeOrcLksDXRudgZGSEKMptbN2B31LILzHzFiof24GBvmBNbLdoz1/LtfhYkPeSIDgnK1cZnUc7v08mL++69XQzZp5A2LpHvvGbEzvQImbWt4E+eMpTH0pvX+r3gGsqS92CwRjNE5/4KCo9NQjNTzud3+eXhb37bJBgucoyUMDVqFYn2Lyll4987Ef80/u+yeqNCaoyTKpKOOKw/URn7B1KZf5dTHNYa3Zk98r/nSrkzwvfQvw7y2hb+W7/adEith9xPRi1jR6dcOJRw3z6ky/gJS86hqXD4/T1hEnoMCkdOulMD0/7Sq7u5C5UEtQ2hbgS1vWTSj9VF7GlCmu2NPjO5b/iXe+6hJe+5HyuvkqzYWMv1XHb6iDmMyCfjTm6HNr96XzXzu+LgJHRERpJ6rdxmaSUT4GoLnMDfZncnQKAEvGr+bT2w6WEyWk7BRXUBQENIxNjWD/nJxfDzvOzv6ELo6WrL8cZWYRlcS50lr65kZ+3E0I+CbNjbUHyF4ZlWuGqzoLVmR2Te7vbR7ZxZ34Tz9yv08SATq8ec8bP0xJKJuIxjzmOA/capF/KYbPS1kbH2pXREqFdGeN66YlSHndasMA0Z311MB+Js6jJ8sOvcBlxmtWbenj7m7/Bt8+/FTteIkkTakmKbe4r5rdnIl8Gld/MUgPGaYxoIidEjhAU2mqUjZE0wtoYJyVS5yfBW6OxRuMig4uiKYNEBhNXEKcwpo52JbQeR00jHFvKv2r19IRFJ2Valj6N1hU0hsgolq+wvOVtZ/A/n34Jjz6lzP7LFFFUwUSKyAhKUoxYjLiQ1hrtvF3Id5CyXOpGSI+sJ5VprBD+5q8Nw8lhvpwBSroPl/SwbkODW24b5zWv/TRvf+clXH9zL1vGU+qpJk01OBCXG+KcKjp7CM0kz4dFSL2qgzvqrNmaLJ/zSJD1SdJgaKjO0qGtLB9MWDYASwdgyQAMD8DQAAwOtsLQoD82nAtDA/542zmDrd+b99rBc5rPy50zOAgDg7C8p8Le/aOsGNxMT7mGdXaGFJhfTEj3RuKdovpn5+tfRqf1wlukQUJ7P3t25vstFDvkjPRTX7iBj3/ipySNim/MdILDTGrr88qrAEceOsCPL30FwxULSrAaPxzQpXbnj7QZSAUQR6M6weU/Xdd0Rjqblym5mNQ49t+/zNe+9CqOO6KBoWNbiYxMAoWI+H6yZ3JsZ4H4m1gtrNuW8ImP/orzv3wDDRV1vL8BVUfEYHSFFcu38Isf/g3LVqwIJuPc/fKorL3xD1r7gOGMp53L6pwz0mgWzkgX396F4htNFyG6Si0x3Hmf5tz3f5crfnw3TvoRXcUGZ7iedkGgJcu0BNAoF/l9s5QDVWvOOVBEaK0xRtCRYEoObVJiqYGC1HklKRt6nA6REgpFrT5KY3x/nNkK4j08Z/HM15fm1jNSJpbF5Yy0k0x0ZEOs/lgDpyw2KbNtk+bnv7mT8y/9HTf85X4a4z04iUka7fMoBReGxcWnS66+Nc/JTzRuKldTpUc+17ucFfp4RmkiYxge0rzhdY/lOc99CEsHhLIW0H5YZjZ5vJjYY52RClz03Vt4y7m/YXxbHZFG2AZtapkk2UCUm+B1b3gUpz5sKWVTIbKx97EVXjJbdORUa4qEyhsQpHWOqFZ71nlOdo/Ocpe/52zPkVy8whnEaQmtRzClcY45dn+WLulDY7y7mGmtvfODkKIQ1q6PefRTPs6GrTEmrDicoqABCqW0306onHLOmXvzkQ+9kP4Bf8GkeOeqdectF/4NF4YdUrI+/fkb+K9P/JRGo4JVLSUrv0eZCgUnzxGH9vPjS1/JcCX1hW4aJStP2wR3YTuVLEUsMam2HLB/ha9++ZUcd3iDaA5KVkb4ZW6IDw7Y0tjKr67cxNve9B3Gk4GOewnoKs4pYtPL407fnwv+72lYZzGm5B/c7WWz43ukkiUImgkZ5b4HEt7zrz/m5z9dRU//IKMjEygNFt0cXsqEVUZzjlnAG0MUyggqqqKNRmgwPNzPAcsHOeSg5Rxw0BIOO2IfVh66N3sPDTI01ErjTvkwCQXWQcPBtX/Zwhteez51a/28oVxpzitZkn13FeKmx/c37yZKlvhVnvgedpo6Uh0zbmP+cNVavnPh7/jTn1axYXOJRqqwzuLEr0AKOo1feexv1YY/FspsmwWrG/l06nKeAChsainFMS5NGSjDc5/3EN7xzkezYrhCZAyoKSzGi5g9X8m6momRGs7VZ1SyALCOgUqN//70Czn99GF6S4qyxCiiSRVYFuF7t/CdD0UaYqm92UH850nKyoLgRwC8kvUxNmwrYSTEJy/EoC0llVIYUUHJ2ouPfOhF9A9kv7Wu6KSzls+Q04uWHYq3FjAOv6N5549dUJl8DHQXAHNjuu0vpiQUzExNmvP1O4RPLKWhpxJz8qn785ATNYNL1zOQD0s2MrBkE0NLt7LPATWe8ayTcc5isqUZ85F4uxPi51ZYYKRe4Ytf+S1X/vJeUtvHyMgEyjjQdtpl3V4oG5TEKIkoxTHliibSY6wYrPHE0w/ga196Mz+47K1cdOEr+MLHnsX73vloXvac/XjsScIRhzZYsbTBXksT9l6WsNeyhL2WpVOHpSn7LG2wfHCCwd6NmGgcpf1WUvlYemXPh2wxRlY2W8JrZ5bR2dFaCJAhwTOV32VUR5aSnmBJPM6TTh/mEx99Jj+97E18/jOv5dST+lg2OEKP3spwLzTq9bAPmjSVoObwYLOuzoacyWEqlJfuJoq8SxdtGK8bvvH1O/mPc7/Lhq0K21SvOkV990O7O6G0NcOiZY5pbxTY6gQVFVN2ZSLXi3MVnItwzuCcwUre8YoPqfjjtuOcRrdzpHVOKrM/JztvqnPy93Ciw+azCqwJ+k7mPmFn4euVf+JMmdBZovJjQLO4fA9i+5UsBaUYeiqOnh5Lb4+lr+Lo7XH0VdpD65ilr8dSKWWbI7du1pklM1b48KNSijiC3kpKb8XSW3EdwbaHnpTenga9PQ0MNarVlJEJGB2VrmFkzIdtY8K2URgdg9FxSMM2dnNGSXNJq6QlBvoVn/30y/nND9/K1T94Wy68i6sv+2euuuzd/PCCN/DXT97Xq/3Tqf57NH64KEnhV1feywVfv4FGEqN0DW2892cRi1Ku2cb67AnzAII/tEwpVwqMTjju2P35+EfeyA+/90987hMv4oxTBjhkqaO3N0WVHaI10IfIEFBCwoR6i1/+L0RTBkeEkhJx6oU7rhfneqYr1TDjr4sZjSJBEaFshHEDGOklkjKxqzDQoxkchDMeN8G3vv5CfnLpO/jGl1/PG151Oo9+zLEMD2oqZUekGyiSZj0Bb3GcRGZlbra8wdN+Jh6av7fO6NwGRMJtxAipGeGKH49x7rnfZWw0VFXosF8X7FKCFXjWiKIU9aGtQYv2bnJFMhd53lCQMxaYbF5m7lh2Tv54lDvfuHaDw2zPicI9pzone7bvYghKwrilVWCDE9+dKi18PfA72WVzFqd6fj6P/AiTN3xPdf6ey3YNFwpgneXu+0a56951lIwhckG4hXTvHJppIkLP4ARHHrmSnnJPkIZe8M3N5Ok3/HRW2LRhjFtvXQ1SDgUvQ3zB6HD+qEyVVDs2bt2H//3Uj6g3NGmceA/fmVDO9kTMGmVRRM4QpQPES3t4xbP35UXnHE80xSjjbJlr8k+fRs7rzZJlxJ4zXCjiJ1zece8mXvXKi7hrzTpSG6FsOczpCRuqKsEFT8QA6BqiEpztpZHUiSsD9FSEow4Z4lUvehTPetIhDA8mEHkP/E0UHYKixfR50E6KpZFqrrl6C89/xbdIG3XsNNa2DEuPHy48qMEl338TK4Z3qEu0S+gs253p5pzDOUeaJiRpwshm4Ya/bOS2m9bzl9Uj3HzPJh64f4yk1kOt6hC2+IKgggYtkVfE9ES4YQ+uS964oHhDqyfemQWh1hOjGBgY4D3vfBwvOPsw4l7CMEkXC/LkR+1ytme48DvnTR4uXGzvJgLf+eYNvP0Dv2d8pAbMPCcL/P7hvarGeV94Oac/Zgmm4qd++Nfr9pL+fmqKX2evcE99B49MKV+mwi/VmeLO2a0m/TCfNEA0m7dFPPLJH2fN+oicM6HuhLTWoqhHKS84Y2/+66Mv9sOFU9gMpkqZ6XN68bLdSpYg1BxYpdBOMKGnqQkGlymGEEVAaeeXcjrxWwPsgJKlRCNOYTufN+lWrddUSpFquOseeM6z3s/YWC8NKQUlq6VckVt9qEiJqGLSPmxF8W9vOo3XveYEojlsw7Tw7JlKln8TYayW8F8fuZYvf+0qJtIqqSuhpMcvbVZ+vpYo1zYnUIhxTmF0lTiq0Vse5syn9fNP734+y/trDMcGTQlcub39nKdXTbGkzvDDS+7gNe/8KVJvkBj/pM6GPs+eoGTNDglzuBKcU9hUkzhNPdWMjKds2ljltlvu49o/3sqNq6qsXzvK5o01alVDI7VYC+IqKBRKNQC/Jyi0lLzZK1muKbeOOiDm8//3Mo46JsZQ8b935tckGbPr2ZOVrO99+ybe+J5fUxv3K0ZFqTAnKR/d9kUTykGvrvKlz76c0x+7FNXj87+lrnQyjSIDi0LJ6ioGsltN98gdpg5oNm2JedJzP8udq1K0bTdgTEJ5tw3aGcap8fRT+znvC69nYDBkVZf4TpUyXd97N2A74+2ToaShrISyEUqRJY5StElQultIUdr6LUoUaOVnugomU806HzID2nu81QoVQRQ7TD5EDhNZTJRioqQZdNRAmTpKeTcA4mKs1VjXwLkGThoICULq/0oSvjcQNYaoMZQZBzM27bSPgvnFiuKBdSnf+u61VFNoiEGU70e5sMdg1pBm2SL4idTKaXpMheWlFbzqpQfxgX/5a/YbLDFghlHS45frtz9u3tBonBO2bt3aXBxHFrcu5ScTzV1+2kNRQISSHoyqEMclesqKwV7L/ssVxx/Zz3Offizvffdz+X+ffDGXfOV1/Oj8t/Lvr384j3xoDyuWK3rKlp6SRmw9bD/ih4Z0mBCswwR2Dc3PSnn50QoqdFAUThSrNyq+9p3f0KASmtWFKiGLhM7Xm6ql20UooL+/H3F+4Ui22XrmdzALnZH2i0zAhg0PW0sZ5ruGzdf9pr6P3v4Gex4IaSsKJ1Apl6aJaTtZvlRKFdav39Dce/LBwnbmmRdIqllkdXOyK0TejJ/9bQa/YawKfot8wyZYm9Co16nX6zTqjVmHJLG0vPonQBUYzxXFTHHrVqEaKPGOLR0GqwwuNHqTq2mObCPb5rt1nrAn4a1D4JNv4dSQbmTzp/xTHdCwcMG3rmLD1q00HIjqwRHlljn7PM4rLv6wRbSltxTxmledxptfcwZ7VXroEb+s2A85lVoXdCsuO4SgBDZv3uy/quaWxTMQSmLT5LIz038XENLcv2VQoLVBRWDKjnKvZXg4YZ99HEceBa96zWl86Yuv5cL/9xo++5ln87yz9+LA/UcYKEWU8AoWTcU7y1Y/+J89yx/LMtt3q73lGkYa8JNf3cGa9eFn0vbyMa9lZBHQ7X26HdtVKBgcHmagv6/N/NFSrrojOFLboNFImlaohXmt+brrdhau7bxszihFFEF/Xx8ubc2ZnBofMUFw1gtrnxeLjazlny5sH9upZPmky6syzWOivOmeYAvsDCi/IzdgbcJ1N23gs1/+M5/+0k186ryb+J8QPp2FL97EZ/LhvBv4zBdv5YvnX8/6TVmDlVfy8mSxzH6Lg4+isu/ZCjic9/6htDfHhjhKmOPjGzkJ27j0+fu5nh1Jut0CP+wWtNidrGRlqpVg/f+do1aFn//8AVKJvfVT8k4J89e2V4eS1vTrhIcfv5S/fckJDA15/0iGvEeEzGeOy4X5wzrFqnUbwPmJ4bOrsGHCfjMus7lmd8Xntx9a8EeaMiRIF39c+T3rjCOqCCuWwIlH9fDUM/fj3H/7K757wbt4y+tP4PhjDZXKGCZOQFvS1IKk/hmStJStEMIT/Delgn8seOCBCX7zq7tQDmyazJPIXXw0OyZZcreSfdEgQNwzSmRsrvvlLVXeCtkdEe9UZHM9xm897l9u6n+zef3OhOpMwLnQ7T754zOzc8qigrBNTqxgr+H+kFozECInehyxhjE7yHh95hjP/u3nCy9r8xt1Tw7bl9aTW6lZkC8SeSUL8OOvYRy2eyBM2BKStMFv/3Av7//oz/n3D/2K93zol63wQR/eOyn8gvd98Dd89NO/4r61fosZr0RVgJ6OhjJLklZMFRrlSt76JuBIsUpa2/+FydP+n09cP/EVP7EeF55nd3Yp2Kl4JTNLBbdTlSxQIe0FQVFPHatWj3LbDdvA9mCIWq4OctbHbjFUqWOfQcWH3vc8lgzgh6iN8/mNv97r/63y6WX2VHecI8G8fs+azaGAzfa+WfnNLC/zq/gtLvL106e/DqvAjHgPet42ZQjqcVPqGKCiFUNlOOKQHt7y5odx3pdfxStf/UiGl/i5h5oYsdoPFRoX/G3lS00o3QqvfimFTeokNcfVv7kOWwel4nyOTFvmdkfyVvzF+k5LVghxlMllz4yNsQIxMfdtSXDOj150qlXt/2a4X5PszPwVnd9novMenfeZglwGZR93Tr4Zr2QBK5b0Emmmjyc030V0HUQzkvQxMuGtwlNdOcPbLwDtKddZv3e0XmyXktVJZxGZCa+4+E9Z1Ke8doofsgoxZ8JF2QRXb7HxQnZHE3NPwqdDq8e485UscGhSHNZEnH/hj6joMBRkHTqM8wuhgWgOG7ZTrmge+dgjOfDwkPeu7BXtvHG1a67PT2nI2vPNG2t+Dglxe6HuEucWueHCac/bs2jKkxmsFG2E07RR7L2iwZve/Fg+/79/z0H79dJTKmHcEGKF1IbViV3y1itYvhwppbCJ5s5b15BUwbnJV8xPCSmYLUNDwyi/0iEnmaZHaYWIY8P6DaRiu8qI3Z2dVw61l8oRDAwolKp3ntAdUYiUsbZOvZ4wNjbht57NNnFY+Ii3MfmRKrzbTGH7xPC8KFl0EYzThtAT9RNQfdAh5M/zCw8FpTMLQ/53/30yzZjkQrdTQiw67tktC5pMcXhPRIkKQ79+1t12qrQ7gEZhEBxJHa79/b1Yl/qh5ma5CFYP5Z0OmmYZ0T5oTamnxt++5ik4xvzwm1Pe36V0E9Pz/44KxdgIjI/iGwbp8WmLCys8/TBZRktghrKbtQqLoHXYmba09nrZHrrjLRwOjbgSvRpOOk749P+8kHJlPREKceJdfrgwGbct+5Uvc+H+RhmU9LBpY0K9CiY2e7QAmEFiLgqM1hx8wP4orUPjPIv8EKHeqHP//feR2HTSqtLdimmapp2D8pbhqMHKw1cQxWGe4rSEjqIrE5mU6kSNDes3YNMUcc5bh3fyOyl8/bfWhjl92VEfpvqXn/s3lyjPm5K1XcyYQdMzlxdtI3tu7vndV6nkQ9ZItq4pWEi0zx4Rtm0ZZ9O6Rqs2ZvnWtGb5LlHnBFijNQet7OPQIxUurYBuKWk7WvZmiygY2ZaS1A1C2u7Da7dBoOlBbnEjQXmNBPrQHHvkMO/8t6dRqoyhVBdLYigzwSbaOg6IM4xsTUhTSNPZLlgoWAgUEJuIo488ilh7i8rsNCahp1LmjttuJ7HpTu0o7HEIgMJEwsGH7kVcctN0enKI8rtsaEFEc+ONN5OmDbTJOpo7l9RZNm3axB133MEdd9zB7bffzm233uPDbVOHO++8ky2bRrChk9bZ3kzFrlWyuuCna0mYYhfEYXiXTkWo+0u2NNLZt6ST79P5HGn2fKXVq+68aA9CoVDiFR2vx++8opI9EwAbsXWTZjzpZzwSakZIlPjd8bxJClyWP+35qI3hiANWsDSCOI78XZXfjLhlCc2uyXKz00S8Y7mcppbxVLNmrBenGl13oRdaO2ZkwT83xCEIt12JV7NSUvH7MfrhM784wYf2yaH+fBAsqaviJOySkLvf5Fq3IwTrtxiU0mjl0Ebojy3Pf+JJ7HPAFgxlRKfhuWGVM+K31pHwUs6vBlWAUpaqsqxPvDzYU8nW9rSFRUgUa44+NiaO6nOKpqLEtg0VNm8R0snVb47MRibMpu2Z7Tk5WZRdkrus44ydgME5zXGHrGBwVm2gCjMnY+9ZwMGPfv0A42kvStmwq0M7c8nb7aFRN/zHJ67kmS+/mGe98Ic844U/5WnnXMJZ51zC05/fHp7x/Et45tk+PP9vfsWq1QliCf7ZptJB2tk5+TIHWmVo+qSe+pc5EnSnSQpVF/xvnUcLFgqF3+B1dGSckZHx1sTjvJDpdEKbQyvN0uEhcKEiKwNhqHEyU91lx3DOcM21tzJWrQIgys6yB95ZyBcmfrPDP9u6hDX3C/ffB/euhrvu09yxxnD7GsPta3QI+HA/3Hk/3Hav4c57K9xxl6PuIKWKCwtJ5v+NfJOjfCaD9tbMpUOGgw/aF6MrzWe342VNJnvyMsiKY+2GKlpn/pUKdhUKy+FHDFMuC6CQWTVfCsSQNhw337iZelJYJHcEQeFEGByEw1bug9I+D6aWaKE2ia+bIor77h9h3bpG8F3WXeudvvXfMWwC6zckbNyi2LhJ2LjZsXGLY+Nmy6aOsDGETVss28YS79+TVv9+NsymlC4APsER3ytQorIFh4uGToHbEq+LKJILiM+TUDzCipydjcKX5rUPrPPlJGsIvYzNzJ5toS3PJvk9CwfayL/X/L6jiFCvwk9//AfKZe9bzQ9tzqashzrSrKLzG7e5olAkjYT//q9v8NSzPswTzvo4Z571YZ5w1gd54lkf4olnfdiHp37Uh7M+zhPO+jhPeeZHOetZH+IFL/wQd961mSTN3LwsBNl8PIXSYf6WAqNh7xV7oU3krW0dC11a5I+2av34xHhb2crIl7U9jhnL585GiLSwZFlCuezQOvLOqGdCNGCwacotN44jYdJ8wXagvNzSxndgHv7I44mjiDTnXdTXuXwgWOFatSWpx9x602rExT4nVLsFPM9C1DHnYMO6KthSeJ537ZGXG/mgw7H+QShXwnQTmHWsdpGSlSOYJToT0id4+NetbZwHpPX4yXS030hwaZDZ1mduJXdzcg3hNNaihaKZugLVag1j4qYy3jqjM7SubB5VoGJvwcr/3o7quL7bObOjLSYibNxguePW9WFwyuT2xGydPz0zn7FTEMDFjI9X2DrWz6aJMlvHY8ZGYsZGY8ZGSyFkn/3frWNlNk/0s3mkwsZtm3d6tRHACpSiGLSZRXL6HAwz/VBKMTg0hNJhjmAH3Y4VLAwKy/K9YWiojBJwdsbM9NYuiXBiufGGrUzUqmF4e8fq+YMTG8q7oOMGDz3xMKx1lEolr4wERapNSUFNcuM0MSZc+7ubERuFbnO7ktWZK526wZzI3VSC3Wx8HFbdsxFxpjntgaa/vOxhXkH0M0o0IAwvjSn3KLzntc5YTs0uUrKy18knXzYYlCPThtmRVO5CeGzmLXxWyaVAlO8Fu+Coc0oFrWCHyfJEcFTrdZzkj9LMxMytZEZbJRXvCFQUWNJwjYRin8+8fPXOh7kh+GXJjgaCkKSO2+5cw8bNjqSeNAubND2LZ+XfK7T5ANKxsXl3s/pOQyk0MYoYpxSp0ogYlBiUMyinfAhzKn1waBwiCQ1psGnrmF8xqtpX6iwkgiMFxmoJDu+IdjYVN9sDUwnss7S1r197Hu0hdHuZbsd2OYqhSoknPPphGFf3zoQ7LScZuWrsp9sJN928gY1bnZ/PSTZU1aXdKZgB72/uiMOXstcyQxwbRNd83Z40jia5NPbpraXM1b+9na0jkKZJ2/SNyXJwfkkl5Zo/byZJy4yPh43lyZWj7MnNsuTCnDLF3sMRPZUYVNJ1LtlULBIlKzMrtjIkU34Wovg35/ZkIZebTetV7vyWJStTrlpDDgULQzap2lrLmrVraaQ5k7Q/IWRc3nd3OyKOLVuqTEz4Peu8bpUNU88/gmAFUiZoSJWJmvDTK69jPO0hIkIRNkZvDXw260LeaupPaQml1gvvQgSMVixZUsaYGsYBUsJJGSclH4jwahU45XNQpUI/EfUk4d67NS4pZTP7d5Ki5XAabl21gdSlmLCYYyZECaIVfeUSe/VDrV5ry7FWzu3+5KrSon4xwdDjBnj+WY9msFQHm0wa2skax+ariN/CzTpYu2GcSy67lVodv01SW0tQMDN+8r136xNx2MFw0D4TVKujODWGiLf8tim7WSYQdnRAMCrmznsSfv27e6kntbZ1y53FcF6KYi6LnSguv/I6tk4o+vt6vYU6m2LQZoHLHzNERvPQg5fQVylDthdz/j2nYWFamznTyoT8Ef/NJ7WQX3m1Y2T3lsya1XlCN5RvOJoKlsoav4KFwO92JTglxKUSWvneRNfqN4UzUmsdax/YghWCUiN+/8muZL3ajFAmZ1c6AFCiMAKaMlYl3Ht/lR/95AYQwvw2h5IwUJFtat35Ltm9mnUixKnLqsSdigJjFIesXEGsNUalKOVdYjRDUyls/VUmpWEnMNLP1VfdQ7KTty0TDFu2eD9lVhqhTzo9Xj44lE4YXFLCRBBFnVt27Vk0ZWKb7F1MKIQSaDjgEDhs5V5EufZiavw5giGxNX70o9+zbYsCKYXyOnuLREFWc/zm6trAi1/6TIaHl+BsP6KCL8I2OuQY4FSdat3wre/8iobrJXVTyeT5oFVGGtaxdbPh1ptX4axFKe9Zr33BWzjfW1YgrEiPKxM89vHHU+4toVQmC2ZT/naRkqXDY30Uuzcy7czmnNmTDRYpwW/P0nlCN8K52flNi0PBgtBUpZRhn733Q2PCnKwu2lSObKq4BnCO+1ZvYf3aTJSmYKcq8vkr82HqZ2UEMYICjIPIlqjXejn/G1ezaYt/XntZyRZ7NN8y90v4tujKloCxLN87xhiNUo3g8dmGkPrNr1X4HpZnC4JWGusS/vLHB7j91vXM2eXUDrT6NtU8cH+ddWtTrNRx0j683A0FvocbJeyzXw9x7D2H78k0ZeKkErmY8CMevYNw7LEHEM+qbfZWF0UJpVNW313jluu2Ua/qMC/L7apmcLcls/CIgkc86jCGlyqUlFup2DZkqFrpK94xkzUWayr85urV3HmnkMxkEdqO+t88XRy4mhdPSnHz9ePcv3orkTFIcP3THOJss6wHpVAUyiTsc4Di0CP6MCbJtf2zU9B3g9LV2QDNkCGzYI75NZkdvkHBTPh89kJwyfAQkZ6dwpNHRJgYd1x88e9JJZvKvDA0e0GSgku59w648spbaPjJZHsAgtKWAw5aSqUUo1WKotOSFdKgqSUKSIwTg47HGR0b5/vf/hW1WnAKObfsnB1+1xUAHEK1Yfn8F85nfFQQySY8T48AiGCM5dCV+2Di+ZE7BfOAgqic8rSnPY7+vt5mw9hs+Lplr7Je2dKO+kSZ8778fdZumMAFe3nXa+YTwU/WbNWKPQIN7LUXHP/Q5fT3GSS3m0JWh9qVLR+sAxVrxsbLvP8DF7Bu80aUdQukjAg4v/3P6Pg4n/rUNxgf96sJM+WqGd88KqxCVRHW1nnEY45mcIkLk94ze/ikq7qyMO81I6NZp4QeUfQpAyoGFaG0D9k/g1+1YBR+ex2t0KWUxJYQt50mfEeYmOsLggpjyUqFPVp0x2TqUD5UVpkXqH0oyONTOCpp9jtgOUp5P1Ndyc9l8hINJT7UaorLf3At20YNKaVQPzJRN7tK0knr6swBpxfUTiyQsHWb8L73ncc99zRA97c22w7PVUEhy6btK5lq0HCxkbL/gRXKfVuIVAmxcUc6hrcQ7Ydlxa8echisHcBKxLd/up512yLGGwmCBepAFZrr+QJzzSJpdSzFQSowkVqu/ctmfvrTjQgJSoHMwiKlANswDOsyZ556ArpZZrIh0blGbnHTVh0W8WtlOSfO8dCTS5z88OXUJwQjpjnK4BvHMG8WAW1D8AuVEiv87pqt/PaajYzVylh6pjOM7zih2IgoEiYYa9SoO0vDurZQd/mQUkvrVJMqtaTGeHWi6QB4cVHH6ITXvfrpDMQV4ihutY9Ccw6Hl80ty71REdYlWK353R828Ier+xgb94qQiPVuNrIyOMuyOHXxNaSpAQ2//MV6br9nG400J2uac7A8QphCpBWJCOiIUini8Y8/nJ5K2c/fClu+ofweuDOxi5QsP/8qimNiUyOOttAzVPVhcIKewQl6hybozf4OjdM7NE7P0Di9gxOUBrfRO5SgzFzHHTyKKXOkRbfEy/IlK0QFC4PQXDOojWJouEL/wNSuBzMFpdtAkCNm1f0JV1yxmpqDRBMq2ORz54K/g0WoI1RxqkEqMJ5qfvjje7juhnWgS6RpMkkEZMoVOYVrd0Ch6B+MeerTTwXXQNOtk5PlRssPnl9GX0aUYdM2zd//wwWs31iiYSOEckiNHbc0epnuVxGmClatqfGJT13G6EQZpQXQQdGdAaXo6+1lvxVlHvuoA/BeN8q5E3afPNsTUUrT15/wuMcdyd4renGJgJ5AMD5nsnm2zTm01s+hRQFCtdHHpz7zA+64u0FDsoGh9jo6fzhEJzRUjZtvm+CHP7yfSy7dwMWXbeTiyzZy0eU+fL8tbOKiyzdx0WUb+d6l6/nhTzdz6Y/vodHYyRMaZ0ABJVKOPTLmkafuT4T18qxbUgYFK1O2NBBFGqMH+NB/ns8Nt25itDaGX/tp/XY7c8iGzkc2H60UYmJWrany+c9ezZZRi83imJNWnVgRImPATXDcQ/bmpGMPIS6V26eRzFI7V7Jzlvi0IdRQlGk0GjywusGfb9xAveSXSevs5Zuxak1+8ttJgihHb18vjzxhX5YNZufNHnGgsNx2h+HJz30/m/9/e2ceZ1lRHf5vVd37lt579gGGddgHUBYVUBAEQSMKIkgEokElxohrYgJGgrjFBVyikai/JP6i0RC3uEYxgoiKC4JgiAKCMDDMwsx0z3S/7d6qkz+q7nuv3/QMzYSe6YH68rm8ea/vu+/eWk6dOnXOqdoQqlhjCEynSCkHRlqYgVH++g0H8ZpXPpOkE909Bwj+BVLcvGP1w4aTnv8uVk6MYLIWVvpIVPaoe0aNDvTzwfeczItfuC84cNq1felmHQFUjnj7A6vXG84663088MAgNtRTYRXy7bwTn+JnGh2rlnMJ5VKJhQsm+f+ffQ3771OlogXT9qLq7Si977ekKDnfPFo4amRWsPkAN9+8hje/4QusGxMyrXEhGWIv0vVL07U1qxISHMv3avGNr7yZBSPWZ6zfaQhCiyaan94yyWsu/ifWPqJwRYfdgs52QEp8ygrv/C8MVIWXnncob7zkJEaGoJRmmFBG7Ta2tcuy9SoSEXJymlnK6vVw1Ye+wpe+/Gus80LCtpXa4t62hqJcznjVHx3G29/6XBLjZ7dJqPmpUUVbuZmdxI9/uIE/vOTTbBov+XLvPQEAg6YFqfCMpw/zpX98OZV0GkE2tx4Nh6BROBy5OFauSjjnJVexZm2VplqPtQtwOvNqfdeDewuK+BVtBVr3kZgGp566G+9+99nMH4Kq6S2px+fhBYslZ/2mFpf+5bf58Y9W0cgK+dVRCHt/3f/B+q1npMHJp+3DR95xFiPDA71n7kQEnxwn4TvXPcTbLv0qq9a1yEX5KO5Qhr6/hHwo7Sd1iDhKaYq4CY47rsJfv+189turn2oCJVPy1iJ/gfAavjoN041mDr8F2OYJuOLtX+M7313J+s0TiElBhwlij/rjfc28ZSs1oOw6PvLh13LKKQsZ6cv8xLIrOl2FvIfb4tHPmAUUJQBKpRJ77TvIi87Yl3NOW8a5py3jJeE4+/RlnH3aMs4+fU/OPm0vzj5tL845fR/OOX0fzj19P17wrELB2qJ5PirtAa3bfeRRmeqo7P/9xKZdLDONwHxc8RJRoxkYUBx1zL6dzhqYrgqm9GMUyihqrZzV6xIuv/yLPDLmw3g71qzi5KnXfjSknX8n8bm4pJ+f/uRB3vSWzzOZLyApzyNvORLTCQUucrGokLiv/TldeX6KcwiypR0RPfN7mx0UQoLBsN8BIyzbax5a+xmhEJYIt4KvJ1/WomCikfNv197M+z/4HdZtAEXI/Px/RIBmnvLAg/DmN32Gr37lLvJ8AFHgQu6ureGtYOGNgoHhJqc9/zCg0RUY1en/j7W9zHmmK5rpPtupeKVRoTEqYffFcMlrzwDGcbpEK8kRFdw/CtkevgcalEaUpkVGPU/5znV38773f4N6Bq5nkv14IRiarsyHP3wdP/jhSjaOV5ioaSYnfVLM9lGbetRrKY16Sr0OjabjjBcfS6XabU3dyUhYV3OGRITjnr4bz3n2gZRTX+rCdF2lu78otNLk1mFJuPGnD3L5FV/nnt81EZd2lJ/t7GIaaLYyxiYsH/jot7juv9YxPrE5bAPU8cfyFDfo3yu8KVzqdQ7cd4hjjhplsGL859vRJ7YuGWeVrp9VWTiCkOuyxrVf2//uCRXHdQbK7vFylvDJS31Hn3Hqh10UaQ+Onfc7Fp8t2KDR5Lz6lS/BmBwVkkQW5V8sMrWP0I4kzHwdGWlJyJ3hJz96hEte/w8+CZ4kwYXH+m8WTSl8c1tP7JukgHI4chrNAb547Z285XX/xoOrRxmbrDFZr5OmVXLnlToNKJP5JY2uS0u4YHHf7TQlyj+ZE58rfhu3s8NQeEEz2C+c8+JnUq0QnscrYNPTKU+/ZONwYqg1+vj852/nVa/8JKvWQpZpRIqIr0LSbYVwiuDlpHNgLUxMKL5w7c2cc85V/OTm9TRb/ZikH3FF65ieIjGsoEBZKqniD887gRWHLcUUqRtUq/drTyx6y7uQu3OMoiYVkCh47nMOYNFCi1JV8sRbsYrb9ivx/l13/xKV4XROo9XPF79yB1e+50uMbxJy1BQ7dbds2Rbd53S/WqCewzXX3MR/fOUhJmsai4/I1SIYEYwTkikHJA6M0xinSFzOM45azlMPWxQ2uZ9DFMEuohgaEP7owuOZP9+gdRakb8iXF3LmSXvlwZeYEBzPSbGylF/8cj1vev1nue22cRpZilWKHB/FWPjYdTup974W7ySU+9i44i/+4mNc++W7eGRzg1wnXgkPW2R1o51GhxFB8HVT0U3e/943Mn9UMMFCuj3sJCWLrl6chh26e/p08cGUPxQaWO9ReKV3X2D7aKdp6CrPwjoi+BBUp/QMNyfddWkLpEJA9Z4wmyi6NGwYLGsO3FNz+BEjVKoG0X750ucq81PWKZaIbsSRZw3ENRFR3HJzzksvvIZvfe9eNk5kNF2dVqvW0dlp+MAM6sEpuwm0wuH/rcRi8xatluLXd07w5rd8gcvf8W0e2DSESmogGUYrlAgtZ3BpmWMO25/d9/IZ04vnKoRzd9l23nshNL4pCzrKDq2BaSm62KASXnjqAIccXEVUPyIGlAnbUxQUUYfdorAIFMjInSNXKbf/tsVzTvl7LnvTd/jZz1YxPtmk0fLBWKK2fPRir2+AWgM2jsMtv8x433t/wrkv/gSXv/MXrN6YkKFwqkWWT24hhrsp2rbWBuWgmtY5eJ8qrzjnWAYS65cDjKClazltuorbhZFeWTtdP5qDJAaWLIF3vfdCBvtb9KsSSdh1QASMMn4QD88jU4YJhyRCvVXlC9f+novf8jl+fc96NtsGdZvTwtIko4mQhdSl3lvI4cL/cyBrH46cDAXUW7CpDnethDdd9nX+7uP3sH7MkLkSojSiNdYocqOwxmCNwenUHyrF6hJWa3KnGR5NufwdZ7JoMJ1x8ssdgsIvCSqfoNRozYGHJrzxzacwPDSBchaRNIwhRV7JYrLl68QBzjmcFZxNyfIKDzxU4WUXfY3L3nMjt/xmE2OTRfnm5OK8suZCBHHX3NiFyVbucjZusvzH1+/hgj/6Z667QTE2XsIawVLCknrnkq6i1AL9mSF1oKzBiaWsNvOSMw5k/wMs1aSC0NzuDm+uuOKKK3o/fMIjXnNav0Hz2Wt/SD0rt7XU6Zqx/8wnj1TKokpVTnjGAo5+6jLMznST2QLpPIHy7ycmNJ/53I1salXQziKkaOWDl7dFuVzi1FP24ZADRtvDYzCY7iC8kqUAjUIZaCnhR9etIbMt0Hl7PW3KXYX8LeGf0KUwKxGsc2zcMMlN16/kgXty+ipDDI0aTNmhjAJVQlEOyn/hvB0qWZq4Bow/YrnjF7/nXz77G9531U3cdZdlfHM/mmG0qnWc2hUkSYnh/gYf/dD5/Pz2B1nzkEUk+MQp2vXVHrf9I+FdRByVquOiC4+nv+qCn8POxpvalVKMLlzK97//c2yzD3SGqLw9b5taJ92HrwwdooZxQqsl3PabdVx3/X1878b7uW9Vi7HVjzC2eiObxhy1Sc2G9TkrH9zEffdv5LY7VnLjj/6Hr35zFVd97Fb++XM38aNbHmDluiZ5LrgQhuUHVTVNBFDXB6Gf4Bqk6UZ2X9TP5Vc+n0MOHqEUIhEVXigX35rS4ra49s5l5QN1vvTtX9Jsmt6e0YVGYcHAHntUOPfMp2BMcBrvGkq29u2dRbfE8tUm5NJi4eIRVq+a4O47J3BdmxXTtoIUPbK4hreSF8vxziasvj/hh9f/jrKZz8IF86hWFCUNCQR5GVKWkPmyC87ZfrqtSAAlhpbLeHhtg29++y7e9e5/59Zb17Npc451XRsLt30aik7h78nPJkI0PSkl3eQNbziVE08eoq9U9X1mrtVK1yNorVm6Z5lVD2T8/t4N5LkCDE7n7fPashlff15OhghEoNlqkec5d96+ju99+7esfGiSyZZDV0sM0II8B2dxLsdKHSs1Wi1hfGPOnb9+iO98ex0f+Nsb+fIX72LtOmi2rLepFfepdKj97kdQ3kdOORRDpInh8EMTrrzyTOYvBKNLoLw8935inW/PRPHdKY7vOx0LKMtddxtOO/vdrJ8cajuCT+fHq4LfiYhBk2EGRnn76w/iT1/5dKbzF915uKAYhJv+Pzi+Dw/28/73nsw5L9gXgBxHuhOtdxZYO57xZ6/6Gj+6+X5sorC6nX4moPALc6E/CV6YBX8LLT40VyGkStGXpAyUU0bnKZ516r4cd/xTGZ2f0NcnJLoPJQmIkGUZ9UbOyvWWW35xJzfecBsPr9pMvVmikelgtCkEtwURnAhaK6rlnLNesD+Xv/2FvPLPv8YPb1qLa9VR4pBu/6vuQUBBEtro0FCLH3/vrSyZZ5Gg0T96t549BEGcV2TGJxQf/8QN/PNnfsmkTWm5BGuLJZtCuSmertPeiuEFvE+acxYRUK5KWs5Iqusp6yqp7sMYjdG+TsVZnHPkNiezOc08p5UrnCsjropCe7uDslMmEVuWV8fxvbgXnecsntfPxa85gvPPP5bRYdD4fDptkTyNbGiz5Y/sFLbH8f2L//RyqnNLkE1Lr8RyLkOkSS4Ja9aWeOsbruWndzxCveXIM4s2YZ/MKZWjO+kdOksUiIUSgwxXNIvm57zoRYfzzBOWs2T3MsPzMpJSjSFZgHUOk6bkzvoB3jZpNQYY31Bm9cMtvvvd/+a663/GI2MZm+tCZg0SttTy+MHcP03RN/x9+RvRaKWpaDj22EE+9JGXMX9ISEyGUZUZDeo7k5waj2zs47V/8jlu/ukaWlqTF0NSOMcrVEGpDIplMYH0JaJRYtAISZpTqmaUK8LSRSPsu+8+zFuwAKUVOU2cyrnn7lXc97v1NBuavC60ss3kzQa4Ek4qOOONBAUdlbsLEVKtKJsqo/M2c9UHz+HYZ4xQSgyaflA2rGMVdRYk2AzqIypZM1ay/Ia4hha6f5S3v+GJr2R94L0n85I5omQBNK3w3evu47K3Xs+a8RotMkT72Wp70O6ygLXTI4iPPlSiEK1xyg+gJaVxTaFSLkMpg8RS7TdonQcfIy/4FF4YbBiv4XKHEkM9s9iQ1007jXYKLQrdZalx1nHc05fwqY+fx7xROO+SL/Nf1z8IuXfO9UpW0WX9ExS1kjoAy9BQxo3X/QW7L7Q4XUTg7TwKcSEiuNyx7hHhdW/+JD+6pUXdDoHbjOC8Yz90CaSe9iZFfhrtr+kE5VK0ycDUwtK8HyRpK0Nd9mbxQlo5wqbT3lJgUeHsYmgNG0J3b+ja46RvEEarOZe88SQuOP8oBqqQaFAhHC0qWXODXoklkvncdLaEdQ3uuGUDl/zl97n/oXEERe6KPXCLyhFf21OULAcIWZ6RqgFUS9FX1ig9Sd+AMG9hH/0DsNfeS1g8b4T58+dTbzTYuHGMjWPjPPTwWtY/UidrlsiasGmsSZKW2VRrYCpVmllOqtMgi1QYTIr2KB3lqq3Oa4yDAw9I+NjHX8HyfQxlZdAqB5X4PI5zlbBvK6qP+38P55//IR5aY5gUn+zG+1912YKEqUqW6vil+D7n60cpH2VpEoVzgtEGpTQWC1rRysUnPQ5FoxF07lAZXo5oHzE45UbpWkZG+R0ebJ2Fw5q/ufx0nv+85fRXErSYtsWz8NnqyLSZKVk7U17vPEK5+DFu6nKCNy9P9e/pNnEWw/icM9s+zhSdQFHMPHYy4jA0eeZxy7joVUcyMgRKSiBVlOtDSwmlx8PJRSforjFvNfJ9XWHRNB241DBhc8ZrMD6hWb0OVq02PLxGWLXGsmq148GHHavWQLNRpdksY10ZrcptqxnKoUI+HhCUytGSsHRxmYv/5FSG5/ufXzLQj8Yn2pPuhtZ1tEVtSFAqKNauJQiGUBTTHDuKQqgopXwOs2HLB//21Rxz0CDzrEMZnwDSK2MJiPHtR3wKh/aBam+zARqnNM44rNI46UOk7BVdMahwiJigEoX6DIqYv59wf+GKXkj7c/xyb3B4lSraaJy1pCiqKmfR0ASXvP4ILrjgSPqqxVJ6MfDNgBmdtGOZgexvo9jBjWg7mdqrCWExCUZrKuWUw49cwl/99UksWtTwS3uS0GoaxFa9v6Cy7cpSBLkmyvtvmRRHC1JLLW9SsyXWjCXcdW/Orbc3+fo37ufTn/k5H/jwd/n4NTfyuc/fzje+9QC33Nri3t8bVq5yPLzeUpeETa0WkiZkzmJMUKJU9xP4dgzKJ+21fYjzy47kk+y3r+KdV57LAfuWqSR+77zp0sDMRQxlDLBsD7j66vPYZ88J+nUVk+co50fNtgFLiv7Z1VhDEXUncHaAFUOzCdaltDJNoym0GppWXWOkHHbycpDn2NxhnY8k9Yp0Ue6FvM4R3UR0A9EZTjkyyRhd1OKv3nY6p568P/0ljVEa3Z4EExSGzj3PRMHiSa1kqc6afHclb3XQUp2+4mfgU5rGE4+5oFgVhK0PtBL6BxTnv3wFLz77EAb7chJRaCeIs17pKpp0W0sOS3Khsn1ECyAaJ4ZMFFaBBEd6cWHDUOetK74cNCLK57zSJXKrQRIMCSZYsBRe0VIolFUM9zle8+oTeeaz50OaoQwMlRVJoegRyrgtbPxRJOtrty+B8fHGlFa5sxSsgkK4OHLScsbSpTkfverlPPuYQappQkmXfD4Z55WjbeMtkU6B04LTGqsSHClICtKlaOEVNL9/aHh67aNNRRlE+RmuaqfHCF4s4bMiQz8iVJKEStJi+d6at/3V87noj4+lbGokymJ0ULTE31/XrU5/7OoUjWhnNqoZ0FvsCo0OGxVDSqmScNLJu3PZ285kt6UG5XJKiQY9AarWtmYWk8ailyk0WrT3EVQWjI+IQ2ucKJCUPE+weZVWs0SrUSJvlbB5CbFJ8MjSKK39pu/GgA7tMNx4u0hFhQLOuyLrc5RAWSsOObDE1VdfwDFHjVJOwoShfcNzvLEpUMp77hptOebohVxzzZ9y+KEJAyWfCU8VMlV8ocxsJA31pBPEaZxTWKd8Kh4HNrPBEq5DVCZe0k9bZsXETKMEtAiJtixdLFx62emc8YIDWTBUoqSD77JyIRqxmLz5LO9bXHYbPDmVrBkyxchQVHUxMD7B6W36j6FNPe6IUohWOO2VoerAZl7/+mdzxvP2pD/ZQEk1UDQRO+rTCYT66fbF6IwdoVsXJ23xYKFDh/86JdF1qGDJCRZNf4nwDWsYHdCcd94Kzj//ECqlJpAhCoZGjM+bVXxni9/u4JUvQCnvdzZLeXy2n/DU2qFNnT32Fd774XM587mHMpDkGOe31nFTvBF6y7LrL50Zz2Nnytd6r+ELUqQErux9BfIWo0OWZx2/mE//v9fzspcdSUkbKmmpqz4jcx0/2HXVlnNUlOIPTt+fSy99McuWKdJ0M9psRGkbBLlfHuyoPZ2+rnzH7hpBt9IStvIxxZjR++EWhNFEBKUbaDNOf0mz4tAq11xzMU89oo9qKQv32tUvtvG7cw2FQpywfPkwn/rHV3DiibszWGmQSMtvI6Z8AI23ePu0C1NzV22bTqkUZdMpnO669MuEXQUnBpEEXIXUVSjZFst3U/zdB8/nzOevoL/kUAJ5FibMvfkMt4MntZK1reosuuGU7tieAT1JCq6rwW9f83r8kBBG7dBUjLBwuM7f/OXzeN2rjyWVMSR3JGWLE+kyR/dUcttPKLwtlKQgrP3R+bfuSRraexQCuX0uMFBRnPeyA3jDm05gsFpklm8BwoLFKc41O7/RfS9Bee9W4hUgTrFxbHzWkiX+XzC6hGYAowfRSpi3yPK+d5/Ih953IUsWQpLmiLagu8u3V3DhxeKU9zNhuuv0UiwRgZIKaTKA0bDv3gNcffVFfPijF7D3XoJYSykpoYP1ATr3+2SgbfSd8uGjCMg5iEJRMpayhtOeuzufuOZiVhw6D02Ka1bAloJDbqcveXm+ZbtsH364bo/h7Xah1RRrhm8rxeDe+bxzQjhQfhnd9aGkH2sFYyY495yj+dg1Z7H3XpZETYLLgyVdtlxS2wXQWlNKU0pas2ie4iMfO4NLL30hixcKhgyjDKpb0ZqpghUqo1NH3R9PU3/hKCIzlRJKlZwkaZAwxnlnPYV//5c3cuLRVYYpU1YpaPFuD51K6/rxx86TQlfYAgUo6SR+7Hy4jYMpUmcGzWEO4df7C4Hh6X2+3sMnkgMJM6ru0Ogdi7RjccT7YFAlVWUWzU949UXHcPWHzmf33avk+WSwAG2rwxYl0F2nWzvXKzxaBB2u2Y5VCvl3fIZpQSvHvKEyV1x5Jpe88UQG+rR3hAcUPtptcDillCQQcmDNpDU561i3dj06mdpVe1vmjqez3KGURWkfCTQ04Hje84b43Bf/mFdcfCQLFyckOgQP9BxCZya7tTKY8oxbnFa8Ca/tv3fVsYAmp9pX4ylPS7jigyfw+WtfzUknDTM6JCSKrQ+MuyBtB9/pizPgzwm93PsS9j7/zm1c24lCiSHROSUtrDi0xMeuuYBXv+YE5o22MManXqAdQOHxXXubBeZp74XovL+Q/3AbheXPseI3hPPLVw5RddCTiN7EwYeMctWHL+Itl+7H0gX9lI1Bk3pHcB2s+GE1ZVdErOAyy2Al4byXLedTn7yEc19yMsMDJZKQFqGoj3b33ar8LsyExd96yl6k06rD2O5HvM7VtXIonXHIEWX+/tMv5R3vOJXdRqFPhki1AyvYXFCPY4DBkzJPllBDqLNhU8Kn//UGmgyT5ARdt+cQv15vRFCqgVYZVVPlqBMHOfLIvQm7CMwROo3Mvwgb65rPfP77uE1DkPeTUEMToiYkAUlQknYOV0ZRolxNeM5z9+Tg/YfR1qFcjqgkzNh2LKo9l9MhA3zqHaDNJEklY7+9lnDaCUegJ9fw4CqNy2poyRDxVpQQSxR8JxI0PoeQ96NqO0aB6swZvS+RX7c3ODSWRBKMmiShhRFvozJaMWgyjl8xxEffcxbHn7KU/r4SJlzI5AmJTdHKsH7M8bUvP0DNTSK6FQSuT9jpB/piqcLQSC1Oa6qUOPKQvXna05a2ty5s3+NUEbNDKfbs8v0kDdngU5QyJCZh/qDhuKcu5NRTDiNPNfVanfpkk1Y2idJgxWLFtmeigh9JtBi0BL8HvAKrCIouQiKCEUJ6BYdoi6iivnyOM6NTtNL091XZb7cqf3D8Av78kufwuouP4+gV/cwfSihphSn8LKYZxArLxK7GnWst3/za7dhWilE+x1NCQqIMJTQlNImklJSjBCxbWuJFL15BSZd8Xeyaj+1R/n9KFImGRPtM5Mc9bRkHH7qcdevG2TzRpGmdj0LVCdYZlE7wrU3Q4v3xVGHlDj5UXk4EK5ZSIeWADj5IwRfQe2+226YKbVqc9vvlGrDSopxo9lxW4ZUvP5p3vf10VhwII30jVEzikxWrIN/CNkBti3nv8+4CKK18GhYMZWNYsqTECccv5vDDFuDEsmFDRqOpseR+b9FQExrfz41LSFwZ5VKviqlisgtaErSrhGr3eRMtJawqecVUO5x23oGdJgNDliOP2Zs3X3wIb33tiTzloN2oVAxJxSccVipEGQYr5fRjXbeCNzOelCkcckBhuff3hjPOu5J1m/owUuk9DQqZI/jOIhqlM/orVf7kLQfzZxc+k/7eL+xUnB+Cpbhpx8r1hpPPeCeTGxbgbILoGkoVQbThELqMmv51dKjOO991EmeedhCJA8QiuryVhrdj8AHAxe8LTrKwpUtK7mBsXLj73kk+9Ylv8YufP0SzlTI2kYPp89/VXplSZgMKEEkRSj5VAD6SxafwKJYTwo7xon2elaSGuD5srkhSTaWk2GMpnP3CA7jo/Ocw0qcwfZ1M+VpAO+eT8CjNr367lpdf+Hke2qyxyqCLzuxdbztPqcCSoIDB1PHSsw7jssuOoz9o9HPF/NyxQvn76m4bDp8rLLOazMLGMcdNP7iNr371B/x2lbB+LKfVTMlaJZyjs21NmEGWvP9xEGkdy4NXl7WPOtIW0YLWGmO8crdHNaV/IOOwIxZz6ulH8ZSn7Ma80RwEKiWvSBRW3W2JPq9E7ry2/lgoakEQ7r7/Ef71Mz+nNgkiuU/M6kBLkaRXvOOv3gTKcdDBu/HHF52ECa2qkArFv3d1LC0sDSayJvXmIDdcfw+f/IdbWfVwzuRkk6xlyfM8TIp8epaiFDptzv/NR7v5wAlfNp0zdGhLyo74j5RFYbHW4pxQ6Xf0D+YsXFThhS9awfNPP4bdFw8wVAWX1yiZErpoc6Ef9ToIPFHqJZOcpsuZqMO6dYZ/+dwN/ODWB1j98EYmx8FlCc4WQUAJ2hW7ZAiCQ7kwDohBiUZUhlJNP+HShOhjgzYp1QQW9dU55NAFvOwVJ7Li8EWMDClSoyglaQiomlqiXi5MXQ7u0F0r3b1l6zwplawMIaHB2FjKt2+4nVoTKt62uAUqCHq/oUITkQqpGeCIo0c4aL/FW92xbefQq2QJ47Wc/7zuFhqTfWFu5nOOeHnrG0kx4/L4PFJpWuXIZyxgn2V9+Lgub9mZO/ilJidhXyzlLVGZNNi0Ce67p8F/fvMOvvSNX7G+3g8CjVoThSIxZVyeI+L3QkQFa5dSXskShbOaSrlK3mqhlcK6DG1aaF2i0mfZY1mFP3zpsZx64iEsmZcyUAWj8M75oe+pYHFBfCb0TY2M6/9zIz/71S8YGh5kdGSYSqVCkmi/X0xwxhIlOOU90FJdZvk+izj8iH5KZm6p9NtCCmEV2pUoyFpQ2wwPrWswOen4xc9/yx233ct//3Yl62p+sLPisOJI9Sg207Qy78NWKqUYrciyJlopSqWEkmuyYDhhj72GeOoxe7HisH3YY8FCRkb76B9QpCVIUjA6zH6Dku2HskcTkI/297lFUd6ZmyRvDbTjJFRXry0kvRN8CLtyKJVQqSh0MJMWk5hd6+mnRwBHhqWFooo4TZY3GRsvc+dvxvn6127ihzfcxkRNMa5SxBpsliA2BecQZxFxGGPaSpZTris4QuGco9lokpoEk2q0rvgoOAGFYbhPMVLeyJHH7M5Z5x3N8uUDLBpZSjUFk3i1QZFjVLCaBcuVb6nTM5ek8PaQ0fSR2pKiHGQ5PLIJ1q2b5Kc//h9+ctOv+O+7xqjn/WS50GoKNkQUCoJ1fsnXWYfWitQ4NE2ca1JJR0iSnPkL4elPX84Jx6/g0P2WsnBBmf5BS7kS9kC0Cq396K3YmtVqOqKSNSNaQAkLuaEFOGNJbTAFT4Mown5VObgKrRZUqqCwfn+swPTf3pF0K1n+fY6QO7DOgBKMUxhlMUEKe3O4P3xD8EqWiN9ZRqtW2FDHz3V3/jMWCOKfLrzz9++wIZdShXoNVq+HX/92Fb/65T3c+vO7uffuh5nIFmKtCRmjfadzqstOJhpDGYVFJzV0UqOUNDhgj2GOe9ZhHHXM/hx04B4snu+XXZTTaOPlo+uVgGFGq1AITWxzoMgMQZJ4dzfrhESrIFb9Fg+FQHc2AQVpOkHKQM/F5y5+kAjP07ZIhYEvdyRGIzk0G9DKYPWGjMlajbGNY2wcHwdVoVbPqdVrGG0YGOgnSRLK5ZSRoWEGBgaYN6gZ7INqP+iSIKqONgatQi6tkA6jyHVTZNruiu/cBjMToHOBdtmKYFULoYyiFZ4yDVvC+DMFAVdYa/x7pXz/KbYcYZd6+q3j5UGwfojBdOWSzARqNVizpsmv7/wdd9y9lttvu5u7fvMwrXpCxgj1VkKz2SRJ/L6cogXRFh3KTmtNlmVUKxVcnqMMjCTCQL9i/wOXcsiKvTjogKUccfASFiwsUekTymkTZSv+5jRY7e9SBcv3E13JEsCS4wQSZ9DOl2WehHQMFvLM70n60OoG69ZtYP36zWzaVGf9hs1M1OqMbdyIc4KIZdHCURJjmT+/n/nzBxnoH2HvfXZjt937qfZBKfWJhbUiGBdcZxYceogiKlmPOzmQ+LYN2lsN/HLN9AXWafC+UK0Do72I8h3OM/23dyR+UPMvXoAW3gHF3zUqRLz1VntxztRXn6tXBSvYXHjGAoeEPcRAhdwnfjUf8aZeX8U5Ck2jrskaMLYB/ufecX7/wMM8+OBDbNw4zkStRpZ7ZQ181xseHKF/oMwee85j+f5LWL7fUnYbSaj0W9IklKhTKGd8YlGtQHu1otvO7FPpFduzNHF5HyiHMTpEDIY9/Hrqo1AaO+NCi4TSlHPmMsXThGE9tCH8FEd0mJwIPjQAbN7yUUA6weY5aFDa50ECr4iqEI3kX/3VtN+fFoKCbbFhGPJtodNyixIt3j8aMzlnbtBb1r51ZqHNmfbuDt4OAEq0z3ovhE28w/dDNJ0/94lBIek0DmX9FlgQmojxf8szB6JpNGFiHO67bw133TfOyofHWbNmDbVajXqz7vtrkTdJQbVaIU1LjIwMs3TJYpbtuQcHLFnAbku94l+pgiiHE+vzeSntk2YioYmKz6sV/A+9q4Jv0L3SuZtdvW4sXkEyAsoZEIUNS0JOfJ5Cow06TPaLDaGdT11GnvlzvauFb/cmEZy00Eb5KGEShJD+hqLQChng90psf8xjUbJ6a+bRv/ekVLIglNWjl882mVJ/c47H4+66m0bRPOcKU5+vtxnPpNMUA3aj0WhvKgxgjGn7+eiwb96jIkHBwvf+Tll59aLbm+zJwuPRAreH7rbQ3Sp6fS+eaBRtrmjXW6O3rxRs6zu7Mt3lwnY8Z57nWGtRXSkBCNcxYT/R7UEKmRH8BHvZWf1ntily5033XI+1bmZCb713//5s/F4vT14l6wnPE7WLPr5kWUaaplOULELn294O2N2htu8KkUhkLtE9TG6vXNgahRLYS5TgTwyikhWJRCKRyKPQaxGJRGZCVLIikUgkEolEZoEZOpxEIpFIJBKJRB4LUcmKRCKRSCQSmQWikhWJRCKRSCQyC0QlKxKJRCKRSGQWiEpWJBKJRCKRyCwQlaxIJBKJRCKRWSAqWZFIJBKJRCKzQFSyIpFIJBKJRGaBqGRFIpFIJBKJzAJRyYpEIpFIJBKZBaKSFYlEIpFIJDILRCUrEolEIpFIZBaISlYkEolEIpHILBCVrEgkEolEIpFZICpZkUgkEolEIrNAVLIikUgkEolEZoGoZEUikUgkEonMAlHJikQikUgkEpkFopIViUQikUgkMgtEJSsSiUQikUhkFohKViQSiUQikcgsEJWsSCQSiUQikVkgKlmRSCQSiUQis0BUsiKRSCQSiURmgahkRSKRSCQSicwCUcmKRCKRSCQSmQWikhWJRCKRSCQyC0QlKxKJRCKRSGQWiEpWJBKJRCKRyCwQlaxIJBKJRCKRWSAqWZFIJBKJRCKzQFSyIpFIJBKJRGaBqGRFIpFIJBKJzAJRyYpEIpFIJBKZBaKSFYlEIpFIJDIL/C/jfPbVcS4hJQAAAABJRU5ErkJggg=="


def _pv_display(v):
    """Every empty cell renders as an explicit NA. A blank cell is indistinguishable from
    'the log didn't report it', which is exactly the thing the report exists to surface."""
    if v is None:
        return "NA"
    t = str(v).strip()
    return t if t else "NA"


def _pv_collect(node_results, key):
    """The report is organised by technology, not by node: every 4G sector at the site appears
    in one 4G table and every 5G sector in one 5G table, regardless of which node owns it.
    Rows are sorted by sector name so a site's sectors read in a stable, predictable order."""
    rows = []
    for _node, res in node_results.items():
        for r in res.get(key, []) or []:
            rows.append(r)
    return sorted(rows, key=lambda r: str(r.get("sector") or ""))


def build_parameter_verification_pdf(scope, node_results, has_pre=True, fa_code=None, site_ids=None):
    """Site-wide (not node-wise) parameter verification report.

    Layout rules, all of which exist to stop the "data everywhere" look:
    - One row per sector; each parameter is a Pre/CIQ/On-site column group placed side by side.
    - A parameter group is NEVER split across two tables. When the groups don't fit the page,
      the table continues in a "(cont.)" block, and the Comments column is repeated on every
      block so a reader never has to cross-reference back to an earlier chunk.
    - Column widths are derived from the real printable width, and the number of value columns
      per block is capped, so columns are always legible and nothing runs off the right edge.
    - Empty values render as NA rather than blank.
    - The MasTec logo and the site banner are drawn by the page callback, so they repeat on
      every page including continuation pages; header rows repeat with repeatRows=2.
    """
    import base64
    from reportlab.lib import colors as pv_colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
                                     CondPageBreak)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.utils import ImageReader
    from reportlab.lib.enums import TA_LEFT, TA_CENTER

    COLOR_MAP = {
        "green": pv_colors.HexColor("#DCF1E2"), "red": pv_colors.HexColor("#FBD8DB"),
        "amber": pv_colors.HexColor("#FDEDC6"), "gray": pv_colors.HexColor("#ECEEF1"),
    }
    NAVY = pv_colors.HexColor("#012A4E")
    ACCENT = pv_colors.HexColor("#FF5B24")
    SUBHEADER_BG = pv_colors.HexColor("#4B7BB0")
    GRIDC = pv_colors.HexColor("#DBE1E8")
    RULEC = NAVY
    ZEBRA = pv_colors.HexColor("#F6F8FB")

    PAGESIZE = landscape(letter)
    PAGE_W, PAGE_H = PAGESIZE
    BAND_H = 0.60 * inch
    L_MARGIN = R_MARGIN = 0.30 * inch

    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=6.2, leading=7.4,
                                 alignment=TA_CENTER)
    sector_style = ParagraphStyle("sector", parent=styles["Normal"], fontSize=6.2, leading=7.4,
                                   alignment=TA_LEFT, fontName="Helvetica-Bold")
    note_style = ParagraphStyle("note", parent=styles["Normal"], fontSize=5.6, leading=6.9,
                                 alignment=TA_LEFT)
    hdr_style = ParagraphStyle("hdr", parent=styles["Normal"], fontSize=6.2, leading=7.4,
                                textColor=pv_colors.white, alignment=TA_CENTER,
                                fontName="Helvetica-Bold")
    sub_style = ParagraphStyle("sub", parent=hdr_style, fontSize=5.4, leading=6.4)
    tbl_title_style = ParagraphStyle("tt", parent=styles["Normal"], fontSize=8.5, leading=10.5,
                                      fontName="Helvetica-Bold", textColor=pv_colors.white,
                                      alignment=TA_LEFT)
    legend_style = ParagraphStyle("lg", parent=styles["Normal"], fontSize=5.9, leading=7,
                                   textColor=pv_colors.HexColor("#4A5563"))

    def P(text, style=cell_style):
        if text is None or (isinstance(text, str) and text.strip() == ""):
            return Paragraph("", style)
        return Paragraph(str(text).replace("&", "&amp;").replace("<", "&lt;"), style)

    def V(value):
        return P(_pv_display(value), cell_style)

    site_id_text = ", ".join([str(s) for s in site_ids]) if site_ids else "NA"
    try:
        _logo_reader = ImageReader(io.BytesIO(base64.b64decode(LOGO_B64)))
    except Exception:
        _logo_reader = None

    def _page_furniture(canvas, doc_):
        canvas.saveState()
        canvas.setFillColor(NAVY)
        canvas.rect(0, PAGE_H - BAND_H, PAGE_W, BAND_H, stroke=0, fill=1)
        canvas.setStrokeColor(ACCENT)
        canvas.setLineWidth(1.6)
        canvas.line(0, PAGE_H - BAND_H, PAGE_W, PAGE_H - BAND_H)
        x_text = L_MARGIN
        if _logo_reader is not None:
            try:
                canvas.drawImage(_logo_reader, L_MARGIN, PAGE_H - BAND_H + 0.12 * inch,
                                 width=0.92 * inch, height=0.35 * inch,
                                 mask="auto", preserveAspectRatio=True, anchor="sw")
                x_text = L_MARGIN + 1.10 * inch
            except Exception:
                pass
        canvas.setFillColor(pv_colors.white)
        canvas.setFont("Helvetica-Bold", 9.5)
        canvas.drawString(x_text, PAGE_H - BAND_H + 0.32 * inch,
                          f"{scope} \u2014 Parameter Verification Report")
        canvas.setFont("Helvetica", 6.8)
        canvas.setFillColor(pv_colors.HexColor("#C7DAEE"))
        canvas.drawString(x_text, PAGE_H - BAND_H + 0.14 * inch,
                          f"FA Code: {fa_code or 'NA'}     Site ID(s): {site_id_text}")
        canvas.setFont("Helvetica", 6.8)
        canvas.setFillColor(pv_colors.HexColor("#C7DAEE"))
        canvas.drawRightString(PAGE_W - R_MARGIN, PAGE_H - BAND_H + 0.14 * inch,
                               f"Page {canvas.getPageNumber()}")
        canvas.setStrokeColor(GRIDC)
        canvas.setLineWidth(0.5)
        canvas.line(L_MARGIN, 0.22 * inch, PAGE_W - R_MARGIN, 0.22 * inch)
        canvas.restoreState()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=PAGESIZE,
                             topMargin=BAND_H + 0.16 * inch, bottomMargin=0.34 * inch,
                             leftMargin=L_MARGIN, rightMargin=R_MARGIN,
                             title=f"{scope} Parameter Verification")
    story = []

    _legend_items = [("green", "Matching"), ("red", "Mismatch"),
                      ("amber", "Attention / NA / confirm manually"), ("gray", "No data")]
    legend = Table([[P(lbl, legend_style) for _, lbl in _legend_items]],
                   colWidths=[doc.width / len(_legend_items)] * len(_legend_items), hAlign="LEFT")
    _lg = [("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
           ("BOX", (0, 0), (-1, -1), 0.4, GRIDC),
           ("INNERGRID", (0, 0), (-1, -1), 0.4, GRIDC),
           ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
           ("TOPPADDING", (0, 0), (-1, -1), 2), ("BOTTOMPADDING", (0, 0), (-1, -1), 2)]
    for i, (c, _) in enumerate(_legend_items):
        _lg.append(("BACKGROUND", (i, 0), (i, 0), COLOR_MAP[c]))
    legend.setStyle(TableStyle(_lg))
    story.append(legend)
    story.append(Spacer(1, 6))

    CAT_A_LTE_PARAMS = ("rachRootSequence", "PCI", "Cellrange")
    CAT_A_NR_PARAMS = ("rachRootSequence", "nRPCI", "Cellrange")

    def _cat_a_comment(row, params):
        """Comment wording for PCI / nRPCI / Cellrange / TAC / NRTAC / nCI / rachRootSequence.
        Mismatch -> '<PARAMS> Mismatch found on the Pre existing sectors' (or '... on the newly
        adding sectors'). Match -> 'Matching as per pre' for a pre-existing sector, and no
        comment at all for a newly-added one, which has no Pre state to have matched."""
        reds = [p for p in params if row.get(f"{p}_color") == "red"]
        if reds:
            tail = "newly adding" if row.get("is_new") else "Pre existing"
            return f"{', '.join(reds)} Mismatch found on the {tail} sectors"
        if any(row.get(f"{p}_color") in ("gray", "amber", None) for p in params):
            return ""
        return "" if row.get("is_new") else "Matching as per pre"

    def section_title(text):
        t = Table([[P(text, tbl_title_style)]], colWidths=[doc.width], hAlign="LEFT")
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), NAVY),
            ("LINEBELOW", (0, 0), (-1, -1), 1.4, ACCENT),
            ("LEFTPADDING", (0, 0), (-1, -1), 7), ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 3.5), ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
        ]))
        return t

    # Cap on value columns per block. 12 keeps every value column at roughly 0.58in, which is
    # legible at 6.2pt; letting a block grow past this is what made the old tables unreadable.
    MAX_SUB_COLS = 12
    SECTOR_W = 1.38 * inch
    COMMENT_W = 2.00 * inch
    MAX_SUB = 1.15 * inch

    def build_wide_table(wide_rows, param_groups, title, comment_groups=None, cat_a_params=None):
        """param_groups: [(display_label, key_prefix, sub_cols), ...] with sub_cols
        ['pre','ciq','post'] or ['ciq','post']."""
        if not wide_rows:
            return
        want_comments = bool(comment_groups) or bool(cat_a_params)
        avail = doc.width - 2

        # Split the groups into evenly-sized blocks rather than filling each block to the brim
        # and leaving a one-parameter runt at the end.
        total_sub = sum(len(g[2]) for g in param_groups)
        n_blocks = max(1, -(-total_sub // MAX_SUB_COLS))
        target = -(-total_sub // n_blocks)
        blocks, cur, cur_n = [], [], 0
        for g in param_groups:
            n = len(g[2])
            if cur and cur_n + n > target and len(blocks) < n_blocks - 1:
                blocks.append(cur)
                cur, cur_n = [], 0
            cur.append(g)
            cur_n += n
        if cur:
            blocks.append(cur)

        def _comments_for(block_params):
            """A row's Comments cell only ever mentions parameters that are actually shown in
            the same block. A block whose columns are Earfcn/TX/RX must not report a TAC
            mismatch — TAC lives in a different block, and that block reports it itself."""
            out = []
            for row in wide_rows:
                if not want_comments:
                    out.append("")
                    continue
                if cat_a_params:
                    out.append(_cat_a_comment(row, [p for p in cat_a_params if p in block_params]))
                elif comment_groups == "row":
                    out.append(row.get("comment", "") or "")
                else:
                    parts = []
                    for params, fixed_text in comment_groups:
                        trig = [p for p in params
                                if p in block_params and row.get(f"{p}_color") == "red"]
                        if not trig:
                            continue
                        parts.append(group_mismatch_comment(trig, fixed_text)
                                     if fixed_text is not None
                                     else row.get(f"{trig[0]}_note", "Mismatch found"))
                    out.append(" / ".join(dict.fromkeys(parts)))
            return out

        block_comments = [_comments_for({kp for _, kp, _ in blk}) for blk in blocks]
        any_comment = any(t.strip() for texts in block_comments for t in texts)
        base_comment_w = (COMMENT_W if any_comment else 0.95 * inch) if want_comments else 0

        # Never leave a section header stranded at the foot of a page with a row or two under
        # it — push the whole section to the next page if it can't get a reasonable start here.
        story.append(CondPageBreak(1.5 * inch))
        story.append(section_title(title))
        story.append(Spacer(1, 3))

        for b_i, block in enumerate(blocks):
            comment_texts = block_comments[b_i]
            n_sub = sum(len(sc) for _, _, sc in block)
            comment_w = base_comment_w
            sub_w = min(MAX_SUB, (avail - SECTOR_W - comment_w) / n_sub)
            # Any width left over after capping the value columns widens Comments first, then
            # the Sector column — never left as dead space between columns.
            slack = max(0.0, avail - (SECTOR_W + n_sub * sub_w + comment_w))
            if want_comments and any_comment:
                comment_w += slack * 0.6
                sector_w = SECTOR_W + slack * 0.4
            else:
                sector_w = SECTOR_W + slack

            top_header = [P("Sector", hdr_style)]
            sub_header = [P("", sub_style)]
            col_widths = [sector_w]
            span_cmds, group_edges, col_idx = [], [], 1
            for label, key_prefix, sub_cols in block:
                span_start = col_idx
                for sc in sub_cols:
                    top_header.append(P("", hdr_style))
                    sub_header.append(P("ON SITE" if sc == "post" else sc.upper(), sub_style))
                    col_widths.append(sub_w)
                    col_idx += 1
                span_cmds.append(("SPAN", (span_start, 0), (col_idx - 1, 0)))
                top_header[span_start] = P(label, hdr_style)
                group_edges.append(col_idx - 1)
            if want_comments:
                top_header.append(P("Comments", hdr_style))
                sub_header.append(P("", sub_style))
                col_widths.append(comment_w)

            data = [top_header, sub_header]
            painted = []
            for r_i, row in enumerate(wide_rows, start=2):
                data_row = [P(row.get("sector"), sector_style)]
                c_i = 1
                for label, key_prefix, sub_cols in block:
                    color = row.get(f"{key_prefix}_color")
                    for sc in sub_cols:
                        val = row.get(f"{key_prefix}_{sc}")
                        data_row.append(V(val))
                        # A newly-added sector has no Pre state at all: its Pre cell reads NA
                        # and is amber, whatever the parameter's own verdict is.
                        if sc == "pre" and _pv_display(val) == "NA":
                            painted.append((r_i, c_i, "amber"))
                        elif color:
                            painted.append((r_i, c_i, color))
                        c_i += 1
                if want_comments:
                    data_row.append(P(comment_texts[r_i - 2], note_style))
                data.append(data_row)

            tbl = Table(data, repeatRows=2, colWidths=col_widths, hAlign="LEFT")
            cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), NAVY),
                ("BACKGROUND", (0, 1), (-1, 1), SUBHEADER_BG),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, GRIDC),
                ("BOX", (0, 0), (-1, -1), 0.8, RULEC),
                ("LINEBELOW", (0, 1), (-1, 1), 0.8, RULEC),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 2), ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 2), ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("SPAN", (0, 0), (0, 1)),
                ("ROWBACKGROUNDS", (0, 2), (-1, -1), [pv_colors.white, ZEBRA]),
            ] + span_cmds
            if want_comments:
                last = len(col_widths) - 1
                cmds.append(("SPAN", (last, 0), (last, 1)))
                cmds.append(("LINEBEFORE", (last, 0), (last, len(data) - 1), 0.8, RULEC))
            for edge in group_edges:
                cmds.append(("LINEAFTER", (edge, 0), (edge, len(data) - 1), 0.6, RULEC))
            cmds.append(("LINEAFTER", (0, 0), (0, len(data) - 1), 0.8, RULEC))
            for r_i, c_i, color in painted:
                cmds.append(("BACKGROUND", (c_i, r_i), (c_i, r_i),
                              COLOR_MAP.get(color, pv_colors.white)))
            tbl.setStyle(TableStyle(cmds))
            if b_i:
                story.append(Spacer(1, 3))
                story.append(P("(cont.)", legend_style))
                story.append(Spacer(1, 2))
            story.append(tbl)
            story.append(Spacer(1, 9))

    c2 = ["ciq", "post"]
    c3 = ["pre", "ciq", "post"] if has_pre else ["ciq", "post"]

    naming_lte_groups = [("FDD naming", "FDD_naming", c2), ("Rilinks (RiPort)", "Rilinks", c2),
                         ("Sector carrier", "sector_carrier", c2), ("ISDLONLY", "ISDLONLY", c2)]
    naming_nr_groups = [("Rilinks (RiPort)", "Rilinks", c2),
                        ("NR sector carrier", "sector_carrier", c2),
                        ("NRCELL CU naming", "NRCELL_CU_naming", c2)]
    lte_groups_1 = [("EarfcnDL", "EarfcnDL", c3), ("EarfcnUL", "EarfcnUL", c3), ("TX", "TX", c3),
                    ("RX", "RX", c3), ("Bandwidth", "Bandwidth", c3),
                    ("ConfiguredOutputPower", "ConfiguredOutputPower", c3),
                    ("CellID", "CellID", c3), ("TAC", "TAC", c3)]
    lte_groups_2 = [("rachRootSequence", "rachRootSequence", c3), ("PCI", "PCI", c3),
                    ("Cellrange", "Cellrange", c3)]
    nr_groups_1 = [("arfcnDL", "arfcnDL", c3), ("arfcnUL", "arfcnUL", c3),
                   ("bSChannelBwDL", "bSChannelBwDL", c3), ("bSChannelBwUL", "bSChannelBwUL", c3),
                   ("ConfiguredOutputPower", "ConfiguredOutputPower", c3),
                   ("cellLocalId", "cellLocalId", c3), ("ssbFrequency", "ssbFrequency", c3),
                   ("TX", "TX", c3), ("RX", "RX", c3),
                   ("NRTAC", "NRTAC", c3), ("nCI", "nCI", c3)]
    nr_groups_2 = [("rachRootSequence", "rachRootSequence", c3), ("nRPCI", "nRPCI", c3),
                   ("Cellrange", "Cellrange", c3)]

    naming_lte = _pv_collect(node_results, "_wide_naming_lte")
    naming_nr = _pv_collect(node_results, "_wide_naming_nr")
    lte_rows = _pv_collect(node_results, "_wide_lte")
    nr_rows = _pv_collect(node_results, "_wide_nr")

    # Blueprint row order: naming first, then all "Earfcn/TX-RX/power/cell id" tables grouped
    # together (4G then 5G), then all "PCI/Cellrange/TAC" tables (4G then 5G) — type-first.
    build_wide_table(naming_lte, naming_lte_groups, "4G Naming & Configuration",
                      comment_groups="row")
    build_wide_table(naming_nr, naming_nr_groups, "5G Naming & Configuration",
                      comment_groups="row")
    build_wide_table(lte_rows, lte_groups_1, "4G Sectors \u2014 Earfcn / TX-RX / Power / Cell ID / TAC",
                      comment_groups=[g for g in COMMENT_GROUPS_LTE if g[1] is not None])
    build_wide_table(nr_rows, nr_groups_1,
                      "5G Sectors \u2014 arfcn / Bandwidth / Power / cellLocalId / SSB / TX-RX / NRTAC / nCI",
                      comment_groups=[g for g in COMMENT_GROUPS_NR if g[1] is not None])
    build_wide_table(lte_rows, lte_groups_2, "4G Sectors \u2014 rachRootSequence / PCI / Cellrange",
                      cat_a_params=CAT_A_LTE_PARAMS)
    build_wide_table(nr_rows, nr_groups_2, "5G Sectors \u2014 rachRootSequence / nRPCI / Cellrange",
                      cat_a_params=CAT_A_NR_PARAMS)

    doc.build(story, onFirstPage=_page_furniture, onLaterPages=_page_furniture)
    return buf.getvalue()

    
# ============================================================
# UI
# ============================================================

if "qkx_page" not in st.session_state:
    st.session_state.qkx_page = "home"
if "qkx_scope" not in st.session_state:
    st.session_state.qkx_scope = None

def _qkx_go(page, scope=None, report_only=False):
    st.session_state.pop("qkx_results", None)
    st.session_state.qkx_page = page
    st.session_state.qkx_report_only = report_only
    if scope is not None:
        st.session_state.qkx_scope = scope
    st.rerun()

# ---- sticky top bar + shared styling (stays put on scroll — every page) ----
# MasTec brand palette used as ACCENTS only now: Prussian Blue #00284e, Endeavour #024ea4, Orange #ff5b24.
# Main content area is a light background — this is a deliberate fix: the previous dark theme fought
# with Streamlit's native (light) widget chrome and, combined with a buggy fill-mode:both entrance
# animation, left several result sections stuck at near-zero opacity (confirmed via screenshots).
st.markdown("""
<style>
  .stApp {
      background: linear-gradient(180deg, #eef3fa 0%, #f7f9fc 100%);
  }
  .qkx-topbar {
      position: sticky; top: 0; z-index: 999;
      display: flex; justify-content: space-between; align-items: center;
      padding: 0.9rem 1.75rem; margin: -1rem -1rem 1.5rem -1rem;
      background: linear-gradient(90deg, #011b36 0%, #012a4e 100%);
      border-bottom: 1px solid rgba(255,91,36,0.55);
      box-shadow: 0 4px 18px rgba(0,0,0,0.2);
  }
  .qkx-topbar .qkx-logo { font-size: 1.4rem; font-weight: 900; color: #ffffff; letter-spacing: 1px; }
  .qkx-topbar .qkx-logo span { color: #ffffff; }
  .qkx-topbar .qkx-credit { font-size: 0.78rem; color: #cfe0f5; text-align: right; line-height: 1.3; }

  .qkx-hero { text-align: center; margin: 1rem 0 2.5rem 0; }
  .qkx-hero h1 {
      font-size: 3.2rem; font-weight: 900; letter-spacing: 3px; margin-bottom: 0.3rem;
      color: #012a4e;
  }
  .qkx-hero p { color: #4a5b70; font-size: 1.02rem; }

  div[data-testid="stButton"] button {
      border-radius: 10px; font-weight: 700; border: 1.5px solid #013a6b;
      background: linear-gradient(135deg, #024ea4, #013a6b); color: #ffffff;
      box-shadow: 0 3px 8px rgba(1,42,78,0.25);
      transition: transform 0.15s ease, box-shadow 0.15s ease, border-color 0.15s ease;
  }
  div[data-testid="stButton"] button:hover {
      border-color: #ff5b24; color: #ffffff; transform: translateY(-1px);
      box-shadow: 0 6px 14px rgba(255,91,36,0.35);
  }
  div[data-testid="stButton"] button:active { transform: translateY(0); }

  /* Bordered containers (st.container(border=True)) — clean light cards, not translucent-on-dark */
  div[data-testid="stVerticalBlockBorderWrapper"] {
      background: #ffffff !important;
      border: 1px solid #dde5ef !important;
      border-radius: 12px !important;
      box-shadow: 0 2px 10px rgba(1,42,78,0.06);
  }

  .qkx-checklist { margin: 0.5rem 0 0.5rem 0; }
  .qkx-check-row {
      display: flex; align-items: center; gap: 0.75rem;
      padding: 0.55rem 0.9rem; margin-bottom: 0.4rem; border-radius: 10px;
      background: #f3f6fb; border: 1px solid #e2e8f2;
      opacity: 0; transform: translateX(-14px);
      animation: qkxRowIn 0.4s ease forwards;
  }
  @keyframes qkxRowIn { to { opacity: 1; transform: translateX(0); } }
  .qkx-checks-fadeout { animation: qkxFadeOut 0.55s ease forwards; }
  @keyframes qkxFadeOut { to { opacity: 0; transform: translateY(-8px); } }
  .qkx-check-icon {
      width: 26px; height: 26px; border-radius: 50%; display: flex; align-items: center; justify-content: center;
      font-weight: 900; font-size: 0.95rem; flex-shrink: 0;
  }
  .qkx-check-icon.pass { background: linear-gradient(135deg, #024ea4, #17c3a2); color: #ffffff; }
  .qkx-check-icon.fail { background: #d6dee8; color: #6b7c91; }
  .qkx-check-label { font-weight: 600; color: #1a2c40; }
  .qkx-check-label.dim { color: #7c8ba0; }

  /* Subtle, safe entrance motion for results — never hides content: no opacity/fill-mode tricks,
     so if the animation doesn't fire for any reason the element is simply static and fully visible. */
  @keyframes qkxSettle {
      0%   { transform: translateY(10px); }
      100% { transform: translateY(0); }
  }
  div[data-testid="stVerticalBlockBorderWrapper"] {
      animation: qkxSettle 0.35s ease-out;
  }
</style>
<div class="qkx-topbar">
  <div class="qkx-logo">MAS<span>TEC</span></div>
  <div class="qkx-credit">Made by <b>AKSHATHA KALLUR</b><br>Powered by <b>MASTEC</b></div>
</div>
""", unsafe_allow_html=True)

# ---- HOME ----
if st.session_state.qkx_page == "home":
    st.markdown("""
    <div class="qkx-hero">
      <h1>QUICKIX</h1>
      <p>SOW analysis and Integration templates generator</p>
    </div>
    """, unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3)
    with c1:
        st.caption("Pre-existing sites — MCA, CENM, or CRAN rehome")
        if st.button("MCA", use_container_width=True, key="qkx_card_mca"):
            _qkx_go("family")
        if st.button("\U0001F4CB  MCA - Generate Report", use_container_width=True, key="qkx_card_mca_report"):
            _qkx_go("input", "MCA", report_only=True)
        if st.button("\U0001F50D  MCA - Parameter Verification", use_container_width=True, key="qkx_card_mca_pv"):
            _qkx_go("paramcheck", "MCA")
    with c2:
        st.caption("Nokia to Ericsson site integration")
        if st.button("N2E", use_container_width=True, key="qkx_card_n2e"):
            _qkx_go("input", "N2E")
        if st.button("\U0001F4CB  N2E - Generate Report", use_container_width=True, key="qkx_card_n2e_report"):
            _qkx_go("input", "N2E", report_only=True)
        if st.button("\U0001F50D  N2E - Parameter Verification", use_container_width=True, key="qkx_card_n2e_pv"):
            _qkx_go("paramcheck", "N2E")
    with c3:
        st.caption("New site build")
        if st.button("NSB", use_container_width=True, key="qkx_card_nsb"):
            _qkx_go("input", "NSB")
        if st.button("\U0001F4CB  NSB - Generate Report", use_container_width=True, key="qkx_card_nsb_report"):
            _qkx_go("input", "NSB", report_only=True)
        if st.button("\U0001F50D  NSB - Parameter Verification", use_container_width=True, key="qkx_card_nsb_pv"):
            _qkx_go("paramcheck", "NSB")

    st.divider()
    st.subheader("Instructions")
    st.markdown("""
    1. **Please select your SOW to continue.**
    2. **Upload** the CIQ and EDP & Pre-checks (optional) for the site
    3. **Enter** your User ID
    4. Click **Generate templates**
    5. **Review**, then **download**
    """)

# ---- FAMILY CHOICE (MCA / CENM / CRAN) ----
elif st.session_state.qkx_page == "family":
    if st.button("← Back"):
        _qkx_go("home")
    st.subheader("Choose scope")
    f1, f2, f3 = st.columns(3)
    with f1:
        if st.button("MCA", use_container_width=True, key="qkx_fam_mca"):
            _qkx_go("input", "MCA")
    with f2:
        if st.button("CENM", use_container_width=True, key="qkx_fam_cenm"):
            _qkx_go("input", "CENM")
    with f3:
        if st.button("CRAN", use_container_width=True, key="qkx_fam_cran"):
            _qkx_go("input", "CRAN")

# ---- INPUT PAGE (all scopes land here — same form + results as before) ----
elif st.session_state.qkx_page == "input":
    top_scope = st.session_state.qkx_scope
    if not top_scope:
        st.warning("No scope selected.")
        if st.button("← Back to home"):
            _qkx_go("home")
        st.stop()

    back_target = "family" if top_scope in ("MCA", "CENM", "CRAN") else "home"

    col_left, col_right = st.columns([2, 3])

    with col_left:
        if st.button("← Back"):
            _qkx_go(back_target)

        hide_inputs = st.session_state.get("qkx_report_only") and "qkx_results" in st.session_state
        if not hide_inputs:
            st.subheader(f"Inputs — {top_scope}")
            with st.container(border=True):
                cran_sub = None
                if top_scope == "CRAN":
                    cran_sub = st.selectbox("CRAN scope", ["CRAN SA Rehome Trip 1", "CRAN SA Rehome Trip 2", "CRAN NSA Rehome"])

                ciq_file = st.file_uploader("CIQ (.xlsx / .xls)", type=["xlsx", "xls"])
                edp_file = st.file_uploader("EDP (.xlsx / .xls)", type=["xlsx", "xls"])
                pre_file = None
                if top_scope not in ("N2E", "NSB"):
                    pre_file = st.file_uploader("Pre-checks (.pdf) — optional", type=["pdf"])
                post_file, controller_file = None, None
                if st.session_state.get("qkx_report_only"):
                    # Post-checks and controller-checks: N2E's GPS/SUP/XMU/controller-checks
                    # logic needs these too, confirmed this session — gated on report_only
                    # so plain template generation (no report) doesn't ask for them
                    # unnecessarily.
                    post_file = st.file_uploader("Post-checks (.pdf) — required for report checks", type=["pdf"])
                    controller_file = st.file_uploader(
                        "6610 Controller checks (.pdf) — required only if a 6610 is present",
                        type=["pdf"])
                c1, c2 = st.columns(2)
                with c1:
                    user_id = st.text_input("User ID", placeholder="e.g. pr970b")
                with c2:
                    date_str = st.text_input("Execution date (mmddyyyy)", value=date.today().strftime("%m%d%Y"))
                _report_ready = (ciq_file and edp_file and (post_file if st.session_state.get("qkx_report_only") else True))
                run = st.button("Generate Report \u2192" if st.session_state.get("qkx_report_only") else "Generate templates \u2192",
                                 type="primary", disabled=not _report_ready)
        else:
            run = False

    if run or "qkx_results" in st.session_state:
        report_only = st.session_state.get("qkx_report_only")
        with col_left:
            if report_only:
                ph_log = st.empty()
            else:
                log_card = st.container(border=True)
                with log_card:
                    ph_log = st.empty()
            ph_checks_bottom = st.empty()

        with col_right:
            ph_checks_top = st.empty()
            if not report_only:
                ph_prepost = st.container(border=True)
                ph_sow = st.container(border=True)
                ph_siad = st.container(border=True)
                ph_summary = st.container(border=True)
                ph_outputs = st.container(border=True)
            else:
                ph_prepost = ph_sow = ph_siad = ph_summary = ph_outputs = st.empty()

        if run:
            log_lines = []

            def log(msg):
                log_lines.append(msg)
                ph_log.code("\n".join(log_lines) or "Processing...", language=None)

            log("Starting...")

            _all_templates = {
                "MMBB": TPL_MMBB, "TMBB": TPL_TMBB, "cENM": TPL_CENM, "6610": TPL_6610,
                "CRAN Trip-1": TPL_CRAN_TRIP1, "CRAN Trip-2": TPL_CRAN_TRIP2, "CRAN NSA": TPL_CRAN_NSA,
                "DSS 4-sector": TPL_DSS_4SECTOR, "DSS 3-sector": TPL_DSS_3SECTOR,
            }
            _missing = [f"{label}  (expected: `{path.name}`)" for label, path in _all_templates.items() if not path.exists()]
            if _missing:
                st.error(
                    "Some template files aren't in `templates/MCA/` in the repo — check the exact filenames match "
                    "(GitHub sometimes changes spacing/characters on manual upload):\n\n"
                    + "\n".join(f"- {m}" for m in _missing)
                )
                st.stop()

            log("Reading CIQ workbook...")
            try:
                ciq_wb = load_workbook_any(ciq_file.read(), ciq_file.name)
            except Exception as e:
                st.error(f"This CIQ couldn't be read as either .xlsx or legacy .xls. "
                          f"It may be corrupted, or its content doesn't match its extension — try re-saving "
                          f"it as .xlsx in Excel and re-uploading. Error detail: {e}")
                st.stop()
            if "Mixed Mode Info" not in ciq_wb.sheetnames:
                st.error('Could not find a "Mixed Mode Info" tab in the CIQ.')
                st.stop()
            mm_objs = sheet_objs(ciq_wb["Mixed Mode Info"])
            controller_objs = sheet_objs(ciq_wb["Controller Info"]) if "Controller Info" in ciq_wb.sheetnames else []

            log("Reading EDP workbook...")
            edp_bytes = edp_file.read()
            try:
                edp_wb = load_workbook_any(edp_bytes, edp_file.name)
            except Exception as e:
                st.error(f"This EDP couldn't be read (tried both .xlsx and legacy .xls handling). "
                          f"Try re-saving it as .xlsx in Excel and re-uploading. Error detail: {e}")
                st.stop()
            edp_index = build_edp_index(edp_wb)
            if not edp_index:
                st.error('Could not locate the EDP header row (expected a column "EDP_SITE_ID" and "SITE_NAME").')
                st.stop()

            precheck_text = ""
            if pre_file:
                log("Extracting Pre-checks PDF text...")
                precheck_text = extract_pdf_text(pre_file.read())

            postcheck_text = ""
            if post_file:
                log("Extracting Post-checks PDF text...")
                postcheck_text = extract_pdf_text(post_file.read())

            controller_checks_text = ""
            if controller_file:
                log("Extracting 6610 Controller checks PDF text...")
                controller_checks_text = extract_pdf_text(controller_file.read())

            pre_line = post_line = None
            uid = user_id or "xxUserIDxx"
            dstr = date_str or "xxDatexx"

            if top_scope == "MCA":
                summary_rows, pre_line, post_line, siad_rows, outputs, binary_outputs, scope_lines = generate_mca(
                    ciq_wb, edp_index, controller_objs, mm_objs, uid, dstr, precheck_text, log)
            elif top_scope == "CENM":
                summary_rows, pre_line, post_line, siad_rows, outputs, binary_outputs, scope_lines = generate_cenm(
                    ciq_wb, edp_index, controller_objs, mm_objs, uid, dstr, precheck_text, log)
            elif top_scope == "N2E":
                summary_rows, pre_line, post_line, siad_rows, outputs, binary_outputs, scope_lines = generate_n2e(
                    ciq_wb, edp_index, controller_objs, mm_objs, uid, dstr, log)
            elif top_scope == "NSB":
                summary_rows, pre_line, post_line, siad_rows, outputs, binary_outputs, scope_lines = generate_nsb(
                    ciq_wb, edp_index, controller_objs, mm_objs, uid, dstr, log)
            else:  # CRAN
                cran_opts = {
                    "CRAN SA Rehome Trip 1": (TPL_CRAN_TRIP1, False, False, "CRAN_Trip1"),
                    "CRAN SA Rehome Trip 2": (TPL_CRAN_TRIP2, True, False, "CRAN_Trip2"),
                    "CRAN NSA Rehome": (TPL_CRAN_NSA, True, True, "CRAN_NSA"),
                }
                tpl_path, inc_src, need_6673, out_name = cran_opts[cran_sub]
                summary_rows, pre_line, post_line, siad_rows, outputs, binary_outputs, scope_lines = generate_cran(
                    ciq_wb, edp_index, controller_objs, mm_objs, uid, dstr, precheck_text, log,
                    tpl_path, inc_src, need_6673, out_name)

            log("Done.")

            binary_outputs += get_universal_static_outputs(ciq_wb, mm_objs, log)

            st.session_state.qkx_results = {
                "top_scope": top_scope, "scope_lines": scope_lines, "pre_line": pre_line, "post_line": post_line,
                "siad_rows": siad_rows, "summary_rows": summary_rows, "outputs": outputs, "binary_outputs": binary_outputs,
                "log_lines": log_lines, "mm_objs": mm_objs, "controller_objs": controller_objs, "ciq_wb": ciq_wb,
                "precheck_text": precheck_text, "postcheck_text": postcheck_text,
                "controller_checks_text": controller_checks_text, "edp_index": edp_index,
                "uid": uid, "dstr": dstr,
            }
            if not st.session_state.get("qkx_report_only"):
                render_checks_panel_animated(ph_checks_top, top_scope, scope_lines)
                render_checks_panel_static(ph_checks_bottom, top_scope, scope_lines)
            else:
                ph_log.empty()
                st.rerun()
        else:
            r = st.session_state.qkx_results
            top_scope, scope_lines = r["top_scope"], r["scope_lines"]
            pre_line, post_line = r["pre_line"], r["post_line"]
            siad_rows, summary_rows = r["siad_rows"], r["summary_rows"]
            outputs, binary_outputs = r["outputs"], r["binary_outputs"]
            mm_objs, controller_objs, ciq_wb = r["mm_objs"], r["controller_objs"], r["ciq_wb"]
            precheck_text = r["precheck_text"]
            postcheck_text = r.get("postcheck_text", "")
            controller_checks_text = r.get("controller_checks_text", "")
            edp_index = r.get("edp_index")
            uid = r.get("uid") or "xxUserIDxx"
            dstr = r.get("dstr") or "xxDatexx"
            if not report_only:
                ph_log.code("\n".join(r["log_lines"]), language=None)
            else:
                ph_log.empty()
            ph_checks_top.empty()
            if not st.session_state.get("qkx_report_only"):
                render_checks_panel_static(ph_checks_bottom, top_scope, scope_lines)

        if not st.session_state.get("qkx_report_only"):
            with ph_prepost:
                if pre_line and post_line:
                    st.subheader("Pre / Post configuration")
                    st.code(f"Pre Configuration:  {pre_line}\nPost Configuration: {post_line}", language=None)

            with ph_sow:
                if scope_lines:
                    st.subheader("Scope of work summary")
                    st.code("\n".join(scope_lines_to_readable_text(scope_lines)), language=None)
                    with st.expander("Copy tab-separated version for Excel/Notepad"):
                        st.text_area("Tab-separated (select all, copy, paste into Excel — lands in columns)",
                                      "\n".join(scope_lines), height=150, key="sow_raw")

            with ph_siad:
                if siad_rows:
                    st.subheader("SIAD port assignment")
                    st.dataframe(pd.DataFrame(siad_rows), use_container_width=True, hide_index=True)

            with ph_summary:
                st.subheader("Extraction summary")
                st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

            with ph_outputs:
                if outputs:
                    st.subheader("Generated output")
                    for name, text in outputs:
                        unresolved = highlight_unresolved(text)
                        with st.expander(f"{name}  ({str(len(unresolved)) + ' unresolved' if unresolved else 'fully resolved'})"):
                            st.text_area("Preview", text, height=300, key=name)
                            st.download_button("Download .txt", text, file_name=name, key=f"dl_{name}")

                if binary_outputs:
                    st.subheader("Excel outputs")
                    for name, data in binary_outputs:
                        st.download_button(f"Download {name}", data, file_name=name, key=f"dl_bin_{name}")

                if outputs or binary_outputs:
                    if len(outputs) + len(binary_outputs) > 1:
                        zip_buf = io.BytesIO()
                        with zipfile.ZipFile(zip_buf, "w") as zf:
                            for name, text in outputs:
                                zf.writestr(name, text)
                            for name, data in binary_outputs:
                                zf.writestr(name, data)
                        st.download_button("Download all as .zip", zip_buf.getvalue(), file_name="generated_templates.zip")

        if top_scope == "MCA" and st.session_state.get("qkx_report_only"):
            import importlib
            # Confirmed bug class (already documented in this project's own handoff notes):
            # a plain "import X" reuses whatever's cached in sys.modules from an earlier
            # session even after the file on disk changed — reloading only mca_report_ui
            # itself doesn't refresh ITS dependencies. Every module mca_report_ui imports
            # needs its own explicit reload, in dependency order (leaves first).
            import report_detect, mca_checklist, mca_glue, mca_report_text, mca_completed_logic, mca_xlsm_surgical
            importlib.reload(report_detect)
            importlib.reload(mca_checklist)
            importlib.reload(mca_glue)
            importlib.reload(mca_report_text)
            importlib.reload(mca_completed_logic)
            importlib.reload(mca_xlsm_surgical)
            import mca_report_ui
            importlib.reload(mca_report_ui)
            mca_report_ui.render(sys.modules[__name__], ciq_wb, mm_objs, controller_objs, precheck_text, pre_line, post_line, scope_lines,
                                  postcheck_text=postcheck_text, controller_checks_text=controller_checks_text, edp_index=edp_index)

        if top_scope == "N2E" and st.session_state.get("qkx_report_only"):
            import importlib
            import mca_completed_logic, n2e_completed_logic, mca_xlsm_surgical
            importlib.reload(mca_completed_logic)
            importlib.reload(n2e_completed_logic)
            importlib.reload(mca_xlsm_surgical)
            import n2e_report_ui
            importlib.reload(n2e_report_ui)
            n2e_report_ui.render(sys.modules[__name__], ciq_wb, mm_objs, controller_objs, edp_index, uid, dstr,
                                  postcheck_text=postcheck_text, controller_checks_text=controller_checks_text)

        if top_scope == "NSB" and st.session_state.get("qkx_report_only"):
            import importlib
            import mca_completed_logic, nsb_completed_logic, mca_xlsm_surgical
            importlib.reload(mca_completed_logic)
            importlib.reload(nsb_completed_logic)
            importlib.reload(mca_xlsm_surgical)
            import nsb_report_ui
            importlib.reload(nsb_report_ui)
            nsb_report_ui.render(sys.modules[__name__], ciq_wb, mm_objs, controller_objs, edp_index, uid, dstr,
                                  postcheck_text=postcheck_text, controller_checks_text=controller_checks_text)


# ---- PARAMETER VERIFICATION (MCA / N2E / NSB) ----


# ---- PARAMETER VERIFICATION (MCA / N2E / NSB) ----
elif st.session_state.qkx_page == "paramcheck":
    pv_scope = st.session_state.qkx_scope
    pv_has_pre = pv_scope not in ("N2E", "NSB")
    if st.button("← Back", key="pv_back"):
        _qkx_go("home")
    st.subheader(f"Parameter Verification — {pv_scope}")
    if pv_has_pre:
        st.caption("Compares CIQ vs Pre logs vs Onsite logs, flags mismatches. "
                   "Upload the CIQ, then all Pre logs and Onsite logs for every node at the site — "
                   "each file is matched to its node automatically by filename.")
    else:
        st.caption("Compares CIQ vs Onsite logs, flags mismatches. "
                   f"{pv_scope} has no pre-existing state to compare against, so no Pre log is needed. "
                   "Upload the CIQ, then all Onsite logs for every node at the site — "
                   "each file is matched to its node automatically by filename.")

    with st.container(border=True):
        pv_ciq_file = st.file_uploader("CIQ (.xlsx / .xls)", type=["xlsx", "xls"], key="pv_ciq")
        if pv_has_pre:
            pv_pre_files = st.file_uploader("Pre logs (.log / .txt) — one or more, one per node",
                                             type=["log", "txt"], accept_multiple_files=True, key="pv_pre")
        else:
            pv_pre_files = []
        pv_onsite_files = st.file_uploader("Onsite logs (.log / .txt) — one or more, one per node",
                                            type=["log", "txt"], accept_multiple_files=True, key="pv_onsite")
        pv_ready = pv_ciq_file and pv_onsite_files and (pv_pre_files if pv_has_pre else True)
        pv_run = st.button("Verify parameters", type="primary", key="pv_run", disabled=not pv_ready)

    if pv_run:
        with st.spinner("Parsing logs and comparing against CIQ..."):
            import openpyxl
            pv_wb = openpyxl.load_workbook(io.BytesIO(pv_ciq_file.getvalue()), data_only=True)
            pv_ws = pv_wb["Mixed Mode Info"]
            pv_rows = list(pv_ws.iter_rows(values_only=True))
            pv_headers = [str(h).strip() if h is not None else "" for h in pv_rows[0]]
            pv_mm_objs = [dict(zip(pv_headers, r)) for r in pv_rows[1:] if any(c is not None for c in r)]

            pv_pre_inputs = [(f.name, f.getvalue().decode("utf-8", errors="replace")) for f in pv_pre_files]
            pv_onsite_inputs = [(f.name, f.getvalue().decode("utf-8", errors="replace")) for f in pv_onsite_files]

            pv_result = run_parameter_verification(pv_wb, pv_mm_objs, pv_pre_inputs, pv_onsite_inputs, scope=pv_scope)

            pv_fa_code = None
            if "5G Info" in pv_wb.sheetnames:
                pv_ws5g = pv_wb["5G Info"]
                pv_hdr5g = [c.value for c in pv_ws5g[1]]
                for row in pv_ws5g.iter_rows(min_row=2, values_only=True):
                    d = dict(zip(pv_hdr5g, row))
                    if d.get("FA Code"):
                        pv_fa_code = d.get("FA Code")
                        break
            pv_site_ids = [r.get("Node to be built as") for r in pv_mm_objs if r.get("Node to be built as")]

            st.session_state.pv_results = pv_result
            st.session_state.pv_scope_ran = pv_scope
            st.session_state.pv_fa_code = pv_fa_code
            st.session_state.pv_site_ids = pv_site_ids

    if st.session_state.get("pv_results") and st.session_state.get("pv_scope_ran") == pv_scope:
        pv_result = st.session_state.pv_results

        st.markdown(f"**FA Code:** {st.session_state.get('pv_fa_code') or 'N/A'}  |  "
                    f"**Site ID(s):** {', '.join(st.session_state.get('pv_site_ids', [])) or 'N/A'}")

        if pv_result["unmatched_pre"] or pv_result["unmatched_onsite"]:
            with st.expander("⚠ Files that didn't match any node in the CIQ", expanded=True):
                for fn in pv_result["unmatched_pre"]:
                    st.write(f"Pre log: **{fn}** — no matching node found in Mixed Mode Info")
                for fn in pv_result["unmatched_onsite"]:
                    st.write(f"Onsite log: **{fn}** — no matching node found in Mixed Mode Info")

        if pv_result["nodes_missing_pre"] or pv_result["nodes_missing_onsite"]:
            with st.expander("⚠ CIQ nodes with missing logs", expanded=True):
                for n in pv_result["nodes_missing_pre"]:
                    st.write(f"**{n}** — no Pre log uploaded")
                for n in pv_result["nodes_missing_onsite"]:
                    st.write(f"**{n}** — no Onsite log uploaded")

        total_green = total_amber = total_red = 0
        for node, res in pv_result["node_results"].items():
            for group in ("lte", "nr", "naming_lte", "naming_nr"):
                for cell_id, results in res.get(group, []):
                    for r in results:
                        if r["color"] == "green":
                            total_green += 1
                        elif r["color"] == "amber":
                            total_amber += 1
                        elif r["color"] == "red":
                            total_red += 1

        m1, m2, m3 = st.columns(3)
        m1.metric("Matches", total_green)
        m2.metric("Expected changes", total_amber)
        m3.metric("Mismatches", total_red)

        PV_COLOR_HEX = {"green": "#C6EFCE", "red": "#FFC7CE", "amber": "#FFEB9C", "gray": "#F2F2F2"}

        def render_wide_naming(wide_rows, label, keys):
            if not wide_rows:
                return
            st.markdown(f"**{label}**")
            display_rows = []
            colors_by_row = []
            for row in wide_rows:
                d = {"Sector": row["sector"]}
                row_colors = {}
                for disp, kp in keys:
                    d[f"{disp} (CIQ)"] = _pv_display(row.get(f"{kp}_ciq"))
                    d[f"{disp} (On site)"] = _pv_display(row.get(f"{kp}_post"))
                    row_colors[f"{disp} (CIQ)"] = row.get(f"{kp}_color")
                    row_colors[f"{disp} (On site)"] = row.get(f"{kp}_color")
                d["Comments"] = row.get("comment", "")
                display_rows.append(d)
                colors_by_row.append(row_colors)
            df = pd.DataFrame(display_rows)

            def _style(row):
                rc = colors_by_row[row.name]
                return [f"background-color: {PV_COLOR_HEX.get(rc.get(col), 'white')}" if col in rc else "" for col in row.index]

            st.dataframe(df.style.apply(_style, axis=1), use_container_width=True, hide_index=True)

        def render_wide_params(wide_rows, label, keys):
            if not wide_rows:
                return
            st.markdown(f"**{label}**")
            display_rows = []
            colors_by_row = []
            for row in wide_rows:
                d = {"Sector": row["sector"]}
                row_colors = {}
                for disp, kp in keys:
                    if pv_has_pre:
                        _pre_txt = _pv_display(row.get(f"{kp}_pre"))
                        d[f"{disp} (Pre)"] = _pre_txt
                        # An NA Pre cell is amber wherever it appears — a newly-added sector has
                        # no Pre state, and a missing one is not evidence of a match.
                        row_colors[f"{disp} (Pre)"] = "amber" if _pre_txt == "NA" else row.get(f"{kp}_color")
                    d[f"{disp} (CIQ)"] = _pv_display(row.get(f"{kp}_ciq"))
                    d[f"{disp} (On site)"] = _pv_display(row.get(f"{kp}_post"))
                    row_colors[f"{disp} (CIQ)"] = row.get(f"{kp}_color")
                    row_colors[f"{disp} (On site)"] = row.get(f"{kp}_color")
                d["Comments"] = row.get("comment", "")
                display_rows.append(d)
                colors_by_row.append(row_colors)
            df = pd.DataFrame(display_rows)

            def _style(row):
                rc = colors_by_row[row.name]
                return [f"background-color: {PV_COLOR_HEX.get(rc.get(col), 'white')}" if col in rc else "" for col in row.index]

            st.dataframe(df.style.apply(_style, axis=1), use_container_width=True, hide_index=True)

        NAMING_LTE_KEYS = [("FDD naming", "FDD_naming"), ("Rilinks", "Rilinks"),
                           ("Sector carrier", "sector_carrier"), ("ISDLONLY", "ISDLONLY")]
        NAMING_NR_KEYS = [("Rilinks", "Rilinks"), ("Sector carrier", "sector_carrier"),
                          ("NRCELL CU naming", "NRCELL_CU_naming")]
        LTE_KEYS_1 = [("EarfcnDL", "EarfcnDL"), ("EarfcnUL", "EarfcnUL"), ("TX", "TX"), ("RX", "RX"),
                      ("Bandwidth", "Bandwidth"), ("ConfiguredOutputPower", "ConfiguredOutputPower"), ("CellID", "CellID")]
        LTE_KEYS_2 = [("rachRootSequence", "rachRootSequence"), ("PCI", "PCI"),
                      ("Cellrange", "Cellrange"), ("TAC", "TAC")]
        NR_KEYS_1 = [("arfcnDL", "arfcnDL"), ("arfcnUL", "arfcnUL"), ("bSChannelBwDL", "bSChannelBwDL"),
                     ("bSChannelBwUL", "bSChannelBwUL"), ("ConfiguredOutputPower", "ConfiguredOutputPower"),
                     ("cellLocalId", "cellLocalId"), ("ssbFrequency", "ssbFrequency")]
        NR_KEYS_2 = [("rachRootSequence", "rachRootSequence"), ("nRPCI", "nRPCI"),
                     ("Cellrange", "Cellrange"), ("NRTAC", "NRTAC"), ("nCI", "nCI")]

        # Site-wide, not node-wise — same as the PDF: every 4G sector in one set of tables and
        # every 5G sector in another, sorted by sector name, regardless of owning node.
        pv_naming_lte = _pv_collect(pv_result["node_results"], "_wide_naming_lte")
        pv_naming_nr = _pv_collect(pv_result["node_results"], "_wide_naming_nr")
        pv_lte_rows = _pv_collect(pv_result["node_results"], "_wide_lte")
        pv_nr_rows = _pv_collect(pv_result["node_results"], "_wide_nr")

        render_wide_naming(pv_naming_lte, "4G Naming & Configuration", NAMING_LTE_KEYS)
        render_wide_naming(pv_naming_nr, "5G Naming & Configuration", NAMING_NR_KEYS)
        render_wide_params(pv_lte_rows, "4G Sectors — Earfcn / TX-RX / Power / Cell ID", LTE_KEYS_1)
        render_wide_params(pv_nr_rows, "5G Sectors — arfcn / Bandwidth / Power / cellLocalId / SSB", NR_KEYS_1)
        render_wide_params(pv_lte_rows, "4G Sectors — rachRootSequence / PCI / Cellrange / TAC", LTE_KEYS_2)
        render_wide_params(pv_nr_rows, "5G Sectors — rachRootSequence / nRPCI / Cellrange / NRTAC / nCI", NR_KEYS_2)

        pv_pdf_bytes = build_parameter_verification_pdf(
            pv_scope, pv_result["node_results"], has_pre=pv_has_pre,
            fa_code=st.session_state.get("pv_fa_code"), site_ids=st.session_state.get("pv_site_ids"))
        st.download_button("Download PDF report", pv_pdf_bytes,
                            file_name=f"{pv_scope}_Parameter_Verification.pdf",
                            mime="application/pdf", key="pv_dl_pdf")
