"""
MCA Integration Report — Completed section, new logic built in this session.

Design principle throughout: reuse QUICKIX's existing, already-tested functions
(classify_carriers, band_label, extract_precheck_sectors, generate_port_conversion_checks,
PORT_BY_GEN, DU_TYPE_TO_GEN, sheet_objs, is_populated, etc. — all imported from app.py)
rather than re-deriving anything QUICKIX already knows how to do. Only genuinely new
data sources (GPS Status, Transport SFP, Sidehaul Info, controller-checks PDF, the
Calltest market table) get new parsers here.

Every function below is traceable to a specific confirmed decision in this session's
transcript — see the inline comments for the "why", not just the "what".
"""

import re
from pathlib import Path
import streamlit as st

# IMPORTANT: do NOT "import app as qx" at module level. app.py has Streamlit UI code
# directly at module scope (not wrapped in `if __name__ == "__main__"`), so importing it
# fresh here re-executes the ENTIRE Streamlit script a second time within the same run —
# recreating every widget (e.g. the "← Back" button) and crashing with
# StreamlitDuplicateElementId. This is exactly why the existing codebase passes `app` as
# a parameter into render(app, ...) instead of importing it — same pattern followed here.
qx = None


def set_app_module(app_module):
    """Called once by mca_report_ui.render() with the already-loaded `app` module it
    received as its own parameter — never a fresh import."""
    global qx
    qx = app_module


def _normalize(text):
    """Pre/Post-checks text comes from two genuinely different extraction paths depending
    on the uploaded file's real format (confirmed this session: some "PDF" uploads are
    actually zip archives with a clean pre-extracted text layer, one line per row; others
    are real PDFs run through pdftotext -layout, which pads columns with runs of spaces
    and inserts blank lines between a table's header and its data). Every parser below
    needs to tolerate both. Collapsing all horizontal whitespace to single spaces and
    dropping blank lines makes the two formats equivalent before any table regex runs."""
    if not text:
        return text
    lines = [re.sub(r'[ \t]+', ' ', ln).strip() for ln in text.splitlines()]
    return "\n".join(ln for ln in lines if ln)


# ============================================================
# SECTION 1 — NEW DATA-SOURCE PARSERS
# Every regex here is built and verified against REAL Pre/Post-checks text extracted
# from real project files during this session (ECL02586, SCL01706), not guessed.
# ============================================================

def extract_gps_status(check_text):
    """Parses the 'GPS Status' table (present in both Pre- and Post-checks — same format,
    confirmed against real ECL02586 data in both).

        GPS Status
        Node       Product Designation  Product No.    Product Rev.
        ECL02586   GRU 05 01             NCD 901 89/1   R1D

    Returns {node: product_designation}. 'Product Designation' is what the person
    referred to as 'GPS type'."""
    out = {}
    text = _normalize(check_text)
    if not text:
        return out
    # Confirmed real gap: multi-node Post-/Pre-checks documents can have a SEPARATE "GPS
    # Status" table per node section, not one combined table for the whole site — re.search
    # (single match) only ever found the FIRST one, silently missing every other node's GPS
    # entry. finditer() now merges every block found.
    row_re = re.compile(
        r'^(\S+) ((?:GPS|GRU) \d{2} \d{2}) (\S+(?: \S+)*?/\S+) (\S+)$', re.M)
    for m in re.finditer(r'GPS Status\nNode Product Designation Product No\. Product Rev\.\n(.*?)(?=\n[A-Z][\w /]*\n|\Z)',
                          text, re.S):
        block = m.group(1)
        # Row shape: "NODE  <product designation, 1-3 space-separated tokens>  <product no>  <rev>"
        # Confirmed real values: "GRU 05 01", "GPS 02 01", "GPS 03 01", "GRU 04 01" — always
        # a 2-letter prefix + two 2-digit groups. Product No. always contains a "/" (e.g.
        # "NCD 901 89/1"); Product Rev. is the final bare token (e.g. "R1D").
        for m2 in row_re.finditer(block):
            node, designation, _prodno, _rev = m2.groups()
            out[node.strip()] = designation.strip()
    return out


KNOWN_GPS_TYPES = {"GPS 02 01", "GPS 03 01", "GRU 04 01", "GRU 05 01"}
# "Most likely a GPS 01" — confirmed: not a fallback VALUE, just the person's own label for
# "whatever shows up that isn't one of the 4 known types." We never emit "GPS 01" as a
# value; we just treat anything outside KNOWN_GPS_TYPES as unconfirmed.


@st.cache_data
def extract_sync_status_2(post_text):
    """Parses 'Status of Synchronization 2' — confirmed real structure (ECL02586, SCL01706):

        Status of Synchronization 2
        Node       MO                                            OpState
        ECL02586   Equipment=1,FRU=1,SyncPort=1                  ENABLED
        ECL02586   Transport=1,Synchronization=1,TimeSyncIO=1    ENABLED

    Returns {node: opstate} for the TimeSyncIO=1 row specifically — confirmed this row
    reflects GPS sync status, not the SyncPort row (which we ignore per instruction).
    Confirmed real bug found against real N2E data (WAWN094133/WAWN084133): the value
    after 'TimeSyncIO=' isn't always literally '1' — real example shows
    'TimeSyncIO=GPS ENABLED'. Regex now accepts any sync-source identifier there."""
    out = {}
    text = _normalize(post_text)
    if not text:
        return out
    for m in re.finditer(
            r'(\S+) Transport=1,Synchronization=1,TimeSyncIO=\S+ (ENABLED|DISABLED)', text):
        node, opstate = m.groups()
        out[node.strip()] = opstate
    return out


@st.cache_data
def extract_transport_sfp(post_text):
    """Parses the 'Transport SFP' table — confirmed real structure (ECL02586):

        Transport SFP
        Node      BOARD     PORT  VENDOR    VENDORPROD      REV   SERIAL      DATE      ERICSSONPROD  TEMP  TXbs  TXdBm  RXdBm  BER
        ECL02586  RANP6672  IC    ERICSSON  SPP10ELRIDFKEN  0101  EB30728343  20240621  RDH10265/3    R1A   46C   30%   -2.02  -1.71  0/0

    Confirmed real-data quirk: ERICSSONPROD often carries a trailing revision token
    (e.g. "RDH10265/3 R1A") that pushes the row to 15 space-separated fields against a
    14-column header — same class of PDF-column-misalignment issue already documented
    elsewhere in this project (DL_UL_LOSS_ROW_RE, etc.). Handled by capturing
    ERICSSONPROD non-greedily up to the point where TEMP (a bare "<n>C" token) appears.

    Returns {node: {"ericssonprod": str, "txdbm": float, "rxdbm": float, "ber": str}}."""
    out = {}
    text = _normalize(post_text)
    if not text:
        return out
    # Confirmed real-data quirk: some sites' header stops at ERICSSONPROD, omitting
    # "TEMP TXbs TXdBm RXdBm BER" from the header line entirely — even though the data
    # row itself still carries those values. Header's trailing columns made optional
    # to match both variants; the row-level regex already handles the actual values
    # correctly either way.
    m = re.search(
        r'Transport SFP\nNode BOARD PORT VENDOR VENDORPROD REV SERIAL DATE '
        r'ERICSSONPROD(?: TEMP TXbs TXdBm RXdBm BER)?\n(.*?)(?=\nTransport Fiber link Status|\n[A-Z][\w /]*\n|\Z)',
        text, re.S)
    if not m:
        return out
    block = m.group(1)
    row_re = re.compile(
        r'^(\S+) (\S+) (\S+) (\S+) (\S+) (\S+) (\S+) (\d{8}) '
        r'(.+?) \d+C \d+% (-?\d+\.\d+) (-?\d+\.\d+) (\S+)$',
        re.M)
    for m2 in row_re.finditer(block):
        node = m2.group(1)
        ericssonprod = m2.group(9).strip()
        txdbm, rxdbm, ber = float(m2.group(10)), float(m2.group(11)), m2.group(12)
        out[node.strip()] = {"ericssonprod": ericssonprod, "txdbm": txdbm, "rxdbm": rxdbm, "ber": ber}
    return out


def extract_transport_fiber_opmode_for_node(check_text, node, port_labels):
    """Thin wrapper around QUICKIX's existing extract_transport_fiber_opmode — confirmed
    reusable as-is against Post-checks text (it's generic text parsing, not hardcoded to
    'pre'). Kept as a separate name here only for readability at call sites."""
    return qx.extract_transport_fiber_opmode(check_text, node, port_labels)


def extract_sidehaul_info(ciq_wb):
    """CIQ 'Sidehaul Info' tab — confirmed real structure:

        Switch | SH Switch ID | SH Switch Slot & Port -1 | SH Switch Slot & Port -2 |
        SH Switch Port SFP | Basebands | Baseband Port 1 | Baseband Port 2 | Baseband Port SFP

    One row per switch-port connection — a site can have more than one (confirmed real
    example: FSL00456, 2 rows same switch, different slot/port + baseband).
    'TBD' is a real literal placeholder value some CIQs leave in Baseband Port columns —
    treated as not-populated, same guard class as the NaN-as-"nan" bug fixed elsewhere."""
    rows = []
    if "Sidehaul Info" not in ciq_wb.sheetnames:
        return rows
    for row in qx.sheet_objs(ciq_wb["Sidehaul Info"]):
        switch = row.get("Switch")
        if not qx.is_populated(switch):
            continue
        node_id = row.get("Basebands")
        if str(node_id or "").strip().upper() == "TBD":
            node_id = None
        rows.append({
            "switch_type": switch,
            "switch_id": row.get("SH Switch ID"),
            "slot_port": row.get("SH Switch Slot & Port -1") or row.get("SH Switch Slot & Port -2"),
            "node_id": node_id,
            # Cable part number: confirmed manual — no CIQ/EDP source exists for this.
        })
    return rows


# ---- Controller-checks PDF (separate file from Pre/Post-checks, confirmed real
# structure from FNOC222775_C001_controller_checks.pdf) ----

@st.cache_data
def extract_controller_checks(text):
    """Returns {
        'node_alarm_status': 'OK' | 'NOT OK' | None,
        'controller_state': {'admin': 'UNLOCKED'|'LOCKED', 'fault': bool, 'oper': 'ENABLED'|'DISABLED'} | None,
        'sau_state': {'admin':..., 'fault':..., 'oper':...} | None,
        'alarm_ports': [{'port': str, 'admin': 'UNLOCKED'|'LOCKED', 'slogan': str|None, 'severity': str|None}, ...],
    }"""
    result = {"node_alarm_status": None, "controller_state": None, "sau_state": None, "alarm_ports": []}
    if not text:
        return result

    m = re.search(r'Node Alarm Status\s*[:\s]*\n?.*?\b(OK|NOT OK)\b', text)
    # Confirmed real line: "FNOC222775_C001 OK CXP2010233/2_R26C12 107.253.119.42" under
    # "Node Alarm Status Current Upgrade Package IP Address" header.
    m2 = re.search(r'Node Alarm Status Current Upgrade Package IP Address\r?\n\S+\s+(OK|NOT OK)', text)
    if m2:
        result["node_alarm_status"] = m2.group(1)

    def _fru_state(fru_name):
        # Confirmed real data has (at least) two genuine format variants for the
        # number of status fields before the timestamp/product info — some sites have
        # 3 (administrativeState, faultIndicator, operationalIndicator), others have 5
        # (administrativeState, faultIndicator, isSharedWithExternalMe,
        # maintenanceIndicator, operationalIndicator). Rather than assuming a fixed
        # count, admin/fault are matched at their fixed positions, then the regex
        # specifically searches for whichever field is genuinely "N (ENABLED)" or
        # "N (DISABLED)" — the one distinctive value type among the intermediate
        # fields (UNLOCKED/LOCKED, OFF/ON, STEADY_ON, NOT_APPLICABLE, etc.) — instead
        # of a hardcoded field count that breaks whenever the format varies.
        # Confirmed real gap (FCWC138613_C001_controller_checks.pdf, FCL05583): the fault
        # field isn't always ON/OFF — some sites' SAU row has "0 (LOCKED) 1 (NOT_AVAILABLE)
        # 0 (DISABLED)" instead, which the old ON|OFF-only match silently failed on,
        # returning None (and, downstream, made sau_connections_placement never detect a
        # disabled 6610 SAU at all for that site). Widened to any \w+ value; fault_ok stays
        # conservative (only literal "OFF" counts as no-fault, matching prior behavior for
        # every already-passing case).
        mm = re.search(
            re.escape(f"FieldReplaceableUnit={fru_name}") +
            r'\s+\d+\s+\((UNLOCKED|LOCKED)\)\s+\d+\s+\((\w+)\)'
            r'(?:\s+\d+\s+\(\w+\))*?\s+\d+\s+\((ENABLED|DISABLED)\)',
            text)
        if not mm:
            return None
        admin, fault, oper = mm.groups()
        return {"admin": admin, "fault_ok": fault == "OFF", "oper": oper}

    result["controller_state"] = _fru_state("Controller6610")
    result["sau_state"] = _fru_state("SAU")

    # External alarms table — confirmed real rows (verified against the actual PDF table,
    # column order: MO | activeExternalAlarm | administrativeState | alarmSlogan |
    # normallyOpen | operationalState | perceivedSeverity):
    # "FieldReplaceableUnit=SAU,AlarmPort=1  false 1 (UNLOCKED) RBS PNC DC MJ  false 1 (ENABLED) 3 (MAJOR)"
    # "FieldReplaceableUnit=SAU,AlarmPort=11 true  0 (LOCKED)   RBS HVAC FAIL  false 1 (ENABLED) 2 (CRITICAL)"
    # "FieldReplaceableUnit=SAU,AlarmPort=20 false 0 (LOCKED)                  true  1 (ENABLED) 4 (MINOR)"  <- no slogan
    # Confirmed real bug: the OLD regex never captured activeExternalAlarm at all, and
    # confused operationalState (a static circuit-enable flag — ALWAYS "ENABLED" whether
    # or not there's a real fault, confirmed by inspecting the raw data) for it. That's
    # why every scripted port was being reported as "active" regardless of its real state.
    port_re = re.compile(
        r'FieldReplaceableUnit=SAU,AlarmPort=(\d+)\s+(true|false)\s+\d+\s+\((\w+)\)\s*'
        r'([A-Z][A-Z0-9 +]*?)?\s*(true|false)\s+\d+\s+\((\w+)\)\s+\d+\s+\((\w+)\)')
    for m3 in port_re.finditer(text):
        port, active, admin, slogan, _normally_open, oper, severity = m3.groups()
        slogan = slogan.strip() if slogan and slogan.strip() else None
        result["alarm_ports"].append({
            "port": port, "admin": admin, "slogan": slogan, "severity": severity,
            "active": active == "true", "alarm_state": oper,
        })
    return result


# ---- Calltest market table (Calltest_sheet.xlsx, 'Legacy' tab for MCA) ----

@st.cache_data
def load_calltest_table(xlsx_path, tab_name="Legacy"):
    """Parses Calltest_sheet.xlsx into:
        prefix_to_market: {"NC": "NCSC", "EC": "NCSC", ...}
        rules: {(market, scenario): {"PSAP": bool, "LTE Speed test": bool,
                                       "F-net with F-net SIM": bool, "5G speedtest": bool}}
    Confirmed structure: Market column only populated on the first row of each group;
    it applies to every row until the next non-blank Market cell.
    tab_name: confirmed NSB uses its own dedicated "NSB" tab in the same workbook
    (different market-lookup rules from MCA's "Legacy" tab), not the market-based lookup
    N2E skips entirely — pass tab_name="NSB" for NSB's Call Test logic.
    Confirmed cached: this is a static file bundled with the app that never changes
    between reruns, but was previously being re-read from disk on every single call —
    called from multiple places (GPS, Call Test, DSS, Florida) across every render pass,
    since Streamlit reruns the whole script on every interaction."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[tab_name]
    prefix_to_market, rules = {}, {}
    current_market = None
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        market, sectors, psap, lte_speed, fnet, fiveg_speed, site_name = row[:7]
        if market:
            current_market = market
        if not sectors or not current_market:
            continue
        rules[(current_market, sectors.strip())] = {
            "PSAP": str(psap).strip().upper() == "Y",
            "LTE Speed test": str(lte_speed).strip().upper() == "Y",
            "F-net with F-net SIM": str(fnet).strip().upper() == "Y",
            "5G speedtest": str(fiveg_speed).strip().upper() == "Y",
        }
        if site_name:
            for prefix in str(site_name).split("/"):
                prefix_to_market[prefix.strip().upper()] = current_market
    return prefix_to_market, rules


# ============================================================
# SECTION 2 - INTEGRATION: Post-checks presence verification (warning-only, per confirmed
# decision - never alters what the report shows, only feeds the Warnings tab)
# ============================================================

def verify_integration_against_postcheck(classification, post_text):
    """For every cell in classification['added'], check it's PRESENT in Post-checks
    (any admin state - LOCKED counts as integrated, confirmed: lock state doesn't matter,
    only whether the cell exists at all). Returns a list of warning dicts, one per cell
    that's completely absent from Post-checks. Never touches the report itself."""
    post_pairs, _ = qx.extract_precheck_sectors(post_text)
    post_cells = {cell for (_node, cell) in post_pairs}

    warnings = []
    for node, cells in classification.get("added", {}).items():
        for cell in cells:
            if cell not in post_cells:
                label, sector = qx.band_label(cell)
                warnings.append({
                    "type": "integration_missing",
                    "text": f"{label} {sector} ({cell}) sector is missing on : {node}.",
                })
    return warnings


# ============================================================
# SECTION 3 - RETUNE: sector-tracked whole/partial display (fixes the confirmed real gap
# where the built code silently dropped sector info) + eUtran Parameters/Post-checks
# verification
# ============================================================

WHOLE_BAND_SET = {"Alpha", "Beta", "Gamma"}


def classify_retunes_with_sectors(ciq_wb):
    """Re-derives retune events the same way classify_carriers does (same-node
    Sector Del_Movement rows where channel/BW differ) but keeps the sector this time -
    confirmed fix: qx.classify_carriers's own retuned list drops it."""
    out = []
    if "Sector Del_Movement" not in ciq_wb.sheetnames:
        return out
    for r in qx.sheet_objs(ciq_wb["Sector Del_Movement"]):
        src_node, src_sector = r.get("Source Node name"), r.get("Source Sector")
        tgt_node, tgt_sector = r.get("Target Node name"), r.get("Target Sector")
        if not (qx.is_populated(src_node) and qx.is_populated(tgt_node)):
            continue
        if str(tgt_node).strip().upper() == "DELETE":
            continue
        if str(src_node).strip().upper() != str(tgt_node).strip().upper():
            continue
        src_dl = str(r.get("Source channelNumberDL", "")).strip()
        tgt_dl = str(r.get("Target channelNumberDL", "")).strip()
        src_bw = str(r.get("Source Bandwidth", "")).strip()
        tgt_bw = str(r.get("Target Bandwidth", "")).strip()
        if src_dl == tgt_dl and src_bw == tgt_bw:
            continue
        label, sector = qx.band_label(src_sector)
        if not label:
            continue
        out.append({
            "label": label, "sector": sector, "cell": tgt_sector or src_sector,
            "from": f"{src_dl}/{src_bw}", "to": f"{tgt_dl}/{tgt_bw}",
        })
    return out


def format_retunes(retune_events):
    """Groups by (label, from, to) signature, tracking sectors under each."""
    grouped = {}
    for r in retune_events:
        key = (r["label"], r["from"], r["to"])
        grouped.setdefault(key, set()).add(r["sector"])

    lines = []
    for (label, frm, to), sectors in grouped.items():
        if WHOLE_BAND_SET <= sectors:
            sectors_str = " sectors"
        else:
            ordered = sorted(sectors, key=lambda s: qx.SECTOR_ORDER.index(s) if s in qx.SECTOR_ORDER else 99)
            sectors_str = " " + ", ".join(ordered)
        lines.append(f"Retune on:\t{label}{sectors_str}\tFrom:\t{frm}\tTo:\t{to}")
    return lines


def verify_retune_against_checks(ciq_wb, retune_events, post_text):
    """Verification target = CIQ eUtran Parameters/5G Info (confirmed canonical source)
    vs. Post-checks' actual earfcndl/dlChannelBandwidth (LTE) or arfcnDL/bSChannelBwDL
    (5G). Warning-only, never alters the report line itself."""
    warnings = []
    eutran_by_cell = {row.get("EutranCellFDDId"): row for row in qx.sheet_objs(ciq_wb["eUtran Parameters"])} \
        if "eUtran Parameters" in ciq_wb.sheetnames else {}
    fiveg_by_cell = {row.get("NRCellDU"): row for row in qx.sheet_objs(ciq_wb["5G Info"])} \
        if "5G Info" in ciq_wb.sheetnames else {}

    post_lte = {}
    for m in re.finditer(
            r'(\S+)\s+UNLOCKED\s+NOT_BARRED\s+(\d+)\s+(\d+)\s+(\d+)\s+ENABLED', _normalize(post_text)):
        cell, bw, dl, ul = m.groups()
        post_lte[cell] = {"dl": int(dl), "bw": int(bw)}

    for ev in retune_events:
        cell = ev["cell"]
        target_row = eutran_by_cell.get(cell) or fiveg_by_cell.get(cell)
        if not target_row:
            continue
        target_dl = target_row.get("earfcnDl") or target_row.get("arfcnDL")
        target_bw = target_row.get("dlChannelBandwidth") or target_row.get("bSChannelBwDL")
        actual = post_lte.get(cell)
        if not actual:
            warnings.append({"type": "retune_missing",
                              "text": f"Retune {ev['label']} {ev['sector']} ({cell}) not confirmed - "
                                      f"Post-checks has no entry for this cell."})
            continue
        if str(target_dl).strip() != str(actual["dl"]).strip() or int(str(target_bw or 0)) != actual["bw"]:
            warnings.append({"type": "retune_mismatch",
                              "text": f"Retune {ev['label']} {ev['sector']} ({cell}) not confirmed - "
                                      f"Post-checks shows {actual['dl']}/{actual['bw']}, expected {target_dl}/{target_bw}."})
    return warnings


# ============================================================
# SECTION 4 - MOVED SECTORS: target-node-presence-only verification (confirmed: source-node
# presence is NOT checked, per explicit instruction)
# ============================================================

def verify_moved_sectors_against_postcheck(classification, post_text, ciq_wb=None):
    """Confirmed real bug (found against ALL00640/ALL01340 data): classification['moved']
    only carries the SOURCE cell name (src_sector) — but a move can rename the cell as part
    of moving it (confirmed real example: ALL01340_7A_1 -> ALL00640_7A_1, every single move
    in that CIQ was renamed). Checking the source name against the target node's Post-checks
    entries can never match once a rename happened. Fixed to build a rename map straight from
    Sector Del_Movement (same pattern Radio Swap already uses correctly) and check the
    TARGET cell's identity, falling back to the unrenamed name only if no rename map is
    available (ciq_wb not passed) or the specific cell wasn't renamed."""
    post_pairs, _ = qx.extract_precheck_sectors(post_text)
    post_by_node = {}
    for node, cell in post_pairs:
        post_by_node.setdefault(node, set()).add(cell)

    rename_map = {}
    if ciq_wb is not None and "Sector Del_Movement" in ciq_wb.sheetnames:
        for r in qx.sheet_objs(ciq_wb["Sector Del_Movement"]):
            src_sector, tgt_sector = r.get("Source Sector"), r.get("Target Sector")
            tgt_node = r.get("Target Node name")
            if src_sector and tgt_sector and str(tgt_node).strip().upper() != "DELETE":
                rename_map[src_sector] = tgt_sector

    warnings = []
    for mv in classification.get("moved", []):
        cell, to_node, from_node = mv["cell"], mv["to_node"], mv["from_node"]
        target_identity = rename_map.get(cell, cell)
        if target_identity not in post_by_node.get(to_node, set()):
            label, sector = qx.band_label(cell)
            warnings.append({
                "type": "moved_sector_missing",
                "text": f"Moved sector {label} {sector} ({target_identity}) not confirmed on target node : "
                        f"{to_node} (moved from {from_node}).",
            })
    return warnings


# ============================================================
# SECTION 5 - DELETED SECTOR: absence-only verification (confirmed: mirror-image of
# Moved Sectors - success = cell NOT present in Post-checks at all)
# ============================================================

def verify_deleted_sectors_against_postcheck(classification, post_text):
    post_pairs, _ = qx.extract_precheck_sectors(post_text)
    post_cells = {cell for (_node, cell) in post_pairs}

    warnings = []
    for node, cells in classification.get("deleted_sectors", {}).items():
        for cell in cells:
            if cell in post_cells:
                label, sector = qx.band_label(cell)
                warnings.append({
                    "type": "deleted_sector_still_present",
                    "text": f"Deleted sector {label} {sector} ({cell}) still present in Post-checks - "
                            f"deletion not confirmed on : {node}.",
                })
    return warnings


# ============================================================
# SECTION 6 - RADIO SWAP: Completed vs. Pending is DETERMINED by Post-checks, not a
# report-only warning (confirmed: this is the one item where verification changes report
# placement, replacing the old "always defaults to Pending" rule)
# ============================================================

def classify_radio_swap_placement(precheck_text, postcheck_text, ciq_wb):
    """Reuses qx.classify_radio_swaps for detection, then re-checks the actual installed
    radio type in Post-checks (qx.extract_precheck_radio_types reused as-is - confirmed
    generic). Completed if Post-checks radio == the CIQ target ('to'); Pending otherwise
    (matches Pre-checks 'from', or anything else - no special-case warning, per
    explicit instruction to ignore that edge case)."""
    swaps = qx.classify_radio_swaps(precheck_text, ciq_wb)
    post_radios = qx.extract_precheck_radio_types(postcheck_text)

    completed, pending = [], []
    for sw in swaps:
        cell = sw.get("group_key", (None,))[0]
        post_types = [qx.radio_family(r) for r in post_radios.get(cell, [])] if cell else []
        target_type = sw["to"].replace("RRU ", "").strip()
        if target_type in post_types:
            completed.append(sw)
        else:
            pending.append(sw)
    return completed, pending


def format_radio_swaps(swap_list, label_prefix="Radio Swap on:", stakeholder=None):
    """Same two-step grouping as the confirmed-mature qx logic: physical radio group first,
    then merge across sectors sharing an identical (from, to) signature. Confirmed real gap:
    this never appended a stakeholder tag at all — every other Pending item type gets a
    "(Tower Crew)"/"(MIC PM)" suffix (ROW_MAP already declares "Tower Crew" for radio_swap),
    Radio Swap never did. stakeholder is only meant to be passed for the Pending call."""
    merged = {}
    for r in swap_list:
        sig = (r["from"], r["to"])
        merged.setdefault(sig, {"labels": set(), "sectors": set()})
        merged[sig]["labels"].add(r["label"])
        merged[sig]["sectors"].add(r["sector"])

    lines = []
    for (frm, to), grp in merged.items():
        labels = tuple(sorted(grp["labels"]))
        label_str = labels[0] if len(labels) == 1 else f"[{'|'.join(labels)}]"
        sector_set = grp["sectors"]
        sector_names = sorted(sector_set, key=lambda s: qx.SECTOR_ORDER.index(s) if s in qx.SECTOR_ORDER else 99)
        is_whole = WHOLE_BAND_SET <= sector_set
        sectors_str = " sectors" if is_whole else (f" {', '.join(sector_names)}" if sector_names else "")
        stakeholder_str = f" ({stakeholder})" if stakeholder else ""
        lines.append(f"{label_prefix}\t{label_str}{sectors_str}{stakeholder_str}\tFrom:\t{frm}\tTo:\t{to}")
    return lines


# ============================================================
# SECTION 7 - GPS: Installation (new nodes, grouped by type), Upgrade (existing nodes,
# type changed Pre->Post), and the two site-health checks (unconfirmed type on MMBB,
# sync disabled) - all four pieces confirmed this session, none existed before.
# ============================================================

def gps_installation_lines(new_nodes, post_gps_status):
    """Groups new nodes by shared GPS type; Node IDs '|'-joined per group. Returns
    (dedicated_row_line, overflow_lines) - first group goes on the template's one
    dedicated row, any additional distinct-type groups spill to the buffer pool."""
    by_type = {}
    for node in new_nodes:
        gtype = post_gps_status.get(node, "NOT FOUND")
        by_type.setdefault(gtype, []).append(node)

    lines = [f"GPS Installation: {'|'.join(nodes)}  Version: {gtype}"
             for gtype, nodes in by_type.items()]
    if not lines:
        return None, []
    return lines[0], lines[1:]


def gps_upgrade_lines(existing_nodes, pre_gps_status, post_gps_status):
    """Existing node (not new), GPS type differs Pre vs Post -> Completed, buffer-only
    (no dedicated row exists for this item at all)."""
    lines = []
    for node in existing_nodes:
        pre_t, post_t = pre_gps_status.get(node), post_gps_status.get(node)
        if pre_t and post_t and pre_t != post_t:
            lines.append(f"GPS upgraded from: {pre_t} to: {post_t} on: {node}.")
    return lines


def gps_unconfirmed_type_check(mm_objs, post_gps_status):
    """Every MMBB node, Post-checks only. Unconfirmed = not one of the 4 known types.
    Same text used for BOTH the Pending line and the Warnings tab, per confirmed decision."""
    hits = []
    for row in mm_objs:
        if str(row.get("BBU Mode", "")).strip().upper() != "MMBB":
            continue
        node = row.get("Node to be built as")
        gtype = post_gps_status.get(node)
        if gtype and gtype not in KNOWN_GPS_TYPES:
            hits.append(node)
    if not hits:
        return None
    return f"GPS needs to be upgraded for : {'|'.join(hits)}"


def gps_sync_disabled_check(mm_objs, post_sync_status):
    """Every node, Post-checks TimeSyncIO row. All disabled nodes merge into ONE Pending
    line reusing the existing 'GPS Installation:' Pending label - no warning."""
    hits = [row.get("Node to be built as") for row in mm_objs
            if post_sync_status.get(row.get("Node to be built as")) == "DISABLED"]
    if not hits:
        return None
    return f"GPS Installation: {'|'.join(hits)}"


# ============================================================
# SECTION 8 - TRANSPORT SFP: grouping by shared manual model, plus the new TXdBm/RXdBm/BER
# threshold verification (new node -> Pending; existing+board-swap -> Pending+Warning;
# existing, no swap -> Pre-Existing Issues only)
# ============================================================

SFP_RANGES = {"1G": (-9.0, 9.0), "10G": (-6.2, 6.2)}


def transport_sfp_installation_lines(nodes_needing_sfp, sfp_models_by_node):
    """nodes_needing_sfp: from new_nodes OR Port-Conversion-triggered nodes.
    sfp_models_by_node: {node: (bbu_end_model, siad_end_model)} - both MANUAL entries,
    confirmed. Grouped by shared (bbu,siad) pair."""
    by_model = {}
    for node in nodes_needing_sfp:
        model = sfp_models_by_node.get(node, ("", ""))
        by_model.setdefault(model, []).append(node)
    lines = []
    for (bbu, siad), nodes in by_model.items():
        lines.append(f"Transport SFP Installation on: {'|'.join(nodes)}  "
                      f"SFP Model (BBU End): {bbu}  SFP Model (SIAD End): {siad}")
    return lines


def _sfp_out_of_range_labels(node, board_gen, post_text_for_speed, transport_sfp_data):
    """Determines port speed via PORT_BY_GEN + Transport Fiber link Status OpMode, then
    checks TXdBm/RXdBm against the speed-appropriate range and BER against 0/0.
    Returns (combined_label_string or None)."""
    port_labels = qx.PORT_BY_GEN.get(board_gen)
    if not port_labels:
        return None
    opmode = qx.extract_transport_fiber_opmode(post_text_for_speed, node, port_labels)
    if not opmode:
        return None
    speed = "10G" if "10G" in opmode.upper() else ("1G" if "1G" in opmode.upper() else None)
    if not speed:
        return None
    lo, hi = SFP_RANGES[speed]

    reading = transport_sfp_data.get(node)
    if not reading:
        return None

    bits = []
    if reading["txdbm"] > hi:
        bits.append("High TXDBM")
    elif reading["txdbm"] < lo:
        bits.append("Low TXDBM")
    if reading["rxdbm"] > hi:
        bits.append("High RXDBM")
    elif reading["rxdbm"] < lo:
        bits.append("Low RXDBM")

    ber = reading["ber"]
    ber_note = None
    if not ber or ber.strip() == "":
        ber_note = "BER is not pulling on the transport port"
    elif ber.strip() != "0/0":
        ber_note = "BER pulling on the transport port"

    if not bits and not ber_note:
        return None

    parts = []
    if bits:
        parts.append(f"{'/'.join(bits)} out of ranges")
    if ber_note:
        parts.append(ber_note)
    return ", ".join(parts)


def transport_sfp_verification(ciq_wb, mm_objs, new_nodes, board_swap_nodes, post_text, transport_sfp_data):
    """Returns (pending_lines, warnings, pre_existing_lines) - three separate buckets per
    the confirmed node-classification rule."""
    pending_lines, warnings, pre_existing_lines = [], [], []
    board_swap_node_names = {n for n, _pre, _post in board_swap_nodes} if board_swap_nodes and \
        isinstance(board_swap_nodes[0], tuple) and len(board_swap_nodes[0]) == 3 else set(board_swap_nodes or [])

    for row in mm_objs:
        node = row.get("Node to be built as")
        gen = qx.get_node_generation(ciq_wb, row)
        if not gen:
            continue
        label = _sfp_out_of_range_labels(node, gen, post_text, transport_sfp_data)
        if not label:
            continue
        text = f"{label} on the transport port on : {node}."

        if node in new_nodes:
            pending_lines.append(text)
        elif node in board_swap_node_names:
            pending_lines.append(text)
            warnings.append({"type": "transport_sfp_out_of_range", "text": text})
        else:
            pre_existing_lines.append(text)
    return pending_lines, warnings, pre_existing_lines


# ============================================================
# SECTION 9 - 6610 CONTROLLER: SAU Connections, External alarm Scripting/testing, and the
# cascading "6610 not actually integrated" Pending rule. Manual locked-port classification
# (buckets 1-6) is a UI concern, not encoded here - see module docstring at bottom.
# ============================================================

def sau_connections_placement(controller_checks_data, controller_id):
    """None means 'no 6610 present -> stays manual' (caller's responsibility to only call
    this when a 6610 exists)."""
    sau = controller_checks_data.get("sau_state")
    if not sau:
        return None
    return "Completed" if sau["oper"] == "ENABLED" else "Pending"


def sau_node_state(postcheck_text):
    """Node-level SAU state, parsed from Post-checks' Hardware Status Information table.
    Confirmed real component-name variants: 'SAU-1', 'SAU-2', 'SAU-3', or bare 'SAU'
    (all matched by SAU\\S*). Deliberately does NOT reuse the generic
    _hardware_component_state() helper: that function keeps only ONE oper-state per node
    (last match in the text wins), which silently loses data whenever a node has more than
    one SAU unit — e.g. SAU-1 ENABLED + SAU-2 DISABLED on the same node would collapse to
    just 'DISABLED' and hide the genuinely-enabled unit. This parser instead returns
    {node: [(component, oper), ...]} — every unit, unmerged."""
    out = {}
    if not postcheck_text:
        return out
    text = _normalize(postcheck_text)
    for m in re.finditer(
            r'(\S+) (SAU\S*) (UNLOCKED|LOCKED) (ON|OFF) (\S+) (ENABLED|DISABLED)', text):
        node, comp, _admin, _fault, _steady, oper = m.groups()
        out.setdefault(node, []).append((comp, oper))
    return out


def sau_enabled_nodes(postcheck_text):
    """Confirmed MCA rule: SAU can be enabled on either the 6610 controller or the node
    itself. When the 6610's own SAU is confirmed disabled (via controller-checks), this
    auto-detects which node(s) Post-checks shows SAU as ENABLED on instead, so the
    'SAU enabled on : {Node ID}' Notes line no longer needs a fully manual entry.
    A node counts as enabled if ANY of its SAU units (SAU-1/SAU-2/SAU-3/bare SAU) shows
    ENABLED — a site can have multiple SAU units per node, and only one needs to be live.
    Returns a sorted list of node names (usually 0 or 1, but returns every match)."""
    if not postcheck_text:
        return []
    states = sau_node_state(postcheck_text)
    return sorted(node for node, units in states.items()
                  if any(oper == "ENABLED" for _comp, oper in units))


def external_alarm_scripting_confirmed(controller_checks_data):
    """Any AlarmPort row with a real (non-blank) slogan = confirmed scripted."""
    return any(p["slogan"] for p in controller_checks_data.get("alarm_ports", []))


def external_alarm_testing_placement(controller_checks_data):
    """Among SCRIPTED ports (real slogan present) only: all locked -> Pending + Notes line.
    Confirmed real gap found and fixed: a MIXED scenario (some locked, some unlocked —
    not all) used to just report 'Completed' with zero mention of the ports that are
    still individually locked. Now returns a third value: a Notes line automatically
    listing those still-locked ports (reusing the same 'port[SLOGAN]' format already used
    elsewhere), so nothing gets silently dropped just because the overall status is
    Completed. Returns (section, pending_note, mixed_locked_note)."""
    scripted = [p for p in controller_checks_data.get("alarm_ports", []) if p["slogan"]]
    if not scripted:
        return None, None, None
    all_locked = all(p["admin"] == "LOCKED" for p in scripted)
    if all_locked:
        return "Pending", "All external alarms are kept locked, due to NEA is pending.", None

    still_locked = [p for p in scripted if p["admin"] == "LOCKED"]
    mixed_locked_note = None
    if still_locked:
        port_slogan_map = {p["port"]: p["slogan"] for p in still_locked}
        ports_str = ", ".join(p["port"] for p in still_locked)
        ports_fmt = format_ports_with_slogans(ports_str, port_slogan_map)
        mixed_locked_note = f"Port(s) {ports_fmt} are kept locked (Owner: AT&T)."
    return "Completed", None, mixed_locked_note


def controller_integration_cascade(six610_present_and_edp_published, controller_checks_data, controller_id):
    """Returns True if the cascade fires (6610 present/EDP-published, but controller-checks
    doesn't confirm alarm scripting) - caller then moves all 4 items to Pending:
    6610 Controller Integration, External alarm Scripting on, LKF Installation (6610 portion),
    External alarm testing. No warning, per confirmed decision."""
    if not six610_present_and_edp_published:
        return False
    if not controller_checks_data or not external_alarm_scripting_confirmed(controller_checks_data):
        return True
    return False


# ============================================================
# SECTION 10 - LKF INSTALLATION: 4 independent OR'd triggers (3 pre-existing + 1 new this
# session), per-node Completed/Pending choice, merged by chosen status.
# ============================================================

# ============================================================
# SECTION 10 - LKF INSTALLATION: 4 independent OR'd triggers (3 pre-existing + 1 new this
# session). Confirmed: the Node and the Controller are genuinely independent installation
# points in real site work (one can be Completed while the other is still Pending) — so
# they get tracked and chosen separately, then combined intelligently per section.
# ============================================================

def lkf_node_triggers(new_nodes, board_swap_nodes, precheck_text, mm_objs):
    """3 of the 4 original triggers (new node / board swap / single-tech node gaining a
    second tech) — the node-level ones. The 6610-controller-present trigger is now tracked
    SEPARATELY via lkf_controller_triggered(), not folded into every node at the site."""
    triggered = set(new_nodes) | {n for n, _p, _q in board_swap_nodes} if board_swap_nodes and \
        isinstance(board_swap_nodes[0], tuple) else set(new_nodes)

    pre_pairs, _ = qx.extract_precheck_sectors(precheck_text)
    pre_cells_by_node = {}
    for node, cell in pre_pairs:
        pre_cells_by_node.setdefault(node, set()).add(cell)

    for row in mm_objs:
        node = row.get("Node to be built as")
        has_enb, has_gnb = qx.is_populated(row.get("eNBId")), qx.is_populated(row.get("gNBId"))
        if not (has_enb and has_gnb):
            continue  # target isn't MMBB/TMBB, 4th trigger doesn't apply
        pre_cells = pre_cells_by_node.get(node, set())
        if not pre_cells:
            continue  # not in Pre-checks at all -> that's the "new node" trigger, already covered
        pre_has_5g = any(qx.is_5g_cell(c) for c in pre_cells)
        pre_has_lte = any(not qx.is_5g_cell(c) for c in pre_cells)
        if pre_has_lte and not pre_has_5g:
            triggered.add(node)  # LTE-only -> gaining 5G
        elif pre_has_5g and not pre_has_lte:
            triggered.add(node)  # 5G-only -> gaining LTE
    return sorted(n for n in triggered if n)


def lkf_controller_triggered(controller_id):
    """4th trigger, now tracked independently from any specific node — a 6610 controller
    being present at the site needs its own LKF installation, whether or not any node also
    needs one."""
    return bool(controller_id)


def lkf_trigger_nodes(new_nodes, board_swap_nodes, controller_id, precheck_text, mm_objs):
    """Kept for backward compatibility with anything still calling the old combined
    signature — now just the union of the node triggers and (if a controller is present)
    every node, matching the old behavior exactly. Prefer lkf_node_triggers +
    lkf_controller_triggered for new code, which track them independently."""
    nodes = set(lkf_node_triggers(new_nodes, board_swap_nodes, precheck_text, mm_objs))
    if controller_id:
        nodes |= {row.get("Node to be built as") for row in mm_objs}
    return sorted(n for n in nodes if n)


def lkf_lines_by_choice(node_choices, controller_choice, controller_id):
    """node_choices: {node: 'Completed'|'Pending'}. controller_choice: 'Completed'|
    'Pending'|None (None = controller not triggered or not yet chosen).
    Confirmed: Node and Controller are independent — combined onto one line only when they
    land in the SAME section; otherwise each gets its own line in its own section. Pending
    gets the confirmed (MIC) tag either way."""
    by_choice = {"Completed": [], "Pending": []}
    for node, choice in node_choices.items():
        by_choice[choice].append(node)

    out = {}
    for choice, nodes in by_choice.items():
        node_part = "|".join(nodes)
        controller_here = controller_id if controller_choice == choice else ""
        if not node_part and not controller_here:
            continue
        if node_part and controller_here:
            line = f"LKF Installation: {node_part} | {controller_here}"
        elif node_part:
            line = f"LKF Installation: {node_part}"
        else:
            line = f"LKF Installation: {controller_here}"
        out[choice] = f"{line} (MIC)" if choice == "Pending" else line

    # Controller landed in a section with no nodes sharing that choice, and no line was
    # built for it above (e.g. controller=Pending but zero nodes chose Pending at all) —
    # cover that case explicitly.
    if controller_choice and controller_choice not in out:
        line = f"LKF Installation: {controller_id}"
        out[controller_choice] = f"{line} (MIC)" if controller_choice == "Pending" else line
    return out


# ============================================================
# SECTION 11 - CALL TEST: Market lookup + the standalone moved-LTE PSAP rule
# ============================================================

def determine_market(node_name, prefix_to_market):
    if not node_name:
        return None
    prefix = str(node_name)[:2].upper()
    return prefix_to_market.get(prefix)


def call_test_lines(classification, market, rules, moved_lte_bands, added_bands_by_tech, moved_bands_by_tech):
    """added_bands_by_tech / moved_bands_by_tech: {'lte': set(), '5g': set(), 'cband_dod': set()}
    already computed by the caller from band_label() on classification['added']/['moved'].
    Returns list of report lines."""
    lines = []
    required = {"PSAP": False, "LTE Speed test": False, "F-net with F-net SIM": False, "5G speedtest": False}

    scenarios = set()
    if added_bands_by_tech.get("lte"):
        scenarios.add("Newly adding LTE")
    if added_bands_by_tech.get("5g"):
        scenarios.add("Newly adding 5G")
    if moved_bands_by_tech.get("5g") or moved_bands_by_tech.get("cband_dod"):
        scenarios.add("Moving 5G(Incl C band/DOD/DOD BWE)")
    if added_bands_by_tech.get("cband_dod"):
        scenarios.add("Newly adding  C band/DOD")

    for scenario in scenarios:
        row = rules.get((market, scenario))
        if row:
            for test, flag in row.items():
                if flag:
                    required[test] = True

    # Standalone rule, confirmed: moved LTE ALWAYS fires PSAP, independent of the table -
    # the table has no "Moving LTE" scenario at all.
    psap_from_moved_lte = bool(moved_lte_bands)

    if required["PSAP"] or psap_from_moved_lte:
        # Confirmed: no descriptive label at all ("Moved LTE Sectors:", etc.) — just the
        # bands directly, regardless of whether they're genuinely moved or newly added.
        psap_bands = moved_lte_bands if moved_lte_bands else added_bands_by_tech.get("lte", set())
        psap_str = ", ".join(sorted(psap_bands))
        lines.append(f"PSAP test/Speedtest/VoLTE voice calltest:\t{psap_str}"
                     f"\tPSAP Schedule ID:\t")
    if required["LTE Speed test"]:
        lines.append(f"Speedtest/VoLTE voice calltest:\t{', '.join(sorted(added_bands_by_tech.get('lte', [])))}")
    if required["5G speedtest"]:
        fiveg_all = set(added_bands_by_tech.get("5g", set())) | set(added_bands_by_tech.get("cband_dod", set())) \
            | set(moved_bands_by_tech.get("5g", set())) | set(moved_bands_by_tech.get("cband_dod", set()))
        lines.append(f"Speed test:\t{', '.join(sorted(fiveg_all))}")
    # Confirmed real fix: F-net with F-net SIM should only fire if FNET sectors are
    # genuinely present on the site, even if the CT sheet says Y for this market/scenario
    # — the CT sheet flag alone isn't enough, since a site with no FNET at all has
    # nothing to test.
    fnet_present = "FNET" in added_bands_by_tech.get("lte", set()) or "FNET" in moved_bands_by_tech.get("lte", set())
    if required["F-net with F-net SIM"] and fnet_present:
        lines.append("Calltest with F-NET SIM:\tF-NET Sectors")

    return lines


# ============================================================
# SECTION 12 - PORT CONVERSION: Post-checks verification layer on top of QUICKIX's
# existing generate_port_conversion_checks (which only detects the PLANNED state)
# ============================================================

def verify_port_conversion_against_postcheck(ciq_wb, mm_objs, precheck_text, postcheck_text, edp_index):
    """Reuses qx.generate_port_conversion_checks for detection (unchanged — it deliberately
    fires using Pre-checks' board generation "regardless of what it's being swapped to").
    Re-checks the SAME (Pre-generation's) port in POST-checks. Confirmed bug found against
    real ALL00640 data: when a board swap is ALSO happening (Pre=G2, CIQ target=G3+), the
    physical board changed, so the old generation's port (e.g. TN_A/TN_B) legitimately has
    NO reading in Post-checks anymore — the new board only has TN_IDL_B. That's not a real
    failure, it's the wrong port being checked; check_port_conversion_via_board_swap already
    covers completion correctly for these nodes using the NEW board's port. Skip this stale
    check entirely for any node where a board swap applies."""
    _outputs, summary_rows, _scope_lines = qx.generate_port_conversion_checks(
        ciq_wb, mm_objs, edp_index, precheck_text, lambda *a: None)

    # Nodes where Pre generation != CIQ target generation — a board swap is in progress,
    # so the stale-port check below doesn't apply to them at all.
    swap_node_names = set()
    for row in mm_objs:
        node = row.get("Node to be built as")
        if not node:
            continue
        pre_model = qx.extract_pre_hw(precheck_text, node)
        pre_gen = qx.DU_TYPE_TO_GEN.get(str(pre_model).strip()) if pre_model else None
        post_gen = qx.get_node_generation(ciq_wb, row)
        if pre_gen and post_gen and pre_gen != post_gen:
            swap_node_names.add(node)

    warnings = []
    for s in summary_rows:
        if s.get("Item") != "Port Conversion":
            continue
        node = s.get("Source")
        if node in swap_node_names:
            continue
        note = s.get("Note", "")
        m = re.search(r'\((\w+)\)? board, port: ([\w/]+)', note) or re.search(r'board: (\w+), port: ([\w/]+)', note)
        if not m:
            continue
        gen, ports = m.groups()
        port_labels = ports.split("/")
        opmode = qx.extract_transport_fiber_opmode(postcheck_text, node, port_labels)
        if not opmode or "10G" not in opmode.upper():
            warnings.append({
                "type": "port_conversion_not_confirmed",
                "text": f"Port conversion (1G to 10G) not confirmed on : {node} - "
                        f"Post-checks still shows {opmode or 'no reading'} on {'/'.join(port_labels)}.",
            })
    return warnings


# ============================================================
# SECTION 13 - NGS: port the main-branch fix into report-feature (both the corrected
# both_lte check AND the NodeGroupSync=Y safety-net, which report-feature was missing
# entirely). Safety-net stays log-only, never feeds the Warnings tab, per confirmed decision.
# ============================================================

def ngs_pair_is_pure_lte(a_to_b, b_to_a):
    """Ported verbatim from main branch's _ngs_pair_is_pure_lte - confirmed fix for the
    real bug where report-feature's node-level has_lte check produced false negatives on
    dual-tech TMBB nodes referenced via their 5G side."""
    for _own_cell, ref_cell in a_to_b + b_to_a:
        if qx.nr_band_label(ref_cell)[0] is not None:
            return False
    return True


def ngs_safety_net(ciq_wb, cell_to_node, confirmed_nodes, log):
    """Ported from main branch. Log-only by explicit instruction - never feeds Warnings."""
    hits = []
    for sheet_name, cell_col in (("eUtran Parameters", "EutranCellFDDId"), ("5G Info", "NRCellDU")):
        if sheet_name not in ciq_wb.sheetnames:
            continue
        for row in qx.sheet_objs(ciq_wb[sheet_name]):
            if str(row.get("NodeGroupSync", "")).strip().upper() != "Y":
                continue
            cell = row.get(cell_col)
            if not qx.is_populated(cell):
                continue
            node = cell_to_node.get(str(cell).strip())
            if node and node not in confirmed_nodes:
                log(f"NodeGroupSync=Y flagged for {cell} ({node}) but no confirmed NGS pair "
                    f"was detected - check this cell manually.")
                hits.append(cell)
    return hits


# ============================================================
# SECTION 14 - SHARED BUFFER OVERFLOW POOL (Completed rows 81-90 / Pending rows 158-166)
# First-come-first-served across every item that exceeds its dedicated row count.
# ============================================================

class BufferPool:
    """Confirmed design: shared, first-come-first-served, "Label : Detail" format
    (matches the real Report_MCA formula B&" : "&C already in the template). Pool
    exhaustion -> Warnings tab flag, nothing silently dropped."""

    def __init__(self, completed_capacity=10, pending_capacity=9):
        self.completed_capacity = completed_capacity
        self.pending_capacity = pending_capacity
        self.completed_used = []
        self.pending_used = []
        self.overflow_warnings = []

    def add(self, section, label, detail):
        bucket = self.completed_used if section == "Completed" else self.pending_used
        capacity = self.completed_capacity if section == "Completed" else self.pending_capacity
        if len(bucket) >= capacity:
            self.overflow_warnings.append({
                "type": "buffer_pool_exhausted",
                "text": f"{label} - additional entries could not fit in report, review manually.",
            })
            return False
        bucket.append(f"{label} : {detail}")
        return True


def check_port_conversion_via_board_swap(ciq_wb, mm_objs, precheck_text, postcheck_text):
    """NEW scenario, confirmed against real ECL02586 data: when a board swap explains the
    speed change (Pre board generation != CIQ target generation) rather than a pure
    same-board 1G->10G conversion, the OLD generate_port_conversion_checks logic just
    skips the site entirely (Rule 1: "a board swap is already in progress ... Port
    Conversion doesn't apply"). That's still correct for the PLANNED/pending line - but
    it means a real completed conversion (accomplished as a side-effect of the swap) was
    never being recognized as complete at all. Real example: ECL02586 Pre=G2(6630, port
    TN_B, 1G), CIQ target=G4(6672) -> check the NEW board's relevant port (TN_IDL_C) in
    Post-checks: shows 10G_FULL -> conversion is COMPLETE via the swap.
    Confirmed real fix: this must be a genuine BEFORE/AFTER comparison, not just checking
    the new board's port alone — the OLD board's port in Pre-checks must actually show 1G
    first. Without that check, this could fire even when the old board was already at
    10G, which isn't a real "conversion" at all, just a coincidental match.
    Returns list of {"node":..., "text":...} for every node where this applies."""
    completed = []
    for row in mm_objs:
        node = row.get("Node to be built as")
        if not node:
            continue
        pre_model = qx.extract_pre_hw(precheck_text, node)
        pre_gen = qx.DU_TYPE_TO_GEN.get(str(pre_model).strip()) if pre_model else None
        post_gen = qx.get_node_generation(ciq_wb, row)
        if not pre_gen or not post_gen or pre_gen == post_gen:
            continue  # not a board-swap scenario - the existing same-board check handles this node instead

        # Confirmed fix: verify the OLD board's port actually shows 1G in Pre-checks first.
        pre_port_labels = qx.PORT_BY_GEN.get(pre_gen)
        if not pre_port_labels:
            continue
        pre_opmode = qx.extract_transport_fiber_opmode(precheck_text, node, pre_port_labels)
        if not pre_opmode or "1G" not in pre_opmode.upper() or "10G" in pre_opmode.upper():
            continue  # old board wasn't genuinely at 1G — not a real conversion

        post_port_labels = qx.PORT_BY_GEN.get(post_gen)
        if not post_port_labels:
            continue
        opmode = qx.extract_transport_fiber_opmode(postcheck_text, node, post_port_labels)
        if opmode and "10G" in opmode.upper():
            completed.append({
                "node": node,
                "text": f"Port speed 1G to 10G conversion with MPST: {node}.",
            })
    return completed


# ============================================================
# SECTION 15 - SIDEHAUL INFO auto-fill for Switch type / Switch ID / Slot-Port / Node ID
# (confirmed: Cable part number stays manual - no CIQ/EDP source exists for it)
# ============================================================

def sidehaul_display_rows(ciq_wb):
    """Returns the auto-filled portion of each Switch/Slot-Port connection row - one dict
    per row from Sidehaul Info, ready to combine with a manual Cable part number per row."""
    return extract_sidehaul_info(ciq_wb)


def format_sidehaul_lines(sidehaul_rows, cable_part_numbers=None):
    """cable_part_numbers: {row_index: manual_value} - confirmed manual, keyed by row order
    since Sidehaul Info has no unique cable identifier of its own. One line per connection."""
    cable_part_numbers = cable_part_numbers or {}
    lines = []
    for i, row in enumerate(sidehaul_rows):
        cable = cable_part_numbers.get(i, "")
        lines.append(
            f"Switch type: {row['switch_type']}  Switch ID: {row['switch_id']}  "
            f"Slot/Port: {row['slot_port']}  Cable part number: {cable}  Node ID: {row['node_id']}"
        )
    return lines


# ============================================================
# SECTION 16 - EDP PUBLISH fallback (confirmed: replaces the old generic
# "EDP is not published for the controller" message entirely with the real Pending row)
# ============================================================

def edp_publish_line(node_id, controller_id, switch_id_manual=""):
    """Confirmed format, Pending only, stakeholder (AT&T). Switch ID is manual — no source
    exists for it in this specific context (distinct from the IDL Connections Switch ID,
    which DOES come from Sidehaul Info — this one doesn't, confirmed)."""
    return f"EDP Publish: {node_id} | {controller_id} | {switch_id_manual}"


# ============================================================
# SECTION 17 - FDD RENAMING: proper band-label grouping (confirmed gap — was still using
# raw cell names / report_detect.detect_fdd_renaming's un-grouped tuples instead of the
# band_label()-based (node, old_label, new_label) grouping agreed on)
# ============================================================

def fdd_renaming_lines(ciq_wb):
    """Re-derives renames the same way report_detect.detect_fdd_renaming does (Sector
    Del_Movement rows where Source Node == Target Node but cell name differs), then groups
    by (node, old_label, new_label) using band_label() — old_label/new_label already carry
    the carrier-extension number (AWS_1, AWS_2, AWS_3, etc.), confirmed this is what
    distinguishes real renames (never a different band family, only a carrier-number
    change). Node field stays the Node ID per confirmed decision (not a sector letter)."""
    if "Sector Del_Movement" not in ciq_wb.sheetnames:
        return []
    grouped = {}
    for r in qx.sheet_objs(ciq_wb["Sector Del_Movement"]):
        src_node, tgt_node = r.get("Source Node name"), r.get("Target Node name")
        src_cell, tgt_cell = r.get("Source Sector"), r.get("Target Sector")
        if not (qx.is_populated(src_node) and qx.is_populated(tgt_node) and qx.is_populated(src_cell) and qx.is_populated(tgt_cell)):
            continue
        if str(src_node).strip().upper() != str(tgt_node).strip().upper():
            continue
        if str(src_cell).strip() == str(tgt_cell).strip():
            continue
        old_label, _s1 = qx.band_label(src_cell)
        new_label, _s2 = qx.band_label(tgt_cell)
        if not old_label or not new_label:
            continue
        key = (str(src_node).strip(), old_label, new_label)
        grouped.setdefault(key, []).append((src_cell, tgt_cell))

    lines = []
    for (node, old_label, new_label), _pairs in grouped.items():
        lines.append(f"FDD Renaming on: {node} From: {old_label} To: {new_label}.")
    return lines


# ============================================================
# SECTION 18 - CURRENT CONFIGURATION (confirmed rule, agreed early in this session but
# never actually built): compare Post-checks' actual hardware state against the CIQ's
# target Post Configuration; only populate this field when they DIFFER (equipment still
# missing as per Final Configuration). qx.pre_hw_string is generic text parsing, reused
# directly against postcheck_text, same as elsewhere this session.
# ============================================================

def detect_missing_nodes(postcheck_text, candidate_nodes):
    """New shared check: a node is considered genuinely NOT integrated if its hardware
    string can't be found anywhere in Post-checks at all — same detection signal
    current_configuration_line already uses via pre_hw_string, just applied here to
    decide node presence rather than to flag a hardware mismatch. Returns the subset of
    candidate_nodes with no hardware string found. Confirmed distinct from a hardware
    MISMATCH (node present but wrong config) — this is node ABSENCE entirely."""
    if not postcheck_text:
        return list(candidate_nodes)
    missing = []
    for node in candidate_nodes:
        actual_hw = qx.pre_hw_string(postcheck_text, node)
        if not actual_hw:
            missing.append(node)
    return missing


def _identity_tag_from_checks(check_text, node_name):
    """Confirmed same real rule as Pre Configuration's pre_node_label(): derived purely from
    which cells are actually present in the checks text (works for Pre- or Post-checks alike,
    same 'Summary Status' table format either way — already reused this way elsewhere in this
    file, e.g. verify_integration_against_postcheck() against post_text). NOT from the CIQ's
    row structure or BBU Mode column, since that only reflects the CIQ's target/declared
    state, not what's actually deployed. LTE only -> SMBB. 5G only -> SMBB. LTE + 5G (no
    CBAND/DOD) -> MMBB. LTE + 5G + CBAND/DOD -> TMBB. CBAND/DOD share 5G band code 077
    (_N077[A-F]_n), same pattern app.py's nr_band_label() uses. Returns
    (secondary_name_or_None, mode_tag_or_empty)."""
    pairs, _ = qx.extract_precheck_sectors(check_text)
    node_cells = [cell for (n, cell) in pairs if n == node_name]
    if not node_cells:
        return None, ""
    fiveg_cells = [c for c in node_cells if qx.is_5g_cell(c)]
    lte_cells = [c for c in node_cells if not qx.is_5g_cell(c)]
    has_cband_dod = any(re.search(r'_N077[A-F]_\d+$', c) for c in fiveg_cells)
    if lte_cells and fiveg_cells:
        mode_tag = "TMBB" if has_cband_dod else "MMBB"
    elif lte_cells or fiveg_cells:
        mode_tag = "SMBB"
    else:
        mode_tag = ""
    secondary = None
    if fiveg_cells and lte_cells:
        m = re.match(r"^(.+?)_N\d{3}[A-F]_\d+$", fiveg_cells[0])
        secondary = m.group(1) if m else fiveg_cells[0]
    return secondary, mode_tag


def current_configuration_line(ciq_wb, mm_objs, postcheck_text, missing_nodes=None, dual_identity=False,
                                derive_identity_from_checks=False):
    """Returns the Current Configuration string, or "" if Post-checks already matches the
    CIQ target for every node (nothing missing, field should stay blank/not triggered).
    Confirmed: when any node is missing from Post-checks entirely (missing_nodes
    non-empty), shows EVERY present node's actual hardware unconditionally — whatever
    is actually on site now — not just mismatches, since a missing node changes what
    "current configuration" means for the whole report.
    Confirmed dual_identity=True (N2E/NSB only, not MCA): matches Post Configuration's
    own "{node}(P)/{secondary}(S)({bbu_mode})({hw})" format exactly, for co-located
    LTE+5G nodes — but using the ACTUAL hardware from Post-checks, not the CIQ target,
    since this field reflects what's really deployed now.
    Confirmed derive_identity_from_checks=True (MCA only, separate from dual_identity):
    same output format, but the (P)/(S) pairing AND mode tag come from
    _identity_tag_from_checks() against Post-checks itself, not the CIQ row's
    eNodeB/gNodeB Name fields or BBU Mode column — this is what "current configuration"
    should mean (what's actually deployed right now). Also robust to whichever CIQ Mixed
    Mode Info convention is in play (one combined row per site, or two separate rows, one
    per identity) — Post-checks always lists a dual-identity node's cells under the same
    primary Node column regardless, so a row whose OWN node name turns out to be some
    other row's derived secondary is skipped, rather than appearing as a spurious extra
    entry with no pairing."""
    if not postcheck_text:
        return ""
    missing_nodes = missing_nodes or []
    lines = []

    secondary_of, mode_tag_of = {}, {}
    if derive_identity_from_checks:
        for row in mm_objs:
            n = row.get("Node to be built as")
            if not n:
                continue
            secondary_of[n], mode_tag_of[n] = _identity_tag_from_checks(postcheck_text, n)
    secondary_node_names = {v for v in secondary_of.values() if v}

    for row in mm_objs:
        node = row.get("Node to be built as")
        if not node:
            continue
        if derive_identity_from_checks and node in secondary_node_names:
            continue  # already covered by its primary's combined entry, below
        e_name, g_name = row.get("eNodeB Name"), row.get("gNodeB Name")
        is_lte_primary = str(node).strip().upper() == str(e_name or "").strip().upper()
        target_row = qx.find_row_by_name(ciq_wb, "eNB Info", "eNodeB Name", e_name) if is_lte_primary else \
            qx.find_row_by_name(ciq_wb, "gNB Info", "gNodeB Name", g_name)
        if not target_row:
            target_row = qx.find_row_by_name(ciq_wb, "eNB Info", "eNodeB Name", e_name) or \
                qx.find_row_by_name(ciq_wb, "gNB Info", "gNodeB Name", g_name)
        target_hw = qx.hw_string(target_row) or "NOT FOUND"
        actual_hw = qx.pre_hw_string(postcheck_text, node) or "NOT FOUND"

        include = (node not in missing_nodes) if missing_nodes else (target_hw != actual_hw)
        if not include:
            continue

        if derive_identity_from_checks:
            secondary, mode_tag = secondary_of.get(node), mode_tag_of.get(node)
            if secondary and mode_tag:
                lines.append(f"{node}(P)/{secondary}(S)({mode_tag})({actual_hw})")
            else:
                lines.append(f"{node}({actual_hw})")
        elif dual_identity and qx.is_populated(e_name) and qx.is_populated(g_name):
            secondary = g_name if is_lte_primary else e_name
            bbu_mode = row.get("BBU Mode")
            lines.append(f"{node}(P)/{secondary}(S)({bbu_mode})({actual_hw})")
        else:
            lines.append(f"{node}({actual_hw})")
    if not lines:
        return ""
    return " + ".join(lines) if (dual_identity or derive_identity_from_checks) else "/".join(lines)


# ============================================================
# SECTION 19 - LOCKED ALARM PORTS: the 6 confirmed buckets from the
# 6610_Alarm_Cutover_Process_Reporting_Standards reference doc, reduced to simple text
# generators (confirmed: reuse the shared buffer lines, no dedicated classification UI).
# ============================================================

def write_buffer_with_overflow(row_writes, rows, items, col=2, sep=" | "):
    """Confirmed shared helper (moved from N2E-local, now used by all three scopes):
    when there are more items than available template rows, the overflow no longer
    gets silently dropped — everything from the last available row onward gets
    combined into ONE line, joined by the separator, and written into that final row.
    Every row before the last still gets exactly one item as normal. Appends directly
    to the passed-in row_writes list (matches each scope's existing local convention)."""
    if not rows:
        return
    if len(items) <= len(rows):
        for i, row_num in enumerate(rows):
            if i < len(items):
                row_writes.append((row_num, True, [(col, items[i])]))
            else:
                row_writes.append((row_num, False, []))
    else:
        for i, row_num in enumerate(rows[:-1]):
            row_writes.append((row_num, True, [(col, items[i])]))
        overflow = sep.join(items[len(rows) - 1:])
        row_writes.append((rows[-1], True, [(col, overflow)]))


def write_buffer_2col_with_overflow(row_writes, rows, items, col_a=3, col_b=4, sep=" | "):
    """Same overflow-safe principle as write_buffer_with_overflow, but for the
    2-column row structure used by Integration (bands, node) and Transport SFP
    (node, models): items is a list of (val_a, val_b) tuples. On overflow, the two
    columns' overflow values are joined SEPARATELY (col_a's overflow values joined
    together, col_b's overflow values joined together) into the last row — not
    interleaved — so e.g. Integration's last row would show
    'LTE_B2/PCS_2 | LTE_B4' in the bands column and 'NODE_A | NODE_B' in the node
    column, keeping each column internally consistent rather than mixing bands and
    node names together."""
    if not rows:
        return
    if len(items) <= len(rows):
        for i, row_num in enumerate(rows):
            if i < len(items):
                row_writes.append((row_num, True, [(col_a, items[i][0]), (col_b, items[i][1])]))
            else:
                row_writes.append((row_num, False, []))
    else:
        for i, row_num in enumerate(rows[:-1]):
            row_writes.append((row_num, True, [(col_a, items[i][0]), (col_b, items[i][1])]))
        overflow_items = items[len(rows) - 1:]
        overflow_a = sep.join(str(x[0]) for x in overflow_items)
        overflow_b = sep.join(str(x[1]) for x in overflow_items)
        row_writes.append((rows[-1], True, [(col_a, overflow_a), (col_b, overflow_b)]))


def missing_slogan_ports(ports_str, port_slogan_map):
    """Confirmed real requirement: a slogan is a MUST for every reported port, not
    optional. Returns the list of typed port numbers that have NO matching detected
    slogan (typo, or a port outside the detected/locked set) — used to show a clear
    warning instead of silently generating a report line with a bare, slogan-less port."""
    if not ports_str:
        return []
    port_list = [p.strip() for p in ports_str.split(",") if p.strip()]
    return [p for p in port_list if p not in port_slogan_map]


def format_ports_with_slogans(ports_str, port_slogan_map):
    """ports_str: the engineer's typed comma-separated port numbers (e.g. '2, 4, 10').
    port_slogan_map: {port_number_str: slogan} from the detected locked_ports_list.
    Confirmed format: 'port[SLOGAN]' per port, joined with Oxford comma + 'and' before the
    last one — e.g. '2[RBS INTRUSION], 4[RBS HEX FAIL], and 10[RBS COMMERCIAL]'. A port with
    no matching detected slogan (typo, or a port outside the detected set) is shown bare —
    confirmed this should be flagged via missing_slogan_ports() wherever ports are entered,
    since a slogan is required for every port, not optional."""
    if not ports_str:
        return ""
    port_list = [p.strip() for p in ports_str.split(",") if p.strip()]
    annotated = [f"{p}[{port_slogan_map[p]}]" if p in port_slogan_map else p for p in port_list]
    if len(annotated) == 1:
        return annotated[0]
    if len(annotated) == 2:
        return f"{annotated[0]} and {annotated[1]}"
    return ", ".join(annotated[:-1]) + f", and {annotated[-1]}"


def locked_port_bucket_1(ports, port_slogan_map=None):
    """Pre-existing locked state -> Pre-Existing Issues."""
    if not ports:
        return None
    ports_fmt = format_ports_with_slogans(ports, port_slogan_map or {})
    return f"Alarm Ports {ports_fmt} remain in a locked state, matching the pre\u2011existing condition. (Owner: AT&T PM/OPS)"


def locked_port_bucket_2(ports, port_slogan_map=None):
    """Pre-existing active alarm, kept locked to avoid OPS tickets -> Pre-Existing Issues."""
    if not ports:
        return None
    ports_fmt = format_ports_with_slogans(ports, port_slogan_map or {})
    return f"Pre\u2011existing active alarms on ports {ports_fmt} are kept locked to avoid OPS tickets. (Owner: AT&T PM/OPS)"


def loops_bridge_clips_notes(loops_ports, bridge_clips_ports, no_equip_ports, alarm_ports_data, port_slogan_map=None):
    """Confirmed redesign: one shared set of 3 side-by-side category inputs (Pre-existing
    loops / Bridge clips / No equipment end connections), each generating its own Note
    line regardless of active state. Separately, ALL ports across all 3 categories
    combined are checked against the real activeExternalAlarm field (not admin/LOCKED —
    confirmed that field is a static circuit flag, always ENABLED regardless of a real
    fault) to produce ONE combined "kept locked" line covering only the genuinely active
    ones. Slogans included everywhere.
    alarm_ports_data: the raw controller_checks_data['alarm_ports'] list (each with
    'port', 'slogan', 'active' keys).
    Returns (notes_lines: list[str], active_pending_line: str or None)."""
    port_slogan_map = port_slogan_map or {}
    port_active_map = {p["port"]: bool(p.get("active")) for p in (alarm_ports_data or [])}

    notes_lines = []
    if loops_ports:
        fmt = format_ports_with_slogans(loops_ports, port_slogan_map)
        notes_lines.append(f"Pre\u2011existing loops have been removed from alarm ports {fmt} (Owner: AT&T).")
    if bridge_clips_ports:
        fmt = format_ports_with_slogans(bridge_clips_ports, port_slogan_map)
        notes_lines.append(f"Pre\u2011existing Bridge clips have been removed from alarm ports {fmt} (Owner: AT&T).")
    if no_equip_ports:
        fmt = format_ports_with_slogans(no_equip_ports, port_slogan_map)
        notes_lines.append(f"No equipment connections on port {fmt} (Owner: AT&T).")

    all_ports_str = ",".join(p for p in (loops_ports, bridge_clips_ports, no_equip_ports) if p)
    # Confirmed real bug: a port entered in more than one category (e.g. 16 in loops AND
    # bridge clips AND no-equipment) was appearing that many times over in the combined
    # "Active alarms observed on ports..." line — same port, same slogan, repeated. Dedupe
    # while preserving first-seen order, since it's still the SAME port either way.
    all_port_list = list(dict.fromkeys(p.strip() for p in all_ports_str.split(",") if p.strip()))
    active_ports = [p for p in all_port_list if port_active_map.get(p)]

    active_pending_line = None
    if active_ports:
        fmt = format_ports_with_slogans(",".join(active_ports), port_slogan_map)
        active_pending_line = f"Active alarms observed on ports {fmt} are kept locked (Owner: AT&T)."

    return notes_lines, active_pending_line


def locked_port_bucket_4(ports, owner, port_slogan_map=None):
    """Post-cutover, FE couldn't clear -> Pending."""
    if not ports:
        return None
    ports_fmt = format_ports_with_slogans(ports, port_slogan_map or {})
    return f"Post external alarm cutover, active alarms observed on ports {ports_fmt} have been kept locked (Owner: {owner})."


# ============================================================
# SECTION 20 - XLSM ROW WRITES for every item built this session. build_xlsm_row_writes()
# only ever processed the ORIGINAL mca_checklist.CHECKLIST items — everything new (GPS,
# Transport SFP structured data, Radio Swap split, LKF node/controller split, EDP Publish,
# fixed FDD Renaming, Port-Conversion-via-swap, Current Config, locked-port buckets, and
# the free-text buffer boxes themselves) was NEVER written into the .xlsm at all before this
# — confirmed real gap. Reuses the exact ROW_MAP slots already reserved for these items.
# ============================================================

def build_new_xlsm_row_writes(
        row_map,
        current_config_text="",
        gps_completed_groups=None, gps_pending_lines=None,
        sfp_completed_groups=None, sfp_pending_lines=None,
        radio_swap_completed=None, radio_swap_pending=None,
        lkf_completed_line=None, lkf_pending_line=None,
        fdd_lines=None,
        edp_publish=None,
        ngs_completed_line=None, ngs_pending_line=None,
        buffer_completed_extra=None, buffer_pending_extra=None, buffer_pre_existing_extra=None):
    """Every *_groups / *_lines argument is the STRUCTURED data (lists/tuples), not the
    pre-formatted display strings, so exact column values can be written. Returns a list of
    (row_num, checked, [(col, value), ...]) tuples, ready to combine with whatever
    mca_glue.build_xlsm_row_writes already produced for the untouched original items."""
    rw = []

    # ---- Current Configuration (row 11) — checked only when it actually applies. ----
    if current_config_text:
        rw.append((row_map["current_configuration"], True, [(3, current_config_text)]))

    # ---- GPS Installation (completed=[61], pending=[124]) — first group uses the
    # dedicated row; additional distinct-type groups spill to the Completed buffer pool. ----
    gcompleted_rows = row_map["gps_installation"]["completed"]
    if gps_completed_groups:
        first_nodes, first_type = gps_completed_groups[0]
        rw.append((gcompleted_rows[0], True, [(3, "|".join(first_nodes)), (5, first_type)]))
    else:
        rw.append((gcompleted_rows[0], False, []))
    gpending_rows = row_map["gps_installation"]["pending"]
    if gps_pending_lines:
        rw.append((gpending_rows[0], True, [(3, gps_pending_lines[0])]))
    else:
        rw.append((gpending_rows[0], False, []))

    # ---- Transport SFP (completed=[67,68,69], pending=[132-135]) — one node-group per row,
    # matching the same "one row per instance" pattern already used for Moved Sectors etc. ----
    sfp_c_rows = row_map["transport_sfp"]["completed"]
    sfp_completed_groups = sfp_completed_groups or []
    for i, row_num in enumerate(sfp_c_rows):
        if i < len(sfp_completed_groups):
            nodes, bbu, siad = sfp_completed_groups[i]
            rw.append((row_num, True, [(3, "|".join(nodes)), (4, f"{bbu} / {siad}")]))
        else:
            rw.append((row_num, False, []))
    sfp_p_rows = row_map["transport_sfp"]["pending"]
    sfp_pending_lines = sfp_pending_lines or []
    for i, row_num in enumerate(sfp_p_rows):
        if i < len(sfp_pending_lines):
            rw.append((row_num, True, [(3, sfp_pending_lines[i])]))
        else:
            rw.append((row_num, False, []))

    # ---- Radio Swap (completed=[56,57,58], pending=[119,120,121]) — placement now
    # DETERMINED by Post-checks, not a manual toggle; write both sides independently. ----
    rs_c_rows = row_map["radio_swap"]["completed"]
    radio_swap_completed = radio_swap_completed or []
    for i, row_num in enumerate(rs_c_rows):
        if i < len(radio_swap_completed):
            label, sectors_str, frm, to = radio_swap_completed[i]
            rw.append((row_num, True, [(3, f"{label}{sectors_str}"), (5, frm), (7, to)]))
        else:
            rw.append((row_num, False, []))
    rs_p_rows = row_map["radio_swap"]["pending"]
    radio_swap_pending = radio_swap_pending or []
    for i, row_num in enumerate(rs_p_rows):
        if i < len(radio_swap_pending):
            label, sectors_str, frm, to = radio_swap_pending[i]
            rw.append((row_num, True, [(3, f"{label}{sectors_str}"), (5, frm), (7, to)]))
        else:
            rw.append((row_num, False, []))

    # ---- LKF Installation (completed=[62], pending=[125]) — Node and Controller combine
    # onto one line per section (confirmed: only ever one line max per section). ----
    lkf_c_row = row_map["lkf_installation"]["completed"][0]
    rw.append((lkf_c_row, bool(lkf_completed_line), [(3, lkf_completed_line)] if lkf_completed_line else []))
    lkf_p_row = row_map["lkf_installation"]["pending"][0]
    rw.append((lkf_p_row, bool(lkf_pending_line), [(3, lkf_pending_line)] if lkf_pending_line else []))

    # ---- NGS activation (completed=[60], pending=[123]) — per-pair 3-way choice, confirmed
    # change from auto-Completed-only. Multiple confirmed pairs sharing the same choice
    # combine onto the one dedicated row. Pre-Existing choices don't appear here at all —
    # they only add a Notes line, handled separately in the UI. ----
    ngs_c_row = row_map["ngs_activation"]["completed"][0]
    rw.append((ngs_c_row, bool(ngs_completed_line), [(3, ngs_completed_line)] if ngs_completed_line else []))
    ngs_p_row = row_map["ngs_activation"]["pending"][0]
    rw.append((ngs_p_row, bool(ngs_pending_line), [(3, ngs_pending_line)] if ngs_pending_line else []))

    # ---- FDD Renaming, corrected (completed=[54,55], pending=[117,118]) — band-label
    # grouped, not raw cell tuples. ----
    fdd_c_rows = row_map["fdd_renaming"]["completed"]
    fdd_lines = fdd_lines or []
    for i, row_num in enumerate(fdd_c_rows):
        if i < len(fdd_lines):
            node, old_label, new_label = fdd_lines[i]
            rw.append((row_num, True, [(3, node), (5, old_label), (8, new_label)]))
        else:
            rw.append((row_num, False, []))

    # ---- EDP Publish (pending=[113], no Completed counterpart) ----
    edp_row = row_map["edp_publish"]["pending"][0]
    rw.append((edp_row, bool(edp_publish), []))  # no VALUE_COLUMNS entry (structural row) — text lives in the plain-text report

    # Port Conversion (including the via-board-swap completion path) is now merged into ONE
    # line at the scope_lines level before the checklist ever runs — it flows through the
    # NORMAL mca_glue.build_xlsm_row_writes path via its existing dedicated row, same as any
    # other checklist item. No separate writer needed here (would double-write row 46).

    # ---- Buffer pools: Completed (81-90), Pending (158-166), Pre-Existing (169-178) —
    # first-come-first-served, "Label : Detail" already split into (label, detail) pairs
    # by the caller. Whatever doesn't fit is simply not written (Warnings tab already
    # flags pool exhaustion — confirmed decision, nothing silently duplicated here). ----
    def _fill_buffer(rows, entries):
        entries = entries or []
        for i, row_num in enumerate(rows):
            if i < len(entries):
                label, detail = entries[i]
                rw.append((row_num, True, [(2, label), (3, detail)]))
            else:
                rw.append((row_num, False, []))

    def _fill_single_column_buffer(rows, entries):
        """Pre-Existing Issues rows are single-column (just B) — confirmed different from
        the Completed/Pending buffer rows' B&' : '&C combined format."""
        entries = entries or []
        for i, row_num in enumerate(rows):
            if i < len(entries):
                rw.append((row_num, True, [(2, entries[i])]))
            else:
                rw.append((row_num, False, []))

    _fill_buffer(row_map["additional_completed"]["completed"], buffer_completed_extra)
    _fill_buffer(row_map["additional_pending"]["pending"], buffer_pending_extra)
    _fill_single_column_buffer(row_map["pre_existing_issues"], buffer_pre_existing_extra)

    return rw


# ============================================================
# SECTION 21 - FLORIDA-ONLY: newly added CBAND/DOD/DOD_BWE individual cell names
# (confirmed this session — row block 92-104, previously untouched all session)
# ============================================================

def florida_newly_added_cells(market, classification):
    """Every individual newly-added CBAND/DOD/DOD_BWE cell (NRCellDU), only when market is
    Florida (same Calltest_sheet.xlsx lookup already used for Call Test). Confirmed: one
    cell per row up to the template's 12 available slots (93-104); anything beyond that
    gets appended onto the LAST row, '|'-joined, rather than dropped."""
    if market != "Florida":
        return []
    cells = []
    for node, node_cells in classification.get("added", {}).items():
        for cell in node_cells:
            label, _sector = qx.band_label(cell)
            if label in ("CBAND", "DOD", "DOD_BWE"):
                cells.append(cell)
    return cells


def florida_cells_to_rows(cells, capacity=12):
    """Splits into up to `capacity` row values — cell 1..capacity-1 get their own row,
    anything from capacity onward is joined with '|' onto the final row."""
    if not cells:
        return []
    if len(cells) <= capacity:
        return list(cells)
    rows = cells[:capacity - 1]
    rows.append("|".join(cells[capacity - 1:]))
    return rows



# ============================================================
# SUP / XMU shared helpers — confirmed real logic, verified against real ALL00640
# Post-checks data (SUP-1 ... ENABLED, XMU03-1-1 ... ENABLED both present in the real
# Hardware Status table). Shared between N2E and NSB, which both use the same mechanism.
# ============================================================

def _hardware_component_state(post_text, component_prefix):
    """Generic Hardware Status Information row parser: '{Node} {Component} {admin} {fault}
    {steady} {oper} {description}...' — confirmed real format. Returns
    {node: oper_state} for any component whose name starts with component_prefix
    (e.g. 'SUP' matches 'SUP-1', 'XMU' matches 'XMU03-1-1')."""
    out = {}
    text = _normalize(post_text) if post_text else ""
    for m in re.finditer(
            r'(\S+) (' + component_prefix + r'\S*) (UNLOCKED|LOCKED) (ON|OFF) (\S+) (ENABLED|DISABLED)', text):
        node, _comp, _admin, _fault, _steady, oper = m.groups()
        out[node] = oper
    return out


def detect_site_mismatch(mm_objs, controller_objs=None, **labeled_texts):
    """New safety check: confirmed requirement — at least ONE node ID must appear
    across ALL uploaded documents together (not just some overlap independently per
    document, which could hide a sneaker mismatch — e.g. CIQ has nodes A+B,
    Pre-checks has A, Post-checks has B, but A and B never co-occur in the same
    document). Confirmed fix: "controller_checks_text" is validated separately
    against the CIQ's "6610 Controller" tab (controller_objs, Controller=="6610" ->
    Controller ID) rather than the radio node names — a Controller-checks file is
    tied to a controller ID (e.g. "AZPC102030_C001"), which would never genuinely
    appear in a radio node's Pre/Post-checks text, so validating it against the same
    node-name set as everything else would always incorrectly flag a mismatch even
    on a genuinely correct file. Every OTHER labeled document (Pre-checks,
    Post-checks) still validates against radio node names as before, requiring one
    common node across all of them together.
    Returns (is_mismatch, offending_labels) — offending_labels lists every uploaded
    document that individually contains none of its expected identifiers. Note:
    offending_labels can be empty even when is_mismatch is True, if each
    radio-node document individually has SOME CIQ node but no single node is common
    to all of them — in that case the mismatch is about the documents not agreeing
    with EACH OTHER, not any one document being wrong on its own."""
    node_names = [row.get("Node to be built as") for row in mm_objs if row.get("Node to be built as")]
    controller_ids = [row.get("Controller ID") for row in (controller_objs or [])
                       if str(row.get("Controller", "")).strip() == "6610" and row.get("Controller ID")]

    uploaded = {label: text for label, text in labeled_texts.items() if text}
    if not uploaded:
        return False, []

    offending_labels = []
    is_mismatch = False

    controller_text = uploaded.pop("controller_checks_text", None)
    if controller_text is not None:
        if controller_ids:
            if not any(cid in controller_text for cid in controller_ids):
                offending_labels.append("controller_checks_text")
                is_mismatch = True
        # else: no controller IDs in the CIQ at all — nothing to validate against,
        # not treated as a mismatch on its own.

    if uploaded and node_names:
        common_nodes = set(node_names)
        for label, text in uploaded.items():
            nodes_in_this_doc = {n for n in node_names if n in text}
            if not nodes_in_this_doc:
                offending_labels.append(label)
            common_nodes &= nodes_in_this_doc
        if not common_nodes:
            is_mismatch = True

    return is_mismatch, offending_labels
    return not bool(common_nodes)


def xmu_sup_locked_warning(postcheck_text, integrated_nodes):
    """New NSB check: flags any XMU or SUP found in Post-checks with admin state
    LOCKED — separate from sup_capacity_warning (which checks capacity/counts, not
    lock state). A LOCKED XMU/SUP needs to be unlocked before it can actually
    function, regardless of whether SUP capacity is otherwise sufficient. Returns one
    warning line per affected node+component combination."""
    if not postcheck_text or not integrated_nodes:
        return []
    warnings = []
    for node in integrated_nodes:
        if re.search(re.escape(node) + r"\s+XMU\S*\s+LOCKED\s+OFF\s+(?:(?:true|false)\s+)?STEADY_ON\s+(?:ENABLED|DISABLED)",
                      postcheck_text, re.I):
            warnings.append(f"XMU is in locked state on {node}, please unlock.")
        if re.search(re.escape(node) + r"\s+SUP\S*\s+LOCKED\s+OFF\s+(?:(?:true|false)\s+)?STEADY_ON\s+(?:ENABLED|DISABLED)",
                      postcheck_text, re.I):
            warnings.append(f"SUP is in locked state on {node}, please unlock.")
    return warnings


def sup_capacity_warning(postcheck_text, integrated_nodes):
    """New site-wide capacity check: each SUP accommodates up to 2 XMU/5216 boards
    (confirmed pooled across the whole site, not per-node — a lone board on one node
    can share a SUP slot with a lone board on another). Counts total XMU+5216 boards
    across every integrated node (from Post-checks — base model via extract_pre_hw for
    5216, but XMU count uses a local, state-agnostic regex rather than
    extract_pre_xmu_count, since a DISABLED XMU is still physically present and
    confirmed to still need SUP capacity, unlike extract_pre_xmu_count's ENABLED-only
    matching used elsewhere). Counts total SUP instances actually found in Post-checks
    (still ENABLED-only — this check is specifically about the XMU/5216 side). If found
    SUP < required (ceil(total_boards / 2)), fires a warning. Confirmed: since the
    exact board-to-SUP pairing is determined in the field (not something this tool can
    predict), the warning lists every node that has an XMU or 5216 board, not a single
    attributed node."""
    if not postcheck_text or not integrated_nodes:
        return []
    total_boards = 0
    boards_nodes = []
    has_5216 = False
    has_xmu = False
    for node in integrated_nodes:
        base = qx.extract_pre_hw(postcheck_text, node)
        xmu_count = len(re.findall(
            re.escape(node) + r"\s+XMU\S*\s+(?:UNLOCKED|LOCKED)\s+OFF\s+(?:(?:true|false)\s+)?STEADY_ON\s+(?:ENABLED|DISABLED)",
            postcheck_text, re.I))
        node_boards = xmu_count + (1 if base == "5216" else 0)
        if node_boards:
            total_boards += node_boards
            boards_nodes.append(node)
            if base == "5216":
                has_5216 = True
            if xmu_count:
                has_xmu = True
    if not total_boards:
        return []
    total_sup_found = sum(
        len(re.findall(re.escape(node) + r"\s+SUP\S*\s+(?:UNLOCKED|LOCKED)\s+OFF\s+(?:(?:true|false)\s+)?STEADY_ON\s+ENABLED",
                       postcheck_text, re.I))
        for node in integrated_nodes)
    required_sup = -(-total_boards // 2)  # ceil without importing math
    if total_sup_found < required_sup:
        if has_5216 and has_xmu:
            reason = "5216 or the node with XMU"
        elif has_5216:
            reason = "5216"
        else:
            reason = "the node with XMU"
        return [f"SUP is not scripted for the: {', '.join(boards_nodes)} ({reason})."]
    return []


def nodes_expecting_sup(mm_objs, ciq_wb):
    """New check, confirmed per-node (not site-wide): SUP Connections is expected on any
    node whose CIQ TARGET hardware string contains either '5216' (a specific DU model)
    or 'XMU' — not just XMU alone. Confirmed example: a 3-node site where only one node
    has 5216 should only trigger SUP for that one node, not the whole site."""
    expecting = set()
    for row in mm_objs:
        node = row.get("Node to be built as")
        e_name, g_name = row.get("eNodeB Name"), row.get("gNodeB Name")
        is_lte_primary = str(node).strip().upper() == str(e_name or "").strip().upper()
        target_row = qx.find_row_by_name(ciq_wb, "eNB Info", "eNodeB Name", e_name) if is_lte_primary else \
            qx.find_row_by_name(ciq_wb, "gNB Info", "gNodeB Name", g_name)
        if not target_row:
            target_row = qx.find_row_by_name(ciq_wb, "eNB Info", "eNodeB Name", e_name) or \
                qx.find_row_by_name(ciq_wb, "gNB Info", "gNodeB Name", g_name)
        target_hw = qx.hw_string(target_row) or ""
        if "5216" in target_hw or "XMU" in target_hw:
            expecting.add(node)
    return expecting


def nodes_expecting_xmu(mm_objs, ciq_wb):
    """New check: which specific nodes' CIQ TARGET hardware string (not the combined
    site-wide post_line) contains 'XMU' — confirmed distinct from xmu_in_ciq(), which
    only checks the whole site at once and can't identify which node(s) specifically
    expect it. Used to catch the case where a node is genuinely present in Post-checks
    but its expected XMU component is missing from Post-checks' Hardware Status
    entirely (not found at all, not just DISABLED) — that node would otherwise be
    silently dropped from XMU Installation reporting altogether."""
    expecting = set()
    for row in mm_objs:
        node = row.get("Node to be built as")
        e_name, g_name = row.get("eNodeB Name"), row.get("gNodeB Name")
        is_lte_primary = str(node).strip().upper() == str(e_name or "").strip().upper()
        target_row = qx.find_row_by_name(ciq_wb, "eNB Info", "eNodeB Name", e_name) if is_lte_primary else \
            qx.find_row_by_name(ciq_wb, "gNB Info", "gNodeB Name", g_name)
        if not target_row:
            target_row = qx.find_row_by_name(ciq_wb, "eNB Info", "eNodeB Name", e_name) or \
                qx.find_row_by_name(ciq_wb, "gNB Info", "gNodeB Name", g_name)
        target_hw = qx.hw_string(target_row) or ""
        if "XMU" in target_hw:
            expecting.add(node)
    return expecting


def xmu_in_ciq(post_configuration_text):
    """Confirmed trigger: XMU appears in the CIQ target (Post Configuration string)."""
    return "XMU" in (post_configuration_text or "")


def gps_pending_stakeholder(market):
    """Confirmed rule, applies across MCA/NSB (N2E has no market lookup at all and
    always uses Tower Crew regardless): GPS Installation Pending's stakeholder is
    'MIC PM' specifically in NCSC, 'Tower Crew' in every other market."""
    return "MIC PM" if market == "NCSC" else "Tower Crew"


def sa_conversion_nodes(ciq_wb, mm_objs):
    """Confirmed generic (not scope-specific — shared between N2E and NSB): checks CIQ's
    NR_SA tab presence per node."""
    return [row.get("Node to be built as") for row in mm_objs
            if qx.check_sa_conversion(ciq_wb, row.get("Node to be built as"))]


def sa_conversion_note(sa_nodes):
    """Confirmed Notes addition when SA Conversion is detected — fires only if the CIQ's
    NR_SA tab is present AND SA Conversion is detected for at least one node."""
    if not sa_nodes:
        return None
    return "Termpointtoamf is in unlocked state."


# ============================================================
# WARNING CHECKS (Transport SFP, LTE/5G sector params, SCTP, DigitalTilt, AMF) —
# confirmed genuinely generic (Post-checks vs CIQ comparisons, no scope-specific
# assumptions baked in), moved here from n2e_completed_logic.py so both N2E and NSB
# can share the exact same tested implementation instead of duplicating it.
# ============================================================
def transport_sfp_threshold_warnings(ciq_wb, mm_objs, post_text, transport_sfp_data):
    """Confirmed N2E-specific check (different wording from MCA's transport_sfp_verification):
    checks TXdBm/RXdBm against the speed-appropriate range (reusing SFP_RANGES) and
    BER against 0/0, for every node. Confirmed exact wording:
    'High/low RXdBm/TXdBm on Transport SFP: {node}.',
    'BER not reporting on the Transport port: {node}.' (BER empty), and
    'BER NZ reporting on the Transport port: {node}.' (BER present but non-zero).
    All fire independently (a node can trigger any combination, or none).
    Confirmed: all go into the Warnings tab AND get reported as Pending to MIC PM
    (via the buffer, since there's no dedicated template row for this).
    Returns (warning_texts: list[str], pending_lines: list[str])."""
    warning_texts, pending_lines = [], []
    for row in mm_objs:
        node = row.get("Node to be built as")
        gen = qx.get_node_generation(ciq_wb, row)
        if not gen:
            continue
        port_labels = qx.PORT_BY_GEN.get(gen)
        if not port_labels:
            continue
        opmode = qx.extract_transport_fiber_opmode(post_text, node, port_labels)
        if not opmode:
            continue
        speed = "10G" if "10G" in opmode.upper() else ("1G" if "1G" in opmode.upper() else None)
        if not speed:
            continue
        lo, hi = SFP_RANGES[speed]

        reading = transport_sfp_data.get(node)
        if not reading:
            continue

        out_of_range = (reading["txdbm"] > hi or reading["txdbm"] < lo
                         or reading["rxdbm"] > hi or reading["rxdbm"] < lo)
        if out_of_range:
            text = f"High/low RXdBm/TXdBm on Transport SFP: {node}."
            warning_texts.append(text)
            pending_lines.append(f"{text} (MIC PM)")

        ber = reading["ber"]
        if not ber or ber.strip() == "":
            text = f"BER not reporting on the Transport port: {node}."
            warning_texts.append(text)
            pending_lines.append(f"{text} (MIC PM)")
        elif ber.strip() != "0/0":
            text = f"BER NZ reporting on the Transport port: {node}."
            warning_texts.append(text)
            pending_lines.append(f"{text} (MIC PM)")

    return warning_texts, pending_lines


def sa_conversion_amf_warning(post_text, sa_conversion_nodes_list):
    """Confirmed check: for nodes with SA Conversion present, verify each
    TermPointToAmf entry's admin state. Confirmed real format:
    '{Node} GNBCUCPFunction=1,TermPointToAmf={amf_name} {UNLOCKED|LOCKED} {OpState}'.
    If ANY TermPointToAmf entry for the node is LOCKED (not all — any single one is
    enough), fires the warning. Confirmed exact wording:
    'TermpointtoAmf is in locked state on {node}, please unlock.'
    Returns list of warning texts (one per affected node, deduplicated)."""
    if not sa_conversion_nodes_list or not post_text:
        return []
    pattern = re.compile(r'(\S+) GNBCUCPFunction=1,TermPointToAmf=(\S+) (UNLOCKED|LOCKED) (\w+)')
    locked_nodes = set()
    for m in pattern.finditer(post_text):
        node, _amf_name, admin, _oper = m.groups()
        if node in sa_conversion_nodes_list and admin == "LOCKED":
            locked_nodes.add(node)
    return [f"TermpointtoAmf is in locked state on {node}, please unlock." for node in sorted(locked_nodes)]


# ============================================================
# LTE/5G SECTOR PARAMETER VERIFICATION — confirmed CIQ column mapping from the design
# conversation, cross-checked against real WAL94133 data. Compares Post-checks' actual
# on-air values against the CIQ's intended target values, per cell.
# ============================================================

_LTE_CELL_ROW_RE = re.compile(
    r'(\S+) (LOCKED|UNLOCKED) (?:\d+ \S+ )?(BARRED|UNBARRED) (\d+) (\d+) (\d+) '
    r'(ENABLED|DISABLED) (\d+) (true|false) (\S+) (\d+) (\d+)')


@st.cache_data
def extract_lte_cell_status(post_text):
    """Parses the real 'LTE FDD Cell Status Information' table. Confirmed real header:
    'Cells adminState availabilityStatus cellBarred dlChannelBandwidth earfcndl earfcnul
    OpState PCI PLMNStatus sectorCarrierRef tac ulChannelBandwidth' — availabilityStatus
    is a two-token value (e.g. '3 OFF_LINE'), confirmed by column-count cross-check
    against real data. Returns {cell: {field: value}}.
    Confirmed fix: availabilityStatus made non-capturing/optional since some sites
    genuinely omit it entirely (row jumps straight from adminState to cellBarred with
    no availabilityStatus value at all) — matching it as required caused the ENTIRE row
    to silently fail to match, making a genuinely-present cell look "missing" from
    Post-checks in the cell-presence warnings."""
    out = {}
    for m in _LTE_CELL_ROW_RE.finditer(post_text or ""):
        (cell, _admin, _barred, dlbw, earfcndl, earfcnul,
         _opstate, pci, _plmn, sector, tac, ulbw) = m.groups()
        out[cell] = {
            "dlChannelBandwidth": dlbw, "earfcndl": earfcndl, "earfcnul": earfcnul,
            "PCI": pci, "sectorCarrierRef": sector, "tac": tac, "ulChannelBandwidth": ulbw,
        }
    return out


def lte_sector_param_warnings(ciq_wb, mm_objs, post_text, eutran_rows=None):
    """Confirmed CIQ mapping (cross-checked against real data, corrected from the
    original ask: PCI compares against CIQ's own PCI column, not cellId):
    dlChannelBandwidth->dlChannelBandwidth, earfcndl->earfcnDl, earfcnul->earfcnUl,
    PCI->PCI, sectorCarrierRef->sectorId, ulChannelBandwidth->ulChannelBandwidth (all in
    'eUtran Parameters'); tac->tac (in 'eNB Info', matched via eNBId, same value for
    every cell under that eNB). Returns list of warning texts, one per mismatched field.
    Confirmed perf fix: eutran_rows can be passed in pre-computed (once per render) to
    avoid re-parsing the same sheet repeatedly across this and other warning checks —
    falls back to computing it here if not provided, for backward compatibility."""
    warnings = []
    if not post_text or "eUtran Parameters" not in ciq_wb.sheetnames:
        return warnings
    post_cells = extract_lte_cell_status(post_text)
    if not post_cells:
        return warnings

    if eutran_rows is None:
        eutran_rows = qx.sheet_objs(ciq_wb["eUtran Parameters"])
    ciq_rows = {r.get("EutranCellFDDId"): r for r in eutran_rows
                if r.get("EutranCellFDDId")}
    field_map = [
        ("dlChannelBandwidth", "dlChannelBandwidth"), ("earfcndl", "earfcnDl"),
        ("earfcnul", "earfcnUl"), ("PCI", "PCI"), ("sectorCarrierRef", "sectorId"),
        ("ulChannelBandwidth", "ulChannelBandwidth"),
    ]
    for cell, post_vals in post_cells.items():
        ciq_row = ciq_rows.get(cell)
        if not ciq_row:
            continue
        for post_key, ciq_key in field_map:
            post_val = str(post_vals.get(post_key, "")).strip()
            ciq_val = str(ciq_row.get(ciq_key, "")).strip()
            if ciq_val and post_val:
                ciq_parts = [p.strip() for p in ciq_val.split("/")]
                if post_val not in ciq_parts:
                    warnings.append(f"{post_key} mismatch on {cell}: Post-checks={post_val}, CIQ={ciq_val}.")

    # tac — matched via eNBId, same value expected for every cell under that eNB.
    if "eNB Info" in ciq_wb.sheetnames:
        enb_tac = {r.get("eNBId"): r.get("tac") for r in qx.sheet_objs(ciq_wb["eNB Info"]) if r.get("eNBId")}
        cell_to_enbid = {r.get("EutranCellFDDId"): r.get("eNBId") for r in eutran_rows if r.get("EutranCellFDDId")}
        checked_enbids = set()
        for cell, post_vals in post_cells.items():
            enbid = cell_to_enbid.get(cell)
            if not enbid or enbid in checked_enbids:
                continue
            checked_enbids.add(enbid)
            ciq_tac = str(enb_tac.get(enbid, "")).strip()
            post_tac = str(post_vals.get("tac", "")).strip()
            if ciq_tac and post_tac and post_tac != ciq_tac:
                warnings.append(f"tac mismatch on eNBId {enbid}: Post-checks={post_tac}, CIQ={ciq_tac}.")
    return warnings


@st.cache_data
def extract_5g_cell_du_status(post_text):
    """Parses '5G NR Cell DU Status' table. Confirmed real complication: nCI is
    sometimes empty (variable-length row), so pure positional splitting is unreliable —
    uses the fact that nRSectorCarrierRef always equals the cell name itself as an
    anchor, confirmed against real data (both cellLocalId and nRPCI verified to match
    CIQ exactly with this approach). Returns {cell: {field: value}}."""
    out = {}
    for line in (post_text or "").splitlines():
        tokens = line.split()
        if len(tokens) < 10 or tokens[1] not in ("LOCKED", "UNLOCKED"):
            continue
        cell = tokens[0]
        try:
            anchor_idx = tokens.index(cell, 1)
        except ValueError:
            continue
        before = tokens[1:anchor_idx]
        after = tokens[anchor_idx + 1:]
        if len(before) < 7 or len(after) < 3:
            continue
        local_id, cell_range = before[2], before[3]
        nrpci = before[-1]
        nrtac = after[0]
        out[cell] = {"cellLocalId": local_id, "cellRange": cell_range, "nRPCI": nrpci, "nRTAC": nrtac}
    return out


@st.cache_data
def extract_5g_sector_carrier(post_text):
    """Parses '5G NR Sector Carrier' table. Confirmed real header:
    'nrSectorCarrier adminState arfcnDL arfcnUL bSChannelBwDL bSChannelBwUL
    configuredMaxTxPower opState txDirection' — fixed structure, no variable-length
    fields. Returns {cell: {field: value}}."""
    out = {}
    pattern = re.compile(
        r'(\S+) (LOCKED|UNLOCKED) (\d+) (\d+) (\d+) (\d+) (\d+) (ENABLED|DISABLED) (\S+)')
    for m in pattern.finditer(post_text or ""):
        cell, _admin, arfcndl, arfcnul, bwdl, bwul, maxtxpower, _opstate, _txdir = m.groups()
        out[cell] = {
            "arfcnDL": arfcndl, "arfcnUL": arfcnul, "bSChannelBwDL": bwdl,
            "bSChannelBwUL": bwul, "configuredMaxTxPower": maxtxpower,
        }
    return out


@st.cache_data
def extract_ssb_frequency(post_text):
    """Parses the real 'NRCellDU={cell} ssbFrequency {value}' lines. Returns
    {cell: ssbFrequency}."""
    out = {}
    for m in re.finditer(r'NRCellDU=(\S+) ssbFrequency (\d+)', post_text or ""):
        cell, val = m.groups()
        out[cell] = val
    return out


@st.cache_data
def extract_5g_cell_cu_status(post_text):
    """Parses '5G NR Cell CU Status' table — confirmed genuinely separate real table
    from '5G NR Cell DU Status', with its own cellLocalId that should independently
    match the CIQ (both CU and DU are checked, since either could diverge on its own).
    Confirmed real header: 'MO cellLocalId cellState nCI serviceState' — cellState and
    serviceState are confirmed blank in the real data, so bounded extraction between the
    section's own header and the next section header ('5G NR Sector Carrier') is used
    for safety, rather than a generic file-wide regex. Returns {cell: cellLocalId}."""
    out = {}
    section_match = re.search(
        r'5G NR Cell CU Status\nMO cellLocalId cellState nCI serviceState\n(.*?)\n5G NR Sector Carrier',
        post_text or "", re.DOTALL)
    if not section_match:
        return out
    for m in re.finditer(r'(\S+) (\d+) (\d+)', section_match.group(1)):
        cell, local_id, _nci = m.groups()
        out[cell] = local_id
    return out


def fiveg_sector_param_warnings(ciq_wb, mm_objs, post_text, fiveg_rows=None):
    """Confirmed CIQ mapping: cellLocalId, CellRange, nRPCI, arfcnDL, arfcnUL,
    bSChannelBwDL, bSChannelBwUL, configuredMaxTxPower, ssbFrequency all compared
    directly against '5G Info' (matched by NRCellDU); nrTAC compared against 'NR_SA'
    (matched by node name, same value expected for every cell on that node — same
    per-node pattern as LTE's tac/eNBId check).
    Confirmed perf fix: fiveg_rows can be passed in pre-computed (once per render) to
    avoid re-parsing the same sheet repeatedly across this and other warning checks —
    falls back to computing it here if not provided, for backward compatibility."""
    warnings = []
    if not post_text or "5G Info" not in ciq_wb.sheetnames:
        return warnings

    cell_du = extract_5g_cell_du_status(post_text)
    cell_cu = extract_5g_cell_cu_status(post_text)
    sector_carrier = extract_5g_sector_carrier(post_text)
    ssb_freq = extract_ssb_frequency(post_text)
    if not cell_du:
        return warnings

    if fiveg_rows is None:
        fiveg_rows = qx.sheet_objs(ciq_wb["5G Info"])
    ciq_rows = {r.get("NRCellDU"): r for r in fiveg_rows if r.get("NRCellDU")}

    # cellLocalId — confirmed checked independently from BOTH the CU Status and DU
    # Status tables, since either could diverge from the CIQ on its own.
    for cell, cu_local_id in cell_cu.items():
        ciq_row = ciq_rows.get(cell)
        if not ciq_row:
            continue
        ciq_val = str(ciq_row.get("cellLocalId", "")).strip()
        if ciq_val and cu_local_id.strip() not in [p.strip() for p in ciq_val.split("/")]:
            warnings.append(f"cellLocalId mismatch on {cell} (5G NR Cell CU Status): Post-checks={cu_local_id}, CIQ={ciq_val}.")

    for cell, post_vals in cell_du.items():
        ciq_row = ciq_rows.get(cell)
        if not ciq_row:
            continue
        combined = dict(post_vals)
        combined.update(sector_carrier.get(cell, {}))
        if cell in ssb_freq:
            combined["ssbFrequency"] = ssb_freq[cell]

        field_map = [
            ("cellLocalId", "cellLocalId"), ("cellRange", "CellRange"), ("nRPCI", "nRPCI"),
            ("arfcnDL", "arfcnDL"), ("arfcnUL", "arfcnUL"), ("bSChannelBwDL", "bSChannelBwDL"),
            ("bSChannelBwUL", "bSChannelBwUL"), ("configuredMaxTxPower", "configuredMaxTxPower"),
            ("ssbFrequency", "ssbFrequency"),
        ]
        for post_key, ciq_key in field_map:
            post_val = str(combined.get(post_key, "")).strip()
            ciq_val = str(ciq_row.get(ciq_key, "")).strip()
            if not ciq_val or not post_val:
                continue
            ciq_parts = [p.strip() for p in ciq_val.split("/")]
            if post_val not in ciq_parts:
                source_label = " (5G NR Cell DU Status)" if post_key == "cellLocalId" else ""
                warnings.append(f"{post_key} mismatch on {cell}{source_label}: Post-checks={post_val}, CIQ={ciq_val}.")

    # nrTAC — matched via node name in NR_SA, same value expected for every cell on that node.
    if "NR_SA" in ciq_wb.sheetnames:
        nrsa_tac = {r.get("Node Name"): r.get("nrTAC") for r in qx.sheet_objs(ciq_wb["NR_SA"]) if r.get("Node Name")}
        checked_nodes = set()
        for cell, post_vals in cell_du.items():
            node = cell.rsplit("_N", 1)[0] if "_N" in cell else None
            if not node or node in checked_nodes:
                continue
            checked_nodes.add(node)
            ciq_tac = str(nrsa_tac.get(node, "")).strip()
            post_tac = str(post_vals.get("nRTAC", "")).strip()
            if ciq_tac and post_tac and post_tac != ciq_tac:
                warnings.append(f"nRTAC mismatch on node {node}: Post-checks={post_tac}, CIQ={ciq_tac}.")
    return warnings


def lte_cell_presence_warnings(ciq_wb, post_text, eutran_rows=None):
    """New check, distinct from lte_sector_param_warnings: that function only compares
    FIELD VALUES for cells present in BOTH the CIQ and the Health Check (Post-checks),
    silently skipping any cell missing from either side. This check instead flags
    presence mismatches themselves — both directions: a cell in the CIQ with no
    matching row in Post-checks, and a cell in Post-checks with no matching row in the
    CIQ. Confirmed scope: checks EVERY cell in the CIQ's 'eUtran Parameters' sheet,
    not just newly-added ones — a genuinely unexpected cell (extra or missing) is worth
    flagging regardless of whether it's part of this specific integration.
    Returns list of warning texts."""
    warnings = []
    if not post_text or "eUtran Parameters" not in ciq_wb.sheetnames:
        return warnings
    post_cells = extract_lte_cell_status(post_text)
    if eutran_rows is None:
        eutran_rows = qx.sheet_objs(ciq_wb["eUtran Parameters"])
    ciq_cell_ids = {r.get("EutranCellFDDId") for r in eutran_rows if r.get("EutranCellFDDId")}
    post_cell_ids = set(post_cells.keys())

    for cell in sorted(ciq_cell_ids - post_cell_ids):
        warnings.append(f"Cell {cell} present in CIQ but missing from Post-checks.")
    for cell in sorted(post_cell_ids - ciq_cell_ids):
        warnings.append(f"Cell {cell} present in Post-checks but missing from CIQ.")
    return warnings


def fiveg_cell_presence_warnings(ciq_wb, post_text, fiveg_rows=None):
    """Same principle as lte_cell_presence_warnings, for 5G: checks EVERY cell in the
    CIQ's '5G Info' sheet against every cell found in Post-checks (via
    extract_5g_cell_du_status), flagging presence mismatches in both directions.
    Returns list of warning texts."""
    warnings = []
    if not post_text or "5G Info" not in ciq_wb.sheetnames:
        return warnings
    post_cells = extract_5g_cell_du_status(post_text)
    if fiveg_rows is None:
        fiveg_rows = qx.sheet_objs(ciq_wb["5G Info"])
    ciq_cell_ids = {r.get("NRCellDU") for r in fiveg_rows if r.get("NRCellDU")}
    post_cell_ids = set(post_cells.keys())

    for cell in sorted(ciq_cell_ids - post_cell_ids):
        warnings.append(f"Cell {cell} present in CIQ but missing from Post-checks.")
    for cell in sorted(post_cell_ids - ciq_cell_ids):
        warnings.append(f"Cell {cell} present in Post-checks but missing from CIQ.")
    return warnings


def sctp_status_warnings(post_text):
    """Confirmed check: every SCTP endpoint should be ENABLED. Confirmed real format:
    '{Node} Transport=1,SctpEndpoint={endpoint} {ENABLED|DISABLED}'. Fires one warning
    per disabled endpoint. Confirmed exact wording:
    'Transport=1,SctpEndpoint={endpoint} SCTP is disabled.'"""
    warnings = []
    if not post_text:
        return warnings
    pattern = re.compile(r'(\S+) Transport=1,SctpEndpoint=(\S+) (ENABLED|DISABLED)')
    for m in pattern.finditer(post_text):
        _node, endpoint, state = m.groups()
        if state == "DISABLED":
            warnings.append(f"Transport=1,SctpEndpoint={endpoint} SCTP is disabled.")
    return warnings


def digital_tilt_warnings(ciq_wb, mm_objs, post_text, classification, fiveg_rows=None):
    """Confirmed check: DigitalTilt in Post-checks (the 'usedDigitalTilt' value — the
    'digitalTilt' field itself is confirmed blank/not printed in the real data) must
    match CIQ's 'Electrical Tilt' column, for CBAND/DOD sectors only. Confirmed real
    format: 'NRSectorCarrier={sector},CommonBeamforming={n} {usedDigitalTilt}'. Matched
    to CIQ via the NRSectorCarrier column. Only checks sectors whose band is CBAND or
    DOD/DOD_BWE (confirmed via classification['added'] + band_label).
    Confirmed perf fix: fiveg_rows can be passed in pre-computed (once per render,
    shared with fiveg_sector_param_warnings) to avoid re-parsing the same '5G Info'
    sheet a third time — falls back to computing it here if not provided."""
    warnings = []
    if not post_text or "5G Info" not in ciq_wb.sheetnames:
        return warnings

    # Confirmed real Cband/DOD sectors on this site, from classification (already band-labeled).
    cband_dod_cells = set()
    for cells in classification.get("added", {}).values():
        for c in cells:
            label, _sector = qx.band_label(c)
            if label in ("CBAND", "DOD", "DOD_BWE"):
                cband_dod_cells.add(c)
    if not cband_dod_cells:
        return warnings

    post_tilt = {}
    for m in re.finditer(r'NRSectorCarrier=(\S+?),CommonBeamforming=\d+ (\d+)', post_text):
        sector, used_tilt = m.groups()
        post_tilt[sector] = used_tilt

    if fiveg_rows is None:
        fiveg_rows = qx.sheet_objs(ciq_wb["5G Info"])
    ciq_rows = {r.get("NRSectorCarrier"): r for r in fiveg_rows if r.get("NRSectorCarrier")}

    for sector in cband_dod_cells:
        if sector not in post_tilt or sector not in ciq_rows:
            continue
        post_val = post_tilt[sector].strip()
        ciq_val = str(ciq_rows[sector].get("Electrical Tilt", "")).strip()
        if ciq_val and post_val != ciq_val:
            warnings.append(f"DigitalTilt mismatch on {sector}: Post-checks={post_val}, CIQ={ciq_val}.")
    return warnings


def sort_bands_lte_first(bands):
    """Confirmed ordering: LTE bands first (alphabetically among themselves), then
    5G/CBAND/DOD bands (alphabetically among themselves) — NOT a single pure alphabetical
    sort, which mixes them together since '5G_850' alphabetically precedes 'AWS_1'.
    Confirmed real example: 'AWS_1/AWS_2/FNET/LTE_700/LTE_700_E/PCS_1/5G_850/CBAND/DOD',
    not '5G_850/AWS_1/AWS_2/CBAND/DOD/FNET/LTE_700/LTE_700_E/PCS_1'."""
    lte = sorted(b for b in bands if not (b.startswith("5G_") or b in ("CBAND", "DOD", "DOD_BWE")))
    fiveg = sorted(b for b in bands if b.startswith("5G_") or b in ("CBAND", "DOD", "DOD_BWE"))
    return lte + fiveg


def scripted_locked_bands(ciq_wb):
    """Confirmed check: any 5G cell with configuredMaxTxPower=5 in the CIQ means that
    band is scripted and locked. Returns the raw set of band labels (deduplicated),
    or an empty set if none match. Used both for the Notes line itself and for the
    DSS-splitting rule (bands that are scripted/locked go straight to Pending/AT&T)."""
    if "5G Info" not in ciq_wb.sheetnames:
        return set()
    bands = set()
    for row in qx.sheet_objs(ciq_wb["5G Info"]):
        if str(row.get("configuredMaxTxPower", "")).strip() == "5":
            cell = row.get("NRCellDU")
            if cell:
                label, _sector = qx.band_label(cell)
                if label:
                    bands.add(label)
    return bands


def scripted_locked_bands_note(ciq_wb):
    """Confirmed exact format: '{Band1} & {Band2} is scripted and Locked.' — band names
    (deduplicated), not cell names, joined with '&'. Returns the note text, or None if
    no cells match."""
    bands = scripted_locked_bands(ciq_wb)
    if not bands:
        return None
    sorted_bands = sort_bands_lte_first(bands)
    return f"{' & '.join(sorted_bands)} is scripted and Locked."


def split_dss_bands_by_scripted_locked(dss_bands, scripted_locked_bands, market):
    """Confirmed rule, applies across MCA/N2E/NSB: DSS bands that are ALSO scripted and
    locked (configuredMaxTxPower=5) go DIRECTLY to Pending with stakeholder AT&T,
    bypassing the normal Completed/Pending user choice. Any DSS bands that DON'T overlap
    with the scripted/locked set still go through the normal user choice.
    Confirmed NTX rule: DSS is excluded from the report ENTIRELY in NTX market —
    regardless of scripted/locked status, nothing gets reported at all (not just the
    overlapping portion).
    Confirmed real bug fix: each entry in dss_bands is actually a 5G|LTE PAIR (e.g.
    '5G_PCS_1|PCS_1', confirmed via app.py's generate_dss docstring and
    dss_activation_labels.append(f"{nr_label}|{lte_label}")), not a single band — so the
    overlap check must look at whether EITHER band inside the pair is scripted/locked,
    not compare the whole pair-string directly (which could never match a set of
    individual band names). The pair itself stays intact in the returned sets, since DSS
    is always reported as a pair, never split apart.
    Returns (auto_pending_bands: set, user_choice_bands: set) — each still containing
    whole pair-strings like '5G_PCS_1|PCS_1'."""
    if market == "NTX":
        return set(), set()
    overlap, remainder = set(), set()
    for pair in dss_bands:
        constituent_bands = set(pair.split("|"))
        if constituent_bands & set(scripted_locked_bands):
            overlap.add(pair)
        else:
            remainder.add(pair)
    return overlap, remainder
