import io
import openpyxl
from mca_row_map import ROW_MAP


def fill_legacy_mca(template_path, values, report_text=None):
    """values: dict of row_key -> either a plain value (single-column rows) or a list of
    column values [colB, colC, colD, ...] (multi-column rows), or for multi-slot items,
    a list of such value-lists (one per used slot).
    checked: dict of row_number -> True/False, explicit checkbox states to set.
    Always preserves the VBA project (keep_vba=True) so the existing macro (Report_MCA
    generation + DeleteMatchingRows cleanup) still works when the user opens the file.

    report_text: confirmed fix — the raw template's Report_MCA sheet only resolves
    cleanly (no "to_be_deleted" placeholders) after the engineer manually runs the
    legacy_mca_report macro in Excel; until then, Excel's own formula recalculation on
    open still shows "to_be_deleted" for every unchecked row (checked correctly, but not
    yet macro-cleaned). Rather than reverse-engineering ~190 rows of the VBA's relative-
    reference formulas to replicate DeleteMatchingRows exactly, this writes the SAME
    already-clean text used for the "Download report (.txt)" output directly into
    Report_MCA as plain static values — guaranteed consistent with the .txt report since
    both come from the identical source, and no manual macro step needed at all.
    Returns the filled workbook as raw bytes (in-memory) — no filesystem output path at all,
    since a hardcoded sandbox-only path previously caused a misleading 'template not found'
    error on deployment (the actual failure was writing the OUTPUT, not reading the template)."""
    wb = openpyxl.load_workbook(template_path, keep_vba=True)
    ws = wb["Legacy_MCA"]

    def set_row(row_num, checked, col_value_pairs=None):
        ws.cell(row=row_num, column=1, value=checked)
        if col_value_pairs:
            for col, v in col_value_pairs:
                if v is not None:
                    ws.cell(row=row_num, column=col, value=v)

    for row_num, checked, col_value_pairs in values.get("row_writes", []):
        set_row(row_num, checked, col_value_pairs)

    if report_text is not None and "Report_MCA" in wb.sheetnames:
        report_ws = wb["Report_MCA"]
        # Clear every existing formula/value first so no stale "to_be_deleted" formula
        # rows survive alongside the new static content.
        report_ws.delete_rows(1, report_ws.max_row + 5)
        for i, line in enumerate(report_text.split("\n"), start=1):
            report_ws.cell(row=i, column=1, value=line)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
